from __future__ import annotations

import hashlib
import importlib.util
import json
import re
import subprocess
import sys
import tempfile
import unittest
import xml.etree.ElementTree as ET
from pathlib import Path
from urllib.parse import unquote
from zipfile import ZipFile

import openpyxl


REPO_ROOT = Path(__file__).resolve().parents[1]
HAN_CHARACTER = re.compile(r"[\u4e00-\u9fff]")
SPREADSHEET_NS = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
STRUCTURED_REFERENCE = re.compile(
    r"(?<![A-Za-z0-9_.])([A-Za-z_][A-Za-z0-9_.]*)\[([^\]]+)\]"
)

SCHEMA_SHA256 = {
    "skills/generate/contracts/common.schema.json": "c76aaebc6b683a98a30ba09c7ee5f8d18bf328648d9ca0fd3c59b4faf4750a57",
    "skills/generate/contracts/current.schema.json": "1c5fd52b9c1e3094ffaa95c01fa2ab2867c4306be551a1ff1f89172bb714c51a",
    "skills/generate/contracts/delivery-bundle.schema.json": "ef50c386ea95f7513cb423d832ddc243015dc461c46b22525d22db203c6753c7",
    "skills/generate/contracts/delivery-slice.schema.json": "b9c0cad211f01668af67cf92b8664932eec2441d16ec1d07e597151e9e817993",
    "skills/generate/contracts/final-review.schema.json": "5aa7d403911b3e0a43a19a8537773c0e2823db43713e7e36fea805684a34b4c5",
    "skills/generate/contracts/generation-manifest.schema.json": "8bd9152faf3ba224363fcb4e971a6399df32d0609e80d6ed83e39e3da54a8074",
    "skills/generate/contracts/id-decisions.schema.json": "dd8a0cffae6dc3017ba7c8b0845d496a598866717a43607fef536c5b5f5347a6",
    "skills/generate/contracts/input-manifest.schema.json": "eec15f30713865a42f13e086ac2db2f7c30fddaa2831f8a88223fa623a2b98bb",
    "skills/generate/contracts/question.schema.json": "ee4409e67a599f7e7680b48b502ad1aa9acdbc8396347d8425fdb9691a3d40e2",
    "skills/generate/contracts/request.schema.json": "ac292683facff487ce7df58c500eb54243f0d6ff132a6881e3ffeef4c46b2e74",
    "skills/generate/contracts/run-plan.schema.json": "be60a53300af421d6747f3cb5bf982940eaca897d210dcb9c1a7464691a8caec",
    "skills/generate/contracts/scope-bundle.schema.json": "147d9d911aa86e45ce075c1d9cc5d9b4954e5cb6d28879e782f2e47985d2f740",
    "skills/generate/contracts/scope-slice.schema.json": "eaf23cb307cb18bfde2282e7bb40330f5f91d2a4edee33c1ea8add3389ca6ccb",
}

TEMPLATE_SHA256 = "51f88c98a6f68fb2b95b58c28b95a7d68897df38d685532ef89a5de19727bac9"

CURRENT_USER_DOCS = (
    "README.md",
    "docs/architecture/ai-plugin-marketplace-design.md",
    "plugins/ai-sow/README.md",
    "plugins/ai-sow/docs/AI_SOW_PLUGIN_DESIGN.md",
    "plugins/ai-sow/docs/CONTEXT.md",
)

TASK_STANDARD_DOCS = (
    "README.md",
    "CHANGELOG.md",
    "plugins/ai-sow/README.md",
    "plugins/ai-sow/docs/AI_SOW_PLUGIN_DESIGN.md",
    "plugins/ai-sow/docs/CONTEXT.md",
    "plugins/ai-sow/docs/PRD_HLD_AUTOMATED_SOW_WORKFLOW_PLAN.md",
    "plugins/ai-sow/docs/reference/SOW任务分类与开发交付人天标准_v1.3.md",
)

DELIVERY_REFERENCES = (
    "acceptance-criteria.md",
    "delivery-authoring.md",
    "delivery-compilation.md",
    "delivery-decomposition.md",
    "delivery-examples.md",
    "delivery-work-classification.md",
    "effective-start-matching.md",
    "epic-authoring.md",
    "feature-authoring.md",
    "question-authoring.md",
    "story-authoring.md",
    "task-authoring.md",
    "technical-work-classification.md",
)


def enum_arrays(value: object, path: str = "$") -> dict[str, list[object]]:
    if isinstance(value, dict):
        result = {path: value["enum"]} if "enum" in value else {}
        for key, child in value.items():
            result.update(enum_arrays(child, f"{path}.{key}"))
        return result
    if isinstance(value, list):
        result: dict[str, list[object]] = {}
        for index, child in enumerate(value):
            result.update(enum_arrays(child, f"{path}[{index}]"))
        return result
    return {}


class RepositoryLayoutTests(unittest.TestCase):
    def test_user_docs_describe_one_automatic_generate_flow(self) -> None:
        for relative in CURRENT_USER_DOCS:
            text = (REPO_ROOT / relative).read_text(encoding="utf-8")
            with self.subTest(document=relative):
                self.assertIn("ai-sow:generate", text)
                self.assertNotIn("逐阶段批准", text)
                self.assertNotIn("Owner receipt", text)

    def test_manifest_prompts_only_advertise_generate(self) -> None:
        manifest = json.loads(
            (REPO_ROOT / "plugins/ai-sow/.codex-plugin/plugin.json").read_text(
                encoding="utf-8"
            )
        )
        self.assertTrue(
            manifest["interface"]["longDescription"].startswith(
                "一次提供 PRD、HLD"
            )
        )
        prompts = manifest["interface"]["defaultPrompt"]
        self.assertEqual(len(prompts), 3)
        self.assertTrue(all("ai-sow:generate" in prompt for prompt in prompts))
        self.assertTrue(all("下一阶段" not in prompt for prompt in prompts))

    def test_target_workflow_plan_is_marked_implemented(self) -> None:
        text = (
            REPO_ROOT
            / "plugins/ai-sow/docs/PRD_HLD_AUTOMATED_SOW_WORKFLOW_PLAN.md"
        ).read_text(encoding="utf-8")
        self.assertIn("状态：已实现", text)
        self.assertIn("执行日期：2026-09-02", text)
        self.assertIn("d6738ee25cace4eb97db1cd204f769c6c63b7128", text)

    def test_runtime_has_no_unsupported_document_parser_dependencies(self) -> None:
        plugin_root = REPO_ROOT / "plugins/ai-sow"
        dependency_text = "\n".join(
            (plugin_root / relative).read_text(encoding="utf-8")
            for relative in ("pyproject.toml", "uv.lock")
        ).lower()
        for dependency in (
            "pypdf",
            "pdfplumber",
            "python-docx",
            "python-pptx",
            "pymupdf",
        ):
            self.assertNotIn(dependency, dependency_text)

    def test_xlsx_formulas_only_reference_existing_table_columns(self) -> None:
        plugin_root = REPO_ROOT / "plugins/ai-sow"
        workbooks = [
            plugin_root / "skills/generate/assets/sow-template.xlsx",
            plugin_root / "docs/reference/SOW估算与生成示例_v1.3.xlsx",
        ]

        for workbook in workbooks:
            with self.subTest(workbook=workbook.name), ZipFile(workbook) as archive:
                tables: dict[str, set[str]] = {}
                for member in archive.namelist():
                    if not member.startswith("xl/tables/table") or not member.endswith(
                        ".xml"
                    ):
                        continue
                    table = ET.fromstring(archive.read(member))
                    tables[table.attrib["name"]] = {
                        column.attrib["name"]
                        for column in table.findall(
                            "x:tableColumns/x:tableColumn", SPREADSHEET_NS
                        )
                    }

                invalid_references: set[str] = set()
                for member in archive.namelist():
                    if not member.endswith(".xml") or not member.startswith(
                        ("xl/worksheets/", "xl/tables/")
                    ):
                        continue
                    root = ET.fromstring(archive.read(member))
                    formula_nodes = root.findall(".//x:f", SPREADSHEET_NS)
                    formula_nodes.extend(
                        root.findall(".//x:calculatedColumnFormula", SPREADSHEET_NS)
                    )
                    for node in formula_nodes:
                        for table_name, column_name in STRUCTURED_REFERENCE.findall(
                            node.text or ""
                        ):
                            if column_name not in tables.get(table_name, set()):
                                invalid_references.add(f"{table_name}[{column_name}]")

                self.assertEqual(set(), invalid_references, workbook)

    def test_ai_sow_package_is_self_contained(self) -> None:
        plugin_root = REPO_ROOT / "plugins/ai-sow"
        required = [
            ".codex-plugin/plugin.json",
            ".claude-plugin/plugin.json",
            "pyproject.toml",
            "uv.lock",
            "README.md",
            "runtime/diagnostics.py",
            "runtime/project_io.py",
            "skills/generate/SKILL.md",
            "skills/generate/scripts/orchestrator.py",
            "skills/generate/assets/sow-template.xlsx",
            "skills/generate/contracts/question.schema.json",
            "tests/support/smoke_plugin.py",
            "docs/reference/SOW任务分类与开发交付人天标准_v1.3.md",
            "docs/reference/SOW估算与生成示例_v1.3.xlsx",
        ]
        for relative in required:
            self.assertTrue((plugin_root / relative).is_file(), relative)
        for name in DELIVERY_REFERENCES:
            relative = f"skills/generate/references/{name}"
            self.assertTrue((plugin_root / relative).is_file(), relative)

        for document in plugin_root.rglob("*.md"):
            if ".venv" in document.parts:
                continue
            for raw_target in re.findall(
                r"!?(?:\[[^\]]*\])\(([^)]+)\)",
                document.read_text(encoding="utf-8"),
            ):
                target = raw_target.strip()
                if target.startswith("<") and ">" in target:
                    target = target[1 : target.index(">")]
                else:
                    target = target.split(maxsplit=1)[0]
                if target.startswith(("#", "http://", "https://", "mailto:")):
                    continue
                target = unquote(target.split("#", 1)[0])
                if not target:
                    continue
                resolved = (document.parent / target).resolve()
                with self.subTest(document=document, target=target):
                    self.assertTrue(resolved.is_relative_to(plugin_root.resolve()))
                    self.assertTrue(resolved.exists(), resolved)

    def test_manifest_identity_and_contract_version_match(self) -> None:
        plugin_root = REPO_ROOT / "plugins/ai-sow"
        release_version = "0.1.0-beta.1"
        runtime_version = "0.1.0b1"
        manifest = json.loads(
            (plugin_root / ".codex-plugin/plugin.json").read_text(encoding="utf-8")
        )
        package_schema = json.loads(
            (plugin_root / "skills/generate/contracts/generation-manifest.schema.json").read_text(encoding="utf-8")
        )
        request = json.loads(
            (plugin_root / "skills/generate/fixtures/greenfield/request.json").read_text(encoding="utf-8")
        )
        pyproject_text = (plugin_root / "pyproject.toml").read_text(encoding="utf-8")
        lock_text = (plugin_root / "uv.lock").read_text(encoding="utf-8")
        self.assertEqual(manifest["name"], "ai-sow")
        self.assertEqual(manifest["version"], release_version)
        self.assertEqual(package_schema["$id"], "urn:ai-sow:generate:generation-manifest:1")
        self.assertEqual(request["contract"], "ai-sow-generate-request-v1")
        self.assertRegex(
            pyproject_text,
            rf'(?ms)^\[project\].*?^version = "{re.escape(runtime_version)}"$',
        )
        self.assertRegex(
            lock_text,
            rf'(?ms)^\[\[package\]\]\nname = "ai-sow-plugin-runtime"\nversion = "{re.escape(runtime_version)}"$',
        )
        self.assertIn("SOW 标准 1.3", (plugin_root / "README.md").read_text(encoding="utf-8"))
        for relative in (
            "README.md",
            "CHANGELOG.md",
            "SECURITY.md",
            "docs/architecture/ai-plugin-marketplace-design.md",
            "plugins/ai-sow/README.md",
            "plugins/ai-sow/docs/AI_SOW_PLUGIN_DESIGN.md",
            "plugins/ai-sow/docs/reference/SOW任务分类与开发交付人天标准_v1.3.md",
        ):
            self.assertIn(
                release_version,
                (REPO_ROOT / relative).read_text(encoding="utf-8"),
                relative,
            )
    def test_user_install_docs_match_bootstrapped_runtime(self) -> None:
        expected = {
            "README.md": "无需预装 Git、Python",
            "CONTRIBUTING.md": "普通插件用户由 `ai-sow:generate` 的 bootstrap 自动准备隔离运行时",
            "docs/architecture/ai-plugin-marketplace-design.md": "普通插件用户无需预装 uv、Python",
            "plugins/ai-sow/README.md": "不要求 uv 位于 PATH",
            "plugins/ai-sow/docs/CONTEXT.md": "普通用户无需预装 Python/uv",
            "plugins/ai-sow/references/runtime-environment.md": "普通插件用户无需预装 Python、`uv`",
        }
        for relative, statement in expected.items():
            self.assertIn(
                statement,
                (REPO_ROOT / relative).read_text(encoding="utf-8"),
                relative,
            )

        runtime_contract = (
            REPO_ROOT / "plugins/ai-sow/references/runtime-environment.md"
        ).read_text(encoding="utf-8")
        self.assertIn("<plugin-root>/.venv/bin/python", runtime_contract)
        self.assertIn("<plugin-root>/.venv/Scripts/python.exe", runtime_contract)

    def test_task_estimation_contract_has_no_removed_shape_or_modes(self) -> None:
        plugin_root = REPO_ROOT / "plugins/ai-sow"
        delivery = json.loads(
            (plugin_root / "skills/generate/contracts/delivery-bundle.schema.json").read_text(encoding="utf-8")
        )
        task_properties = delivery["$defs"]["task"]["properties"]
        for field in (
            "professionalDomain",
            "activity",
            "quantity",
            "taskFamily",
            "baseEffort",
            "complexityFactor",
            "personDays",
        ):
            self.assertNotIn(field, task_properties)
        self.assertNotIn("sitEstimates", delivery["properties"])
        self.assertEqual(
            task_properties["workMode"]["enum"],
            ["新建", "调整", "接入复用"],
        )

    def test_release_assets_and_smoke_implementation_use_plugin_scope(self) -> None:
        plugin_root = REPO_ROOT / "plugins/ai-sow"
        self.assertFalse((REPO_ROOT / "scripts" / "smoke_plugin.py").exists())
        self.assertTrue((plugin_root / "tests/support/smoke_plugin.py").is_file())

    def test_manifest_user_content_is_chinese(self) -> None:
        manifest = json.loads(
            (REPO_ROOT / "plugins/ai-sow/.codex-plugin/plugin.json").read_text(
                encoding="utf-8"
            )
        )
        visible_values = [
            manifest["description"],
            manifest["interface"]["shortDescription"],
            manifest["interface"]["longDescription"],
            *manifest["interface"]["defaultPrompt"],
        ]
        claude_manifest = json.loads(
            (REPO_ROOT / "plugins/ai-sow/.claude-plugin/plugin.json").read_text(
                encoding="utf-8"
            )
        )
        claude_marketplace = json.loads(
            (REPO_ROOT / ".claude-plugin/marketplace.json").read_text(encoding="utf-8")
        )
        visible_values.append(claude_manifest["description"])
        visible_values.extend(
            entry["description"] for entry in claude_marketplace["plugins"]
        )
        for value in visible_values:
            self.assertRegex(value, HAN_CHARACTER)

    def test_all_skills_resolve_shared_output_language_reference(self) -> None:
        plugin_root = REPO_ROOT / "plugins/ai-sow"
        shared_reference = plugin_root / "references/output-language.md"
        self.assertTrue(shared_reference.is_file(), shared_reference)

        skill_paths = sorted((plugin_root / "skills").glob("*/SKILL.md"))
        self.assertEqual(
            {path.parent.name for path in skill_paths},
            {"generate"},
        )
        for skill_path in skill_paths:
            declared_references = [
                Path(target)
                for target in re.findall(
                    r"\[[^\]]+\]\(([^)]+)\)",
                    skill_path.read_text(encoding="utf-8"),
                )
                if target.endswith("output-language.md")
            ]
            self.assertEqual(len(declared_references), 1, skill_path)
            resolved = (skill_path.parent / declared_references[0]).resolve()
            self.assertTrue(resolved.is_relative_to(plugin_root.resolve()), resolved)
            self.assertEqual(resolved, shared_reference.resolve())

    def test_open_source_release_files_exist(self) -> None:
        required = [
            "LICENSE",
            "NOTICE",
            "README.md",
            "CHANGELOG.md",
            "CONTRIBUTING.md",
            "CODE_OF_CONDUCT.md",
            "SECURITY.md",
            "plugins/ai-sow/LICENSE",
            "plugins/ai-sow/NOTICE",
        ]
        for relative in required:
            self.assertTrue((REPO_ROOT / relative).is_file(), relative)

    def test_repository_relative_markdown_links_resolve(self) -> None:
        completed = subprocess.run(
            ["git", "ls-files", "*.md"],
            cwd=REPO_ROOT,
            check=True,
            capture_output=True,
            text=True, encoding="utf-8",
        )
        for relative in completed.stdout.splitlines():
            document = REPO_ROOT / relative
            if not document.is_file():
                continue
            for raw_target in re.findall(
                r"!?(?:\[[^\]]*\])\(([^)]+)\)",
                document.read_text(encoding="utf-8"),
            ):
                target = raw_target.strip()
                if target.startswith("<") and ">" in target:
                    target = target[1 : target.index(">")]
                else:
                    target = target.split(maxsplit=1)[0]
                if target.startswith(("#", "http://", "https://", "mailto:")):
                    continue
                target = unquote(target.split("#", 1)[0])
                if not target:
                    continue
                resolved = (document.parent / target).resolve()
                with self.subTest(document=relative, target=target):
                    self.assertTrue(resolved.is_relative_to(REPO_ROOT.resolve()))
                    self.assertTrue(resolved.exists(), resolved)

    def test_public_readme_documents_plugin_lifecycle(self) -> None:
        text = (REPO_ROOT / "README.md").read_text(encoding="utf-8")
        self.assertNotIn("<repository-url>", text)
        for command in (
            "codex plugin marketplace add InspireChina/ai-plugin-marketplace",
            "codex plugin add ai-sow@ai-plugin-marketplace",
            "codex plugin marketplace upgrade ai-plugin-marketplace",
            "codex plugin remove ai-sow@ai-plugin-marketplace",
            "codex plugin marketplace remove ai-plugin-marketplace",
        ):
            self.assertIn(command, text)

    def test_repository_pins_line_endings_for_windows_checkouts(self) -> None:
        """Git for Windows defaults to core.autocrlf=true; a CRLF checkout breaks
        bootstrap.sh and changes the schema bytes the SHA-256 assertions pin."""
        attributes = REPO_ROOT / ".gitattributes"
        self.assertTrue(attributes.is_file(), attributes)
        self.assertIn("* text=auto eol=lf", attributes.read_text(encoding="utf-8"))

        probes = (
            "plugins/ai-sow/skills/generate/scripts/bootstrap.sh",
            "plugins/ai-sow/skills/generate/contracts/request.schema.json",
            "plugins/ai-sow/skills/generate/assets/sow-template.xlsx",
        )
        completed = subprocess.run(
            ["git", "check-attr", "eol", "binary", "--", *probes],
            cwd=REPO_ROOT,
            check=True,
            capture_output=True,
            text=True,
            encoding="utf-8",
        )
        resolved = completed.stdout
        for relative in probes[:2]:
            self.assertIn(f"{relative}: eol: lf", resolved)
        self.assertIn(f"{probes[2]}: binary: set", resolved)

        for relative in probes[:2]:
            self.assertNotIn(b"\r\n", (REPO_ROOT / relative).read_bytes(), relative)

    def test_publisher_identity_is_uniform(self) -> None:
        publisher = "Inspire"
        codex = json.loads(
            (REPO_ROOT / "plugins/ai-sow/.codex-plugin/plugin.json").read_text(
                encoding="utf-8"
            )
        )
        claude = json.loads(
            (REPO_ROOT / "plugins/ai-sow/.claude-plugin/plugin.json").read_text(
                encoding="utf-8"
            )
        )
        marketplace = json.loads(
            (REPO_ROOT / ".claude-plugin/marketplace.json").read_text(encoding="utf-8")
        )
        self.assertEqual(codex["author"]["name"], publisher)
        self.assertEqual(codex["interface"]["developerName"], publisher)
        self.assertEqual(claude["author"]["name"], publisher)
        self.assertEqual(marketplace["owner"]["name"], publisher)

        for relative in ("NOTICE", "plugins/ai-sow/NOTICE"):
            notice = (REPO_ROOT / relative).read_text(encoding="utf-8")
            self.assertIn(f"Copyright 2026 {publisher}", notice, relative)

    def test_public_readme_documents_claude_code_lifecycle(self) -> None:
        text = (REPO_ROOT / "README.md").read_text(encoding="utf-8")
        for command in (
            "/plugin marketplace add InspireChina/ai-plugin-marketplace",
            "/plugin install ai-sow@ai-plugin-marketplace",
            "/plugin marketplace update ai-plugin-marketplace",
            "/plugin uninstall ai-sow@ai-plugin-marketplace",
            "/plugin marketplace remove ai-plugin-marketplace",
        ):
            self.assertIn(command, text)

    def test_claude_marketplace_matches_codex_marketplace(self) -> None:
        claude = json.loads(
            (REPO_ROOT / ".claude-plugin/marketplace.json").read_text(encoding="utf-8")
        )
        codex = json.loads(
            (REPO_ROOT / ".agents/plugins/marketplace.json").read_text(encoding="utf-8")
        )
        self.assertEqual(claude["name"], codex["name"])
        self.assertTrue(claude["owner"]["name"].strip())
        self.assertEqual(
            {entry["name"] for entry in claude["plugins"]},
            {entry["name"] for entry in codex["plugins"]},
        )
        ai_sow_entries = [
            entry for entry in claude["plugins"] if entry.get("name") == "ai-sow"
        ]
        self.assertEqual(len(ai_sow_entries), 1)
        self.assertEqual(ai_sow_entries[0]["source"], "./plugins/ai-sow")

    def test_plugin_manifests_declare_one_release_identity(self) -> None:
        plugin_root = REPO_ROOT / "plugins/ai-sow"
        codex = json.loads(
            (plugin_root / ".codex-plugin/plugin.json").read_text(encoding="utf-8")
        )
        claude = json.loads(
            (plugin_root / ".claude-plugin/plugin.json").read_text(encoding="utf-8")
        )
        for field in ("name", "version", "description"):
            self.assertEqual(codex[field], claude[field], field)
        self.assertEqual(claude["name"], "ai-sow")

    def test_public_docs_exclude_internal_execution_plans(self) -> None:
        tracked = subprocess.run(
            ["git", "ls-files", "-z"],
            cwd=REPO_ROOT,
            check=True,
            capture_output=True,
        ).stdout.decode("utf-8").split("\0")
        self.assertFalse(
            any(
                path.startswith((".superpowers/", "docs/superpowers/"))
                for path in tracked
            )
        )
        self.assertTrue(
            (REPO_ROOT / "docs/architecture/ai-plugin-marketplace-design.md").is_file()
        )

    def test_ci_covers_all_supported_operating_systems(self) -> None:
        text = (REPO_ROOT / ".github/workflows/ci.yml").read_text(encoding="utf-8")
        for runner in ("ubuntu-latest", "macos-latest", "windows-latest"):
            self.assertIn(runner, text)

    def test_ci_installs_real_office_engine_on_every_supported_os(self) -> None:
        text = (REPO_ROOT / ".github/workflows/ci.yml").read_text(encoding="utf-8")
        for required in (
            "libreoffice-calc",
            "brew install --cask libreoffice",
            "choco install libreoffice-fresh",
            "AI_SOW_OFFICE_BIN",
        ):
            with self.subTest(required=required):
                self.assertIn(required, text)

    def test_public_text_has_no_private_paths_or_internal_plan(self) -> None:
        completed = subprocess.run(
            ["git", "ls-files", "-z"],
            cwd=REPO_ROOT,
            check=True,
            capture_output=True,
        )
        home_prefix = str(Path.home().resolve()) + "/"
        forbidden_plan = "2026-08-19-" + "as-is-output-contract.md"
        for raw_path in completed.stdout.split(b"\0"):
            if not raw_path:
                continue
            path = REPO_ROOT / raw_path.decode()
            if not path.is_file():
                continue
            try:
                text = path.read_text(encoding="utf-8")
            except UnicodeDecodeError:
                continue
            self.assertNotIn(home_prefix, text, path)
            self.assertNotIn(forbidden_plan, text, path)

    def test_generate_owns_the_only_bundled_template(self) -> None:
        plugin_root = REPO_ROOT / "plugins/ai-sow"
        paths = [
            path
            for path in plugin_root.rglob("sow-template.xlsx")
            if ".venv" not in path.parts
        ]
        self.assertEqual(paths, [plugin_root / "skills/generate/assets/sow-template.xlsx"])
        self.assertEqual(
            hashlib.sha256(paths[0].read_bytes()).hexdigest(), TEMPLATE_SHA256
        )

    def test_markdown_reference_describes_the_v13_estimation_model(self) -> None:
        path = (
            REPO_ROOT
            / "plugins/ai-sow/docs/reference/"
            / "SOW任务分类与开发交付人天标准_v1.3.md"
        )
        text = path.read_text(encoding="utf-8")
        for required in (
            "Epic → Feature → Story → Task",
            "一行 Task 对应一个基础单元实例",
            "新建 / 调整 / 接入复用",
            "K_COMPLEXITY_S",
            "affectsEstimate = true",
            "基础单元目录与 M 档基础人天合并为一张表",
        ):
            with self.subTest(required=required):
                self.assertIn(required, text)
        self.assertFalse(path.with_suffix(".docx").exists())

    def test_markdown_reference_task_fields_match_delivery_schema(self) -> None:
        document_path = (
            REPO_ROOT
            / "plugins/ai-sow/docs/reference/"
            / "SOW任务分类与开发交付人天标准_v1.3.md"
        )
        schema_path = (
            REPO_ROOT
            / "plugins/ai-sow/skills/generate/contracts/delivery-bundle.schema.json"
        )
        text = document_path.read_text(encoding="utf-8")
        schema = json.loads(schema_path.read_text(encoding="utf-8"))
        task_schema = schema["$defs"]["task"]
        task_section = text.split("## 3. Task 最小字段", 1)[1].split("## 4.", 1)[0]
        documented_fields = {
            match.group(1)
            for line in task_section.splitlines()
            if (match := re.match(r"\| `([^`]+)` \|", line))
        }

        self.assertEqual(documented_fields, set(task_schema["properties"]))
        complexity_values = " / ".join(task_schema["properties"]["complexity"]["enum"])
        self.assertIn(
            f"稳定 `complexity` 只允许 `{complexity_values}`",
            text,
        )
        self.assertIn("`X/拆分条件` 不是稳定 `complexity` 值", text)

    def test_public_docs_do_not_copy_the_current_template_catalog_size(self) -> None:
        template = REPO_ROOT / "plugins/ai-sow/skills/generate/assets/sow-template.xlsx"
        workbook = openpyxl.load_workbook(template, read_only=False, data_only=False)
        try:
            sheet = workbook["90-估算标准"]
            table = sheet.tables["BaseUnitCatalogTable"]
            min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(
                table.ref
            )
            headers = {
                str(sheet.cell(min_row, column).value): column
                for column in range(min_col, max_col + 1)
            }
            rows = range(min_row + 1, max_row + 1)
            catalog_size = sum(
                bool(sheet.cell(row, headers["基础单元ID"]).value) for row in rows
            )
            task_family_count = len(
                {
                    sheet.cell(row, headers["任务族ID"]).value
                    for row in rows
                    if sheet.cell(row, headers["任务族ID"]).value
                }
            )
        finally:
            workbook.close()

        copied_counts = (
            re.compile(rf"{catalog_size}\s*(?:个|项|行)?\s*基础单元"),
            re.compile(rf"{task_family_count}\s*个任务族"),
        )
        for relative in TASK_STANDARD_DOCS:
            text = (REPO_ROOT / relative).read_text(encoding="utf-8")
            for copied_count in copied_counts:
                with self.subTest(document=relative, pattern=copied_count.pattern):
                    self.assertIsNone(copied_count.search(text))

    def test_markdown_reference_defers_live_catalog_and_effort_to_template(self) -> None:
        path = (
            REPO_ROOT
            / "plugins/ai-sow/docs/reference/"
            / "SOW任务分类与开发交付人天标准_v1.3.md"
        )
        text = path.read_text(encoding="utf-8")
        self.assertIn("`90-估算标准`", text)
        self.assertIn("运行时模板", text)
        self.assertNotIn("### 8.1 前端", text)
        self.assertNotIn("### 12.3 推荐 M 档基础人天矩阵", text)
        self.assertNotIn("| 任务族 | 基础单元 | 计数口径 | 具体工作内容 |", text)

    def test_public_docs_explain_template_runs_and_transparent_questions(self) -> None:
        required_by_document = {
            "README.md": ("本轮专用副本", "重新编译 Delivery", "为什么要问"),
            "plugins/ai-sow/README.md": ("当前只支持 XLSX 模板", "本轮专用副本", "未回答后果"),
            "plugins/ai-sow/docs/AI_SOW_PLUGIN_DESIGN.md": ("当前只支持 XLSX 模板", "重新编译 Delivery", "可读文件"),
            "plugins/ai-sow/docs/CONTEXT.md": ("本轮专用副本", "问题、为什么要问、答案决定什么和未回答后果"),
            "plugins/ai-sow/docs/PRD_HLD_AUTOMATED_SOW_WORKFLOW_PLAN.md": ("重新编译 Delivery", "自然语言结论", "可打开的 Markdown 或 Excel 文件"),
        }
        for relative, required in required_by_document.items():
            text = (REPO_ROOT / relative).read_text(encoding="utf-8")
            for fragment in required:
                with self.subTest(document=relative, fragment=fragment):
                    self.assertIn(fragment, text)

    def test_public_docs_do_not_describe_template_changes_as_render_only(self) -> None:
        forbidden = (
            "模板单独变化时只重新渲染",
            "仅模板变化时跳过语义编译并完整重渲染",
        )
        for relative in TASK_STANDARD_DOCS:
            text = re.sub(
                r"\s+", "", (REPO_ROOT / relative).read_text(encoding="utf-8")
            )
            for fragment in forbidden:
                with self.subTest(document=relative, fragment=fragment):
                    self.assertNotIn(fragment, text)

    def test_public_docs_do_not_require_user_packet_approval(self) -> None:
        for relative in TASK_STANDARD_DOCS:
            text = re.sub(
                r"\s+", "", (REPO_ROOT / relative).read_text(encoding="utf-8")
            )
            with self.subTest(document=relative):
                self.assertNotRegex(
                    text,
                    r"用户(?:必须)?批准(?:精确)?(?:hash-bound)?(?:review)?packet",
                )

    def test_copy_smoke_checks_exact_template_path_and_complete_generation_map(
        self,
    ) -> None:
        path = REPO_ROOT / "plugins/ai-sow/tests/support/smoke_plugin.py"
        spec = importlib.util.spec_from_file_location("ai_sow_smoke_support", path)
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        self.assertTrue(hasattr(module, "_verify_generation_template_path"))
        self.assertTrue(hasattr(module, "_generation_file_digests"))

        with tempfile.TemporaryDirectory() as temp_dir:
            generation_root = Path(temp_dir) / ".ai-sow/generations/000123"
            for relative in (
                "manifest.json",
                "data/scope.json",
                "data/delivery.json",
                "input/sow-template.xlsx",
                "output/sow.xlsx",
                "output/sow-notes.md",
            ):
                target = generation_root / relative
                target.parent.mkdir(parents=True, exist_ok=True)
                target.write_bytes(relative.encode())

            module._verify_generation_template_path(
                {
                    "generationId": "000123",
                    "templatePath": (
                        ".ai-sow/generations/000123/input/sow-template.xlsx"
                    ),
                },
                generation_root,
            )
            for invalid_path in (
                ".ai-sow/templates/sow-template.xlsx",
                ".ai-sow/work/run-template.xlsx",
                "input/sow-template.xlsx",
            ):
                with self.subTest(invalid_path=invalid_path), self.assertRaises(
                    RuntimeError
                ):
                    module._verify_generation_template_path(
                        {
                            "generationId": "000123",
                            "templatePath": invalid_path,
                        },
                        generation_root,
                    )

            before = module._generation_file_digests(generation_root.parent)
            (generation_root / "output/sow-notes.md").write_bytes(b"changed")
            self.assertNotEqual(
                module._generation_file_digests(generation_root.parent), before
            )

    def test_generate_schema_hashes_are_fixed(self) -> None:
        plugin_root = REPO_ROOT / "plugins/ai-sow"
        for relative, expected_hash in SCHEMA_SHA256.items():
            with self.subTest(schema=relative):
                path = plugin_root / relative
                self.assertEqual(hashlib.sha256(path.read_bytes()).hexdigest(), expected_hash)

    def test_marketplace_points_to_ai_sow(self) -> None:
        marketplace = json.loads(
            (REPO_ROOT / ".agents/plugins/marketplace.json").read_text(encoding="utf-8")
        )
        self.assertEqual(marketplace["name"], "ai-plugin-marketplace")
        self.assertEqual(
            marketplace["interface"]["displayName"], "AI Plugin Marketplace"
        )
        ai_sow_entries = [
            entry
            for entry in marketplace["plugins"]
            if entry.get("name") == "ai-sow"
        ]
        self.assertEqual(len(ai_sow_entries), 1)
        self.assertEqual(
            ai_sow_entries[0],
            {
                "name": "ai-sow",
                "description": "一次提供 PRD、HLD 和适用的往期 SOW，自动生成或增量更新可追溯的 SOW 工作簿，并用 LibreOffice 回算后发布。",
                "source": {
                    "source": "local",
                    "path": "./plugins/ai-sow",
                },
                "policy": {
                    "installation": "AVAILABLE",
                    "authentication": "ON_INSTALL",
                },
                "category": "Productivity",
            },
        )

    def test_repository_validator_reports_no_errors(self) -> None:
        completed = subprocess.run(
            [sys.executable, str(REPO_ROOT / "scripts/validate_repository.py")],
            cwd=REPO_ROOT,
            capture_output=True,
            text=True, encoding="utf-8",
            check=False,
        )
        self.assertEqual(completed.returncode, 0, completed.stderr)


if __name__ == "__main__":
    unittest.main()
