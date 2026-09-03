#!/usr/bin/env python3
"""Exercise the installed AI SOW generate Skill from a standalone copy."""

from __future__ import annotations

import argparse
import copy
import hashlib
import json
import os
import shutil
import subprocess
import tempfile
from collections.abc import Mapping, Sequence
from pathlib import Path

import jsonschema
import openpyxl


CASE_MANIFEST_PATH = "tests/fixtures/explicit-architecture/case-manifest.json"
CASE_SCHEMA_PATH = "tests/contracts/case-manifest.schema.json"
ORCHESTRATOR_PATH = "skills/generate/scripts/orchestrator.py"
EXPECTED_GENERATION_FILES = {
    "data/delivery.json",
    "data/scope.json",
    "input/sow-template.xlsx",
    "manifest.json",
    "output/sow-notes.md",
    "output/sow.xlsx",
}
EXPECTED_TABLES = {
    "SOWStoryTable",
    "TaskTable",
    "ProjectSummaryTable",
    "ProjectParameterTable",
    "BaseUnitCatalogTable",
}
EXPECTED_SHEETS = ["01-需求故事", "02-任务清单", "03-工作量汇总", "90-估算标准"]
SCOPE_COLLECTIONS = (
    "epics",
    "features",
    "commitments",
    "effectiveStartItems",
    "designItems",
    "designDecisions",
    "integrations",
    "nfrs",
    "assumptions",
)
SCOPE_ID_FIELDS = {
    "epics": ("EPIC", "epicId"),
    "features": ("FEATURE", "featureId"),
    "commitments": ("COMMITMENT", "commitmentId"),
    "effectiveStartItems": ("EFFECTIVE_START_ITEM", "effectiveStartItemId"),
    "designItems": ("DESIGN_ITEM", "designItemId"),
    "designDecisions": ("DESIGN_DECISION", "designDecisionId"),
    "integrations": ("INTEGRATION", "integrationId"),
    "nfrs": ("NFR", "nfrId"),
    "assumptions": ("ASSUMPTION", "assumptionId"),
}
DELIVERY_COLLECTIONS = (
    "stories",
    "acceptanceCriteria",
    "tasks",
    "dependencies",
)
DELIVERY_ID_FIELDS = {
    "stories": ("STORY", "storyId"),
    "acceptanceCriteria": ("ACCEPTANCE_CRITERION", "acceptanceCriterionId"),
    "tasks": ("TASK", "taskId"),
    "dependencies": ("DEPENDENCY", "dependencyId"),
}
QUESTION_FIELDS = {
    "questionId",
    "subjectIds",
    "question",
    "reason",
    "decisionImpact",
    "unansweredEffect",
}


def _canonical_json_bytes(value: object) -> bytes:
    return (
        json.dumps(
            value,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        )
        + "\n"
    ).encode("utf-8")


def _sha256(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _generation_file_digests(root: Path) -> dict[str, str]:
    return {
        path.relative_to(root).as_posix(): _sha256(path.read_bytes())
        for path in sorted(root.rglob("*"))
        if path.is_file()
    }


def _verify_generation_template_path(
    manifest: Mapping[str, object], generation_root: Path
) -> None:
    generation_id = manifest.get("generationId")
    expected = f".ai-sow/generations/{generation_id}/input/sow-template.xlsx"
    if (
        not isinstance(generation_id, str)
        or generation_root.name != generation_id
        or manifest.get("templatePath") != expected
        or not (generation_root / "input/sow-template.xlsx").is_file()
    ):
        raise RuntimeError(
            "generation templatePath must identify its exact immutable "
            "input/sow-template.xlsx member"
        )


def _load_json(path: Path) -> dict[str, object]:
    value = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise RuntimeError(f"expected JSON object: {path}")
    return value


def _write_json(path: Path, value: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(_canonical_json_bytes(value))


def _complete_id_decisions(
    decisions: dict[str, object],
    candidate: Mapping[str, object],
    id_fields: Mapping[str, tuple[str, str]],
    *,
    has_baseline: bool,
) -> None:
    entries = decisions.get("decisions")
    if not isinstance(entries, list):
        raise RuntimeError("ID decisions fixture must contain a decisions list")
    known = {
        (entry.get("objectType"), entry.get("objectId"))
        for entry in entries
        if isinstance(entry, Mapping)
    }
    for collection, (object_type, id_field) in id_fields.items():
        values = candidate.get(collection)
        if not isinstance(values, list):
            raise RuntimeError(f"candidate collection must be a list: {collection}")
        for item in values:
            if not isinstance(item, Mapping) or not isinstance(item.get(id_field), str):
                continue
            object_id = str(item[id_field])
            if (object_type, object_id) not in known:
                entries.append(
                    {
                        "objectType": object_type,
                        "objectId": object_id,
                        "disposition": "NEW",
                        "meaningPreserved": False,
                        "rationale": "首个 revision 中的新对象。",
                    }
                )
    if has_baseline:
        for decision in entries:
            if isinstance(decision, dict):
                decision.update(
                    {
                        "disposition": "UNCHANGED",
                        "previousId": decision["objectId"],
                        "meaningPreserved": True,
                        "rationale": "模板变化未改变对象语义，保留稳定 ID。",
                    }
                )


def run_command(command: list[str], cwd: Path) -> dict[str, object]:
    """Run a support command and return its final JSON object when present."""
    completed = subprocess.run(
        command,
        cwd=cwd,
        text=True,
        encoding="utf-8",
        capture_output=True,
        check=False,
    )
    if completed.returncode != 0:
        rendered = " ".join(command)
        raise RuntimeError(
            f"command failed ({completed.returncode}): {rendered}\n"
            f"stdout:\n{completed.stdout}\n"
            f"stderr:\n{completed.stderr}"
        )
    lines = [line for line in completed.stdout.splitlines() if line.strip()]
    if lines:
        try:
            payload = json.loads(lines[-1])
        except json.JSONDecodeError:
            payload = None
        if isinstance(payload, dict):
            return payload
    return {"outcome": "OK", "stdout": completed.stdout}


def plugin_uv_command(plugin_root: Path) -> str:
    local_uv = plugin_root / ".ai-sow-tools" / "bin" / (
        "uv.exe" if os.name == "nt" else "uv"
    )
    if local_uv.is_file():
        return str(local_uv)
    return shutil.which("uv") or "uv"


def plugin_python_command(plugin_root: Path) -> str:
    return str(
        plugin_root
        / ".venv"
        / ("Scripts/python.exe" if os.name == "nt" else "bin/python")
    )


def _install_read_guard(audit_root: Path) -> Path:
    audit_root.mkdir(parents=True, exist_ok=True)
    guard = audit_root / "sitecustomize.py"
    guard.write_text(
        """from __future__ import annotations
import os
from pathlib import Path

_ORIGINAL_OPEN = Path.open
_ALLOWED = tuple(
    Path(item).resolve()
    for item in os.environ["AI_SOW_ALLOWED_READ_ROOTS"].split(os.pathsep)
    if item
)
_LOG = Path(os.environ["AI_SOW_FORBIDDEN_READ_LOG"])

def _within(path: Path, root: Path) -> bool:
    try:
        path.relative_to(root)
        return True
    except ValueError:
        return False

def _guarded_open(self, mode="r", buffering=-1, encoding=None, errors=None, newline=None):
    if str(mode).startswith("r"):
        resolved = self.resolve()
        if not any(_within(resolved, root) for root in _ALLOWED):
            with open(_LOG, "a", encoding="utf-8") as stream:
                stream.write(str(resolved) + "\\n")
            raise RuntimeError(f"runtime read escaped installed plugin/project roots: {resolved}")
    return _ORIGINAL_OPEN(self, mode, buffering, encoding, errors, newline)

Path.open = _guarded_open
""",
        encoding="utf-8",
    )
    return guard


def _orchestrator_environment(
    active_plugin: Path,
    project: Path,
    audit_root: Path,
    audit_log: Path,
) -> dict[str, str]:
    return {
        **os.environ,
        "PYTHONPATH": str(audit_root),
        "PYTHONDONTWRITEBYTECODE": "1",
        "PYTHONUTF8": "1",
        "AI_SOW_ALLOWED_READ_ROOTS": os.pathsep.join(
            (str(active_plugin), str(project))
        ),
        "AI_SOW_FORBIDDEN_READ_LOG": str(audit_log),
    }


def _run_orchestrator(
    active_plugin: Path,
    project: Path,
    audit_root: Path,
    audit_log: Path,
    mode: str,
    *arguments: str,
    expected: str,
) -> dict[str, object]:
    command = [
        plugin_python_command(active_plugin),
        str(active_plugin / ORCHESTRATOR_PATH),
        "--project-root",
        str(project),
        "--mode",
        mode,
        *arguments,
    ]
    completed = subprocess.run(
        command,
        cwd=project,
        capture_output=True,
        check=False,
        env=_orchestrator_environment(active_plugin, project, audit_root, audit_log),
    )
    try:
        stdout = completed.stdout.decode("utf-8")
        stderr = completed.stderr.decode("utf-8")
    except UnicodeDecodeError as error:
        raise RuntimeError(f"orchestrator output was not UTF-8: {command}") from error
    lines = [line for line in stdout.splitlines() if line.strip()]
    if len(lines) != 1:
        raise RuntimeError(
            f"orchestrator did not emit exactly one JSON result: {command}\n"
            f"stdout:\n{stdout}\nstderr:\n{stderr}"
        )
    try:
        payload = json.loads(lines[0])
    except json.JSONDecodeError as error:
        raise RuntimeError(f"orchestrator output was not JSON: {stdout}") from error
    if not isinstance(payload, dict) or payload.get("outcome") != expected:
        raise RuntimeError(
            f"orchestrator {mode} expected {expected}, got {payload}\n"
            f"returncode={completed.returncode}\nstderr:\n{stderr}"
        )
    expected_code = 2 if expected == "BLOCKED" else 0
    if completed.returncode != expected_code:
        raise RuntimeError(
            f"orchestrator {mode} returned {completed.returncode}, expected {expected_code}"
        )
    return payload


def _case_manifest(plugin_root: Path) -> list[dict[str, object]]:
    manifest = _load_json(plugin_root / CASE_MANIFEST_PATH)
    schema = _load_json(plugin_root / CASE_SCHEMA_PATH)
    jsonschema.Draft202012Validator(schema).validate(manifest)
    cases = manifest["cases"]
    if not isinstance(cases, list):
        raise RuntimeError("case manifest cases must be a list")
    return [dict(case) for case in cases if isinstance(case, Mapping)]


def _prepare_request(
    active_plugin: Path,
    project: Path,
    case: Mapping[str, object],
    *,
    omit_prior_sow: bool = False,
) -> str:
    project.mkdir(parents=True, exist_ok=True)
    request = _load_json(active_plugin / str(case["requestPath"]))
    sources = request.get("sources")
    if not isinstance(sources, list):
        raise RuntimeError("request sources must be a list")
    prepared_sources: list[dict[str, object]] = []
    for raw_source in sources:
        if not isinstance(raw_source, Mapping):
            continue
        source = dict(raw_source)
        role = str(source["role"])
        if omit_prior_sow and role == "PRIOR_SOW":
            continue
        relative = Path(str(source["path"]))
        target = project / relative
        target.parent.mkdir(parents=True, exist_ok=True)
        if role == "PRIOR_SOW":
            shutil.copy2(
                active_plugin / "skills/generate/assets/sow-template.xlsx",
                target,
            )
        else:
            marker = f"SOURCE_ORIGINAL_{case['caseId']}_{role}"
            target.write_text(
                f"# {role}\n\n{marker}\n\n退款申请、审核、结果通知、异常处理与交付边界。\n",
                encoding="utf-8",
            )
        prepared_sources.append(source)
    request["sources"] = prepared_sources
    request_path = project / "request.json"
    _write_json(request_path, request)
    return request_path.name


def _scope_candidate(
    active_plugin: Path,
    project: Path,
    case: Mapping[str, object],
    prepared: Mapping[str, object],
) -> tuple[str, str]:
    plan = prepared["runPlan"]
    if not isinstance(plan, Mapping):
        raise RuntimeError("prepare result lacks runPlan")
    bundle = _load_json(active_plugin / str(case["scopeSlicePath"]))
    anchors_value = json.loads(
        (project / ".ai-sow/inputs/pending/anchors.json").read_text(encoding="utf-8")
    )
    if not isinstance(anchors_value, list):
        raise RuntimeError("pending anchors must be a list")
    first_anchor: dict[str, Mapping[str, object]] = {}
    for anchor in anchors_value:
        if isinstance(anchor, Mapping) and isinstance(anchor.get("sourceId"), str):
            first_anchor.setdefault(str(anchor["sourceId"]), anchor)
    for collection in SCOPE_COLLECTIONS:
        values = bundle.get(collection, [])
        if not isinstance(values, list):
            raise RuntimeError(f"invalid Scope fixture collection: {collection}")
        for item in values:
            if not isinstance(item, dict):
                continue
            refs = item.get("sourceRefs", [])
            item["sourceRefs"] = [
                {
                    "sourceId": first_anchor[str(ref["sourceId"])]["sourceId"],
                    "anchorId": first_anchor[str(ref["sourceId"])]["anchorId"],
                    "locator": first_anchor[str(ref["sourceId"])]["locator"],
                    "sha256": first_anchor[str(ref["sourceId"])]["sha256"],
                }
                for ref in refs
                if isinstance(ref, Mapping)
            ]
    candidate = {
        "contract": "ai-sow-scope-slice-v1",
        "inputRevisionId": plan["targetRevisionId"],
        "impactPlanSha256": _sha256(_canonical_json_bytes(plan["impact"])),
        "replacesFeatureIds": (
            []
            if plan["impact"]["baselineGenerationId"] is None
            else list(plan["impact"]["affectedFeatureIds"])
        ),
        "newAnchorMappings": [],
        **{name: copy.deepcopy(bundle[name]) for name in SCOPE_COLLECTIONS},
        "responsibilityBoundaries": copy.deepcopy(bundle["responsibilityBoundaries"]),
    }
    candidate_name = "scope-slice.json"
    ids_name = "scope-id-decisions.json"
    _write_json(project / candidate_name, candidate)
    decisions = _load_json(active_plugin / str(case["scopeIdDecisionsPath"]))
    _complete_id_decisions(
        decisions,
        candidate,
        SCOPE_ID_FIELDS,
        has_baseline=plan["impact"]["baselineGenerationId"] is not None,
    )
    _write_json(project / ids_name, decisions)
    return candidate_name, ids_name


def _delivery_candidate(
    active_plugin: Path,
    project: Path,
    case: Mapping[str, object],
) -> tuple[str, str]:
    plan = _load_json(project / ".ai-sow/work/run-plan.json")
    scope = _load_json(project / ".ai-sow/work/scope.candidate.json")
    bundle = _load_json(active_plugin / str(case["deliverySlicePath"]))
    candidate = {
        "contract": "ai-sow-delivery-slice-v3",
        "inputRevisionId": plan["targetRevisionId"],
        "scopeSha256": _sha256(_canonical_json_bytes(scope)),
        "impactPlanSha256": _sha256(_canonical_json_bytes(plan["impact"])),
        "replacesFeatureIds": (
            []
            if plan["impact"]["baselineGenerationId"] is None
            else list(plan["impact"]["affectedFeatureIds"])
        ),
        **{name: copy.deepcopy(bundle[name]) for name in DELIVERY_COLLECTIONS},
    }
    candidate_name = "delivery-slice.json"
    ids_name = "delivery-id-decisions.json"
    _write_json(project / candidate_name, candidate)
    decisions = _load_json(active_plugin / str(case["deliveryIdDecisionsPath"]))
    _complete_id_decisions(
        decisions,
        candidate,
        DELIVERY_ID_FIELDS,
        has_baseline=plan["impact"]["baselineGenerationId"] is not None,
    )
    _write_json(project / ids_name, decisions)
    return candidate_name, ids_name


def _review_candidate(
    active_plugin: Path,
    project: Path,
    case: Mapping[str, object],
    packet: Mapping[str, object],
) -> str:
    plan = _load_json(project / ".ai-sow/work/run-plan.json")
    review = _load_json(active_plugin / str(case["finalReviewPath"]))
    review.update(
        {
            "runId": plan["runId"],
            "inputRevisionId": plan["targetRevisionId"],
            "scopeSha256": _sha256(
                (project / ".ai-sow/work/scope.candidate.json").read_bytes()
            ),
            "deliverySha256": _sha256(
                (project / ".ai-sow/work/delivery.candidate.json").read_bytes()
            ),
            "packetSha256": packet["packetSha256"],
        }
    )
    review_name = "final-review.json"
    _write_json(project / review_name, review)
    return review_name


def _blocked_review_candidate(
    active_plugin: Path,
    project: Path,
    case: Mapping[str, object],
    packet: Mapping[str, object],
) -> str:
    review_name = _review_candidate(active_plugin, project, case, packet)
    review = _load_json(project / review_name)
    delivery = _load_json(project / ".ai-sow/work/delivery.candidate.json")
    stories = delivery.get("stories")
    if not isinstance(stories, list) or not stories or not isinstance(stories[0], Mapping):
        raise RuntimeError("blocked review smoke requires one Story")
    review.update(
        {
            "decision": "BLOCKED",
            "notes": [],
            "questions": [
                {
                    "questionId": "question-smoke-review-boundary",
                    "subjectIds": [stories[0]["storyId"]],
                    "question": "是否确认该故事的验收边界包含异常处理结果？",
                    "reason": "当前终审材料无法确认异常处理结果是否属于本次可验收范围。",
                    "decisionImpact": "确认包含会保留对应 AC 与 Task；确认不包含会重新编译相关 Delivery 切片。",
                    "unansweredEffect": "未回答时无法固定验收和估算边界，本轮保持 BLOCKED 且不发布工作簿。",
                }
            ],
        }
    )
    _write_json(project / review_name, review)
    return review_name


def _verify_transparent_questions(result: Mapping[str, object]) -> None:
    questions = result.get("questions")
    if not isinstance(questions, list) or not questions:
        raise RuntimeError(f"BLOCKED result omitted user questions: {result}")
    for question in questions:
        if not isinstance(question, Mapping) or set(question) != QUESTION_FIELDS:
            raise RuntimeError(f"BLOCKED question is not self-contained: {question}")
        if any(not question[field] for field in QUESTION_FIELDS):
            raise RuntimeError(f"BLOCKED question contains an empty field: {question}")


def _verify_review_material(project: Path, packet: Mapping[str, object]) -> None:
    relative = packet.get("reviewMaterialPath")
    if not isinstance(relative, str) or not relative:
        raise RuntimeError(f"review result omitted readable material: {packet}")
    text = (project / relative).read_text(encoding="utf-8")
    first_line = text.partition("\n")[0]
    if (
        not first_line.startswith("# ")
        or not first_line.endswith("终审材料")
        or "## 下一步" not in text
    ):
        raise RuntimeError(f"review material lacks natural-language summary: {relative}")
    packet_hash = packet.get("packetSha256")
    if isinstance(packet_hash, str) and packet_hash in text:
        raise RuntimeError("review material exposed the internal packet hash")


def _verify_manifest_hashes(project: Path, manifest: Mapping[str, object]) -> None:
    pairs = (
        ("inputManifestPath", "inputManifestSha256"),
        ("scopePath", "scopeSha256"),
        ("deliveryPath", "deliverySha256"),
        ("templatePath", "templateSha256"),
        ("workbookPath", "workbookSha256"),
        ("notesPath", "notesSha256"),
    )
    for path_key, hash_key in pairs:
        target = project / str(manifest[path_key])
        if _sha256(target.read_bytes()) != manifest[hash_key]:
            raise RuntimeError(f"generation hash mismatch: {path_key}")
    if _sha256(_canonical_json_bytes(manifest["finalReview"])) != manifest["finalReviewSha256"]:
        raise RuntimeError("generation final review hash mismatch")


def _verify_workbook(path: Path) -> None:
    workbook = openpyxl.load_workbook(path, data_only=False)
    try:
        if workbook.sheetnames != EXPECTED_SHEETS:
            raise RuntimeError(f"generated workbook sheets changed: {workbook.sheetnames}")
        tables = {name for sheet in workbook.worksheets for name in sheet.tables}
        if tables != EXPECTED_TABLES:
            raise RuntimeError(f"generated workbook tables incomplete: {tables}")
        formulas = [
            cell.value
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if cell.data_type == "f"
        ]
        if not formulas:
            raise RuntimeError("generated workbook contains no formulas")
    finally:
        workbook.close()
    calculated = openpyxl.load_workbook(path, data_only=True)
    try:
        values = [calculated["03-工作量汇总"][f"B{row}"].value for row in range(5, 9)]
        if not all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in values):
            raise RuntimeError(f"generated workbook lacks calculated summary: {values}")
    finally:
        calculated.close()


def _verify_generation(
    project: Path,
    case: Mapping[str, object],
    result: Mapping[str, object],
) -> dict[str, object]:
    expected = case["expectedCounts"]
    if not isinstance(expected, Mapping):
        raise RuntimeError("case expectedCounts must be an object")
    if result.get("decision") != case["expectedDecision"]:
        raise RuntimeError(f"unexpected review decision: {result}")
    change_counts = result.get("changeCounts")
    if not isinstance(change_counts, Mapping):
        raise RuntimeError(f"missing change counts: {result}")
    feature_counts = change_counts.get("features")
    story_counts = change_counts.get("stories")
    task_counts = change_counts.get("tasks")
    if not isinstance(feature_counts, Mapping) or feature_counts.get("recomputed") != expected["features"]:
        raise RuntimeError(f"unexpected Feature counts: {result}")
    if not isinstance(story_counts, Mapping) or story_counts.get("recomputed") != expected["stories"]:
        raise RuntimeError(f"unexpected Story counts: {result}")
    if not isinstance(task_counts, Mapping) or task_counts.get("recomputed") != expected["tasks"]:
        raise RuntimeError(f"unexpected Task counts: {result}")

    current = _load_json(project / ".ai-sow/current.json")
    generation_root = project / ".ai-sow/generations" / str(current["generationId"])
    files = {
        path.relative_to(generation_root).as_posix()
        for path in generation_root.rglob("*")
        if path.is_file()
    }
    if files != EXPECTED_GENERATION_FILES:
        raise RuntimeError(f"generation leaked or omitted files: {files}")
    manifest = _load_json(generation_root / "manifest.json")
    verification = manifest.get("workbookVerification")
    if not isinstance(verification, Mapping) or verification.get("trustState") != "VERIFIED":
        raise RuntimeError(f"generation workbook is not verified: {verification}")
    engine = verification.get("engine")
    if (
        not isinstance(engine, Mapping)
        or not isinstance(engine.get("name"), str)
        or not engine["name"]
        or not isinstance(engine.get("version"), str)
        or not engine["version"]
    ):
        raise RuntimeError(f"generation lacks a real Office engine record: {verification}")
    _verify_generation_template_path(manifest, generation_root)
    _verify_manifest_hashes(project, manifest)
    workbook_path = project / str(manifest["workbookPath"])
    notes_path = project / str(manifest["notesPath"])
    _verify_workbook(workbook_path)
    notes = notes_path.read_text(encoding="utf-8")
    final_review = manifest["finalReview"]
    if not isinstance(final_review, Mapping):
        raise RuntimeError("manifest final review must be an object")
    for note in final_review.get("notes", []):
        if isinstance(note, Mapping) and notes.count(str(note["sowNotesText"])) != 1:
            raise RuntimeError(f"review note was not rendered exactly once: {note}")
    generated_bytes = b"\n".join(
        path.read_bytes() for path in generation_root.rglob("*") if path.is_file()
    )
    if b"SOURCE_ORIGINAL_" in generated_bytes:
        raise RuntimeError("source original content leaked into generation output")
    for hash_key in ("expectedWorkbookSha256", "expectedNotesSha256"):
        if hash_key in case:
            actual = manifest[
                "workbookSha256" if hash_key == "expectedWorkbookSha256" else "notesSha256"
            ]
            if actual != case[hash_key]:
                raise RuntimeError(f"fixed output hash mismatch: {hash_key}")
    return {
        "workbookPath": str(workbook_path.resolve()),
        "templateSha256": manifest["templateSha256"],
        "officeEngine": {"name": engine["name"], "version": engine["version"]},
    }


def _complete_prepared_case(
    active_plugin: Path,
    project: Path,
    case: Mapping[str, object],
    prepared: Mapping[str, object],
    audit_root: Path,
    audit_log: Path,
) -> tuple[dict[str, object], dict[str, object]]:
    plan = prepared.get("runPlan")
    if not isinstance(plan, Mapping):
        raise RuntimeError(f"prepare result lacks a run plan: {prepared}")
    snapshot = project / str(plan["templateSnapshotPath"])
    if _sha256(snapshot.read_bytes()) != plan["templateSha256"]:
        raise RuntimeError("run template snapshot hash does not close")

    scope_name, scope_ids = _scope_candidate(active_plugin, project, case, prepared)
    _run_orchestrator(
        active_plugin,
        project,
        audit_root,
        audit_log,
        "accept-scope",
        "--candidate",
        scope_name,
        "--ids",
        scope_ids,
        expected="READY_FOR_DELIVERY",
    )
    delivery_name, delivery_ids = _delivery_candidate(active_plugin, project, case)
    _run_orchestrator(
        active_plugin,
        project,
        audit_root,
        audit_log,
        "accept-delivery",
        "--candidate",
        delivery_name,
        "--ids",
        delivery_ids,
        expected="REVIEW_REQUIRED",
    )
    packet = _run_orchestrator(
        active_plugin,
        project,
        audit_root,
        audit_log,
        "prepare-review",
        expected="REVIEW_REQUIRED",
    )
    _verify_review_material(project, packet)
    review_name = _review_candidate(active_plugin, project, case, packet)
    _run_orchestrator(
        active_plugin,
        project,
        audit_root,
        audit_log,
        "accept-review",
        "--review",
        review_name,
        expected="READY_TO_RENDER",
    )
    published = _run_orchestrator(
        active_plugin,
        project,
        audit_root,
        audit_log,
        "publish",
        expected="PUBLISHED",
    )
    verified = _verify_generation(project, case, published)
    if verified["templateSha256"] != plan["templateSha256"]:
        raise RuntimeError("published generation did not retain the run template")
    return published, verified


def _run_case(
    active_plugin: Path,
    project: Path,
    case: Mapping[str, object],
    audit_root: Path,
    audit_log: Path,
) -> tuple[dict[str, object], dict[str, object]]:
    request_name = _prepare_request(active_plugin, project, case)
    prepared = _run_orchestrator(
        active_plugin,
        project,
        audit_root,
        audit_log,
        "prepare",
        "--request",
        request_name,
        expected="READY_FOR_SCOPE",
    )
    return _complete_prepared_case(
        active_plugin, project, case, prepared, audit_root, audit_log
    )


def _mutate_project_template(path: Path) -> None:
    workbook = openpyxl.load_workbook(path)
    try:
        sheet = workbook["90-估算标准"]
        table = sheet.tables["BaseUnitCatalogTable"]
        min_col, min_row, max_col, _ = openpyxl.utils.range_boundaries(table.ref)
        headers = {
            str(sheet.cell(min_row, column).value): column
            for column in range(min_col, max_col + 1)
        }
        effort_column = headers["新建M档人天"]
        value = sheet.cell(min_row + 1, effort_column).value
        if not isinstance(value, (int, float)) or isinstance(value, bool):
            raise RuntimeError("template smoke could not find a numeric effort cell")
        sheet.cell(min_row + 1, effort_column).value = value + 0.25
        workbook.save(path)
    finally:
        workbook.close()


def _run_blocked_review_case(
    active_plugin: Path,
    project: Path,
    case: Mapping[str, object],
    audit_root: Path,
    audit_log: Path,
) -> dict[str, object]:
    request_name = _prepare_request(active_plugin, project, case)
    prepared = _run_orchestrator(
        active_plugin,
        project,
        audit_root,
        audit_log,
        "prepare",
        "--request",
        request_name,
        expected="READY_FOR_SCOPE",
    )
    scope_name, scope_ids = _scope_candidate(active_plugin, project, case, prepared)
    _run_orchestrator(
        active_plugin,
        project,
        audit_root,
        audit_log,
        "accept-scope",
        "--candidate",
        scope_name,
        "--ids",
        scope_ids,
        expected="READY_FOR_DELIVERY",
    )
    delivery_name, delivery_ids = _delivery_candidate(active_plugin, project, case)
    _run_orchestrator(
        active_plugin,
        project,
        audit_root,
        audit_log,
        "accept-delivery",
        "--candidate",
        delivery_name,
        "--ids",
        delivery_ids,
        expected="REVIEW_REQUIRED",
    )
    packet = _run_orchestrator(
        active_plugin,
        project,
        audit_root,
        audit_log,
        "prepare-review",
        expected="REVIEW_REQUIRED",
    )
    _verify_review_material(project, packet)
    review_name = _blocked_review_candidate(active_plugin, project, case, packet)
    result = _run_orchestrator(
        active_plugin,
        project,
        audit_root,
        audit_log,
        "accept-review",
        "--review",
        review_name,
        expected="BLOCKED",
    )
    _verify_transparent_questions(result)
    if (project / ".ai-sow/current.json").exists():
        raise RuntimeError("BLOCKED review replaced last-known-good state")
    return result


def run_smoke(
    plugin_root: Path,
    work_dir: Path,
    copy_plugin: bool,
) -> dict[str, object]:
    source_plugin = plugin_root.resolve(strict=True)
    work_dir = work_dir.resolve()
    work_dir.mkdir(parents=True, exist_ok=True)
    if copy_plugin:
        active_plugin = work_dir / "installed" / "ai-sow"
        active_plugin.parent.mkdir(parents=True, exist_ok=True)
        shutil.copytree(
            source_plugin,
            active_plugin,
            ignore=shutil.ignore_patterns(".venv", ".ai-sow-tools", ".pytest_cache", "__pycache__", "*.pyc"),
        )
    else:
        active_plugin = source_plugin

    manifest = _load_json(active_plugin / ".codex-plugin/plugin.json")
    if manifest.get("name") != "ai-sow":
        raise RuntimeError(f"unexpected plugin manifest: {manifest}")
    sync = run_command(
        [
            plugin_uv_command(source_plugin),
            "sync",
            "--project",
            str(active_plugin),
            "--locked",
        ],
        cwd=work_dir,
    )
    if sync.get("outcome") != "OK":
        raise RuntimeError(f"plugin environment sync failed: {sync}")

    cases = {str(case["caseId"]): case for case in _case_manifest(active_plugin)}
    greenfield_case = cases["greenfield"]
    brownfield_case = cases["brownfield"]
    projects_root = work_dir / "customer-projects"
    greenfield = projects_root / "greenfield"
    brownfield = projects_root / "brownfield"
    blocked_resume = projects_root / "blocked-resume"
    blocked_review = projects_root / "blocked-review"
    audit_root = work_dir / "read-guard"
    audit_log = work_dir / "forbidden-reads.log"
    _install_read_guard(audit_root)

    greenfield_result, greenfield_verification = _run_case(
        active_plugin, greenfield, greenfield_case, audit_root, audit_log
    )
    brownfield_result, brownfield_verification = _run_case(
        active_plugin, brownfield, brownfield_case, audit_root, audit_log
    )

    blocked_review_result = _run_blocked_review_case(
        active_plugin, blocked_review, greenfield_case, audit_root, audit_log
    )

    incomplete_request = _prepare_request(
        active_plugin,
        blocked_resume,
        brownfield_case,
        omit_prior_sow=True,
    )
    blocked = _run_orchestrator(
        active_plugin,
        blocked_resume,
        audit_root,
        audit_log,
        "prepare",
        "--request",
        incomplete_request,
        expected="BLOCKED",
    )
    if not any(
        diagnostic.get("code") in {
            "REQUEST_BROWNFIELD_PRIOR_SOW_REQUIRED",
            "BROWNFIELD_PRIOR_SOW_REQUIRED",
        }
        for diagnostic in blocked.get("diagnostics", [])
        if isinstance(diagnostic, Mapping)
    ):
        raise RuntimeError(f"missing-prior-SOW did not return the expected diagnostic: {blocked}")
    blocked_resume_result, blocked_resume_verification = _run_case(
        active_plugin, blocked_resume, brownfield_case, audit_root, audit_log
    )

    reuse = _run_orchestrator(
        active_plugin,
        greenfield,
        audit_root,
        audit_log,
        "prepare",
        "--request",
        "request.json",
        expected="REUSED",
    )
    generations = sorted((greenfield / ".ai-sow/generations").iterdir())
    revisions = sorted((greenfield / ".ai-sow/inputs/revisions").iterdir())
    if len(generations) != 1 or len(revisions) != 1:
        raise RuntimeError("identical replay created a new immutable artifact")

    first_generation = _load_json(greenfield / ".ai-sow/current.json")
    first_generation_root = (
        greenfield / ".ai-sow/generations" / str(first_generation["generationId"])
    )
    first_template = first_generation_root / "input/sow-template.xlsx"
    _mutate_project_template(greenfield / ".ai-sow/templates/sow-template.xlsx")
    template_recompile = _run_orchestrator(
        active_plugin,
        greenfield,
        audit_root,
        audit_log,
        "prepare",
        "--request",
        "request.json",
        expected="READY_FOR_SCOPE",
    )
    recompile_plan = template_recompile.get("runPlan")
    if (
        not isinstance(recompile_plan, Mapping)
        or recompile_plan.get("action") != "FULL_COMPILE"
        or "TEMPLATE_CHANGED"
        not in recompile_plan.get("impact", {}).get("reasonCodes", [])
    ):
        raise RuntimeError(
            f"changed template did not force full Delivery compilation: {template_recompile}"
        )
    frozen_template = greenfield / str(recompile_plan["templateSnapshotPath"])
    frozen_template_bytes = frozen_template.read_bytes()
    _mutate_project_template(greenfield / ".ai-sow/templates/sow-template.xlsx")
    if frozen_template.read_bytes() != frozen_template_bytes:
        raise RuntimeError("live template change altered the active run snapshot")
    template_recompile_result, template_recompile_verification = (
        _complete_prepared_case(
            active_plugin,
            greenfield,
            greenfield_case,
            template_recompile,
            audit_root,
            audit_log,
        )
    )
    if template_recompile_verification["templateSha256"] != _sha256(
        frozen_template_bytes
    ):
        raise RuntimeError("template recompile published a drifting template")
    if _sha256(first_template.read_bytes()) == template_recompile_verification["templateSha256"]:
        raise RuntimeError("template recompile reused the previous generation template")
    if len(list((greenfield / ".ai-sow/generations").iterdir())) != 2:
        raise RuntimeError("template recompile did not create exactly one new generation")

    brownfield_current = (brownfield / ".ai-sow/current.json").read_bytes()
    brownfield_generations_root = brownfield / ".ai-sow/generations"
    brownfield_generation_files = _generation_file_digests(
        brownfield_generations_root
    )
    incomplete_after_success = _prepare_request(
        active_plugin,
        brownfield,
        brownfield_case,
        omit_prior_sow=True,
    )
    blocked_after_success = _run_orchestrator(
        active_plugin,
        brownfield,
        audit_root,
        audit_log,
        "prepare",
        "--request",
        incomplete_after_success,
        expected="BLOCKED",
    )
    if (brownfield / ".ai-sow/current.json").read_bytes() != brownfield_current:
        raise RuntimeError("BLOCKED update replaced current.json")
    after_blocked_files = _generation_file_digests(brownfield_generations_root)
    if after_blocked_files != brownfield_generation_files:
        changed = sorted(
            path
            for path in set(brownfield_generation_files) | set(after_blocked_files)
            if brownfield_generation_files.get(path) != after_blocked_files.get(path)
        )
        raise RuntimeError(
            f"BLOCKED update changed last-known-good generation files: {changed}"
        )

    forbidden_reads = (
        audit_log.read_text(encoding="utf-8").splitlines()
        if audit_log.is_file()
        else []
    )
    return {
        "pluginName": manifest["name"],
        "pluginVersion": manifest.get("version"),
        "pluginRoot": str(active_plugin),
        "workDir": str(work_dir),
        "publicSkills": ["generate"],
        "greenfieldOutcome": greenfield_result["outcome"],
        "brownfieldOutcome": brownfield_result["outcome"],
        "blockedResumeOutcome": blocked_resume_result["outcome"],
        "blockedReviewOutcome": blocked_review_result["outcome"],
        "blockedAfterSuccessOutcome": blocked_after_success["outcome"],
        "reuseOutcome": reuse["outcome"],
        "templateRecompileOutcome": template_recompile_result["outcome"],
        "templateRecompileAction": recompile_plan["action"],
        "runTemplateStable": True,
        "lastKnownGoodPreserved": True,
        "lastKnownGoodFileCount": len(brownfield_generation_files),
        "marketplaceReadCount": len(forbidden_reads),
        "projectRoots": [
            str(greenfield),
            str(brownfield),
            str(blocked_resume),
            str(blocked_review),
        ],
        "workbookPaths": [
            greenfield_verification["workbookPath"],
            template_recompile_verification["workbookPath"],
            brownfield_verification["workbookPath"],
            blocked_resume_verification["workbookPath"],
        ],
        "officeEngines": [
            greenfield_verification["officeEngine"],
            template_recompile_verification["officeEngine"],
            brownfield_verification["officeEngine"],
            blocked_resume_verification["officeEngine"],
        ],
    }


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--plugin-root",
        type=Path,
        default=Path(__file__).resolve().parents[2],
    )
    parser.add_argument("--work-dir", type=Path)
    parser.add_argument("--copy-plugin", action="store_true")
    args = parser.parse_args(argv)
    work_dir = args.work_dir or Path(tempfile.mkdtemp(prefix="ai-sow-smoke-"))
    report = run_smoke(args.plugin_root, work_dir, args.copy_plugin)
    print(json.dumps(report, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
