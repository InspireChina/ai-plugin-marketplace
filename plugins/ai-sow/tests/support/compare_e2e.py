#!/usr/bin/env python3
"""Compare two AI SOW E2E run roots without embedding machine-specific paths."""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
from collections import Counter
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


DEFECT_CHECKS = {
    "D3": "emptyEstimateRejected",
    "D4": "realOfficeRoundtrip",
    "D5": "formulaAuthorityReread",
    "D6": "verifiedTrustBoundary",
    "D7": "lastKnownGood",
    "D8": "exactReplay",
    "D9": "earlyScopeConflict",
    "D10": "reviewSubjectInventory",
    "D11": "unambiguousCounts",
    "D12": "postRenderSemanticAudit",
}


def read_json(path: Path) -> dict[str, Any]:
    value = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise ValueError(f"expected JSON object: {path}")
    return value


def sha256_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def canonical_json_sha256(value: object) -> str:
    payload = (
        json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        + "\n"
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def input_hashes(root: Path) -> dict[str, str]:
    paths = [root / "request.json"]
    inputs = root / "inputs"
    if inputs.is_dir():
        paths.extend(path for path in inputs.rglob("*") if path.is_file())
    return {
        path.relative_to(root).as_posix(): sha256_file(path)
        for path in sorted(paths)
        if path.is_file()
    }


def latest_generation(root: Path) -> tuple[dict[str, Any], Path] | None:
    current_path = root / ".ai-sow/current.json"
    if not current_path.is_file():
        return None
    current = read_json(current_path)
    generation_id = str(current["generationId"])
    generation_root = root / ".ai-sow/generations" / generation_id
    return read_json(generation_root / "manifest.json"), generation_root


def _optional_json(path: Path) -> dict[str, Any] | None:
    try:
        return read_json(path)
    except (FileNotFoundError, json.JSONDecodeError, ValueError):
        return None


def workbook_shape(path: Path) -> tuple[int, int]:
    workbook = load_workbook(path, data_only=False, read_only=False)
    try:
        return len(workbook.sheetnames), sum(
            len(worksheet.tables) for worksheet in workbook.worksheets
        )
    finally:
        workbook.close()


def workbook_formula_error_count(path: Path) -> int:
    workbook = load_workbook(path, data_only=True, read_only=False)
    try:
        return sum(
            1
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows()
            for cell in row
            if cell.data_type == "e"
            or (
                isinstance(cell.value, str)
                and cell.value.startswith(
                    ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "Err:")
                )
            )
        )
    finally:
        workbook.close()


def workbook_effort(path: Path) -> dict[str, float | None]:
    workbook = load_workbook(path, data_only=True, read_only=False)
    try:
        for worksheet in workbook.worksheets:
            for table in worksheet.tables.values():
                if table.displayName not in {
                    "ProjectSummaryTable",
                    "ProjectSummary",
                }:
                    continue
                min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                headers = [
                    worksheet.cell(min_row, column).value
                    for column in range(min_col, max_col + 1)
                ]
                rows = [
                    {
                        str(headers[offset]): worksheet.cell(row, min_col + offset).value
                        for offset in range(len(headers))
                    }
                    for row in range(min_row + 1, max_row + 1)
                ]
                labels = {
                    str(row.get("工作量项") or row.get("估算项")): row.get("人天")
                    for row in rows
                }
                if labels:
                    return {
                        "directDays": _number(
                            labels.get("直接开发人天")
                            or labels.get("直接开发工作量")
                        ),
                        "sitDays": _number(
                            labels.get("SIT支持人天") or labels.get("SIT支持")
                        ),
                        "uatDays": _number(
                            labels.get("UAT支持人天") or labels.get("UAT支持")
                        ),
                        "totalDays": _number(
                            labels.get("项目总人天")
                            or labels.get("总开发人天")
                            or labels.get("总计")
                        ),
                    }
        return {
            "directDays": None,
            "sitDays": None,
            "uatDays": None,
            "totalDays": None,
        }
    finally:
        workbook.close()


def workbook_story_days(path: Path) -> list[float]:
    workbook = load_workbook(path, data_only=True, read_only=False)
    try:
        for worksheet in workbook.worksheets:
            for table in worksheet.tables.values():
                if table.displayName != "SOWStoryTable":
                    continue
                min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                headers = [
                    worksheet.cell(min_row, column).value
                    for column in range(min_col, max_col + 1)
                ]
                if "故事人天" not in headers:
                    return []
                day_column = min_col + headers.index("故事人天")
                return [
                    float(value)
                    for row in range(min_row + 1, max_row + 1)
                    if isinstance(
                        value := worksheet.cell(row, day_column).value,
                        (int, float),
                    )
                    and not isinstance(value, bool)
                ]
        return []
    finally:
        workbook.close()


def workbook_packaging_invariant(path: Path) -> bool:
    formula_workbook = load_workbook(path, data_only=False, read_only=False)
    value_workbook = load_workbook(path, data_only=True, read_only=False)
    try:
        task_values: list[float] | None = None
        summary_formulas: dict[str, object] = {}
        summary_values: dict[str, object] = {}
        for workbook, summary in (
            (formula_workbook, summary_formulas),
            (value_workbook, summary_values),
        ):
            for worksheet in workbook.worksheets:
                for table in worksheet.tables.values():
                    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                    headers = [
                        worksheet.cell(min_row, column).value
                        for column in range(min_col, max_col + 1)
                    ]
                    if table.displayName == "TaskTable" and workbook is value_workbook:
                        if "任务人天" not in headers:
                            return False
                        day_column = min_col + headers.index("任务人天")
                        task_values = [
                            float(value)
                            for row in range(min_row + 1, max_row + 1)
                            if isinstance(
                                value := worksheet.cell(row, day_column).value,
                                (int, float),
                            )
                            and not isinstance(value, bool)
                        ]
                    if table.displayName == "ProjectSummaryTable":
                        if "工作量项" not in headers or "人天" not in headers:
                            return False
                        label_column = min_col + headers.index("工作量项")
                        value_column = min_col + headers.index("人天")
                        for row in range(min_row + 1, max_row + 1):
                            label = worksheet.cell(row, label_column).value
                            if isinstance(label, str):
                                summary[label] = worksheet.cell(row, value_column).value
        direct_formula = summary_formulas.get("直接开发人天")
        uat_formula = summary_formulas.get("UAT支持人天")
        direct = _number(summary_values.get("直接开发人天"))
        sit = _number(summary_values.get("SIT支持人天"))
        uat = _number(summary_values.get("UAT支持人天"))
        total = _number(summary_values.get("总开发人天"))
        return (
            task_values is not None
            and bool(task_values)
            and isinstance(direct_formula, str)
            and "02-任务清单" in direct_formula
            and "故事人天" not in direct_formula
            and isinstance(uat_formula, str)
            and "TaskTable[任务人天]" in uat_formula
            and "SOWStoryTable[UAT适用]" in uat_formula
            and "SOWStoryTable[故事]" in uat_formula
            and "SOWStoryTable[故事人天]" not in uat_formula
            and direct is not None
            and sit is not None
            and uat is not None
            and total is not None
            and math.isclose(direct, sum(task_values), abs_tol=1e-9)
            and math.isclose(total, direct + sit + uat, abs_tol=1e-9)
        )
    finally:
        formula_workbook.close()
        value_workbook.close()


def story_granularity(
    delivery: Mapping[str, Any], story_days: list[float]
) -> dict[str, float | int | None]:
    stories = [
        item for item in delivery.get("stories", []) if isinstance(item, Mapping)
    ]
    task_counts = {str(story.get("storyId")): 0 for story in stories}
    for task in delivery.get("tasks", []):
        if isinstance(task, Mapping):
            story_id = str(task.get("storyId"))
            task_counts[story_id] = task_counts.get(story_id, 0) + 1
    feature_link_counts = [
        1
        if isinstance(story.get("featureId"), str)
        else len(story.get("featureIds", []))
        if isinstance(story.get("featureIds"), list)
        else 0
        for story in stories
    ]
    return {
        "maxFeatureLinksPerStory": max(feature_link_counts, default=0),
        "maxTasksPerStory": max(task_counts.values(), default=0),
        "maxStoryDays": max(story_days, default=None),
        "averageStoryDays": (
            round(sum(story_days) / len(story_days), 2) if story_days else None
        ),
    }


def _number(value: object) -> float | None:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None
    return float(value)


def baseline_timing(root: Path) -> int | None:
    report_path = root / "e2e-report.md"
    if not report_path.is_file():
        return None
    report = report_path.read_text(encoding="utf-8")
    match = re.search(r"实测\s*(\d+)\s*分\s*(\d+)\s*秒", report)
    return int(match.group(1)) * 60 + int(match.group(2)) if match else None


def reported_effort(report: str) -> dict[str, float] | None:
    """Read the numeric sentence used by the legacy 12-Sheet E2E report."""
    patterns = {
        "directDays": r"直接开发\s*([0-9]+(?:\.[0-9]+)?)\s*人天",
        "sitDays": r"SIT\s*支持\s*([0-9]+(?:\.[0-9]+)?)\s*人天",
        "uatDays": r"UAT\s*支持\s*([0-9]+(?:\.[0-9]+)?)\s*人天",
        "totalDays": r"总计\s*([0-9]+(?:\.[0-9]+)?)\s*人天",
    }
    matches = {key: re.search(pattern, report) for key, pattern in patterns.items()}
    if not all(matches.values()):
        return None
    return {
        key: float(match.group(1))
        for key, match in matches.items()
        if match is not None
    }


def _object_ids(
    bundle: Mapping[str, Any], fields: tuple[tuple[str, str], ...]
) -> set[str]:
    return {
        str(item[id_field])
        for collection, id_field in fields
        for item in bundle.get(collection, [])
        if isinstance(item, Mapping) and isinstance(item.get(id_field), str)
    }


def _review_subjects_valid(
    scope: Mapping[str, Any],
    delivery: Mapping[str, Any],
    review: Mapping[str, Any],
) -> bool:
    assumption_ids = _object_ids(scope, (("assumptions", "assumptionId"),))
    responsibility_ids = _object_ids(
        scope, (("responsibilityBoundaries", "responsibilityBoundaryId"),)
    )
    exclusion_ids = {
        str(item["featureId"])
        for item in scope.get("features", [])
        if isinstance(item, Mapping)
        and isinstance(item.get("featureId"), str)
        and isinstance(item.get("scopeDecision"), Mapping)
        and item["scopeDecision"].get("decision") != "IN_SCOPE"
    }
    design_task_ids = {
        str(item["taskId"])
        for item in delivery.get("tasks", [])
        if isinstance(item, Mapping)
        and item.get("taskKind") == "DESIGN"
        and isinstance(item.get("taskId"), str)
    }
    estimate_ids: set[str] = set()
    trigger_ids: set[str] = set()
    for collection, id_field in (
        ("commitments", "commitmentId"),
        ("effectiveStartItems", "effectiveStartItemId"),
        ("integrations", "integrationId"),
        ("nfrs", "nfrId"),
        ("assumptions", "assumptionId"),
    ):
        for item in scope.get(collection, []):
            if not isinstance(item, Mapping) or not isinstance(item.get(id_field), str):
                continue
            identifier = str(item[id_field])
            if isinstance(item.get("estimateBoundary"), str) and item[
                "estimateBoundary"
            ].strip():
                estimate_ids.add(identifier)
            if (
                isinstance(item.get("changeTrigger"), str)
                and item["changeTrigger"].strip()
            ) or (
                isinstance(item.get("trigger"), str) and item["trigger"].strip()
            ):
                trigger_ids.add(identifier)
    allowed = {
        "ASSUMPTION": assumption_ids,
        "RESPONSIBILITY": responsibility_ids,
        "EXCLUSION": exclusion_ids,
        "DESIGN_TASK": design_task_ids,
        "ESTIMATE_BOUNDARY": estimate_ids,
        "CHANGE_TRIGGER": trigger_ids,
    }
    for note in review.get("notes", []):
        if not isinstance(note, Mapping):
            return False
        subjects = note.get("subjectIds")
        category = note.get("category")
        if (
            not isinstance(subjects, list)
            or not subjects
            or not isinstance(category, str)
            or not set(subjects) <= allowed.get(category, set())
        ):
            return False
    all_ids = _object_ids(
        scope,
        (
            ("epics", "epicId"),
            ("features", "featureId"),
            ("commitments", "commitmentId"),
            ("effectiveStartItems", "effectiveStartItemId"),
            ("designItems", "designItemId"),
            ("designDecisions", "designDecisionId"),
            ("integrations", "integrationId"),
            ("nfrs", "nfrId"),
            ("assumptions", "assumptionId"),
            ("responsibilityBoundaries", "responsibilityBoundaryId"),
        ),
    ) | _object_ids(
        delivery,
        (
            ("stories", "storyId"),
            ("acceptanceCriteria", "acceptanceCriterionId"),
            ("tasks", "taskId"),
            ("dependencies", "dependencyId"),
        ),
    )
    return all(
        isinstance(question, Mapping)
        and isinstance(question.get("subjectIds"), list)
        and bool(question["subjectIds"])
        and set(question["subjectIds"]) <= all_ids
        for question in review.get("questions", [])
    )


def _blocked_review_shape_valid(review: Mapping[str, Any]) -> bool:
    if set(review) != {
        "contract",
        "runId",
        "inputRevisionId",
        "scopeSha256",
        "deliverySha256",
        "packetSha256",
        "decision",
        "notes",
        "questions",
    }:
        return False
    if (
        review.get("contract") != "ai-sow-final-review-v1"
        or review.get("decision") != "BLOCKED"
        or not isinstance(review.get("runId"), str)
        or not re.fullmatch(r"[a-z0-9]+(?:-[a-z0-9]+)*", review["runId"])
        or not isinstance(review.get("inputRevisionId"), str)
        or not re.fullmatch(r"[0-9]{6}", review["inputRevisionId"])
        or any(
            not isinstance(review.get(field), str)
            or not re.fullmatch(r"[0-9a-f]{64}", review[field])
            for field in ("scopeSha256", "deliverySha256", "packetSha256")
        )
        or not isinstance(review.get("notes"), list)
        or not isinstance(review.get("questions"), list)
        or not review["questions"]
    ):
        return False
    return all(
        isinstance(question, Mapping)
        and set(question)
        == {"blockingConditionId", "subjectIds", "summary", "question"}
        and isinstance(question.get("blockingConditionId"), str)
        and bool(
            re.fullmatch(
                r"[a-z0-9]+(?:-[a-z0-9]+)*", question["blockingConditionId"]
            )
        )
        and isinstance(question.get("subjectIds"), list)
        and bool(question["subjectIds"])
        and len(question["subjectIds"]) == len(set(question["subjectIds"]))
        and all(isinstance(subject, str) and subject for subject in question["subjectIds"])
        and isinstance(question.get("summary"), str)
        and bool(question["summary"].strip())
        and isinstance(question.get("question"), str)
        and bool(question["question"].strip())
        for question in review["questions"]
    )


def _blocked_review_evidence_valid(root: Path) -> bool:
    evidence_root = root / "e2e-evidence"
    packet_path = evidence_root / "blocked-review-packet.json"
    review_path = evidence_root / "blocked-final-review.json"
    result_path = evidence_root / "blocked-result.json"
    packet = _optional_json(packet_path)
    review = _optional_json(review_path)
    result = _optional_json(result_path)
    if (
        packet is None
        or review is None
        or result is None
        or packet.get("contract") != "ai-sow-final-review-packet-v1"
        or not _blocked_review_shape_valid(review)
    ):
        return False
    artifacts = packet.get("artifacts")
    bundles = packet.get("bundles")
    scope_artifact = artifacts.get("scope") if isinstance(artifacts, Mapping) else None
    delivery_artifact = (
        artifacts.get("delivery") if isinstance(artifacts, Mapping) else None
    )
    scope = bundles.get("scope") if isinstance(bundles, Mapping) else None
    delivery = bundles.get("delivery") if isinstance(bundles, Mapping) else None
    questions = review.get("questions")
    if (
        review.get("packetSha256") != sha256_file(packet_path)
        or review.get("runId") != packet.get("runId")
        or review.get("inputRevisionId") != packet.get("inputRevisionId")
        or not isinstance(scope_artifact, Mapping)
        or review.get("scopeSha256") != scope_artifact.get("sha256")
        or not isinstance(delivery_artifact, Mapping)
        or review.get("deliverySha256") != delivery_artifact.get("sha256")
        or not isinstance(scope, Mapping)
        or not isinstance(delivery, Mapping)
        or not _review_subjects_valid(scope, delivery, review)
        or result.get("outcome") != "BLOCKED"
        or result.get("diagnostics") != []
        or result.get("reviewSha256") != canonical_json_sha256(review)
        or result.get("questions")
        != [question["question"] for question in questions]
    ):
        return False
    return True


def _replay_evidence_valid(root: Path, manifest: Mapping[str, Any]) -> bool:
    evidence = _optional_json(root / "e2e-evidence/replay-evidence.json")
    if evidence is None or set(evidence) != {
        "contract",
        "inputSha256",
        "resultPath",
        "resultSha256",
        "currentPath",
        "currentSha256",
        "generationManifestPath",
        "generationManifestSha256",
    }:
        return False
    expected_paths = {
        "resultPath": "e2e-evidence/replay-result.json",
        "currentPath": ".ai-sow/current.json",
        "generationManifestPath": (
            f".ai-sow/generations/{manifest.get('generationId')}/manifest.json"
        ),
    }
    if (
        evidence.get("contract") != "ai-sow-e2e-replay-evidence-v1"
        or evidence.get("inputSha256") != input_hashes(root)
        or any(evidence.get(key) != value for key, value in expected_paths.items())
    ):
        return False
    result_path = root / expected_paths["resultPath"]
    current_path = root / expected_paths["currentPath"]
    manifest_path = root / expected_paths["generationManifestPath"]
    if any(
        not isinstance(evidence.get(field), str)
        or not re.fullmatch(r"[0-9a-f]{64}", evidence[field])
        for field in (
            "resultSha256",
            "currentSha256",
            "generationManifestSha256",
        )
    ):
        return False
    if (
        not result_path.is_file()
        or evidence.get("resultSha256") != sha256_file(result_path)
        or not current_path.is_file()
        or evidence.get("currentSha256") != sha256_file(current_path)
        or not manifest_path.is_file()
        or evidence.get("generationManifestSha256") != sha256_file(manifest_path)
    ):
        return False
    result = _optional_json(result_path)
    current = _optional_json(current_path)
    evidence_manifest = _optional_json(manifest_path)
    if result is None or current is None or evidence_manifest is None:
        return False
    return (
        evidence_manifest == manifest
        and current.get("generationId") == manifest.get("generationId")
        and current.get("revisionId") == manifest.get("revisionId")
        and current.get("generationManifestPath")
        == expected_paths["generationManifestPath"]
        and current.get("generationManifestSha256")
        == evidence.get("generationManifestSha256")
        and result.get("outcome") == "REUSED"
        and result.get("diagnostics") == []
        and result.get("generationId") == manifest.get("generationId")
        and result.get("revisionId") == manifest.get("revisionId")
        and result.get("workbookPath") == manifest.get("workbookPath")
    )


def derive_workflow_evidence(
    root: Path, manifest: Mapping[str, Any]
) -> dict[str, int]:
    published = len(list((root / ".ai-sow/generations").glob("*/manifest.json")))
    return {
        "published": published,
        "blocked": 1 if _blocked_review_evidence_valid(root) else 0,
        "reused": 1 if _replay_evidence_valid(root, manifest) else 0,
    }


def _bundle_objects(
    bundle: Mapping[str, Any] | None, collection: str, id_field: str
) -> dict[str, Mapping[str, Any]]:
    if bundle is None:
        return {}
    return {
        str(item[id_field]): item
        for item in bundle.get(collection, [])
        if isinstance(item, Mapping) and isinstance(item.get(id_field), str)
    }


def _expected_collection_counts(
    previous: Mapping[str, Any] | None,
    current: Mapping[str, Any],
    collection: str,
    id_field: str,
    affected_before_ids: set[str],
) -> dict[str, int]:
    before = _bundle_objects(previous, collection, id_field)
    after = _bundle_objects(current, collection, id_field)
    reusable = {
        object_id
        for object_id in set(before) & set(after)
        if object_id not in affected_before_ids and before[object_id] == after[object_id]
    }
    return {
        "affected": len(set(before) & affected_before_ids),
        "recomputed": len(after) - len(reusable),
        "reused": len(reusable),
        "deleted": len(set(before) - set(after)),
        "final": len(after),
    }


def expected_change_counts(
    root: Path,
    manifest: Mapping[str, Any],
    scope: Mapping[str, Any],
    delivery: Mapping[str, Any],
) -> dict[str, dict[str, int]]:
    impact = manifest.get("impact")
    if not isinstance(impact, Mapping):
        impact = {}
    baseline_id = impact.get("baselineGenerationId")
    previous_scope: Mapping[str, Any] | None = None
    previous_delivery: Mapping[str, Any] | None = None
    if isinstance(baseline_id, str):
        baseline_root = root / ".ai-sow/generations" / baseline_id / "data"
        previous_scope = _optional_json(baseline_root / "scope.json")
        previous_delivery = _optional_json(baseline_root / "delivery.json")
    affected_features = {
        str(feature_id)
        for feature_id in impact.get("affectedFeatureIds", [])
        if isinstance(feature_id, str)
    }
    previous_stories = _bundle_objects(previous_delivery, "stories", "storyId")
    affected_stories = {
        story_id
        for story_id, story in previous_stories.items()
        if story.get("featureId") in affected_features
    }
    affected_criteria = {
        criterion_id
        for criterion_id, criterion in _bundle_objects(
            previous_delivery, "acceptanceCriteria", "acceptanceCriterionId"
        ).items()
        if criterion.get("storyId") in affected_stories
    }
    affected_tasks = {
        task_id
        for task_id, task in _bundle_objects(
            previous_delivery, "tasks", "taskId"
        ).items()
        if task.get("storyId") in affected_stories
    }
    return {
        "features": _expected_collection_counts(
            previous_scope, scope, "features", "featureId", affected_features
        ),
        "stories": _expected_collection_counts(
            previous_delivery, delivery, "stories", "storyId", affected_stories
        ),
        "acceptanceCriteria": _expected_collection_counts(
            previous_delivery,
            delivery,
            "acceptanceCriteria",
            "acceptanceCriterionId",
            affected_criteria,
        ),
        "tasks": _expected_collection_counts(
            previous_delivery, delivery, "tasks", "taskId", affected_tasks
        ),
    }


def derive_manifest_checks(
    manifest: Mapping[str, Any],
    scope: Mapping[str, Any],
    delivery: Mapping[str, Any],
    workbook: Mapping[str, Any],
    effort: Mapping[str, Any],
    *,
    scope_sha256: str,
    delivery_sha256: str,
    workbook_sha256: str,
    expected_change_counts: Mapping[str, Mapping[str, int]],
) -> dict[str, bool]:
    verification = manifest.get("workbookVerification")
    if not isinstance(verification, Mapping):
        verification = {}
    engine = verification.get("engine")
    real_office = (
        isinstance(engine, Mapping)
        and engine.get("name") == "LibreOffice"
        and isinstance(engine.get("version"), str)
        and bool(engine["version"].strip())
    )
    workbook_hash_bound = manifest.get("workbookSha256") == workbook_sha256
    verified = verification.get("trustState") == "VERIFIED"
    formula_reread = (
        real_office
        and verified
        and verification.get("formulaErrors") == []
        and workbook.get("formulaErrors") == 0
        and workbook.get("packagingInvariant") is True
        and workbook_hash_bound
    )

    review = manifest.get("finalReview")
    review_valid = False
    if isinstance(review, Mapping):
        decision = review.get("decision")
        notes = review.get("notes")
        questions = review.get("questions")
        decision_shape = (
            decision == "PASS" and notes == [] and questions == []
        ) or (
            decision == "PASS_WITH_NOTES"
            and isinstance(notes, list)
            and bool(notes)
            and questions == []
        )
        review_valid = (
            decision_shape
            and review.get("scopeSha256") == scope_sha256
            and review.get("deliverySha256") == delivery_sha256
            and manifest.get("finalReviewSha256") == canonical_json_sha256(review)
            and _review_subjects_valid(scope, delivery, review)
        )

    expected_counts = {
        "features": len(scope.get("features", [])),
        "stories": len(delivery.get("stories", [])),
        "acceptanceCriteria": len(delivery.get("acceptanceCriteria", [])),
        "tasks": len(delivery.get("tasks", [])),
    }
    counts = manifest.get("changeCounts")
    counts_valid = isinstance(counts, Mapping) and counts == expected_change_counts
    effort_fields = ("directDays", "sitDays", "uatDays", "totalDays")
    effort_valid = all(
        isinstance(effort.get(field), (int, float))
        and not isinstance(effort.get(field), bool)
        and isinstance(verification.get(field), (int, float))
        and not isinstance(verification.get(field), bool)
        and math.isclose(
            float(effort[field]), float(verification[field]), abs_tol=1e-9
        )
        for field in effort_fields
    )
    semantic_audit = (
        formula_reread
        and verification.get("storyCount") == expected_counts["stories"]
        and verification.get("taskCount") == expected_counts["tasks"]
        and effort_valid
    )
    return {
        "realOfficeRoundtrip": real_office and workbook_hash_bound,
        "formulaAuthorityReread": formula_reread,
        "verifiedTrustBoundary": verified and workbook_hash_bound,
        "reviewSubjectInventory": review_valid,
        "unambiguousCounts": counts_valid,
        "postRenderSemanticAudit": semantic_audit,
    }


def derive_summary(root: Path) -> dict[str, Any]:
    published = latest_generation(root)
    if published is None:
        raise ValueError(f"run has no summary or current generation: {root}")
    manifest, generation_root = published
    scope = read_json(generation_root / "data/scope.json")
    delivery = read_json(generation_root / "data/delivery.json")
    workbook_path = generation_root / "output/sow.xlsx"
    sheet_count, table_count = workbook_shape(workbook_path)
    report = (
        (root / "e2e-report.md").read_text(encoding="utf-8")
        if (root / "e2e-report.md").is_file()
        else ""
    )
    effort = workbook_effort(workbook_path)
    if any(value is None for value in effort.values()):
        effort = reported_effort(report) or effort
    workbook_summary = {
        "sheetCount": sheet_count,
        "tableCount": table_count,
        "formulaErrors": workbook_formula_error_count(workbook_path),
        "packagingInvariant": workbook_packaging_invariant(workbook_path),
    }
    return {
        "inputSha256": input_hashes(root),
        "workflow": derive_workflow_evidence(root, manifest),
        "objects": {
            "features": len(scope.get("features", [])),
            "stories": len(delivery.get("stories", [])),
            "acceptanceCriteria": len(delivery.get("acceptanceCriteria", [])),
            "tasks": len(delivery.get("tasks", [])),
            "integrations": len(scope.get("integrations", [])),
            "nfrs": len(scope.get("nfrs", [])),
        },
        "granularity": story_granularity(
            delivery, workbook_story_days(workbook_path)
        ),
        "effort": effort,
        "workbook": workbook_summary,
        "timingSeconds": baseline_timing(root),
        "checks": {},
        "scope": scope,
        "delivery": delivery,
        "manifest": manifest,
        "derivedChecks": derive_manifest_checks(
            manifest,
            scope,
            delivery,
            workbook_summary,
            effort,
            scope_sha256=sha256_file(generation_root / "data/scope.json"),
            delivery_sha256=sha256_file(generation_root / "data/delivery.json"),
            workbook_sha256=sha256_file(workbook_path),
            expected_change_counts=expected_change_counts(
                root, manifest, scope, delivery
            ),
        ),
    }


def load_summary(root: Path) -> dict[str, Any]:
    summary_path = root / "e2e-run-summary.json"
    reported = read_json(summary_path) if summary_path.is_file() else {}
    if latest_generation(root) is None:
        return reported
    summary = derive_summary(root)
    for observational_key in ("timingSeconds", "checks"):
        if observational_key in reported:
            summary[observational_key] = reported[observational_key]
    return summary


def change_map(before: Mapping[str, object], after: Mapping[str, object]) -> dict[str, object]:
    keys = sorted(set(before) | set(after))
    return {
        "before": dict(before),
        "after": dict(after),
        "changed": [key for key in keys if before.get(key) != after.get(key)],
    }


def defect_dispositions(current: Mapping[str, Any]) -> dict[str, dict[str, str]]:
    workbook = current.get("workbook", {})
    effort = current.get("effort", {})
    checks = current.get("checks", {})
    derived = current.get("derivedChecks", {})
    workflow = current.get("workflow", {})
    four_sheet = workbook.get("sheetCount") == 4 and workbook.get("tableCount") == 5
    sit_valid = isinstance(effort.get("sitDays"), (int, float)) and effort["sitDays"] > 0
    result = {
        "D1": {
            "status": "COVERED_BY_TEMPLATE" if four_sheet and sit_valid else "REGRESSED",
            "evidence": "四 Sheet 模板由 Task 行工作方式直接计算 SIT。",
        },
        "D2": {
            "status": "COVERED_BY_TEMPLATE" if four_sheet else "REGRESSED",
            "evidence": "简化模板不再投影易误导的 Integration 方向列。",
        },
    }
    for defect, check in DEFECT_CHECKS.items():
        if isinstance(derived, Mapping) and check in derived:
            passed = derived.get(check) is True
            evidence = f"derived:{check}"
        elif (
            check == "exactReplay"
            and isinstance(workflow, Mapping)
            and isinstance(workflow.get("reused"), int)
            and isinstance(workflow.get("published"), int)
        ):
            passed = workflow["reused"] >= 1 and workflow["published"] >= 1
            evidence = "derived:workflow.reused/published"
        elif (
            check == "earlyScopeConflict"
            and isinstance(workflow, Mapping)
            and isinstance(workflow.get("blocked"), int)
            and isinstance(workflow.get("published"), int)
        ):
            passed = workflow["blocked"] >= 1 and workflow["published"] >= 1
            evidence = "derived:workflow.blocked/published"
        else:
            passed = isinstance(checks, Mapping) and checks.get(check) is True
            evidence = f"reported:{check}"
        result[defect] = {
            "status": "FIXED" if passed else "REGRESSED",
            "evidence": evidence,
        }
    return result


def task_rows(summary: Mapping[str, Any]) -> Counter[tuple[str, str, str, str, str]]:
    delivery = summary.get("delivery", {})
    if not isinstance(delivery, Mapping):
        return Counter()
    stories = {
        str(item.get("storyId")): str(item.get("name", ""))
        for item in delivery.get("stories", [])
        if isinstance(item, Mapping)
    }
    return Counter(
        (
            stories.get(str(task.get("storyId")), str(task.get("storyId", ""))),
            str(task.get("name", "")),
            str(task.get("baseUnit", task.get("baseUnitId", ""))),
            str(task.get("workMode", "")),
            str(task.get("complexity", "")),
        )
        for task in delivery.get("tasks", [])
        if isinstance(task, Mapping)
    )


def task_semantic_rows(
    summary: Mapping[str, Any],
) -> Counter[tuple[str, str, str, str]]:
    delivery = summary.get("delivery", {})
    if not isinstance(delivery, Mapping):
        return Counter()
    return Counter(
        (
            str(task.get("name", "")),
            str(task.get("baseUnit", task.get("baseUnitId", ""))),
            str(task.get("workMode", "")),
            str(task.get("complexity", "")),
        )
        for task in delivery.get("tasks", [])
        if isinstance(task, Mapping)
    )


SEMANTIC_FIELDS = {
    "requirements": ("requirement",),
    "subRequirements": ("requirement", "subRequirement", "scopeDecision"),
    "stories": ("subRequirement", "story"),
    "acceptanceCriteria": ("story", "acceptanceCriterion"),
    "tasks": ("story", "task", "baseUnit", "workMode", "complexity"),
}


def _semantic_rows(
    summary: Mapping[str, Any], collection: str
) -> Counter[tuple[str, ...]]:
    scope = summary.get("scope", {})
    delivery = summary.get("delivery", {})
    if not isinstance(scope, Mapping):
        scope = {}
    if not isinstance(delivery, Mapping):
        delivery = {}
    epics = {
        str(item.get("epicId")): str(item.get("name", ""))
        for item in scope.get("epics", [])
        if isinstance(item, Mapping)
    }
    features = {
        str(item.get("featureId")): item
        for item in scope.get("features", [])
        if isinstance(item, Mapping)
    }
    stories = {
        str(item.get("storyId")): item
        for item in delivery.get("stories", [])
        if isinstance(item, Mapping)
    }

    def feature_names(story: Mapping[str, Any]) -> str:
        feature_ids = (
            [story.get("featureId")]
            if isinstance(story.get("featureId"), str)
            else story.get("featureIds", [])
        )
        if not isinstance(feature_ids, list):
            return ""
        return " / ".join(
            sorted(
                str(features.get(str(feature_id), {}).get("name", feature_id))
                for feature_id in feature_ids
            )
        )

    if collection == "requirements":
        return Counter((name,) for name in epics.values())
    if collection == "subRequirements":
        rows: list[tuple[str, ...]] = []
        for feature in features.values():
            decision = feature.get("scopeDecision")
            decision_value = (
                decision.get("decision", "")
                if isinstance(decision, Mapping)
                else decision if isinstance(decision, str) else ""
            )
            rows.append(
                (
                    epics.get(str(feature.get("epicId")), ""),
                    str(feature.get("name", "")),
                    str(decision_value),
                )
            )
        return Counter(rows)
    if collection == "stories":
        return Counter(
            (feature_names(story), str(story.get("name", "")))
            for story in stories.values()
        )
    if collection == "acceptanceCriteria":
        return Counter(
            (
                str(stories.get(str(item.get("storyId")), {}).get("name", "")),
                str(item.get("name", "")),
            )
            for item in delivery.get("acceptanceCriteria", [])
            if isinstance(item, Mapping)
        )
    if collection == "tasks":
        return Counter(
            (
                str(stories.get(str(task.get("storyId")), {}).get("name", "")),
                str(task.get("name", "")),
                str(task.get("baseUnit", task.get("baseUnitId", ""))),
                str(task.get("workMode", "")),
                str(task.get("complexity", "")),
            )
            for task in delivery.get("tasks", [])
            if isinstance(task, Mapping)
        )
    raise ValueError(f"unknown semantic collection: {collection}")


def _semantic_changes(
    baseline: Mapping[str, Any], current: Mapping[str, Any]
) -> dict[str, dict[str, list[dict[str, str]]]]:
    result: dict[str, dict[str, list[dict[str, str]]]] = {}
    for collection, fields in SEMANTIC_FIELDS.items():
        before = _semantic_rows(baseline, collection)
        after = _semantic_rows(current, collection)
        result[collection] = {
            "removed": [
                dict(zip(fields, row, strict=True))
                for row in sorted((before - after).elements())
            ],
            "added": [
                dict(zip(fields, row, strict=True))
                for row in sorted((after - before).elements())
            ],
        }
    return result


def _source_refs(value: Mapping[str, Any]) -> list[dict[str, str]]:
    return [
        {
            key: str(ref[key])
            for key in ("sourceId", "anchorId", "locator")
            if isinstance(ref.get(key), str)
        }
        for ref in value.get("sourceRefs", [])
        if isinstance(ref, Mapping)
        and isinstance(ref.get("sourceId"), str)
        and isinstance(ref.get("anchorId"), str)
    ]


def _semantic_records(
    summary: Mapping[str, Any], collection: str
) -> dict[tuple[str, ...], list[dict[str, Any]]]:
    scope = summary.get("scope", {})
    delivery = summary.get("delivery", {})
    if not isinstance(scope, Mapping):
        scope = {}
    if not isinstance(delivery, Mapping):
        delivery = {}
    epics = {
        str(item.get("epicId")): item
        for item in scope.get("epics", [])
        if isinstance(item, Mapping)
    }
    features = {
        str(item.get("featureId")): item
        for item in scope.get("features", [])
        if isinstance(item, Mapping)
    }
    stories = {
        str(item.get("storyId")): item
        for item in delivery.get("stories", [])
        if isinstance(item, Mapping)
    }
    records: dict[tuple[str, ...], list[dict[str, Any]]] = {}

    def add(row: tuple[str, ...], evidence: dict[str, Any]) -> None:
        records.setdefault(row, []).append(evidence)

    if collection == "requirements":
        for epic in epics.values():
            add(
                (str(epic.get("name", "")),),
                {
                    "objectId": str(epic.get("epicId", "")),
                    "sourceRefs": _source_refs(epic),
                    "classification": "AUTHORING_STANDARD",
                    "reason": "按完整业务域或长期技术能力域重新归并需求层级。",
                },
            )
    elif collection == "subRequirements":
        for feature in features.values():
            decision = feature.get("scopeDecision")
            decision_value = (
                decision.get("decision", "")
                if isinstance(decision, Mapping)
                else decision if isinstance(decision, str) else ""
            )
            add(
                (
                    str(epics.get(str(feature.get("epicId")), {}).get("name", "")),
                    str(feature.get("name", "")),
                    str(decision_value),
                ),
                {
                    "objectId": str(feature.get("featureId", "")),
                    "sourceRefs": _source_refs(feature),
                    "classification": (
                        "EVIDENCE_REINTERPRETATION"
                        if decision_value != "IN_SCOPE"
                        else "AUTHORING_STANDARD"
                    ),
                    "reason": "按可独立纳入、排除和评审的功能边界重新拆分子需求。",
                },
            )
    elif collection == "stories":
        for story in stories.values():
            feature = features.get(str(story.get("featureId")), {})
            add(
                (
                    str(feature.get("name", story.get("featureId", ""))),
                    str(story.get("name", "")),
                ),
                {
                    "objectId": str(story.get("storyId", "")),
                    "sourceRefs": _source_refs(feature),
                    "classification": "AUTHORING_STANDARD",
                    "reason": "按单一、可交付且可独立验收的业务或技术结果重新拆分 Story。",
                },
            )
    elif collection == "acceptanceCriteria":
        for criterion in delivery.get("acceptanceCriteria", []):
            if not isinstance(criterion, Mapping):
                continue
            story = stories.get(str(criterion.get("storyId")), {})
            feature = features.get(str(story.get("featureId")), {})
            add(
                (
                    str(story.get("name", "")),
                    str(criterion.get("name", "")),
                ),
                {
                    "objectId": str(criterion.get("acceptanceCriterionId", "")),
                    "sourceRefs": _source_refs(feature),
                    "classification": "AUTHORING_STANDARD",
                    "reason": "按同一 Story 的可观察正常流、异常或边界结果重新编写 AC。",
                },
            )
    elif collection == "tasks":
        for task in delivery.get("tasks", []):
            if not isinstance(task, Mapping):
                continue
            story = stories.get(str(task.get("storyId")), {})
            feature = features.get(str(story.get("featureId")), {})
            add(
                (
                    str(story.get("name", "")),
                    str(task.get("name", "")),
                    str(task.get("baseUnit", task.get("baseUnitId", ""))),
                    str(task.get("workMode", "")),
                    str(task.get("complexity", "")),
                ),
                {
                    "objectId": str(task.get("taskId", "")),
                    "sourceRefs": _source_refs(feature),
                    "baseUnit": str(
                        task.get("baseUnit", task.get("baseUnitId", ""))
                    ),
                    "rationale": str(task.get("rationale", "")),
                    "classification": "EVIDENCE_REINTERPRETATION",
                    "reason": "依据模板基础单元计数口径与当前有效起点重新拆分 Task。",
                },
            )
    else:
        raise ValueError(f"unknown semantic collection: {collection}")
    return records


def _semantic_evidence(
    baseline: Mapping[str, Any], current: Mapping[str, Any]
) -> list[dict[str, Any]]:
    evidence: list[dict[str, Any]] = []
    for collection in SEMANTIC_FIELDS:
        before = _semantic_rows(baseline, collection)
        after = _semantic_rows(current, collection)
        before_records = _semantic_records(baseline, collection)
        after_records = _semantic_records(current, collection)
        for change, rows, records in (
            ("REMOVED", sorted((before - after).elements()), before_records),
            ("ADDED", sorted((after - before).elements()), after_records),
        ):
            for row in rows:
                record = records[row].pop(0)
                evidence.append(
                    {
                        "collection": collection,
                        "change": change,
                        **record,
                    }
                )
    return evidence


def compare_runs(baseline_root: Path, current_root: Path) -> dict[str, object]:
    baseline = load_summary(baseline_root.resolve())
    current = load_summary(current_root.resolve())
    before_inputs = baseline.get("inputSha256", {})
    after_inputs = current.get("inputSha256", {})
    if not isinstance(before_inputs, Mapping) or not isinstance(after_inputs, Mapping):
        raise ValueError("inputSha256 must be an object")
    input_keys = sorted(set(before_inputs) | set(after_inputs))
    changed_inputs = [
        key for key in input_keys if before_inputs.get(key) != after_inputs.get(key)
    ]
    defects = defect_dispositions(current)
    regressions = [
        defect
        for defect, finding in defects.items()
        if finding["status"] in {"STILL_OPEN", "REGRESSED"}
    ]
    before_tasks = task_rows(baseline)
    after_tasks = task_rows(current)
    before_task_semantics = task_semantic_rows(baseline)
    after_task_semantics = task_semantic_rows(current)
    granularity = current.get("granularity", {})
    current_workbook = current.get("workbook", {})
    acceptance = {
        "singleFeatureStories": not isinstance(
            granularity.get("maxFeatureLinksPerStory"), (int, float)
        )
        or granularity["maxFeatureLinksPerStory"] <= 1,
        "storyTaskLimit": not isinstance(
            granularity.get("maxTasksPerStory"), (int, float)
        )
        or granularity["maxTasksPerStory"] <= 4,
        "packagingInvariant": current_workbook.get("packagingInvariant") is True,
    }
    if not acceptance["singleFeatureStories"]:
        regressions.append("STORY_FEATURE_CARDINALITY_EXCEEDED")
    if not acceptance["storyTaskLimit"]:
        regressions.append("STORY_TASK_LIMIT_EXCEEDED")
    if not acceptance["packagingInvariant"]:
        regressions.append("PACKAGING_INVARIANCE_UNVERIFIED")
    return {
        "inputs": {
            "sameContent": not changed_inputs,
            "changed": changed_inputs,
            "before": dict(before_inputs),
            "after": dict(after_inputs),
        },
        "workflow": change_map(
            baseline.get("workflow", {}), current.get("workflow", {})
        ),
        "objects": change_map(
            baseline.get("objects", {}), current.get("objects", {})
        ),
        "granularity": change_map(
            baseline.get("granularity", {}), current.get("granularity", {})
        ),
        "effort": change_map(
            baseline.get("effort", {}), current.get("effort", {})
        ),
        "workbook": {
            "sheets": {
                "before": baseline.get("workbook", {}).get("sheetCount"),
                "after": current.get("workbook", {}).get("sheetCount"),
            },
            "tables": {
                "before": baseline.get("workbook", {}).get("tableCount"),
                "after": current.get("workbook", {}).get("tableCount"),
            },
            "formulaErrors": {
                "before": baseline.get("workbook", {}).get("formulaErrors"),
                "after": current.get("workbook", {}).get("formulaErrors"),
            },
            "packagingInvariant": {
                "before": baseline.get("workbook", {}).get("packagingInvariant"),
                "after": current.get("workbook", {}).get("packagingInvariant"),
            },
        },
        "timingSeconds": {
            "before": baseline.get("timingSeconds"),
            "after": current.get("timingSeconds"),
        },
        "taskChanges": {
            "removed": [list(item) for item in sorted((before_tasks - after_tasks).elements())],
            "added": [list(item) for item in sorted((after_tasks - before_tasks).elements())],
        },
        "taskSemantics": {
            "removed": [
                list(item)
                for item in sorted(
                    (before_task_semantics - after_task_semantics).elements()
                )
            ],
            "added": [
                list(item)
                for item in sorted(
                    (after_task_semantics - before_task_semantics).elements()
                )
            ],
        },
        "semanticChanges": _semantic_changes(baseline, current),
        "semanticEvidence": _semantic_evidence(baseline, current),
        "acceptance": acceptance,
        "defects": defects,
        "regressions": regressions,
    }


def markdown_report(result: Mapping[str, Any]) -> str:
    inputs = result["inputs"]
    workbook = result["workbook"]
    granularity = result["granularity"]
    semantic = result["semanticChanges"]
    lines = [
        "# AI SOW E2E 前后对比",
        "",
        f"- 输入内容一致：{'是' if inputs['sameContent'] else '否'}",
        f"- Sheet：{workbook['sheets']['before']} → {workbook['sheets']['after']}",
        f"- Table：{workbook['tables']['before']} → {workbook['tables']['after']}",
        "- 最大 Story Task 数："
        f"{granularity['before'].get('maxTasksPerStory')} → "
        f"{granularity['after'].get('maxTasksPerStory')}",
        f"- 回归项：{', '.join(result['regressions']) if result['regressions'] else '无'}",
        "- 语义差异：需求、子需求、Story、AC、Task 均逐项列出增删",
        "- 人天差异：仅记录模板计算结果，不评价基准人天或 Story 人天合理性",
        "",
        "## 五层语义差异数量",
        "",
        "| 层级 | 删除 | 新增 |",
        "|---|---:|---:|",
        *(
            f"| {collection} | {len(changes['removed'])} | {len(changes['added'])} |"
            for collection, changes in semantic.items()
        ),
        "",
        "## D1–D12",
        "",
        "| 缺陷 | 状态 | 证据 |",
        "|---|---|---|",
    ]
    lines.extend(
        f"| {defect} | {finding['status']} | {finding['evidence']} |"
        for defect, finding in result["defects"].items()
    )
    lines.extend(
        [
            "",
            "## 结构化结果",
            "",
            "```json",
            json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True),
            "```",
            "",
        ]
    )
    return "\n".join(lines)


def _markdown_cell(value: object) -> str:
    return str(value).replace("|", "\\|").replace("\n", "<br>")


def delta_evidence_markdown(result: Mapping[str, Any]) -> str:
    labels = {
        "requirements": "需求",
        "subRequirements": "子需求",
        "stories": "Story",
        "acceptanceCriteria": "AC",
        "tasks": "Task",
    }
    evidence = result.get("semanticEvidence", [])
    if not isinstance(evidence, list):
        evidence = []
    lines = [
        "# 同输入 E2E 五层差异依据",
        "",
        "本报告逐项覆盖需求、子需求、Story、AC、Task 的全部新增与删除。",
        "每项必须给出来源锚点或模板基础单元依据；Story 人天不作为拆分正确性的完成门禁。",
        "",
        "| 层级 | 变化 | 对象 ID | 分类 | 来源锚点 / 基础单元 | 说明 |",
        "|---|---|---|---|---|---|",
    ]
    for item in evidence:
        if not isinstance(item, Mapping):
            continue
        refs = item.get("sourceRefs")
        anchors = (
            [
                f"{ref.get('sourceId')}#{ref.get('anchorId')}"
                + (f" ({ref.get('locator')})" if ref.get("locator") else "")
                for ref in refs
                if isinstance(ref, Mapping)
            ]
            if isinstance(refs, list)
            else []
        )
        base_unit = item.get("baseUnit")
        if isinstance(base_unit, str) and base_unit:
            anchors.append(f"baseUnit:{base_unit}")
        reason = str(item.get("reason", ""))
        rationale = item.get("rationale")
        if isinstance(rationale, str) and rationale.strip():
            reason = f"{reason} {rationale}"
        lines.append(
            "| "
            + " | ".join(
                _markdown_cell(value)
                for value in (
                    labels.get(
                        str(item.get("collection")), item.get("collection", "")
                    ),
                    item.get("change", ""),
                    item.get("objectId", ""),
                    item.get("classification", ""),
                    "<br>".join(anchors),
                    reason,
                )
            )
            + " |"
        )
    lines.extend(
        [
            "",
            f"共 {len(evidence)} 项差异；机器可核对原始记录见 `e2e-comparison.json.semanticEvidence`。",
            "",
        ]
    )
    return "\n".join(lines)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--baseline", type=Path, required=True)
    parser.add_argument("--current", type=Path, required=True)
    parser.add_argument("--json-output", type=Path, required=True)
    parser.add_argument("--markdown-output", type=Path, required=True)
    parser.add_argument("--rationale-output", type=Path)
    arguments = parser.parse_args(argv)
    result = compare_runs(arguments.baseline, arguments.current)
    arguments.json_output.parent.mkdir(parents=True, exist_ok=True)
    arguments.markdown_output.parent.mkdir(parents=True, exist_ok=True)
    arguments.json_output.write_text(
        json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    arguments.markdown_output.write_text(markdown_report(result), encoding="utf-8")
    if arguments.rationale_output is not None:
        arguments.rationale_output.parent.mkdir(parents=True, exist_ok=True)
        arguments.rationale_output.write_text(
            delta_evidence_markdown(result), encoding="utf-8"
        )
    return 1 if result["regressions"] or not result["inputs"]["sameContent"] else 0


if __name__ == "__main__":
    raise SystemExit(main())
