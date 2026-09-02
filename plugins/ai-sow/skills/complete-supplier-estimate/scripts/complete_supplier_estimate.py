from __future__ import annotations

import argparse
import copy
import datetime as dt
import hashlib
import json
import math
import os
import re
import sys
import tempfile
import unicodedata
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.formula import ArrayFormula


SKILL_ROOT = Path(__file__).resolve().parents[1]
PLUGIN_ROOT = SKILL_ROOT.parents[1]
SUPPLIER_REFERENCE = SKILL_ROOT / "assets/supplier-estimate-input.xlsx"
FORMAL_TEMPLATE = PLUGIN_ROOT / "assets/sow-template.xlsx"
SUPPLIER_CONTRACT = "supplier-estimate-input-v1"
SUPPLIER_SHEETS = ("01-需求故事", "02-任务清单", "90-填写选项")
FORMAL_SHEETS = (
    "01-需求故事",
    "02-任务清单",
    "03-工作量汇总",
    "90-估算标准",
)
SUPPLIER_TABLES = {
    "SOWStoryTable",
    "TaskTable",
    "SupplierTaskOptionTable",
    "SupplierComplexityOptionTable",
    "SupplierUATOptionTable",
}
FORMAL_TABLES = {
    "SOWStoryTable",
    "TaskTable",
    "ProjectSummaryTable",
    "ProjectParameterTable",
    "BaseUnitCatalogTable",
}
STORY_INPUT_HEADERS = ("需求", "子需求", "故事", "UAT适用", "验收条件", "备注")
TASK_INPUT_HEADERS = ("所属故事", "任务名称", "任务类型", "工作方式", "复杂度", "备注")
SUPPLIER_HEADERS = {
    "SOWStoryTable": [*STORY_INPUT_HEADERS, "任务列表", "校验结果", "故事路径"],
    "TaskTable": [*TASK_INPUT_HEADERS, "校验结果"],
    "SupplierTaskOptionTable": ["基础单元ID", "任务类型", "任务族", "允许工作方式"],
    "SupplierComplexityOptionTable": ["复杂度"],
    "SupplierUATOptionTable": ["UAT适用"],
}
FORMAL_HEADERS = {
    "SOWStoryTable": [
        *STORY_INPUT_HEADERS,
        "任务列表",
        "故事人天",
        "校验结果",
        "故事路径",
    ],
    "TaskTable": [
        *TASK_INPUT_HEADERS,
        "M档标准人天",
        "复杂度系数",
        "任务人天",
        "SIT支持人天",
        "校验结果",
    ],
}
SUPPLIER_FORMULA_HEADERS = {
    "SOWStoryTable": {"任务列表", "校验结果", "故事路径"},
    "TaskTable": {"校验结果"},
}
FORMAL_FORMULA_HEADERS = {
    "SOWStoryTable": {"任务列表", "故事人天", "校验结果", "故事路径"},
    "TaskTable": {"M档标准人天", "复杂度系数", "任务人天", "SIT支持人天", "校验结果"},
}
UNSAFE_MEMBER_PREFIXES = (
    "xl/externalLinks/",
    "xl/embeddings/",
    "xl/activeX/",
    "xl/ctrlProps/",
    "xl/drawings/",
    "xl/media/",
    "xl/charts/",
    "xl/queryTables/",
    "xl/pivotTables/",
    "xl/pivotCache/",
    "customXml/",
    "customUI/",
)
UNSAFE_MEMBERS = {
    "xl/vbaProject.bin",
    "xl/connections.xml",
    "docProps/custom.xml",
}
RISKY_TEXT = re.compile(r"^[=+\-@]")
DETERMINISTIC_TIME = dt.datetime(2000, 1, 1, 0, 0, 0)
DETERMINISTIC_ZIP_TIME = (2000, 1, 1, 0, 0, 0)
DETERMINISTIC_CREATE_SYSTEM = 3
DETERMINISTIC_UNIX_MODE = 0o600
WRAPPED_LINE_HEIGHT = 15
WRAPPED_ROW_PADDING = 4
MAX_EXCEL_ROW_HEIGHT = 409.5


@dataclass(frozen=True)
class InputRow:
    row: int
    values: dict[str, str]


class CompletionError(Exception):
    def __init__(self, diagnostics: list[dict[str, object]]) -> None:
        super().__init__(str(diagnostics[0]["code"]) if diagnostics else "completion failed")
        self.diagnostics = diagnostics


def diagnostic(
    code: str,
    message: str,
    *,
    sheet: str | None = None,
    row: int | None = None,
    field: str | None = None,
) -> dict[str, object]:
    result: dict[str, object] = {"code": code, "message": message}
    if sheet is not None:
        result = {
            "sheet": sheet,
            "row": row,
            "field": field,
            **result,
        }
    return result


def fail(
    code: str,
    message: str,
    *,
    sheet: str | None = None,
    row: int | None = None,
    field: str | None = None,
) -> None:
    raise CompletionError(
        [diagnostic(code, message, sheet=sheet, row=row, field=field)]
    )


def safe_text(value: object) -> object:
    if isinstance(value, str) and RISKY_TEXT.match(value):
        return "'" + value
    return value


def formula_text(value: object) -> str:
    if isinstance(value, ArrayFormula):
        value = value.text
    if not isinstance(value, str):
        raise TypeError("formula value is not text")
    return value


def visible_key(value: str) -> str:
    return unicodedata.normalize("NFC", str(safe_text(value))).casefold()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def inspect_zip(path: Path) -> None:
    try:
        with zipfile.ZipFile(path, "r") as archive:
            names = archive.namelist()
            unsafe = next(
                (
                    name
                    for name in names
                    if name in UNSAFE_MEMBERS
                    or any(name.startswith(prefix) for prefix in UNSAFE_MEMBER_PREFIXES)
                ),
                None,
            )
            if unsafe is not None:
                fail(
                    "UNSAFE_WORKBOOK_CONTENT",
                    f"工作簿包含不受支持的活动或嵌入内容：{unsafe}",
                )
            for name in names:
                if not name.endswith(".rels"):
                    continue
                payload = archive.read(name)
                if re.search(rb"TargetMode\s*=\s*['\"]External['\"]", payload):
                    fail(
                        "UNSAFE_WORKBOOK_CONTENT",
                        f"工作簿包含外部关系：{name}",
                    )
    except zipfile.BadZipFile:
        fail("INPUT_FORMAT_INVALID", "输入文件不是有效的 XLSX 工作簿。")


def table_index(workbook: Any) -> dict[str, tuple[Any, Any]]:
    found: dict[str, tuple[Any, Any]] = {}
    for worksheet in workbook.worksheets:
        for name in worksheet.tables:
            if name in found:
                fail("TABLE_CONTRACT_INVALID", f"工作簿存在重复 Table：{name}")
            found[name] = (worksheet, worksheet.tables[name])
    return found


def table_headers(worksheet: Any, table: Any) -> list[str]:
    min_col, min_row, max_col, _ = range_boundaries(table.ref)
    return [
        str(worksheet.cell(min_row, column).value)
        for column in range(min_col, max_col + 1)
    ]


def table_matrix(workbook: Any, name: str) -> list[list[object]]:
    worksheet, table = table_index(workbook)[name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    return [
        [worksheet.cell(row, column).value for column in range(min_col, max_col + 1)]
        for row in range(min_row, max_row + 1)
    ]


def formula_columns(worksheet: Any, table: Any) -> set[str]:
    min_col, min_row, max_col, _ = range_boundaries(table.ref)
    headers = table_headers(worksheet, table)
    return {
        headers[offset]
        for offset, column in enumerate(range(min_col, max_col + 1))
        if worksheet.cell(min_row + 1, column).data_type == "f"
        and isinstance(worksheet.cell(min_row + 1, column).value, (str, ArrayFormula))
    }


def validate_supplier_structure(workbook: Any, reference: Any) -> None:
    if tuple(workbook.sheetnames) != SUPPLIER_SHEETS:
        fail("UNEXPECTED_SHEETS", "输入工作簿的 Sheet 集合或顺序不受支持。")
    actual_index = table_index(workbook)
    reference_index = table_index(reference)
    if set(actual_index) != SUPPLIER_TABLES or set(reference_index) != SUPPLIER_TABLES:
        fail("TABLE_CONTRACT_INVALID", "输入工作簿的命名 Table 合同不完整。")
    if workbook["90-填写选项"]["B1"].value != SUPPLIER_CONTRACT:
        fail(
            "CONTRACT_VERSION_UNSUPPORTED",
            "供应商模板合同版本不受支持，请使用当前模板重新填写。",
        )

    for name in sorted(SUPPLIER_TABLES):
        worksheet, table = actual_index[name]
        reference_sheet, reference_table = reference_index[name]
        if worksheet.title != reference_sheet.title:
            fail("TABLE_CONTRACT_INVALID", f"Table 所在 Sheet 不匹配：{name}")
        min_col, min_row, max_col, _ = range_boundaries(table.ref)
        ref_min_col, ref_min_row, ref_max_col, _ = range_boundaries(reference_table.ref)
        if (min_col, min_row, max_col) != (ref_min_col, ref_min_row, ref_max_col):
            fail("TABLE_CONTRACT_INVALID", f"Table 起点或宽度不匹配：{name}")
        headers = table_headers(worksheet, table)
        metadata_headers = [column.name for column in table.tableColumns]
        if headers != SUPPLIER_HEADERS[name] or metadata_headers != SUPPLIER_HEADERS[name]:
            fail("HEADER_CONTRACT_INVALID", f"Table 表头不匹配：{name}")

    for name in ("SupplierTaskOptionTable", "SupplierComplexityOptionTable", "SupplierUATOptionTable"):
        if table_matrix(workbook, name) != table_matrix(reference, name):
            fail("TABLE_CONTRACT_INVALID", f"填写选项被修改：{name}")

    for name in ("SOWStoryTable", "TaskTable"):
        worksheet, table = actual_index[name]
        reference_sheet, reference_table = reference_index[name]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        for merged in worksheet.merged_cells.ranges:
            if not (
                merged.max_col < min_col
                or merged.min_col > max_col
                or merged.max_row <= min_row
                or merged.min_row > max_row
            ):
                fail(
                    "MERGED_BUSINESS_CELLS",
                    f"业务 Table 内不能使用合并单元格：{merged}",
                )
        actual_formula_headers = formula_columns(worksheet, table)
        reference_formula_headers = formula_columns(reference_sheet, reference_table)
        if (
            actual_formula_headers != SUPPLIER_FORMULA_HEADERS[name]
            or reference_formula_headers != SUPPLIER_FORMULA_HEADERS[name]
        ):
            fail("FORMULA_CONTRACT_INVALID", f"公式列合同不匹配：{name}")
        ref_min_col, ref_min_row, ref_max_col, _ = range_boundaries(reference_table.ref)
        reference_headers = table_headers(reference_sheet, reference_table)
        reference_formulas: dict[str, tuple[str, str, bool]] = {}
        for offset, header in enumerate(reference_headers):
            cell = reference_sheet.cell(ref_min_row + 1, ref_min_col + offset)
            if header in SUPPLIER_FORMULA_HEADERS[name]:
                if cell.data_type != "f" or not isinstance(cell.value, (str, ArrayFormula)):
                    fail("REFERENCE_ASSET_INVALID", f"参考模板公式缺失：{name}.{header}")
                reference_formulas[header] = (
                    cell.coordinate,
                    formula_text(cell.value),
                    isinstance(cell.value, ArrayFormula),
                )
        for row in range(min_row + 1, max_row + 1):
            for offset, header in enumerate(reference_headers):
                cell = worksheet.cell(row, min_col + offset)
                if header not in reference_formulas:
                    if cell.data_type == "f":
                        fail(
                            "FORMULA_INPUT_NOT_ALLOWED",
                            "输入列不能包含公式，请改为普通文本。",
                            sheet=worksheet.title,
                            row=row,
                            field=header,
                        )
                    continue
                origin, prototype, is_array = reference_formulas[header]
                expected = Translator(prototype, origin=origin).translate_formula(
                    cell.coordinate
                )
                actual = (
                    formula_text(cell.value)
                    if cell.data_type == "f" and isinstance(cell.value, (str, ArrayFormula))
                    else None
                )
                kind_matches = (
                    isinstance(cell.value, ArrayFormula)
                    and cell.value.ref == cell.coordinate
                    if is_array
                    else isinstance(cell.value, str)
                )
                if actual != expected or not kind_matches:
                    fail("FORMULA_CONTRACT_INVALID", f"公式被覆盖：{worksheet.title}!{cell.coordinate}")
        calculated = {
            column.name: (
                column.calculatedColumnFormula.text,
                column.calculatedColumnFormula.array is True,
            )
            for column in table.tableColumns
            if column.calculatedColumnFormula is not None
        }
        reference_calculated = {
            column.name: (
                column.calculatedColumnFormula.text,
                column.calculatedColumnFormula.array is True,
            )
            for column in reference_table.tableColumns
            if column.calculatedColumnFormula is not None
        }
        if calculated != reference_calculated:
            fail("FORMULA_CONTRACT_INVALID", f"Table 公式元数据不匹配：{name}")

    for sheet_name in ("01-需求故事", "02-任务清单"):
        actual = workbook[sheet_name]["A2"]
        expected = reference[sheet_name]["A2"]
        if (
            actual.data_type != "f"
            or expected.data_type != "f"
            or formula_text(actual.value) != formula_text(expected.value)
        ):
            fail("FORMULA_CONTRACT_INVALID", f"问题计数公式被覆盖：{sheet_name}!A2")


def read_input_rows(
    workbook: Any,
    table_name: str,
    input_headers: tuple[str, ...],
) -> list[InputRow]:
    worksheet, table = table_index(workbook)[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = table_headers(worksheet, table)
    result: list[InputRow] = []
    for row in range(min_row + 1, max_row + 1):
        values: dict[str, str] = {}
        for header in input_headers:
            cell = worksheet.cell(row, min_col + headers.index(header))
            value = cell.value
            if value is None:
                values[header] = ""
            elif isinstance(value, str):
                values[header] = value
            else:
                fail(
                    "TEXT_VALUE_REQUIRED",
                    "该字段必须填写为普通文本。",
                    sheet=worksheet.title,
                    row=row,
                    field=header,
                )
        if all(not value.strip() for value in values.values()):
            continue
        result.append(InputRow(row=row, values=values))
    return result


def option_contract(reference: Any) -> tuple[dict[str, set[str]], set[str], set[str]]:
    rows = table_matrix(reference, "SupplierTaskOptionTable")
    allowed_modes = {
        str(row[1]): set(str(row[3]).split("|"))
        for row in rows[1:]
    }
    complexities = {str(row[0]) for row in table_matrix(reference, "SupplierComplexityOptionTable")[1:]}
    uat_values = {str(row[0]) for row in table_matrix(reference, "SupplierUATOptionTable")[1:]}
    return allowed_modes, complexities, uat_values


def validate_business_rows(
    stories: list[InputRow],
    tasks: list[InputRow],
    allowed_modes: dict[str, set[str]],
    complexities: set[str],
    uat_values: set[str],
) -> None:
    diagnostics: list[dict[str, object]] = []
    for entry in stories:
        for field in STORY_INPUT_HEADERS[:-1]:
            if not entry.values[field].strip():
                diagnostics.append(
                    diagnostic(
                        "REQUIRED_VALUE_MISSING",
                        "请填写该必填字段后重试。",
                        sheet="01-需求故事",
                        row=entry.row,
                        field=field,
                    )
                )
                break
        if entry.values["UAT适用"].strip() and entry.values["UAT适用"] not in uat_values:
            diagnostics.append(
                diagnostic(
                    "ENUM_INVALID",
                    "请选择模板下拉列表中的受支持值。",
                    sheet="01-需求故事",
                    row=entry.row,
                    field="UAT适用",
                )
            )
    for entry in tasks:
        for field in TASK_INPUT_HEADERS[:-1]:
            if not entry.values[field].strip():
                diagnostics.append(
                    diagnostic(
                        "REQUIRED_VALUE_MISSING",
                        "请填写该必填字段后重试。",
                        sheet="02-任务清单",
                        row=entry.row,
                        field=field,
                    )
                )
                break
        complexity = entry.values["复杂度"]
        mode = entry.values["工作方式"]
        task_type = entry.values["任务类型"]
        if complexity.strip() and complexity not in complexities:
            diagnostics.append(
                diagnostic(
                    "ENUM_INVALID",
                    "请选择模板下拉列表中的受支持值。",
                    sheet="02-任务清单",
                    row=entry.row,
                    field="复杂度",
                )
            )
        if mode.strip() and mode not in {"新建", "调整", "接入复用"}:
            diagnostics.append(
                diagnostic(
                    "ENUM_INVALID",
                    "请选择模板下拉列表中的受支持值。",
                    sheet="02-任务清单",
                    row=entry.row,
                    field="工作方式",
                )
            )
        if task_type.strip() and task_type not in allowed_modes:
            diagnostics.append(
                diagnostic(
                    "TASK_TYPE_UNKNOWN",
                    "请选择“90-填写选项”中存在的任务类型。",
                    sheet="02-任务清单",
                    row=entry.row,
                    field="任务类型",
                )
            )
        elif mode in {"新建", "调整", "接入复用"} and mode not in allowed_modes.get(task_type, set()):
            diagnostics.append(
                diagnostic(
                    "WORK_MODE_NOT_ALLOWED",
                    "请选择该任务类型允许的工作方式。",
                    sheet="02-任务清单",
                    row=entry.row,
                    field="工作方式",
                )
            )
    if diagnostics:
        raise CompletionError(diagnostics)

    demand_by_feature: dict[str, str] = {}
    story_names: dict[str, int] = {}
    story_paths: dict[str, int] = {}
    for entry in stories:
        demand = entry.values["需求"]
        feature = entry.values["子需求"]
        story = entry.values["故事"]
        path = f"{demand} > {feature} > {story}"
        feature_key = visible_key(feature)
        story_key = visible_key(story)
        path_key = visible_key(path)
        if feature_key in demand_by_feature and demand_by_feature[feature_key] != visible_key(demand):
            diagnostics.append(
                diagnostic(
                    "PARENT_CONFLICT",
                    "同一子需求只能属于一个需求，请统一父级。",
                    sheet="01-需求故事",
                    row=entry.row,
                    field="子需求",
                )
            )
        else:
            demand_by_feature[feature_key] = visible_key(demand)
        if story_key in story_names:
            diagnostics.append(
                diagnostic(
                    "DUPLICATE_STORY",
                    "故事名称必须唯一，请修改重复名称。",
                    sheet="01-需求故事",
                    row=entry.row,
                    field="故事",
                )
            )
        else:
            story_names[story_key] = entry.row
        if path_key in story_paths:
            diagnostics.append(
                diagnostic(
                    "DUPLICATE_STORY_PATH",
                    "故事路径必须唯一，请修改重复的需求层级。",
                    sheet="01-需求故事",
                    row=entry.row,
                    field="故事",
                )
            )
        else:
            story_paths[path_key] = entry.row

    task_names: dict[str, int] = {}
    referenced_paths: set[str] = set()
    for entry in tasks:
        name_key = visible_key(entry.values["任务名称"])
        path_key = visible_key(entry.values["所属故事"])
        if name_key in task_names:
            diagnostics.append(
                diagnostic(
                    "DUPLICATE_TASK",
                    "任务名称必须唯一，请修改重复名称。",
                    sheet="02-任务清单",
                    row=entry.row,
                    field="任务名称",
                )
            )
        else:
            task_names[name_key] = entry.row
        if path_key not in story_paths:
            diagnostics.append(
                diagnostic(
                    "STORY_UNKNOWN",
                    "请将任务关联到工作簿中唯一存在的故事路径。",
                    sheet="02-任务清单",
                    row=entry.row,
                    field="所属故事",
                )
            )
        else:
            referenced_paths.add(path_key)
    for entry in stories:
        path = f"{entry.values['需求']} > {entry.values['子需求']} > {entry.values['故事']}"
        if visible_key(path) not in referenced_paths:
            diagnostics.append(
                diagnostic(
                    "STORY_WITHOUT_TASK",
                    "请至少为该故事填写一个任务。",
                    sheet="01-需求故事",
                    row=entry.row,
                    field="故事",
                )
            )
    if diagnostics:
        raise CompletionError(diagnostics)


def validate_formal_template(workbook: Any) -> None:
    if tuple(workbook.sheetnames) != FORMAL_SHEETS:
        fail("REFERENCE_ASSET_INVALID", "正式模板 Sheet 合同不匹配。")
    index = table_index(workbook)
    if set(index) != FORMAL_TABLES:
        fail("REFERENCE_ASSET_INVALID", "正式模板 Table 合同不匹配。")
    for name in ("SOWStoryTable", "TaskTable"):
        worksheet, table = index[name]
        if table_headers(worksheet, table) != FORMAL_HEADERS[name]:
            fail("REFERENCE_ASSET_INVALID", f"正式模板表头不匹配：{name}")
        if formula_columns(worksheet, table) != FORMAL_FORMULA_HEADERS[name]:
            fail("REFERENCE_ASSET_INVALID", f"正式模板公式原型不匹配：{name}")
        _, min_row, _, max_row = range_boundaries(table.ref)
        if max_row - min_row != 1:
            fail("REFERENCE_ASSET_INVALID", f"正式模板必须只有一个 prototype 行：{name}")


def copy_style(source: Any, target: Any) -> None:
    target.font = copy.copy(source.font)
    target.fill = copy.copy(source.fill)
    target.border = copy.copy(source.border)
    target.alignment = copy.copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy.copy(source.protection)


def wrapped_line_count(value: str, column_width: float) -> int:
    lines = value.splitlines() or [""]
    return sum(
        max(
            1,
            math.ceil(
                sum(
                    2 if unicodedata.east_asian_width(character) in {"W", "F"} else 1
                    for character in line
                )
                / column_width
            ),
        )
        for line in lines
    )


def fill_formal_table(
    workbook: Any,
    table_name: str,
    rows: list[InputRow],
) -> None:
    worksheet, table = table_index(workbook)[table_name]
    min_col, min_row, max_col, old_max_row = range_boundaries(table.ref)
    prototype_row = min_row + 1
    headers = table_headers(worksheet, table)
    prototypes = [
        worksheet.cell(prototype_row, column)
        for column in range(min_col, max_col + 1)
    ]
    formulas = {
        offset: (
            formula_text(cell.value),
            isinstance(cell.value, ArrayFormula),
            cell.coordinate,
        )
        for offset, cell in enumerate(prototypes)
        if cell.data_type == "f" and isinstance(cell.value, (str, ArrayFormula))
    }
    if {headers[offset] for offset in formulas} != FORMAL_FORMULA_HEADERS[table_name]:
        fail("REFERENCE_ASSET_INVALID", f"正式模板公式原型不匹配：{table_name}")
    physical_rows: list[InputRow | None] = rows if rows else [None]
    clear_through = max(old_max_row, min_row + len(physical_rows))
    for row in range(prototype_row, clear_through + 1):
        for column in range(min_col, max_col + 1):
            worksheet.cell(row, column).value = None
    prototype_height = worksheet.row_dimensions[prototype_row].height or 15
    for row_offset, entry in enumerate(physical_rows, start=1):
        row = min_row + row_offset
        values = {} if entry is None else entry.values
        wrapped_lines = 1
        for column_offset, header in enumerate(headers):
            target = worksheet.cell(row, min_col + column_offset)
            source = prototypes[column_offset]
            copy_style(source, target)
            if column_offset in formulas:
                prototype, is_array, origin = formulas[column_offset]
                translated = Translator(prototype, origin=origin).translate_formula(
                    target.coordinate
                )
                target.value = (
                    ArrayFormula(ref=target.coordinate, text=translated)
                    if is_array
                    else translated
                )
                continue
            value = None if entry is None else safe_text(values.get(header, ""))
            target.value = value
            if isinstance(value, str):
                target.data_type = "s"
                if target.alignment.wrap_text:
                    width = worksheet.column_dimensions[
                        get_column_letter(target.column)
                    ].width or 8.43
                    wrapped_lines = max(wrapped_lines, wrapped_line_count(value, width))
        worksheet.row_dimensions[row].height = min(
            MAX_EXCEL_ROW_HEIGHT,
            max(prototype_height, wrapped_lines * WRAPPED_LINE_HEIGHT + WRAPPED_ROW_PADDING),
        )
    new_max_row = min_row + len(physical_rows)
    table.ref = (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{new_max_row}"
    )
    if table.autoFilter is not None:
        table.autoFilter.ref = table.ref
    if worksheet.auto_filter.ref:
        worksheet.auto_filter.ref = table.ref


def style_signature(cell: Any) -> tuple[object, ...]:
    return (
        copy.copy(cell.font),
        copy.copy(cell.fill),
        copy.copy(cell.border),
        copy.copy(cell.alignment),
        cell.number_format,
        copy.copy(cell.protection),
    )


def verify_formula_projection(
    output: Any,
    formal: Any,
    table_name: str,
    row_count: int,
) -> None:
    worksheet, table = table_index(output)[table_name]
    reference_sheet, reference_table = table_index(formal)[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    ref_min_col, ref_min_row, _, _ = range_boundaries(reference_table.ref)
    if max_row - min_row != max(1, row_count):
        fail("OUTPUT_VERIFICATION_FAILED", f"输出 Table 行数不匹配：{table_name}")
    headers = table_headers(worksheet, table)
    reference_headers = table_headers(reference_sheet, reference_table)
    for offset, header in enumerate(headers):
        reference_cell = reference_sheet.cell(ref_min_row + 1, ref_min_col + offset)
        if header not in FORMAL_FORMULA_HEADERS[table_name]:
            continue
        prototype = formula_text(reference_cell.value)
        is_array = isinstance(reference_cell.value, ArrayFormula)
        for row in range(min_row + 1, max_row + 1):
            cell = worksheet.cell(row, min_col + offset)
            expected = Translator(
                prototype,
                origin=reference_cell.coordinate,
            ).translate_formula(cell.coordinate)
            if (
                cell.data_type != "f"
                or not isinstance(cell.value, (str, ArrayFormula))
                or formula_text(cell.value) != expected
                or style_signature(cell) != style_signature(reference_cell)
                or (isinstance(cell.value, ArrayFormula) != is_array)
            ):
                fail(
                    "OUTPUT_VERIFICATION_FAILED",
                    f"输出公式或样式不匹配：{table_name}.{header}",
                )
    if reference_headers != headers:
        fail("OUTPUT_VERIFICATION_FAILED", f"输出表头不匹配：{table_name}")


def verify_output(
    path: Path,
    formal_path: Path,
    stories: list[InputRow],
    tasks: list[InputRow],
) -> None:
    output = openpyxl.load_workbook(path, data_only=False, read_only=False)
    formal = openpyxl.load_workbook(formal_path, data_only=False, read_only=False)
    try:
        validate_formal_template(formal)
        if tuple(output.sheetnames) != FORMAL_SHEETS or set(table_index(output)) != FORMAL_TABLES:
            fail("OUTPUT_VERIFICATION_FAILED", "输出工作簿结构不匹配。")
        for name, entries, input_headers in (
            ("SOWStoryTable", stories, STORY_INPUT_HEADERS),
            ("TaskTable", tasks, TASK_INPUT_HEADERS),
        ):
            verify_formula_projection(output, formal, name, len(entries))
            worksheet, table = table_index(output)[name]
            min_col, min_row, _, _ = range_boundaries(table.ref)
            headers = table_headers(worksheet, table)
            for offset, entry in enumerate(entries, start=1):
                for header in input_headers:
                    actual = worksheet.cell(
                        min_row + offset,
                        min_col + headers.index(header),
                    ).value
                    expected = safe_text(entry.values[header])
                    if expected == "":
                        expected = None
                    if actual != expected:
                        fail(
                            "OUTPUT_VERIFICATION_FAILED",
                            f"输出输入值复读不匹配：{name}.{header}",
                        )
        for name in ("ProjectSummaryTable", "ProjectParameterTable", "BaseUnitCatalogTable"):
            if table_matrix(output, name) != table_matrix(formal, name):
                fail("OUTPUT_VERIFICATION_FAILED", f"正式模板内容发生变化：{name}")
        calculation = output.calculation
        if not (
            calculation.calcMode == "auto"
            and calculation.calcOnSave is True
            and calculation.forceFullCalc is True
            and calculation.fullCalcOnLoad is True
        ):
            fail("OUTPUT_VERIFICATION_FAILED", "输出重算标志不完整。")
    finally:
        output.close()
        formal.close()


def deterministic_external_attr(external_attr: int) -> int:
    return (external_attr & 0xFFFF0000) & ~(0o777 << 16) | (
        DETERMINISTIC_UNIX_MODE << 16
    )


def normalize_xlsx(path: Path) -> None:
    with zipfile.ZipFile(path, "r") as source:
        members = []
        for entry in source.infolist():
            payload = source.read(entry.filename)
            if entry.filename == "docProps/core.xml":
                payload = re.sub(
                    rb"<dcterms:modified[^>]*>.*?</dcterms:modified>",
                    b'<dcterms:modified xsi:type="dcterms:W3CDTF">2000-01-01T00:00:00Z</dcterms:modified>',
                    payload,
                )
            members.append((entry.filename, payload, entry))
    temporary: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{path.name}.",
            suffix=".zip",
            dir=path.parent,
            delete=False,
        ) as stream:
            temporary = Path(stream.name)
        with zipfile.ZipFile(
            temporary,
            "w",
            compression=zipfile.ZIP_DEFLATED,
            compresslevel=9,
        ) as target:
            for name, payload, original in sorted(members, key=lambda item: item[0]):
                entry = zipfile.ZipInfo(name, DETERMINISTIC_ZIP_TIME)
                entry.compress_type = zipfile.ZIP_DEFLATED
                entry.create_system = DETERMINISTIC_CREATE_SYSTEM
                entry.external_attr = deterministic_external_attr(original.external_attr)
                entry.flag_bits = original.flag_bits
                target.writestr(entry, payload)
        os.replace(temporary, path)
        temporary = None
    finally:
        if temporary is not None:
            temporary.unlink(missing_ok=True)


def complete(input_path: Path, output_path: Path) -> dict[str, object]:
    input_path = input_path.resolve()
    output_path = output_path.resolve()
    if input_path == output_path:
        fail("INPUT_OUTPUT_SAME", "输出路径必须与输入路径不同。")
    if output_path.exists():
        fail("OUTPUT_EXISTS", "输出文件已存在；请选择一个新的输出路径。")
    if input_path.suffix.casefold() != ".xlsx" or not input_path.is_file():
        fail("INPUT_FORMAT_INVALID", "输入必须是存在的 .xlsx 文件。")
    if output_path.suffix.casefold() != ".xlsx":
        fail("OUTPUT_FORMAT_INVALID", "输出路径必须使用 .xlsx 扩展名。")
    if not output_path.parent.is_dir():
        fail("OUTPUT_DIRECTORY_MISSING", "输出目录不存在。")
    if not SUPPLIER_REFERENCE.is_file() or not FORMAL_TEMPLATE.is_file():
        fail("REFERENCE_ASSET_INVALID", "插件参考模板资产缺失。")

    input_digest = sha256_file(input_path)
    inspect_zip(input_path)
    inspect_zip(SUPPLIER_REFERENCE)
    inspect_zip(FORMAL_TEMPLATE)
    supplier = openpyxl.load_workbook(input_path, data_only=False, read_only=False)
    reference = openpyxl.load_workbook(
        SUPPLIER_REFERENCE,
        data_only=False,
        read_only=False,
    )
    try:
        validate_supplier_structure(supplier, reference)
        stories = read_input_rows(supplier, "SOWStoryTable", STORY_INPUT_HEADERS)
        tasks = read_input_rows(supplier, "TaskTable", TASK_INPUT_HEADERS)
        allowed_modes, complexities, uat_values = option_contract(reference)
        validate_business_rows(stories, tasks, allowed_modes, complexities, uat_values)
    finally:
        supplier.close()
        reference.close()

    temporary: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{output_path.name}.",
            suffix=".tmp.xlsx",
            dir=output_path.parent,
            delete=False,
        ) as stream:
            temporary = Path(stream.name)
        workbook = openpyxl.load_workbook(
            FORMAL_TEMPLATE,
            data_only=False,
            read_only=False,
        )
        try:
            validate_formal_template(workbook)
            fill_formal_table(workbook, "SOWStoryTable", stories)
            fill_formal_table(workbook, "TaskTable", tasks)
            workbook.calculation.calcMode = "auto"
            workbook.calculation.calcOnSave = True
            workbook.calculation.forceFullCalc = True
            workbook.calculation.fullCalcOnLoad = True
            workbook.properties.created = DETERMINISTIC_TIME
            workbook.properties.modified = DETERMINISTIC_TIME
            workbook.save(temporary)
        finally:
            workbook.close()
        normalize_xlsx(temporary)
        verify_output(temporary, FORMAL_TEMPLATE, stories, tasks)
        if sha256_file(input_path) != input_digest:
            fail("INPUT_MUTATED", "输入文件在补全过程中发生变化，已停止发布。")
        if output_path.exists():
            fail("OUTPUT_EXISTS", "输出文件在发布前已存在，已停止发布。")
        os.replace(temporary, output_path)
        temporary = None
    finally:
        if temporary is not None:
            temporary.unlink(missing_ok=True)

    return {
        "outcome": "OK",
        "input": str(input_path),
        "output": str(output_path),
        "inputSha256": input_digest,
        "storyCount": len(stories),
        "taskCount": len(tasks),
    }


def emit(payload: dict[str, object]) -> None:
    sys.stdout.write(json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + "\n")


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    if hasattr(sys.stderr, "reconfigure"):
        sys.stderr.reconfigure(encoding="utf-8")
    parser = argparse.ArgumentParser(
        description="严格校验供应商估算输入并补全为新的正式工作簿。"
    )
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()
    try:
        emit(complete(args.input, args.output))
        return 0
    except CompletionError as error:
        emit({"outcome": "BLOCKED", "diagnostics": error.diagnostics})
        return 2
    except Exception:
        emit(
            {
                "outcome": "BLOCKED",
                "diagnostics": [
                    diagnostic(
                        "INTERNAL_ERROR",
                        "补全过程遇到未预期错误；未发布输出，请检查模板与运行环境。",
                    )
                ],
            }
        )
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
