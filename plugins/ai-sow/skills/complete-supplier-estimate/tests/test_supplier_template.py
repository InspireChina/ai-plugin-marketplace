from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.formula import ArrayFormula


SKILL_ROOT = Path(__file__).resolve().parents[1]
PLUGIN_ROOT = SKILL_ROOT.parents[1]
SUPPLIER_TEMPLATE = SKILL_ROOT / "assets/supplier-estimate-input.xlsx"
FORMAL_TEMPLATE = PLUGIN_ROOT / "assets/sow-template.xlsx"
CONTRACT_VERSION = "supplier-estimate-input-v1"
STORY_HEADERS = [
    "需求",
    "子需求",
    "故事",
    "UAT适用",
    "验收条件",
    "备注",
    "任务列表",
    "校验结果",
    "故事路径",
]
TASK_HEADERS = [
    "所属故事",
    "任务名称",
    "任务类型",
    "工作方式",
    "复杂度",
    "备注",
    "校验结果",
]
FORBIDDEN_TOKENS = {
    "ProjectParameterTable",
    "BaseUnitCatalogTable",
    "M档标准人天",
    "复杂度系数",
    "任务人天",
    "SIT支持人天",
    "ROUND_STORY",
    "ROUND_PROJECT",
    "K_UAT",
    "SIT_INT_SUPPORT",
    "SIT_EXT_SUPPORT",
    "K_COMPLEXITY_",
    "风险",
    "假设",
}


def index_tables(workbook: openpyxl.Workbook) -> dict[str, tuple[object, object]]:
    return {
        name: (worksheet, worksheet.tables[name])
        for worksheet in workbook.worksheets
        for name in worksheet.tables
    }


def table_values(workbook: openpyxl.Workbook, name: str) -> list[list[object]]:
    worksheet, table = index_tables(workbook)[name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    return [
        [worksheet.cell(row, column).value for column in range(min_col, max_col + 1)]
        for row in range(min_row, max_row + 1)
    ]


def formula_text(value: object) -> str:
    if isinstance(value, ArrayFormula):
        return value.text
    assert isinstance(value, str)
    return value


def validation_covers(worksheet: object, coordinate: str) -> bool:
    return any(
        coordinate in validation.cells
        for validation in worksheet.data_validations.dataValidation
    )


def test_supplier_asset_has_exact_input_only_structure() -> None:
    assert SUPPLIER_TEMPLATE.is_file()
    workbook = openpyxl.load_workbook(SUPPLIER_TEMPLATE, data_only=False)
    try:
        assert workbook.sheetnames == ["01-需求故事", "02-任务清单", "90-填写选项"]
        tables = index_tables(workbook)
        assert {"SOWStoryTable", "TaskTable"}.issubset(tables)
        assert table_values(workbook, "SOWStoryTable")[0] == STORY_HEADERS
        assert table_values(workbook, "TaskTable")[0] == TASK_HEADERS
        assert workbook["90-填写选项"]["B1"].value == CONTRACT_VERSION
        assert workbook["01-需求故事"].column_dimensions["I"].hidden is True

        for table_name, formula_headers in (
            ("SOWStoryTable", {"任务列表", "校验结果", "故事路径"}),
            ("TaskTable", {"校验结果"}),
        ):
            worksheet, table = tables[table_name]
            assert worksheet.auto_filter.ref is None
            assert table.autoFilter is not None
            assert table.autoFilter.ref == table.ref
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            assert max_row - min_row == 1
            headers = [
                worksheet.cell(min_row, column).value
                for column in range(min_col, max_col + 1)
            ]
            calculated = {
                column.name
                for column in table.tableColumns
                if column.calculatedColumnFormula is not None
            }
            assert calculated == formula_headers
            for header in formula_headers:
                cell = worksheet.cell(min_row + 1, min_col + headers.index(header))
                assert cell.data_type == "f"
                assert cell.protection.locked is True
                assert cell.fill.fgColor.rgb[-6:] == "F1F4F6"
            assert table.tableStyleInfo is not None
    finally:
        workbook.close()


def test_supplier_asset_exposes_no_formal_values_or_calculation_contract() -> None:
    workbook = openpyxl.load_workbook(SUPPLIER_TEMPLATE, data_only=False)
    try:
        assert not workbook._external_links
        visible_text: list[str] = []
        formulas: list[str] = []
        comments: list[str] = []
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        visible_text.append(formula_text(cell.value) if cell.data_type == "f" else str(cell.value))
                    if cell.data_type == "f":
                        formulas.append(formula_text(cell.value))
                    if cell.comment is not None:
                        comments.append(cell.comment.text)
        combined = "\n".join([*visible_text, *comments])
        for token in FORBIDDEN_TOKENS:
            assert token not in combined
        assert all("CEILING(" not in formula and "ROUND(" not in formula for formula in formulas)
        assert not list(workbook.defined_names.values())
    finally:
        workbook.close()

    with zipfile.ZipFile(SUPPLIER_TEMPLATE) as archive:
        names = set(archive.namelist())
        assert "xl/vbaProject.bin" not in names
        assert not any(name.startswith("xl/externalLinks/") for name in names)
        assert not any(name.startswith("xl/embeddings/") for name in names)
        assert not any(name.startswith("customXml/") for name in names)
        payload = b"\n".join(archive.read(name) for name in sorted(names))
    for token in FORBIDDEN_TOKENS:
        assert token.encode("utf-8") not in payload


def test_supplier_options_are_a_non_sensitive_projection_of_formal_catalog() -> None:
    formal = openpyxl.load_workbook(FORMAL_TEMPLATE, data_only=False)
    supplier = openpyxl.load_workbook(SUPPLIER_TEMPLATE, data_only=False)
    try:
        catalog = table_values(formal, "BaseUnitCatalogTable")
        headers = [str(value) for value in catalog[0]]
        expected: dict[str, tuple[str, str, str]] = {}
        mode_columns = {
            "新建": headers.index("新建M档人天"),
            "调整": headers.index("调整M档人天"),
            "接入复用": headers.index("接入复用M档人天"),
        }
        for row in catalog[1:]:
            modes = [
                mode
                for mode, column in mode_columns.items()
                if isinstance(row[column], (int, float)) and row[column] > 0
            ]
            expected[str(row[headers.index("基础单元ID")])] = (
                str(row[headers.index("基础单元名称")]),
                str(row[headers.index("任务族名称")]),
                "|".join(modes),
            )

        options = table_values(supplier, "SupplierTaskOptionTable")
        assert options[0] == ["基础单元ID", "任务类型", "任务族", "允许工作方式"]
        actual = {
            str(row[0]): (str(row[1]), str(row[2]), str(row[3]))
            for row in options[1:]
        }
        assert actual == expected
        assert len(actual) == 37
        assert len({value[0] for value in actual.values()}) == 37
        assert table_values(supplier, "SupplierComplexityOptionTable") == [
            ["复杂度"],
            ["S"],
            ["M"],
            ["L"],
        ]
        assert table_values(supplier, "SupplierUATOptionTable") == [
            ["UAT适用"],
            ["是"],
            ["否"],
        ]
    finally:
        formal.close()
        supplier.close()


def test_supplier_tables_keep_formula_and_validation_contract_beyond_legacy_caps() -> None:
    workbook = openpyxl.load_workbook(SUPPLIER_TEMPLATE, data_only=False)
    try:
        story = workbook["01-需求故事"]
        task = workbook["02-任务清单"]
        assert validation_covers(story, "D1000")
        assert validation_covers(task, "A1000")
        assert validation_covers(task, "C1000")
        assert validation_covers(task, "D1000")
        assert validation_covers(task, "E1000")
        assert not validation_covers(story, "E1000")
        assert not validation_covers(story, "F1000")
        assert not validation_covers(task, "F1000")
        assert story["E5"].alignment.wrap_text is True
        assert story["F5"].alignment.wrap_text is True
        assert task["F5"].alignment.wrap_text is True

        story_formula = formula_text(story["H5"].value)
        for diagnostic in (
            "必填项缺失",
            "故事名称重复",
            "故事路径重复",
            "子需求父级冲突",
            "UAT适用值非法",
            "故事缺少任务",
        ):
            assert diagnostic in story_formula
        task_formula = formula_text(task["G5"].value)
        for diagnostic in (
            "必填项缺失",
            "任务名称重复",
            "所属故事未知",
            "复杂度非法",
            "任务类型未知",
            "工作方式非法",
            "任务类型与工作方式不适用",
        ):
            assert diagnostic in task_formula
        assert story["A2"].data_type == "f" and "问题数" in formula_text(story["A2"].value)
        assert task["A2"].data_type == "f" and "问题数" in formula_text(task["A2"].value)

        story_ranges = " ".join(str(item) for item in story.conditional_formatting)
        task_ranges = " ".join(str(item) for item in task.conditional_formatting)
        assert "H5:H1048576" in story_ranges
        assert "G5:G1048576" in task_ranges
    finally:
        workbook.close()
