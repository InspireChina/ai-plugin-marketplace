from __future__ import annotations

import copy
import hashlib
import json
import subprocess
import sys
import zipfile
from pathlib import Path
from typing import Callable

import openpyxl
import pytest
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.formula import ArrayFormula


SKILL_ROOT = Path(__file__).resolve().parents[1]
PLUGIN_ROOT = SKILL_ROOT.parents[1]
SCRIPT = SKILL_ROOT / "scripts/complete_supplier_estimate.py"
SUPPLIER_TEMPLATE = SKILL_ROOT / "assets/supplier-estimate-input.xlsx"
FORMAL_TEMPLATE = PLUGIN_ROOT / "assets/sow-template.xlsx"
STORY_INPUT_HEADERS = ["需求", "子需求", "故事", "UAT适用", "验收条件", "备注"]
TASK_INPUT_HEADERS = ["所属故事", "任务名称", "任务类型", "工作方式", "复杂度", "备注"]


def table_index(workbook: openpyxl.Workbook) -> dict[str, tuple[object, object]]:
    return {
        name: (worksheet, worksheet.tables[name])
        for worksheet in workbook.worksheets
        for name in worksheet.tables
    }


def formula_text(value: object) -> str:
    if isinstance(value, ArrayFormula):
        return value.text
    assert isinstance(value, str)
    return value


def copy_cell_style(source: object, target: object) -> None:
    target.font = copy.copy(source.font)
    target.fill = copy.copy(source.fill)
    target.border = copy.copy(source.border)
    target.alignment = copy.copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy.copy(source.protection)


def replace_table_rows(
    workbook: openpyxl.Workbook,
    table_name: str,
    rows: list[dict[str, str]],
) -> None:
    worksheet, table = table_index(workbook)[table_name]
    min_col, min_row, max_col, old_max_row = range_boundaries(table.ref)
    headers = [
        str(worksheet.cell(min_row, column).value)
        for column in range(min_col, max_col + 1)
    ]
    prototypes = [
        copy.copy(worksheet.cell(min_row + 1, column))
        for column in range(min_col, max_col + 1)
    ]
    physical_rows = rows or [{}]
    for row in range(min_row + 1, max(old_max_row, min_row + len(physical_rows)) + 1):
        for column in range(min_col, max_col + 1):
            worksheet.cell(row, column).value = None
    for row_offset, payload in enumerate(physical_rows, start=1):
        row_number = min_row + row_offset
        for column_offset, header in enumerate(headers):
            source = prototypes[column_offset]
            target = worksheet.cell(row_number, min_col + column_offset)
            copy_cell_style(source, target)
            if source.data_type == "f":
                source_formula = formula_text(source.value)
                translated = Translator(
                    source_formula,
                    origin=source.coordinate,
                ).translate_formula(target.coordinate)
                target.value = (
                    ArrayFormula(ref=target.coordinate, text=translated)
                    if isinstance(source.value, ArrayFormula)
                    else translated
                )
            else:
                target.value = payload.get(header)
                if isinstance(target.value, str):
                    target.data_type = "s"
    max_row = min_row + len(physical_rows)
    table.ref = (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{max_row}"
    )
    if table.autoFilter is not None:
        table.autoFilter.ref = table.ref


def valid_rows() -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    stories = [
        {
            "需求": "需求甲",
            "子需求": "子需求甲",
            "故事": "故事 A",
            "UAT适用": "是",
            "验收条件": "可创建记录\n可查看结果",
            "备注": "=供应商故事备注",
        },
        {
            "需求": "需求甲",
            "子需求": "子需求甲",
            "故事": "故事 B",
            "UAT适用": "否",
            "验收条件": "可完成外部系统交互",
            "备注": "普通故事备注",
        },
    ]
    tasks = [
        {
            "所属故事": "需求甲 > 子需求甲 > 故事 A",
            "任务名称": "任务 1",
            "任务类型": "界面与交互",
            "工作方式": "新建",
            "复杂度": "M",
            "备注": "@供应商任务备注",
        },
        {
            "所属故事": "需求甲 > 子需求甲 > 故事 A",
            "任务名称": "任务 2",
            "任务类型": "内部系统对接",
            "工作方式": "新建",
            "复杂度": "S",
            "备注": "内部对接",
        },
        {
            "所属故事": "需求甲 > 子需求甲 > 故事 B",
            "任务名称": "任务 3",
            "任务类型": "外部系统对接",
            "工作方式": "调整",
            "复杂度": "L",
            "备注": "外部对接",
        },
    ]
    return stories, tasks


def make_supplier(
    path: Path,
    stories: list[dict[str, str]] | None = None,
    tasks: list[dict[str, str]] | None = None,
) -> None:
    default_stories, default_tasks = valid_rows()
    workbook = openpyxl.load_workbook(SUPPLIER_TEMPLATE, data_only=False)
    try:
        replace_table_rows(
            workbook,
            "SOWStoryTable",
            default_stories if stories is None else stories,
        )
        replace_table_rows(
            workbook,
            "TaskTable",
            default_tasks if tasks is None else tasks,
        )
        workbook.save(path)
    finally:
        workbook.close()


def run_cli(input_path: Path, output_path: Path) -> tuple[subprocess.CompletedProcess[str], dict[str, object]]:
    completed = subprocess.run(
        [
            sys.executable,
            str(SCRIPT),
            "--input",
            str(input_path),
            "--output",
            str(output_path),
        ],
        cwd=PLUGIN_ROOT,
        text=True,
        encoding="utf-8",
        capture_output=True,
        check=False,
    )
    payload = json.loads(completed.stdout)
    return completed, payload


def table_rows(workbook: openpyxl.Workbook, name: str) -> list[dict[str, object]]:
    worksheet, table = table_index(workbook)[name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = [
        str(worksheet.cell(min_row, column).value)
        for column in range(min_col, max_col + 1)
    ]
    return [
        {
            header: worksheet.cell(row, column).value
            for header, column in zip(headers, range(min_col, max_col + 1), strict=True)
        }
        for row in range(min_row + 1, max_row + 1)
    ]


def test_cli_completes_valid_supplier_without_mutating_input(tmp_path: Path) -> None:
    supplier = tmp_path / "supplier.xlsx"
    output = tmp_path / "formal.xlsx"
    make_supplier(supplier)
    original_digest = hashlib.sha256(supplier.read_bytes()).hexdigest()

    completed, result = run_cli(supplier, output)

    assert completed.returncode == 0, (completed.stderr, completed.stdout)
    assert completed.stderr == ""
    assert result == {
        "outcome": "OK",
        "input": str(supplier.resolve()),
        "output": str(output.resolve()),
        "inputSha256": original_digest,
        "storyCount": 2,
        "taskCount": 3,
    }
    assert hashlib.sha256(supplier.read_bytes()).hexdigest() == original_digest
    workbook = openpyxl.load_workbook(output, data_only=False)
    formal = openpyxl.load_workbook(FORMAL_TEMPLATE, data_only=False)
    try:
        assert workbook.sheetnames == [
            "01-需求故事",
            "02-任务清单",
            "03-工作量汇总",
            "90-估算标准",
        ]
        assert set(table_index(workbook)) == {
            "SOWStoryTable",
            "TaskTable",
            "ProjectSummaryTable",
            "ProjectParameterTable",
            "BaseUnitCatalogTable",
        }
        stories = table_rows(workbook, "SOWStoryTable")
        tasks = table_rows(workbook, "TaskTable")
        assert [row["故事"] for row in stories] == ["故事 A", "故事 B"]
        assert stories[0]["验收条件"] == "可创建记录\n可查看结果"
        assert stories[0]["备注"] == "'=供应商故事备注"
        assert [row["任务名称"] for row in tasks] == ["任务 1", "任务 2", "任务 3"]
        assert tasks[0]["备注"] == "'@供应商任务备注"
        for table_name, formula_headers in (
            ("SOWStoryTable", {"任务列表", "故事人天", "校验结果", "故事路径"}),
            ("TaskTable", {"M档标准人天", "复杂度系数", "任务人天", "SIT支持人天", "校验结果"}),
        ):
            rows = table_rows(workbook, table_name)
            assert rows
            for row in rows:
                assert all(formula_text(row[header]).startswith("=") for header in formula_headers)
        assert table_rows(workbook, "ProjectParameterTable") == table_rows(
            formal, "ProjectParameterTable"
        )
        assert table_rows(workbook, "BaseUnitCatalogTable") == table_rows(
            formal, "BaseUnitCatalogTable"
        )
        assert workbook.calculation.calcMode == "auto"
        assert workbook.calculation.calcOnSave is True
        assert workbook.calculation.forceFullCalc is True
        assert workbook.calculation.fullCalcOnLoad is True
    finally:
        workbook.close()
        formal.close()


def test_cli_reads_actual_table_boundaries_beyond_legacy_capacity(tmp_path: Path) -> None:
    stories, _ = valid_rows()
    stories = stories[:1]
    tasks = [
        {
            "所属故事": "需求甲 > 子需求甲 > 故事 A",
            "任务名称": f"动态任务 {index:03d}",
            "任务类型": "界面与交互",
            "工作方式": "新建",
            "复杂度": "M",
            "备注": "",
        }
        for index in range(1, 511)
    ]
    supplier = tmp_path / "supplier-large.xlsx"
    output = tmp_path / "formal-large.xlsx"
    make_supplier(supplier, stories, tasks)

    completed, result = run_cli(supplier, output)

    assert completed.returncode == 0, (completed.stderr, completed.stdout)
    assert result["taskCount"] == 510
    workbook = openpyxl.load_workbook(output, data_only=False)
    try:
        rows = table_rows(workbook, "TaskTable")
        assert len(rows) == 510
        assert rows[-1]["任务名称"] == "动态任务 510"
        assert str(rows[-1]["任务人天"]).startswith("=")
    finally:
        workbook.close()


@pytest.mark.parametrize(
    ("arrange", "expected_code"),
    [
        ("existing", "OUTPUT_EXISTS"),
        ("same", "INPUT_OUTPUT_SAME"),
    ],
)
def test_cli_rejects_unsafe_output_paths_without_mutation(
    tmp_path: Path,
    arrange: str,
    expected_code: str,
) -> None:
    supplier = tmp_path / "supplier.xlsx"
    make_supplier(supplier)
    original = supplier.read_bytes()
    output = supplier if arrange == "same" else tmp_path / "formal.xlsx"
    if arrange == "existing":
        output.write_bytes(b"existing-output")
    before_output = output.read_bytes()

    completed, result = run_cli(supplier, output)

    assert completed.returncode == 2
    assert result["outcome"] == "BLOCKED"
    assert result["diagnostics"][0]["code"] == expected_code
    assert supplier.read_bytes() == original
    assert output.read_bytes() == before_output


@pytest.mark.parametrize(
    "member",
    [
        "xl/vbaProject.bin",
        "xl/externalLinks/externalLink1.xml",
        "xl/embeddings/oleObject1.bin",
    ],
)
def test_cli_rejects_active_or_external_zip_content_without_output(
    tmp_path: Path,
    member: str,
) -> None:
    supplier = tmp_path / "unsafe.xlsx"
    output = tmp_path / "formal.xlsx"
    make_supplier(supplier)
    with zipfile.ZipFile(supplier, "a") as archive:
        archive.writestr(member, b"unsafe")

    completed, result = run_cli(supplier, output)

    assert completed.returncode == 2
    assert result["diagnostics"][0]["code"] == "UNSAFE_WORKBOOK_CONTENT"
    assert not output.exists()


def mutate_extra_sheet(workbook: openpyxl.Workbook) -> None:
    workbook.create_sheet("不受支持")


def mutate_missing_table(workbook: openpyxl.Workbook) -> None:
    del workbook["90-填写选项"].tables["SupplierUATOptionTable"]


def mutate_header(workbook: openpyxl.Workbook) -> None:
    workbook["01-需求故事"]["A4"] = "需求变更"


def mutate_merged_business_cells(workbook: openpyxl.Workbook) -> None:
    workbook["01-需求故事"].merge_cells("A5:B5")


def mutate_formula(workbook: openpyxl.Workbook) -> None:
    workbook["01-需求故事"]["H5"] = "通过"


def mutate_contract(workbook: openpyxl.Workbook) -> None:
    workbook["90-填写选项"]["B1"] = "supplier-estimate-input-v0"


@pytest.mark.parametrize(
    ("mutator", "expected_code"),
    [
        (mutate_extra_sheet, "UNEXPECTED_SHEETS"),
        (mutate_missing_table, "TABLE_CONTRACT_INVALID"),
        (mutate_header, "HEADER_CONTRACT_INVALID"),
        (mutate_merged_business_cells, "MERGED_BUSINESS_CELLS"),
        (mutate_formula, "FORMULA_CONTRACT_INVALID"),
        (mutate_contract, "CONTRACT_VERSION_UNSUPPORTED"),
    ],
)
def test_cli_rejects_supplier_contract_tampering(
    tmp_path: Path,
    mutator: Callable[[openpyxl.Workbook], None],
    expected_code: str,
) -> None:
    supplier = tmp_path / "tampered.xlsx"
    output = tmp_path / "formal.xlsx"
    make_supplier(supplier)
    workbook = openpyxl.load_workbook(supplier, data_only=False)
    try:
        mutator(workbook)
        workbook.save(supplier)
    finally:
        workbook.close()

    completed, result = run_cli(supplier, output)

    assert completed.returncode == 2
    assert result["diagnostics"][0]["code"] == expected_code
    assert not output.exists()


def assert_business_error(
    tmp_path: Path,
    stories: list[dict[str, str]],
    tasks: list[dict[str, str]],
    expected: dict[str, object],
) -> None:
    supplier = tmp_path / "invalid.xlsx"
    output = tmp_path / "formal.xlsx"
    make_supplier(supplier, stories, tasks)

    completed, result = run_cli(supplier, output)

    assert completed.returncode == 2
    assert result["outcome"] == "BLOCKED"
    assert result["diagnostics"][0] == expected
    assert not output.exists()


@pytest.mark.parametrize(
    ("case", "expected"),
    [
        (
            "required",
            {
                "sheet": "01-需求故事",
                "row": 5,
                "field": "验收条件",
                "code": "REQUIRED_VALUE_MISSING",
                "message": "请填写该必填字段后重试。",
            },
        ),
        (
            "uat",
            {
                "sheet": "01-需求故事",
                "row": 5,
                "field": "UAT适用",
                "code": "ENUM_INVALID",
                "message": "请选择模板下拉列表中的受支持值。",
            },
        ),
        (
            "complexity",
            {
                "sheet": "02-任务清单",
                "row": 5,
                "field": "复杂度",
                "code": "ENUM_INVALID",
                "message": "请选择模板下拉列表中的受支持值。",
            },
        ),
        (
            "unknown-type",
            {
                "sheet": "02-任务清单",
                "row": 5,
                "field": "任务类型",
                "code": "TASK_TYPE_UNKNOWN",
                "message": "请选择“90-填写选项”中存在的任务类型。",
            },
        ),
        (
            "mode-enum",
            {
                "sheet": "02-任务清单",
                "row": 5,
                "field": "工作方式",
                "code": "ENUM_INVALID",
                "message": "请选择模板下拉列表中的受支持值。",
            },
        ),
        (
            "mode-not-allowed",
            {
                "sheet": "02-任务清单",
                "row": 5,
                "field": "工作方式",
                "code": "WORK_MODE_NOT_ALLOWED",
                "message": "请选择该任务类型允许的工作方式。",
            },
        ),
        (
            "parent-conflict",
            {
                "sheet": "01-需求故事",
                "row": 6,
                "field": "子需求",
                "code": "PARENT_CONFLICT",
                "message": "同一子需求只能属于一个需求，请统一父级。",
            },
        ),
        (
            "duplicate-story",
            {
                "sheet": "01-需求故事",
                "row": 6,
                "field": "故事",
                "code": "DUPLICATE_STORY",
                "message": "故事名称必须唯一，请修改重复名称。",
            },
        ),
        (
            "duplicate-task",
            {
                "sheet": "02-任务清单",
                "row": 6,
                "field": "任务名称",
                "code": "DUPLICATE_TASK",
                "message": "任务名称必须唯一，请修改重复名称。",
            },
        ),
        (
            "unknown-story",
            {
                "sheet": "02-任务清单",
                "row": 5,
                "field": "所属故事",
                "code": "STORY_UNKNOWN",
                "message": "请将任务关联到工作簿中唯一存在的故事路径。",
            },
        ),
        (
            "story-without-task",
            {
                "sheet": "01-需求故事",
                "row": 6,
                "field": "故事",
                "code": "STORY_WITHOUT_TASK",
                "message": "请至少为该故事填写一个任务。",
            },
        ),
    ],
)
def test_cli_reports_literal_business_diagnostics(
    tmp_path: Path,
    case: str,
    expected: dict[str, object],
) -> None:
    stories, tasks = valid_rows()
    if case == "required":
        stories[0]["验收条件"] = ""
    elif case == "uat":
        stories[0]["UAT适用"] = "可能"
    elif case == "complexity":
        tasks[0]["复杂度"] = "X"
    elif case == "unknown-type":
        tasks[0]["任务类型"] = "不存在的任务类型"
    elif case == "mode-enum":
        tasks[0]["工作方式"] = "复制"
    elif case == "mode-not-allowed":
        tasks[0]["工作方式"] = "接入复用"
    elif case == "parent-conflict":
        stories[1]["需求"] = "需求乙"
        tasks[2]["所属故事"] = "需求乙 > 子需求甲 > 故事 B"
    elif case == "duplicate-story":
        stories[1]["故事"] = "故事 A"
        tasks[2]["所属故事"] = "需求甲 > 子需求甲 > 故事 A"
    elif case == "duplicate-task":
        tasks[1]["任务名称"] = "任务 1"
    elif case == "unknown-story":
        tasks[0]["所属故事"] = "需求甲 > 子需求甲 > 不存在"
    elif case == "story-without-task":
        tasks = tasks[:2]
    else:
        raise AssertionError(case)
    assert_business_error(tmp_path, stories, tasks, expected)
