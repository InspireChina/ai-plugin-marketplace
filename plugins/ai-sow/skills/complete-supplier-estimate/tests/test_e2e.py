from __future__ import annotations

import copy
import hashlib
import json
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.formula import ArrayFormula


SKILL_ROOT = Path(__file__).resolve().parents[1]
PLUGIN_ROOT = SKILL_ROOT.parents[1]
SCRIPT = SKILL_ROOT / "scripts/complete_supplier_estimate.py"
SUPPLIER_TEMPLATE = SKILL_ROOT / "assets/supplier-estimate-input.xlsx"
EXCEL_ERRORS = {"#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#N/A"}


def _table(workbook: openpyxl.Workbook, name: str) -> tuple[object, object]:
    matches = [
        (worksheet, worksheet.tables[name])
        for worksheet in workbook.worksheets
        if name in worksheet.tables
    ]
    assert len(matches) == 1
    return matches[0]


def _copy_style(source: object, target: object) -> None:
    target.font = copy.copy(source.font)
    target.fill = copy.copy(source.fill)
    target.border = copy.copy(source.border)
    target.alignment = copy.copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy.copy(source.protection)


def _replace_rows(
    workbook: openpyxl.Workbook,
    table_name: str,
    rows: list[dict[str, str]],
) -> None:
    worksheet, table = _table(workbook, table_name)
    min_col, min_row, max_col, old_max_row = range_boundaries(table.ref)
    headers = [
        str(worksheet.cell(min_row, column).value)
        for column in range(min_col, max_col + 1)
    ]
    prototypes = [
        copy.copy(worksheet.cell(min_row + 1, column))
        for column in range(min_col, max_col + 1)
    ]
    for row_number in range(min_row + 1, max(old_max_row, min_row + len(rows)) + 1):
        for column in range(min_col, max_col + 1):
            worksheet.cell(row_number, column).value = None
    for offset, payload in enumerate(rows, start=1):
        row_number = min_row + offset
        for column_offset, header in enumerate(headers):
            source = prototypes[column_offset]
            target = worksheet.cell(row_number, min_col + column_offset)
            _copy_style(source, target)
            if source.data_type == "f":
                source_formula = (
                    source.value.text
                    if isinstance(source.value, ArrayFormula)
                    else source.value
                )
                translated = Translator(
                    source_formula,
                    origin=source.coordinate,
                ).translate_formula(target.coordinate)
                target.value = (
                    ArrayFormula(ref=target.coordinate, text=translated)
                    if isinstance(source.value, ArrayFormula)
                    else translated
                )
            else:
                target.value = payload.get(header)
    max_row = min_row + len(rows)
    table.ref = (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{max_row}"
    )
    if table.autoFilter is not None:
        table.autoFilter.ref = table.ref


def _summary_values(workbook: openpyxl.Workbook) -> list[float]:
    worksheet, table = _table(workbook, "ProjectSummaryTable")
    min_col, min_row, _, max_row = range_boundaries(table.ref)
    return [
        float(worksheet.cell(row, min_col + 1).value)
        for row in range(min_row + 1, max_row + 1)
    ]


def test_supplier_completion_recalculates_in_libreoffice(tmp_path: Path) -> None:
    soffice = shutil.which("soffice")
    if soffice is None:
        pytest.skip("soffice is unavailable")

    supplier = tmp_path / "supplier.xlsx"
    formal = tmp_path / "formal.xlsx"
    workbook = openpyxl.load_workbook(SUPPLIER_TEMPLATE, data_only=False)
    try:
        _replace_rows(
            workbook,
            "SOWStoryTable",
            [
                {
                    "需求": "需求甲",
                    "子需求": "子需求甲",
                    "故事": "故事 A",
                    "UAT适用": "是",
                    "验收条件": "故事 A 可验收",
                    "备注": "",
                },
                {
                    "需求": "需求甲",
                    "子需求": "子需求乙",
                    "故事": "故事 B",
                    "UAT适用": "否",
                    "验收条件": "故事 B 可验收",
                    "备注": "",
                },
            ],
        )
        _replace_rows(
            workbook,
            "TaskTable",
            [
                {
                    "所属故事": "需求甲 > 子需求甲 > 故事 A",
                    "任务名称": "界面任务",
                    "任务类型": "界面与交互",
                    "工作方式": "新建",
                    "复杂度": "M",
                    "备注": "界面交付",
                },
                {
                    "所属故事": "需求甲 > 子需求甲 > 故事 A",
                    "任务名称": "内部对接任务",
                    "任务类型": "内部系统对接",
                    "工作方式": "新建",
                    "复杂度": "S",
                    "备注": "内部对接",
                },
                {
                    "所属故事": "需求甲 > 子需求乙 > 故事 B",
                    "任务名称": "外部对接任务",
                    "任务类型": "外部系统对接",
                    "工作方式": "调整",
                    "复杂度": "L",
                    "备注": "外部对接",
                },
            ],
        )
        workbook.save(supplier)
    finally:
        workbook.close()

    source_digest = hashlib.sha256(supplier.read_bytes()).hexdigest()
    completed = subprocess.run(
        [
            sys.executable,
            str(SCRIPT),
            "--input",
            str(supplier),
            "--output",
            str(formal),
        ],
        cwd=PLUGIN_ROOT,
        text=True,
        encoding="utf-8",
        capture_output=True,
        check=False,
    )
    assert completed.returncode == 0, (completed.stderr, completed.stdout)
    assert json.loads(completed.stdout)["outcome"] == "OK"
    assert hashlib.sha256(supplier.read_bytes()).hexdigest() == source_digest

    recalculated_dir = tmp_path / "recalculated"
    profile_dir = tmp_path / "libreoffice-profile"
    recalculated_dir.mkdir()
    profile_dir.mkdir()
    recalculated = subprocess.run(
        [
            soffice,
            f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
            "--headless",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(recalculated_dir),
            str(formal),
        ],
        text=True,
        encoding="utf-8",
        capture_output=True,
        check=False,
    )
    assert recalculated.returncode == 0, (recalculated.stdout, recalculated.stderr)
    recalculated_path = recalculated_dir / formal.name
    assert recalculated_path.is_file(), (recalculated.stdout, recalculated.stderr)

    formula_book = openpyxl.load_workbook(recalculated_path, data_only=False)
    value_book = openpyxl.load_workbook(recalculated_path, data_only=True)
    try:
        assert _summary_values(value_book) == [7.0, 2.0, 0.5, 9.5]
        for formula_sheet, value_sheet in zip(
            formula_book.worksheets,
            value_book.worksheets,
            strict=True,
        ):
            for row in formula_sheet.iter_rows():
                for formula_cell in row:
                    if formula_cell.data_type != "f":
                        continue
                    cached = value_sheet[formula_cell.coordinate].value
                    assert cached not in EXCEL_ERRORS, (
                        formula_sheet.title,
                        formula_cell.coordinate,
                        formula_cell.value,
                        cached,
                    )
    finally:
        formula_book.close()
        value_book.close()

    assert hashlib.sha256(supplier.read_bytes()).hexdigest() == source_digest
