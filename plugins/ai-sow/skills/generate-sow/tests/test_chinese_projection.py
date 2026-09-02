from __future__ import annotations

import importlib.util
import json
from pathlib import Path

import openpyxl
import pytest
from openpyxl.utils import range_boundaries


MODULE_PATH = Path(__file__).resolve().parents[1] / "scripts/workbook.py"
SPEC = importlib.util.spec_from_file_location("generate_sow_workbook", MODULE_PATH)
assert SPEC is not None and SPEC.loader is not None
MODULE = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(MODULE)


def test_user_text_that_looks_like_formula_is_written_as_text() -> None:
    assert MODULE.safe_text("=SUM(A1:A2)") == "'=SUM(A1:A2)"
    assert MODULE.safe_text("+1") == "'+1"
    assert MODULE.safe_text("-1") == "'-1"
    assert MODULE.safe_text("@name") == "'@name"
    assert MODULE.safe_text("普通文本") == "普通文本"


def test_projection_rejects_blank_or_ambiguous_visible_names() -> None:
    with pytest.raises(ValueError, match="display name is blank"):
        MODULE.require_unique_names([{"name": ""}], "Story")
    with pytest.raises(ValueError, match="display name is duplicated"):
        MODULE.require_unique_names(
            [{"name": "重复名称"}, {"name": "重复名称"}],
            "Feature",
        )
    with pytest.raises(ValueError, match="after Excel projection"):
        MODULE.require_unique_names(
            [{"name": "=名称"}, {"name": "'=名称"}],
            "Task",
        )


def fixture_data() -> tuple[Path, dict[str, dict[str, object]]]:
    fixture = Path(__file__).resolve().parents[1] / "fixtures/project/.ai-sow"

    def read(relative: str) -> dict[str, object]:
        return json.loads((fixture / relative).read_text(encoding="utf-8"))

    business = read("data/analyze-requirement/requirements.json")
    technical = read("data/generate-design/requirements.json")
    return fixture, {
        "requirements": {
            "epics": [*business["epics"], *technical["epics"]],  # type: ignore[index]
            "features": [*business["features"], *technical["features"]],  # type: ignore[index]
        },
        "asis": read("data/analyze-as-is/asis.json"),
        "design": read("data/generate-design/design.json"),
        "technicalRequirements": technical,
        "delivery": read("data/generate-story/delivery.json"),
        "estimate": read("data/generate-task/estimate.json"),
    }


def table_rows(workbook: openpyxl.Workbook, table_name: str) -> list[dict[str, object]]:
    worksheet = next(
        sheet for sheet in workbook.worksheets if table_name in sheet.tables
    )
    table = worksheet.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = [
        str(worksheet.cell(min_row, column).value)
        for column in range(min_col, max_col + 1)
    ]
    return [
        {
            header: worksheet.cell(row, column).value
            for header, column in zip(
                headers,
                range(min_col, max_col + 1),
                strict=True,
            )
        }
        for row in range(min_row + 1, max_row + 1)
    ]


def test_generated_workbook_preserves_formula_like_user_content_as_text(tmp_path: Path) -> None:
    fixture, data = fixture_data()
    data["requirements"]["epics"][0]["name"] = "=FORMULA_LIKE_NAME"  # type: ignore[index]
    data["delivery"]["stories"][0]["description"] = "+普通备注"  # type: ignore[index]
    data["estimate"]["tasks"][0]["rationale"] = "@任务说明"  # type: ignore[index]
    custom_template = tmp_path / "custom-template.xlsx"
    template_workbook = openpyxl.load_workbook(
        fixture / "templates/sow-template.xlsx",
        data_only=False,
    )
    try:
        template_workbook["01-需求故事"].freeze_panes = "C6"
        template_workbook.save(custom_template)
    finally:
        template_workbook.close()

    output = tmp_path / "safe-text.xlsx"
    MODULE.write_workbook(custom_template, data, output)

    workbook = openpyxl.load_workbook(output, data_only=False)
    try:
        assert workbook["01-需求故事"].freeze_panes == "C6"
        story = table_rows(workbook, "SOWStoryTable")[0]
        assert story["需求"] == "'=FORMULA_LIKE_NAME"
        assert story["备注"] == "'+普通备注"
        task = table_rows(workbook, "TaskTable")[0]
        assert str(task["备注"]).startswith("任务理由：@任务说明")
        story_sheet = workbook["01-需求故事"]
        story_table = story_sheet.tables["SOWStoryTable"]
        min_col, min_row, _, _ = range_boundaries(story_table.ref)
        assert story_sheet.cell(min_row + 1, min_col).data_type == "s"
    finally:
        workbook.close()


def test_generated_workbook_expands_wrapped_rows_for_visible_long_text(tmp_path: Path) -> None:
    fixture, data = fixture_data()
    long_text = "这是一段用于验证最终可见文本自动换行行高的内容。" * 16
    data["delivery"]["stories"][0]["description"] = long_text  # type: ignore[index]
    data["delivery"]["acceptanceCriteria"][0]["name"] = long_text  # type: ignore[index]
    data["estimate"]["tasks"][0]["rationale"] = long_text  # type: ignore[index]

    template = fixture / "templates/sow-template.xlsx"
    output = tmp_path / "long-text.xlsx"
    MODULE.write_workbook(template, data, output)

    workbook = openpyxl.load_workbook(output, data_only=False)
    template_workbook = openpyxl.load_workbook(template, data_only=False)
    try:
        for table_name in ("SOWStoryTable", "TaskTable"):
            worksheet = next(
                sheet for sheet in workbook.worksheets if table_name in sheet.tables
            )
            table = worksheet.tables[table_name]
            _, min_row, _, _ = range_boundaries(table.ref)
            template_height = (
                template_workbook[worksheet.title].row_dimensions[min_row + 1].height
                or 15
            )
            assert (worksheet.row_dimensions[min_row + 1].height or 15) > template_height
    finally:
        workbook.close()
        template_workbook.close()
