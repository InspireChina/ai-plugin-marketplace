from __future__ import annotations

import hashlib
import json
from pathlib import Path
from zipfile import ZipFile

import openpyxl
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.formula import ArrayFormula


SKILL_ROOT = Path(__file__).resolve().parents[1]
PLUGIN_TEMPLATE = SKILL_ROOT.parent.parent / "assets/sow-template.xlsx"
GENERATE_TASK_FIXTURE = SKILL_ROOT.parent / "generate-task/fixtures/sow-template.xlsx"
GENERATE_SOW_FIXTURE = (
    SKILL_ROOT / "fixtures/project/.ai-sow/templates/sow-template.xlsx"
)

EXPECTED_SHEETS = [
    "01-需求故事",
    "02-任务清单",
    "03-工作量汇总",
    "90-估算标准",
]
EXPECTED_HEADERS = {
    "SOWStoryTable": [
        "需求",
        "子需求",
        "故事",
        "UAT适用",
        "验收条件",
        "备注",
        "任务列表",
        "故事人天",
        "校验结果",
        "故事路径",
    ],
    "TaskTable": [
        "所属故事",
        "任务名称",
        "任务类型",
        "工作方式",
        "复杂度",
        "备注",
        "M档标准人天",
        "复杂度系数",
        "任务人天",
        "SIT支持人天",
        "校验结果",
    ],
    "ProjectSummaryTable": ["工作量项", "人天"],
    "ProjectParameterTable": [
        "参数代码",
        "名称",
        "值",
        "单位",
        "适用范围",
        "验证状态/说明",
    ],
    "BaseUnitCatalogTable": [
        "任务族ID",
        "任务族名称",
        "基础单元ID",
        "基础单元名称",
        "计数口径",
        "包含内容",
        "不包含内容",
        "新建M档人天",
        "调整M档人天",
        "接入复用M档人天",
        "S标准",
        "M标准",
        "L标准",
        "X/拆分条件",
    ],
}
FORMULA_COLUMNS = {
    "SOWStoryTable": {"任务列表", "故事人天", "校验结果", "故事路径"},
    "TaskTable": {
        "M档标准人天",
        "复杂度系数",
        "任务人天",
        "SIT支持人天",
        "校验结果",
    },
}
EXPECTED_PARAMETER_ROWS = [
    ["K_COMPLEXITY_S", "S档复杂度系数", 0.6, "倍", "任务人天", "固定规则"],
    ["K_COMPLEXITY_M", "M档复杂度系数", 1, "倍", "任务人天", "固定规则"],
    ["K_COMPLEXITY_L", "L档复杂度系数", 1.5, "倍", "任务人天", "固定规则"],
    ["K_UAT", "UAT支持系数", 0.05, "%", "显式标记 UAT 适用的故事直接人天", "待样本校准"],
    ["SIT_INT_SUPPORT", "内部集成点SIT支持", 0.5, "人天/点", "开发交付估算", "待样本校准"],
    ["SIT_EXT_SUPPORT", "外部集成点SIT支持", 1, "人天/点", "开发交付估算", "待样本校准"],
    ["ROUND_STORY", "故事取整粒度", 0.5, "人天", "任务合计后向上取整", "固定规则"],
    ["ROUND_PROJECT", "项目级取整粒度", 0.5, "人天", "SIT、UAT分别向上取整；总计不再次取整", "固定规则"],
]
EXPECTED_CATALOG_SHA256 = "060ebfe4dc8e5643520b6e25c0e616fbb47d6aa91fc03e6d2c1bc23d92ccf6a7"


def table_location(
    workbook: openpyxl.Workbook,
    table_name: str,
) -> tuple[openpyxl.worksheet.worksheet.Worksheet, object]:
    for worksheet in workbook.worksheets:
        if table_name in worksheet.tables:
            return worksheet, worksheet.tables[table_name]
    raise AssertionError(f"missing table: {table_name}")


def table_rows(workbook: openpyxl.Workbook, table_name: str) -> list[list[object]]:
    worksheet, table = table_location(workbook, table_name)
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    return [
        [worksheet.cell(row, column).value for column in range(min_col, max_col + 1)]
        for row in range(min_row, max_row + 1)
    ]


def formula_text(value: object) -> str:
    if isinstance(value, ArrayFormula):
        return value.text or ""
    return str(value or "")


def test_plugin_template_is_the_only_runtime_authority_and_fixtures_match() -> None:
    assert PLUGIN_TEMPLATE.is_file()
    assert not (SKILL_ROOT.parent / "setup/assets/sow-template.xlsx").exists()
    authoritative = PLUGIN_TEMPLATE.read_bytes()
    assert GENERATE_TASK_FIXTURE.read_bytes() == authoritative
    assert GENERATE_SOW_FIXTURE.read_bytes() == authoritative


def test_formal_template_has_exact_sheets_tables_headers_and_prototypes() -> None:
    workbook = openpyxl.load_workbook(PLUGIN_TEMPLATE, data_only=False)
    try:
        assert workbook.sheetnames == EXPECTED_SHEETS
        actual_tables = {
            name
            for worksheet in workbook.worksheets
            for name in worksheet.tables
        }
        assert actual_tables == set(EXPECTED_HEADERS)
        for table_name, headers in EXPECTED_HEADERS.items():
            rows = table_rows(workbook, table_name)
            assert rows[0] == headers
            if table_name in {"SOWStoryTable", "TaskTable"}:
                assert len(rows) == 2
                assert any(cell is not None for cell in rows[1])
    finally:
        workbook.close()


def test_formal_template_scopes_filters_to_tables_only() -> None:
    workbook = openpyxl.load_workbook(PLUGIN_TEMPLATE, data_only=False)
    try:
        for worksheet in workbook.worksheets:
            if not worksheet.tables:
                continue
            assert worksheet.auto_filter.ref is None
            for table_name in worksheet.tables:
                table = worksheet.tables[table_name]
                assert table.autoFilter is not None
                assert table.autoFilter.ref == table.ref
    finally:
        workbook.close()


def test_business_input_and_formula_columns_have_distinct_editability() -> None:
    workbook = openpyxl.load_workbook(PLUGIN_TEMPLATE, data_only=False)
    try:
        for table_name in ("SOWStoryTable", "TaskTable"):
            worksheet, table = table_location(workbook, table_name)
            min_col, min_row, _, _ = range_boundaries(table.ref)
            headers = table_rows(workbook, table_name)[0]
            for offset, header in enumerate(headers):
                cell = worksheet.cell(min_row + 1, min_col + offset)
                if header in FORMULA_COLUMNS[table_name]:
                    assert cell.protection.locked is True
                    assert cell.fill.fgColor.rgb[-6:] == "F1F4F6"
                    assert formula_text(cell.value).startswith("=")
                else:
                    assert cell.protection.locked is False
                    assert cell.fill.fgColor.rgb[-6:] == "FFFFFF"
            assert worksheet.protection.sheet is True
            assert worksheet.protection.insertRows is False
            assert worksheet.protection.autoFilter is False
            assert worksheet.protection.sort is False
            for merged in worksheet.merged_cells.ranges:
                assert merged.max_row < min_row + 1 or merged.min_row > min_row + 1
        assert workbook["01-需求故事"].column_dimensions["J"].hidden is True
    finally:
        workbook.close()


def test_business_validations_have_no_legacy_row_cap() -> None:
    workbook = openpyxl.load_workbook(PLUGIN_TEMPLATE, data_only=False)
    try:
        expected_columns = {
            "01-需求故事": {4},
            "02-任务清单": {1, 3, 4, 5},
        }
        for sheet_name, columns in expected_columns.items():
            worksheet = workbook[sheet_name]
            covered: set[int] = set()
            for validation in worksheet.data_validations.dataValidation:
                for cell_range in validation.sqref.ranges:
                    if cell_range.max_row == 1_048_576:
                        covered.add(cell_range.min_col)
            assert covered == columns
    finally:
        workbook.close()


def test_formulas_are_template_owned_safe_and_risk_free() -> None:
    workbook = openpyxl.load_workbook(PLUGIN_TEMPLATE, data_only=False)
    try:
        formulas: list[str] = []
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    text = formula_text(cell.value)
                    if text.startswith("="):
                        formulas.append(text)
        assert formulas
        for formula in formulas:
            assert "#REF!" not in formula
            assert "_xlfn._xlws." not in formula
            assert "风险" not in formula
            assert "0.6" not in formula
            assert "1.5" not in formula
            assert "2.5" not in formula
        with ZipFile(PLUGIN_TEMPLATE) as archive:
            xml = b"\n".join(archive.read(name) for name in archive.namelist())
        assert "风险".encode() not in xml
    finally:
        workbook.close()


def test_summary_has_exactly_four_rows_and_no_risk_dimension() -> None:
    workbook = openpyxl.load_workbook(PLUGIN_TEMPLATE, data_only=False)
    try:
        rows = table_rows(workbook, "ProjectSummaryTable")
        assert [row[0] for row in rows[1:]] == [
            "直接开发人天",
            "SIT支持人天",
            "UAT支持人天",
            "总开发人天",
        ]
        assert all(formula_text(row[1]).startswith("=") for row in rows[1:])
    finally:
        workbook.close()


def test_estimation_standard_tables_preserve_v13_values_and_modes() -> None:
    workbook = openpyxl.load_workbook(PLUGIN_TEMPLATE, data_only=False)
    try:
        parameters = table_rows(workbook, "ProjectParameterTable")
        assert parameters[0] == EXPECTED_HEADERS["ProjectParameterTable"]
        assert parameters[1:] == EXPECTED_PARAMETER_ROWS

        catalog = table_rows(workbook, "BaseUnitCatalogTable")
        assert catalog[0] == EXPECTED_HEADERS["BaseUnitCatalogTable"]
        assert len(catalog[1:]) == 37
        assert len({row[1] for row in catalog[1:]}) == 13
        assert len({row[2] for row in catalog[1:]}) == 37
        payload = json.dumps(
            catalog,
            ensure_ascii=False,
            separators=(",", ":"),
            default=str,
        ).encode()
        assert hashlib.sha256(payload).hexdigest() == EXPECTED_CATALOG_SHA256
        for row in catalog[1:]:
            allowed = {
                mode
                for mode, value in zip(("新建", "调整", "接入复用"), row[7:10])
                if isinstance(value, (int, float)) and value > 0
            }
            assert allowed
    finally:
        workbook.close()


def test_generate_sow_runtime_does_not_read_another_skill_asset() -> None:
    forbidden = ("skills/setup/", "skills/generate-task/")
    for path in SKILL_ROOT.joinpath("scripts").glob("*.py"):
        text = path.read_text(encoding="utf-8")
        assert all(fragment not in text for fragment in forbidden)
