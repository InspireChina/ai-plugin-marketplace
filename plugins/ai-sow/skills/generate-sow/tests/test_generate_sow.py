from __future__ import annotations

import hashlib
import importlib.util
import json
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.formula import ArrayFormula


SKILL_ROOT = Path(__file__).resolve().parents[1]
PLUGIN_ROOT = SKILL_ROOT.parents[1]
SCRIPT = SKILL_ROOT / "scripts/generate_sow.py"
FIXTURE = SKILL_ROOT / "fixtures/project"
BUSINESS_TABLES = {"SOWStoryTable", "TaskTable"}
FORMAL_TABLES = {
    *BUSINESS_TABLES,
    "ProjectSummaryTable",
    "ProjectParameterTable",
    "BaseUnitCatalogTable",
}
REMOVED_TABLES = {
    "EpicTable",
    "FeatureTable",
    "AcceptanceCriterionTable",
    "IntegrationTable",
    "AssumptionRiskTable",
    "AsIsDetailTable",
}
SPEC = importlib.util.spec_from_file_location("generate_sow_script", SCRIPT)
assert SPEC is not None and SPEC.loader is not None
GENERATOR = importlib.util.module_from_spec(SPEC)
sys.path.insert(0, str(SCRIPT.parent))
SPEC.loader.exec_module(GENERATOR)
WORKBOOK = sys.modules["workbook"]


def test_skill_uses_current_stage_without_leaf_agents() -> None:
    skill = (SKILL_ROOT / "SKILL.md").read_text(encoding="utf-8")
    assert "当前 Stage Agent 是本 Skill 的唯一用户接口" in skill
    assert "直接运行确定性生成器" in skill
    for forbidden in ("Orchestrator", "Worker", "Validator Agent", "Reviewer"):
        assert forbidden not in skill


def copy_project(tmp_path: Path) -> Path:
    project = tmp_path / "project"
    project.parent.mkdir(parents=True, exist_ok=True)
    shutil.copytree(FIXTURE, project)
    return project


def run_generator(project: Path) -> tuple[subprocess.CompletedProcess[str], dict[str, object]]:
    completed = subprocess.run(
        [sys.executable, str(SCRIPT), "--project-root", str(project)],
        cwd=PLUGIN_ROOT,
        text=True,
        encoding="utf-8",
        capture_output=True,
        check=False,
    )
    payload = json.loads(completed.stdout)
    return completed, payload


def package_tree(root: Path) -> dict[str, str]:
    return {
        path.relative_to(root).as_posix(): hashlib.sha256(path.read_bytes()).hexdigest()
        for path in sorted(root.rglob("*"))
        if path.is_file()
    }


def package_tree_sha256(root: Path) -> str:
    entries = [
        {"path": path, "sha256": digest}
        for path, digest in package_tree(root).items()
    ]
    return hashlib.sha256(GENERATOR.canonical_json_bytes(entries)).hexdigest()


def table_index(workbook: openpyxl.Workbook) -> dict[str, tuple[object, object]]:
    return {
        name: (worksheet, worksheet.tables[name])
        for worksheet in workbook.worksheets
        for name in worksheet.tables
    }


def table_rows(workbook: openpyxl.Workbook, table_name: str) -> list[dict[str, object]]:
    worksheet, table = table_index(workbook)[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = [
        str(worksheet.cell(min_row, column).value)
        for column in range(min_col, max_col + 1)
    ]
    return [
        {
            header: worksheet.cell(row, column).value
            for header, column in zip(
                headers,
                range(min_col, max_col + 1),
                strict=True,
            )
        }
        for row in range(min_row + 1, max_row + 1)
    ]


def load_workbook_data(project: Path) -> dict[str, dict[str, object]]:
    def read(relative: str) -> dict[str, object]:
        return json.loads((project / relative).read_text(encoding="utf-8"))

    business = read(".ai-sow/data/analyze-requirement/requirements.json")
    technical = read(".ai-sow/data/generate-design/requirements.json")
    return {
        "requirements": {
            "epics": [*business["epics"], *technical["epics"]],  # type: ignore[index]
            "features": [*business["features"], *technical["features"]],  # type: ignore[index]
        },
        "asis": read(".ai-sow/data/analyze-as-is/asis.json"),
        "design": read(".ai-sow/data/generate-design/design.json"),
        "technicalRequirements": technical,
        "delivery": read(".ai-sow/data/generate-story/delivery.json"),
        "estimate": read(".ai-sow/data/generate-task/estimate.json"),
    }


def test_build_rows_projects_only_story_and_task_inputs() -> None:
    data = load_workbook_data(FIXTURE)
    workbook = openpyxl.load_workbook(
        FIXTURE / ".ai-sow/templates/sow-template.xlsx",
        data_only=False,
    )
    try:
        rows = WORKBOOK.build_rows(data, WORKBOOK.base_unit_name_map(workbook))
    finally:
        workbook.close()

    assert set(rows) == BUSINESS_TABLES
    delivery = data["delivery"]
    estimate = data["estimate"]
    requirements = data["requirements"]
    epics = {item["epicId"]: item for item in requirements["epics"]}  # type: ignore[index]
    features = {item["featureId"]: item for item in requirements["features"]}  # type: ignore[index]
    acceptance_by_story: dict[str, list[str]] = {}
    for criterion in delivery["acceptanceCriteria"]:  # type: ignore[index]
        acceptance_by_story.setdefault(criterion["storyId"], []).append(criterion["name"])

    first_story = delivery["stories"][0]  # type: ignore[index]
    first_feature = features[first_story["featureId"]]
    first_epic = epics[first_feature["epicId"]]
    assert rows["SOWStoryTable"][0] == {
        "需求": first_epic["name"],
        "子需求": first_feature["name"],
        "故事": first_story["name"],
        "UAT适用": "是",
        "验收条件": "\n".join(acceptance_by_story[first_story["storyId"]]),
        "备注": first_story["description"],
    }

    first_task = estimate["tasks"][0]  # type: ignore[index]
    story = next(
        item for item in delivery["stories"]  # type: ignore[index]
        if item["storyId"] == first_task["storyId"]
    )
    feature = features[story["featureId"]]
    epic = epics[feature["epicId"]]
    assert rows["TaskTable"][0] == {
        "所属故事": f'{epic["name"]} > {feature["name"]} > {story["name"]}',
        "任务名称": first_task["name"],
        "任务类型": "界面与交互",
        "工作方式": first_task["workMode"],
        "复杂度": first_task["complexity"],
        "备注": (
            f'任务理由：{first_task["rationale"]}\n'
            f'工作方式理由：{first_task["workModeRationale"]}\n'
            f'复杂度理由：{first_task["complexityRationale"]}'
        ),
    }


def test_unprojected_names_do_not_create_false_display_collisions() -> None:
    data = load_workbook_data(FIXTURE)
    data["asis"]["effectiveStartItems"][1]["name"] = data["asis"]["effectiveStartItems"][0]["name"]  # type: ignore[index]
    data["delivery"]["assumptions"].append(  # type: ignore[index]
        {**data["delivery"]["assumptions"][0], "assumptionId": "assumption-duplicate"}  # type: ignore[index]
    )
    workbook = openpyxl.load_workbook(
        FIXTURE / ".ai-sow/templates/sow-template.xlsx",
        data_only=False,
    )
    try:
        rows = WORKBOOK.build_rows(data, WORKBOOK.base_unit_name_map(workbook))
    finally:
        workbook.close()
    assert set(rows) == BUSINESS_TABLES


def test_receipt_only_generation_is_deterministic_and_reuses_identical_package(tmp_path: Path) -> None:
    first_project = copy_project(tmp_path / "first")
    second_project = copy_project(tmp_path / "second")

    first, first_result = run_generator(first_project)
    second, second_result = run_generator(second_project)

    assert first.returncode == second.returncode == 0, (
        first.stderr,
        first.stdout,
        second.stderr,
        second.stdout,
    )
    assert first_result["outcome"] == second_result["outcome"] == "OK"
    assert first_result["packageId"] == second_result["packageId"]
    first_package = first_project / str(first_result["packagePath"])
    second_package = second_project / str(second_result["packagePath"])
    assert package_tree(first_package) == package_tree(second_package)
    assert (first_package / "sow.xlsx").read_bytes() == (
        second_package / "sow.xlsx"
    ).read_bytes()
    assert first_result["generatorContract"] == "receipt-only-v4"
    assert first_result["workbookSha256"] == hashlib.sha256(
        (first_package / "sow.xlsx").read_bytes()
    ).hexdigest()
    assert first_result["manifestSha256"] == hashlib.sha256(
        (first_package / "manifest.json").read_bytes()
    ).hexdigest()
    assert first_result["packageTreeSha256"] == package_tree_sha256(first_package)
    assert first_result["fileCount"] == len(package_tree(first_package))

    repeated, repeated_result = run_generator(first_project)
    assert repeated.returncode == 0
    assert repeated_result["publication"] == "REUSED"


def test_package_fingerprint_binds_every_source_name_path_and_hash(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    project_root = copy_project(tmp_path)
    files = GENERATOR.ProjectFiles.open(project_root)
    project = json.loads((project_root / GENERATOR.PROJECT_PATH).read_text(encoding="utf-8"))

    payload = GENERATOR.package_fingerprint_payload(files, project)

    def expected(name: str, package_path: str, source_path: str) -> dict[str, str]:
        return {
            "name": name,
            "path": package_path,
            "sha256": hashlib.sha256((project_root / source_path).read_bytes()).hexdigest(),
        }

    assert payload["generatorContract"] == "receipt-only-v4"
    assert payload["projectIdentity"] == {
        "projectId": project["projectId"],
        "pluginVersion": project["pluginVersion"],
        "sowStandardVersion": project["sowStandardVersion"],
    }
    assert payload["inputs"] == [
        expected(name, GENERATOR.PACKAGE_DATA_PATHS[name], source_path)
        for name, source_path in GENERATOR.DATA_PATHS.items()
    ]
    assert payload["reviews"] == [
        expected(name, GENERATOR.PACKAGE_REVIEW_PATHS[name], source_path)
        for name, source_path in GENERATOR.REVIEW_PATHS.items()
    ]
    assert payload["validationReceipts"] == [
        expected(name, GENERATOR.PACKAGE_VALIDATION_PATHS[name], source_path)
        for name, source_path in GENERATOR.VALIDATION_PATHS.items()
    ]
    assert payload["template"] == expected(
        "template", GENERATOR.PACKAGE_TEMPLATE_PATH, GENERATOR.TEMPLATE_PATH
    )

    original = GENERATOR.package_fingerprint(files, project)
    monkeypatch.setitem(
        GENERATOR.PACKAGE_DATA_PATHS,
        "sourceRequirements",
        "sources/data/analyze-requirement/requirements-v2.json",
    )
    assert GENERATOR.package_fingerprint(files, project) != original


@pytest.mark.parametrize(
    "receipt",
    [
        "analyze-requirement.json",
        "analyze-as-is.json",
        "generate-design.json",
        "generate-story.json",
        "generate-task.json",
    ],
)
def test_every_owner_receipt_is_required_and_exact(tmp_path: Path, receipt: str) -> None:
    project = copy_project(tmp_path)
    (project / ".ai-sow/validation" / receipt).unlink()

    completed, result = run_generator(project)

    assert completed.returncode == 2
    assert result["outcome"] == "BLOCKED"
    assert result["diagnostics"][0]["code"] == "UPSTREAM_HANDOFF_MISSING"


def test_changed_existing_package_fails_closed(tmp_path: Path) -> None:
    project = copy_project(tmp_path)
    completed, result = run_generator(project)
    assert completed.returncode == 0
    package = project / str(result["packagePath"])
    (package / "sow.xlsx").write_bytes(b"different")

    repeated, repeated_result = run_generator(project)

    assert repeated.returncode == 2
    assert repeated_result["diagnostics"][0]["code"] == "PACKAGE_CONTENT_MISMATCH"


def test_unsupported_same_filesystem_publication_has_stable_code(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    staging = tmp_path / "staging"
    final = tmp_path / "final"
    staging.mkdir()

    def unsupported(source: Path, target: Path) -> None:
        raise OSError(18, "cross-device")

    monkeypatch.setattr(GENERATOR.os, "replace", unsupported)
    with pytest.raises(GENERATOR.GenerationError) as captured:
        GENERATOR.publish_staging(staging, final)
    assert captured.value.code == "PACKAGE_PUBLICATION_UNSUPPORTED"
    assert staging.is_dir()
    assert not final.exists()


def test_workbook_projects_only_two_business_tables_with_template_formulas(tmp_path: Path) -> None:
    project = copy_project(tmp_path)
    completed, result = run_generator(project)
    assert completed.returncode == 0, (completed.stderr, completed.stdout)
    package = project / str(result["packagePath"])
    workbook_path = package / "sow.xlsx"

    workbook = openpyxl.load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        assert workbook.sheetnames == [
            "01-需求故事",
            "02-任务清单",
            "03-工作量汇总",
            "90-估算标准",
        ]
        index = table_index(workbook)
        assert set(index) == FORMAL_TABLES
        assert REMOVED_TABLES.isdisjoint(index)
        data = load_workbook_data(project)
        delivery = data["delivery"]
        estimate = data["estimate"]
        assert len(table_rows(workbook, "SOWStoryTable")) == len(delivery["stories"])
        assert len(table_rows(workbook, "TaskTable")) == len(estimate["tasks"])

        story_sheet, story_table = index["SOWStoryTable"]
        task_sheet, task_table = index["TaskTable"]
        for worksheet, table, formula_headers in (
            (
                story_sheet,
                story_table,
                {"任务列表", "故事人天", "校验结果", "故事路径"},
            ),
            (
                task_sheet,
                task_table,
                {"M档标准人天", "复杂度系数", "任务人天", "SIT支持人天", "校验结果"},
            ),
        ):
            assert worksheet.auto_filter.ref is None
            assert table.autoFilter is not None
            assert table.autoFilter.ref == table.ref
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            headers = [
                str(worksheet.cell(min_row, column).value)
                for column in range(min_col, max_col + 1)
            ]
            for header in formula_headers:
                column = min_col + headers.index(header)
                for row in range(min_row + 1, max_row + 1):
                    cell = worksheet.cell(row, column)
                    assert cell.data_type == "f"
                    assert cell.fill.fgColor.rgb[-6:] == "F1F4F6"
                    assert cell.protection.locked is True
            calculated = {
                column.name
                for column in table.tableColumns
                if column.calculatedColumnFormula is not None
            }
            assert calculated == formula_headers

        story_rows = table_rows(workbook, "SOWStoryTable")
        assert all("风险" not in "".join(map(str, row.values())) for row in story_rows)
        task_rows = table_rows(workbook, "TaskTable")
        assert all(not str(row["任务类型"]).startswith("BU-") for row in task_rows)
        assert all(str(row["备注"]).startswith("任务理由：") for row in task_rows)
        assert any("复杂度理由：" in str(row["备注"]) for row in task_rows)
        assert all(" > " in str(row["所属故事"]) for row in task_rows)

        input_hashes = {
            hashlib.sha256((project / path).read_bytes()).hexdigest()
            for path in GENERATOR.DATA_PATHS.values()
        }
        workbook_text = {
            str(cell.value)
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None
        }
        assert input_hashes.isdisjoint(workbook_text)
        assert workbook.calculation.calcMode == "auto"
        assert workbook.calculation.fullCalcOnLoad is True
    finally:
        workbook.close()

    manifest = json.loads((package / "manifest.json").read_text(encoding="utf-8"))
    assert set(manifest["inputs"]) == set(GENERATOR.DATA_PATHS)
    assert manifest["template"]["path"] == "sources/templates/sow-template.xlsx"


def test_unsupported_project_version_is_rejected(tmp_path: Path) -> None:
    project = copy_project(tmp_path)
    metadata = project / ".ai-sow/project.json"
    value = json.loads(metadata.read_text(encoding="utf-8"))
    value["pluginVersion"] = "9.9.9"
    metadata.write_text(json.dumps(value, ensure_ascii=False), encoding="utf-8")

    completed, result = run_generator(project)

    assert completed.returncode == 2
    assert result["diagnostics"][0]["code"] == "PROJECT_SCHEMA_INVALID"


def test_generator_has_no_owner_validator_or_legacy_gate_dependency() -> None:
    source = SCRIPT.read_text(encoding="utf-8")
    assert "runtime.review_gates" not in source
    assert "skills.analyze" not in source
    assert "skills.generate" not in source
    for forbidden in (
        "CARRY_FORWARD",
        "validate_design_gates",
        "shutil.copytree",
        "PosixOutputAnchor",
    ):
        assert forbidden not in source
