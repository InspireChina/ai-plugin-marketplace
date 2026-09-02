from __future__ import annotations

import copy
import datetime as dt
import math
import os
import re
import tempfile
import unicodedata
import zipfile
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.workbook.properties import CalcProperties
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.table import TableFormula


FORMAL_SHEETS = (
    "01-需求故事",
    "02-任务清单",
    "03-工作量汇总",
    "90-估算标准",
)
TABLES = ("SOWStoryTable", "TaskTable")
FORMULA_HEADERS = {
    "SOWStoryTable": {"任务列表", "故事人天", "校验结果", "故事路径"},
    "TaskTable": {"M档标准人天", "复杂度系数", "任务人天", "SIT支持人天", "校验结果"},
}
TABLE_HEADERS = {
    "SOWStoryTable": [
        "需求",
        "子需求",
        "故事",
        "UAT适用",
        "验收条件",
        "备注",
        "任务列表",
        "故事人天",
        "校验结果",
        "故事路径",
    ],
    "TaskTable": [
        "所属故事",
        "任务名称",
        "任务类型",
        "工作方式",
        "复杂度",
        "备注",
        "M档标准人天",
        "复杂度系数",
        "任务人天",
        "SIT支持人天",
        "校验结果",
    ],
}
PROTECTED_SHEETS = {"01-需求故事", "02-任务清单"}
RISKY_TEXT = re.compile(r"^[=+\-@]")
BARE_TEXTJOIN = re.compile(r"(?<![\w.])TEXTJOIN\(")
DETERMINISTIC_TIME = dt.datetime(2000, 1, 1, 0, 0, 0)
DETERMINISTIC_ZIP_TIME = (2000, 1, 1, 0, 0, 0)
# ZIP metadata varies by host platform. Pin the host marker and Unix mode so
# identical approved inputs produce byte-identical workbooks everywhere.
DETERMINISTIC_CREATE_SYSTEM = 3
DETERMINISTIC_UNIX_MODE = 0o600
WRAPPED_LINE_HEIGHT = 15
WRAPPED_ROW_PADDING = 4
MAX_EXCEL_ROW_HEIGHT = 409.5


def safe_text(value: object) -> object:
    if isinstance(value, str) and RISKY_TEXT.match(value):
        return "'" + value
    return value


def formula_text(value: object) -> str:
    """Return a formula's text without discarding legacy array metadata."""
    if isinstance(value, ArrayFormula):
        value = value.text
    if not isinstance(value, str):
        raise TypeError("formula value is not text")
    return value


def normalize_table_formula(formula: object) -> str:
    """Serialize table references and future functions in OOXML form."""
    formula = formula_text(formula)
    parts = formula.split('"')
    for index in range(0, len(parts), 2):
        parts[index] = parts[index].replace("@", "[#This Row],")
        parts[index] = BARE_TEXTJOIN.sub("_xlfn.TEXTJOIN(", parts[index])
    return '"'.join(parts)


def require_unique_names(entries: list[dict[str, Any]], label: str) -> None:
    projected_names: dict[str, str] = {}
    for entry in entries:
        name = entry.get("name")
        if not isinstance(name, str) or not name.strip():
            raise ValueError(f"{label} display name is blank")
        projected = str(safe_text(name))
        key = unicodedata.normalize("NFC", projected).casefold()
        if key in projected_names:
            raise ValueError(
                f"{label} display name is duplicated after Excel projection: "
                f"{projected_names[key]} / {name}"
            )
        projected_names[key] = name


def build_rows(
    data: dict[str, dict[str, Any]],
    base_unit_names: dict[str, str],
) -> dict[str, list[dict[str, object]]]:
    requirements = data["requirements"]
    delivery = data["delivery"]
    estimate = data["estimate"]
    epics = {entry["epicId"]: entry for entry in requirements["epics"]}
    features = {entry["featureId"]: entry for entry in requirements["features"]}
    stories = {entry["storyId"]: entry for entry in delivery["stories"]}

    for label, entries in (
        ("Epic", requirements["epics"]),
        ("Feature", requirements["features"]),
        ("Story", delivery["stories"]),
        ("Task", estimate["tasks"]),
    ):
        require_unique_names(entries, label)

    acceptance_names_by_story: dict[str, list[str]] = {}
    for criterion in delivery["acceptanceCriteria"]:
        acceptance_names_by_story.setdefault(criterion["storyId"], []).append(
            criterion["name"]
        )

    story_rows: list[dict[str, object]] = []
    for story in delivery["stories"]:
        feature = features[story["featureId"]]
        epic = epics[feature["epicId"]]
        story_rows.append(
            {
                "需求": epic["name"],
                "子需求": feature["name"],
                "故事": story["name"],
                "UAT适用": "是" if story["uatRelevant"] else "否",
                "验收条件": "\n".join(
                    acceptance_names_by_story.get(story["storyId"], [])
                ),
                "备注": story["description"],
            }
        )

    task_rows: list[dict[str, object]] = []
    for task in estimate["tasks"]:
        story = stories[task["storyId"]]
        feature = features[story["featureId"]]
        epic = epics[feature["epicId"]]
        base_unit = task["baseUnit"]
        if base_unit not in base_unit_names:
            raise ValueError(f"template base-unit name is missing: {base_unit}")
        notes = [
            f"任务理由：{task['rationale']}",
            f"工作方式理由：{task['workModeRationale']}",
        ]
        if task.get("complexityRationale"):
            notes.append(f"复杂度理由：{task['complexityRationale']}")
        task_rows.append(
            {
                "所属故事": f"{epic['name']} > {feature['name']} > {story['name']}",
                "任务名称": task["name"],
                "任务类型": base_unit_names[base_unit],
                "工作方式": task["workMode"],
                "复杂度": task["complexity"],
                "备注": "\n".join(notes),
            }
        )

    return {"SOWStoryTable": story_rows, "TaskTable": task_rows}


def table_index(workbook: Any) -> dict[str, tuple[Any, Any]]:
    found: dict[str, tuple[Any, Any]] = {}
    for worksheet in workbook.worksheets:
        for table_name in worksheet.tables:
            if table_name in found:
                raise ValueError(f"duplicate template table: {table_name}")
            found[table_name] = (worksheet, worksheet.tables[table_name])
    missing = sorted(set(TABLES) - set(found))
    if missing:
        raise ValueError(f"template tables are missing: {missing}")
    return found


def base_unit_name_map(workbook: Any) -> dict[str, str]:
    index = table_index(workbook)
    if "BaseUnitCatalogTable" not in index:
        raise ValueError("template base-unit catalog is missing")
    worksheet, table = index["BaseUnitCatalogTable"]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = [
        worksheet.cell(min_row, column).value
        for column in range(min_col, max_col + 1)
    ]
    if "基础单元ID" not in headers or "基础单元名称" not in headers:
        raise ValueError("template base-unit name projection columns are missing")
    id_column = min_col + headers.index("基础单元ID")
    name_column = min_col + headers.index("基础单元名称")
    result: dict[str, str] = {}
    used_names: set[str] = set()
    for row in range(min_row + 1, max_row + 1):
        unit_id = worksheet.cell(row, id_column).value
        name = worksheet.cell(row, name_column).value
        if not isinstance(unit_id, str) or not unit_id.strip():
            raise ValueError("template base-unit ID is blank")
        if not isinstance(name, str) or not name.strip():
            raise ValueError(f"template base-unit name is blank: {unit_id}")
        if unit_id in result:
            raise ValueError(f"template base-unit ID is duplicated: {unit_id}")
        if name in used_names:
            raise ValueError(f"template base-unit name is duplicated: {name}")
        result[unit_id] = name
        used_names.add(name)
    return result


def clear_orphan_table_formulas(workbook: Any) -> None:
    for worksheet in workbook.worksheets:
        table_ranges = [
            range_boundaries(worksheet.tables[name].ref)
            for name in worksheet.tables
        ]
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type != "f" or not isinstance(cell.value, str):
                    continue
                if "@" not in cell.value and "[#This Row]," not in cell.value:
                    continue
                inside_table = any(
                    min_col <= cell.column <= max_col
                    and min_row < cell.row <= max_row
                    for min_col, min_row, max_col, max_row in table_ranges
                )
                if not inside_table:
                    cell.value = None


def copy_style(source: Any, target: Any) -> None:
    target.font = copy.copy(source.font)
    target.fill = copy.copy(source.fill)
    target.border = copy.copy(source.border)
    target.alignment = copy.copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy.copy(source.protection)


def wrapped_line_count(value: str, column_width: float) -> int:
    lines = value.splitlines() or [""]
    return sum(
        max(
            1,
            math.ceil(
                sum(
                    2 if unicodedata.east_asian_width(character) in {"W", "F"} else 1
                    for character in line
                )
                / column_width
            ),
        )
        for line in lines
    )


def effective_cell_width(worksheet: Any, cell: Any) -> float:
    for merged in worksheet.merged_cells.ranges:
        if cell.coordinate in merged:
            return sum(
                worksheet.column_dimensions[get_column_letter(column)].width or 8.43
                for column in range(merged.min_col, merged.max_col + 1)
            )
    return worksheet.column_dimensions[get_column_letter(cell.column)].width or 8.43


def fill_table(workbook: Any, table_name: str, rows: list[dict[str, object]]) -> None:
    worksheet, table = table_index(workbook)[table_name]
    min_col, min_row, max_col, old_max_row = range_boundaries(table.ref)
    prototype_row = min_row + 1
    headers = [
        worksheet.cell(min_row, column).value
        for column in range(min_col, max_col + 1)
    ]
    if not all(isinstance(header, str) for header in headers):
        raise ValueError(f"invalid table header: {table_name}")
    expected_headers = TABLE_HEADERS[table_name]
    if headers != expected_headers:
        raise ValueError(
            f"template header mismatch in {table_name}: "
            f"expected {expected_headers}, got {headers}"
        )
    metadata_headers = [column.name for column in table.tableColumns]
    if metadata_headers != expected_headers:
        raise ValueError(
            f"template table metadata mismatch in {table_name}: "
            f"expected {expected_headers}, got {metadata_headers}"
        )

    prototypes = [
        worksheet.cell(prototype_row, column)
        for column in range(min_col, max_col + 1)
    ]
    formulas = {
        offset: (normalize_table_formula(cell.value), isinstance(cell.value, ArrayFormula))
        for offset, cell in enumerate(prototypes)
        if cell.data_type == "f" and isinstance(cell.value, (str, ArrayFormula))
    }
    expected_formula_headers = FORMULA_HEADERS.get(table_name, set())
    actual_formula_headers = {headers[offset] for offset in formulas}
    if actual_formula_headers != expected_formula_headers:
        raise ValueError(f"formula prototype mismatch in {table_name}")
    for column_offset, column in enumerate(table.tableColumns):
        specification = formulas.get(column_offset)
        if specification is None:
            column.calculatedColumnFormula = None
            continue
        formula, is_array = specification
        column.calculatedColumnFormula = TableFormula(
            array=True if is_array else None,
            attr_text=formula.removeprefix("="),
        )

    physical_rows = rows if rows else [{}]
    clear_through = max(old_max_row, min_row + len(physical_rows), prototype_row)
    for row in range(prototype_row, clear_through + 1):
        for column in range(min_col, max_col + 1):
            worksheet.cell(row, column).value = None

    prototype_height = worksheet.row_dimensions[prototype_row].height
    for offset, payload in enumerate(physical_rows, start=1):
        row = min_row + offset
        if prototype_height is not None:
            worksheet.row_dimensions[row].height = prototype_height
        for column_offset, header in enumerate(headers):
            cell = worksheet.cell(row, min_col + column_offset)
            copy_style(prototypes[column_offset], cell)
            if column_offset in formulas:
                formula, is_array = formulas[column_offset]
                translated = Translator(
                    formula,
                    origin=prototypes[column_offset].coordinate,
                ).translate_formula(cell.coordinate)
                cell.value = (
                    ArrayFormula(ref=cell.coordinate, text=translated)
                    if is_array
                    else translated
                )
            else:
                value = None if not rows else safe_text(payload.get(header, ""))
                cell.value = value
                if isinstance(value, str):
                    cell.data_type = "s"
        wrapped_lines = max(
            (
                wrapped_line_count(visible_value, effective_cell_width(worksheet, cell))
                for column_offset, cell in enumerate(
                    worksheet[row][min_col - 1 : max_col]
                )
                for visible_value in [
                    payload.get(headers[column_offset])
                    if column_offset in formulas
                    else cell.value
                ]
                if cell.alignment.wrap_text and isinstance(visible_value, str)
            ),
            default=1,
        )
        worksheet.row_dimensions[row].height = min(
            MAX_EXCEL_ROW_HEIGHT,
            max(
                prototype_height or 15,
                wrapped_lines * WRAPPED_LINE_HEIGHT + WRAPPED_ROW_PADDING,
            ),
        )

    new_max_row = min_row + len(physical_rows)
    for row in range(new_max_row + 1, worksheet.max_row + 1):
        trailing_cells = [
            worksheet.cell(row, column)
            for column in range(min_col, max_col + 1)
        ]
        if any(cell.value not in (None, "") for cell in trailing_cells):
            break
        for cell in trailing_cells:
            cell._style = None
        if worksheet.row_dimensions[row].height == prototype_height:
            worksheet.row_dimensions[row].height = None
    table.ref = (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{new_max_row}"
    )
    if table.autoFilter is not None:
        table.autoFilter.ref = table.ref


def projection_contract(workbook: Any) -> dict[str, dict[str, object]]:
    contract: dict[str, dict[str, object]] = {}
    for table_name, (worksheet, table) in table_index(workbook).items():
        if table_name not in TABLES:
            continue
        min_col, min_row, max_col, _ = range_boundaries(table.ref)
        headers = [
            worksheet.cell(min_row, column).value
            for column in range(min_col, max_col + 1)
        ]
        if not all(isinstance(header, str) for header in headers):
            raise ValueError(f"invalid table header: {table_name}")
        formulas: dict[str, tuple[str, str, bool]] = {}
        styles: dict[str, tuple[object, ...]] = {}
        for offset, header in enumerate(headers):
            cell = worksheet.cell(min_row + 1, min_col + offset)
            styles[str(header)] = style_signature(cell)
            if cell.data_type == "f" and isinstance(cell.value, (str, ArrayFormula)):
                formula = normalize_table_formula(cell.value)
                if "_xlfn._xlws." in formula:
                    raise ValueError(
                        f"unsupported dynamic worksheet formula in {table_name}.{header}"
                    )
                formulas[str(header)] = (
                    cell.coordinate,
                    formula,
                    isinstance(cell.value, ArrayFormula),
                )
        contract[table_name] = {
            "headers": headers,
            "formulas": formulas,
            "styles": styles,
        }
    return contract


def style_signature(cell: Any) -> tuple[object, ...]:
    return (
        copy.copy(cell.font),
        copy.copy(cell.fill),
        copy.copy(cell.border),
        copy.copy(cell.alignment),
        cell.number_format,
        copy.copy(cell.protection),
    )


def verify_workbook(
    path: Path,
    expected: dict[str, list[dict[str, object]]],
    contract: dict[str, dict[str, object]],
) -> None:
    workbook = openpyxl.load_workbook(path, data_only=False, read_only=False)
    try:
        if tuple(workbook.sheetnames) != FORMAL_SHEETS:
            raise ValueError("formal workbook sheet contract changed")
        index = table_index(workbook)
        if workbook.calculation.calcMode != "auto":
            raise ValueError("workbook recalculation is not enabled")
        for table_name, rows in expected.items():
            worksheet, table = index[table_name]
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            if max_row - min_row != max(1, len(rows)):
                raise ValueError(f"table row count mismatch: {table_name}")
            headers = [
                worksheet.cell(min_row, column).value
                for column in range(min_col, max_col + 1)
            ]
            if headers != TABLE_HEADERS[table_name]:
                raise ValueError(f"table header mismatch: {table_name}")
            specification = contract[table_name]
            if headers != specification["headers"]:
                raise ValueError(f"table headers changed: {table_name}")
            formulas = specification["formulas"]
            styles = specification["styles"]
            assert isinstance(formulas, dict) and isinstance(styles, dict)
            physical_rows = rows if rows else [{}]
            for row_offset, payload in enumerate(physical_rows, start=1):
                for column_offset, header in enumerate(headers):
                    cell = worksheet.cell(min_row + row_offset, min_col + column_offset)
                    if style_signature(cell) != styles[header]:
                        raise ValueError(
                            f"prototype style changed in {table_name}.{header}"
                        )
                    if header in formulas:
                        origin, prototype, is_array = formulas[header]
                        expected_formula = Translator(
                            prototype,
                            origin=origin,
                        ).translate_formula(cell.coordinate)
                        actual_formula = (
                            formula_text(cell.value)
                            if isinstance(cell.value, (str, ArrayFormula))
                            else None
                        )
                        formula_kind_matches = (
                            isinstance(cell.value, ArrayFormula)
                            and cell.value.ref == cell.coordinate
                            if is_array
                            else isinstance(cell.value, str)
                        )
                        if (
                            actual_formula != expected_formula
                            or cell.data_type != "f"
                            or not formula_kind_matches
                        ):
                            raise ValueError(
                                f"formula mismatch in {table_name}.{header}"
                            )
                    else:
                        expected_value = (
                            None if not rows else safe_text(payload.get(header, ""))
                        )
                        if expected_value == "":
                            expected_value = None
                        if cell.value != expected_value:
                            raise ValueError(
                                f"projected value mismatch in {table_name}.{header}"
                            )
                        if isinstance(expected_value, str) and cell.data_type != "s":
                            raise ValueError(
                                f"projected text type mismatch in {table_name}.{header}"
                            )
            calculated_headers = {
                column.name
                for column in table.tableColumns
                if column.calculatedColumnFormula is not None
                and column.calculatedColumnFormula.text
            }
            if calculated_headers != FORMULA_HEADERS[table_name]:
                raise ValueError(f"calculated column mismatch in {table_name}")
            formula_columns = {
                column.name: column.calculatedColumnFormula
                for column in table.tableColumns
                if column.calculatedColumnFormula is not None
            }
            for header, (_, _, is_array) in formulas.items():
                if (formula_columns[header].array is True) != is_array:
                    raise ValueError(
                        f"calculated column array mismatch in {table_name}.{header}"
                    )
            if table.autoFilter is not None and table.autoFilter.ref != table.ref:
                raise ValueError(f"autoFilter range mismatch: {table_name}")
        for sheet_name in PROTECTED_SHEETS:
            if not workbook[sheet_name].protection.sheet:
                raise ValueError(f"worksheet protection is missing: {sheet_name}")
    finally:
        workbook.close()


def deterministic_external_attr(external_attr: int) -> int:
    """Keep file-type and DOS bits, but pin an entry's Unix permission mode."""
    return (external_attr & 0xFFFF0000) & ~(0o777 << 16) | (
        DETERMINISTIC_UNIX_MODE << 16
    )


def normalize_xlsx(path: Path) -> None:
    """Normalize ZIP metadata so identical inputs produce identical XLSX bytes."""
    with zipfile.ZipFile(path, "r") as source:
        members = []
        for entry in source.infolist():
            payload = source.read(entry.filename)
            if entry.filename == "docProps/core.xml":
                payload = re.sub(
                    rb"<dcterms:modified[^>]*>.*?</dcterms:modified>",
                    b'<dcterms:modified xsi:type="dcterms:W3CDTF">2000-01-01T00:00:00Z</dcterms:modified>',
                    payload,
                )
            members.append((entry.filename, payload, entry))
    temporary: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{path.name}.",
            suffix=".tmp",
            dir=path.parent,
            delete=False,
        ) as stream:
            temporary = Path(stream.name)
        with zipfile.ZipFile(
            temporary,
            "w",
            compression=zipfile.ZIP_DEFLATED,
            compresslevel=9,
        ) as target:
            for name, payload, original in sorted(members, key=lambda item: item[0]):
                entry = zipfile.ZipInfo(name, DETERMINISTIC_ZIP_TIME)
                entry.compress_type = zipfile.ZIP_DEFLATED
                entry.create_system = DETERMINISTIC_CREATE_SYSTEM
                entry.external_attr = deterministic_external_attr(original.external_attr)
                entry.flag_bits = original.flag_bits
                target.writestr(entry, payload)
        os.replace(temporary, path)
        temporary = None
    finally:
        if temporary is not None:
            temporary.unlink(missing_ok=True)


def write_workbook(
    template_path: Path,
    data: dict[str, dict[str, Any]],
    output_path: Path,
) -> None:
    workbook = openpyxl.load_workbook(
        template_path,
        data_only=False,
        read_only=False,
    )
    if workbook.calculation is None:
        workbook.calculation = CalcProperties()
    try:
        table_index(workbook)
        rows = build_rows(data, base_unit_name_map(workbook))
        contract = projection_contract(workbook)
        clear_orphan_table_formulas(workbook)
        for table_name in TABLES:
            fill_table(workbook, table_name, rows[table_name])
        workbook.calculation.calcMode = "auto"
        workbook.calculation.calcOnSave = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.fullCalcOnLoad = True
        workbook.properties.created = DETERMINISTIC_TIME
        workbook.properties.modified = DETERMINISTIC_TIME
        workbook.save(output_path)
    finally:
        workbook.close()
    normalize_xlsx(output_path)
    verify_workbook(output_path, rows, contract)
