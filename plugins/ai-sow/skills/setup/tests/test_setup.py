from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest


SKILL_ROOT = Path(__file__).parents[1]
SCRIPT = SKILL_ROOT / "scripts/setup.py"
BOOTSTRAP_SH = SKILL_ROOT / "scripts/bootstrap.sh"
BOOTSTRAP_PS1 = SKILL_ROOT / "scripts/bootstrap.ps1"
TEMPLATE = SKILL_ROOT / "assets/sow-template.xlsx"
POWERSHELL = shutil.which("pwsh") or shutil.which("powershell")
POSIX_SHELL = "/bin/sh" if Path("/bin/sh").is_file() else None


def _symlink_supported() -> bool:
    """Windows only allows symlink creation under Developer Mode or elevation."""
    with tempfile.TemporaryDirectory() as temp:
        root = Path(temp)
        try:
            (root / "link").symlink_to(root / "target", target_is_directory=True)
        except (OSError, NotImplementedError):
            return False
    return True


requires_posix_shell = pytest.mark.skipif(
    POSIX_SHELL is None, reason="POSIX shell is unavailable on this platform"
)
requires_symlinks = pytest.mark.skipif(
    not _symlink_supported(),
    reason="creating symlinks requires Developer Mode or elevation on this platform",
)


ENABLE_LONG_PATHS = SKILL_ROOT / "scripts/enable_long_paths.ps1"


def test_long_path_remedy_requires_explicit_consent_and_elevation() -> None:
    """启用长路径会改机器级系统策略，SKILL 必须要求先取得用户明确同意。"""
    skill = (SKILL_ROOT / "SKILL.md").read_text(encoding="utf-8")
    for required in (
        "WINDOWS_LONG_PATH_REQUIRED",
        "取得用户明确同意后才能执行",
        "enable_long_paths.ps1",
        "-Apply",
        "管理员权限",
    ):
        assert required in skill

    script = ENABLE_LONG_PATHS.read_text(encoding="utf-8")
    # 不传 -Apply 只报告；写入前必须先确认已提权。
    assert "[switch]$Apply" in script
    assert "Test-Administrator" in script
    apply_index = script.index("New-ItemProperty")
    assert script.index("Test-Administrator") < apply_index
    assert ENABLE_LONG_PATHS.read_bytes().startswith(b"\xef\xbb\xbf")


@pytest.mark.skipif(POWERSHELL is None, reason="PowerShell is unavailable on this platform")
def test_enable_long_paths_without_apply_makes_no_change() -> None:
    result = subprocess.run(
        [
            str(POWERSHELL),
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(ENABLE_LONG_PATHS),
        ],
        capture_output=True,
        text=True,
        encoding="utf-8",
        check=False,
    )

    payload = json.loads(result.stdout.strip().splitlines()[-1])
    assert payload["outcome"] in {"OK", "NEEDS_INPUT"}
    if payload["outcome"] == "NEEDS_INPUT":
        assert payload["longPathsEnabled"] is False
        assert "明确同意" in payload["nextStep"]
        assert result.returncode == 1
    else:
        assert payload["longPathsEnabled"] is True
        assert result.returncode == 0


def test_setup_blocks_long_project_root_before_writing_anything(tmp_path: Path) -> None:
    """预算不足时必须 fail closed：返回结构化诊断且不创建 .ai-sow。"""
    sys.path.insert(0, str(SKILL_ROOT.parents[1]))
    sys.path.insert(0, str(SKILL_ROOT / "scripts"))
    from runtime.project_io import managed_path_budget
    from setup import DEEPEST_MANAGED_RELATIVE_PATH

    project_root = tmp_path
    for _ in range(3):
        project_root = project_root / ("d" * 60)
    project_root.mkdir(parents=True)

    budget = managed_path_budget(project_root)
    if budget is None or budget >= len(DEEPEST_MANAGED_RELATIVE_PATH):
        pytest.skip("this platform does not enforce MAX_PATH for the probe path")

    result = subprocess.run(
        [
            sys.executable,
            str(SCRIPT),
            "--project-root",
            str(project_root),
            "--project-id",
            "long-path-probe",
            "--name",
            "长路径探针",
        ],
        capture_output=True,
        text=True,
        encoding="utf-8",
        check=False,
    )

    assert result.returncode == 2, result.stdout + result.stderr
    payload = json.loads(result.stdout.strip().splitlines()[-1])
    assert payload["outcome"] == "BLOCKED"
    assert [item["code"] for item in payload["diagnostics"]] == [
        "WINDOWS_LONG_PATH_REQUIRED"
    ]
    assert "enable_long_paths.ps1" in payload["nextStep"]
    assert not (project_root / ".ai-sow").exists()


def test_skill_uses_current_stage_without_leaf_agents() -> None:
    skill = (SKILL_ROOT / "SKILL.md").read_text(encoding="utf-8")
    assert "当前 Stage Agent 是本 Skill 的唯一用户接口" in skill
    assert "直接运行" in skill
    for forbidden in ("Orchestrator Agent", "Worker Agent", "Validator Agent", "Reviewer Agent"):
        assert forbidden not in skill


def test_skill_bootstraps_runtime_without_user_commands() -> None:
    skill = (SKILL_ROOT / "SKILL.md").read_text(encoding="utf-8")
    for required in (
        "<skill-root>/scripts/bootstrap.sh",
        "<skill-root>/scripts/bootstrap.ps1",
        "用户无需手工安装",
        "自动重试",
    ):
        assert required in skill


def test_bootstrap_scripts_pin_official_uv_and_avoid_admin_install() -> None:
    unix = BOOTSTRAP_SH.read_text(encoding="utf-8")
    windows = BOOTSTRAP_PS1.read_text(encoding="utf-8")
    for text in (unix, windows):
        assert "https://astral.sh/uv/0.11.7/install" in text
        assert "UV_UNMANAGED_INSTALL" in text
        # sh 传字面量 `python install 3.12`，PowerShell 传数组 `"python", "install", "3.12"`。
        assert re.search(r'python["\s,]+install["\s,]+3\.12', text)
        assert "sync" in text and "--locked" in text and "--python" in text
        assert "sudo" not in text.lower()
    if POSIX_SHELL is not None:
        assert subprocess.run(
            [POSIX_SHELL, "-n", str(BOOTSTRAP_SH)],
            capture_output=True,
            text=True,
            encoding="utf-8",
            check=False,
        ).returncode == 0
    assert r"(?:\s|$)" in windows
    assert r"(?:\\s|$)" not in windows


def test_windows_bootstrap_is_utf8_bom_encoded_and_pins_console_encoding() -> None:
    """Windows PowerShell 按 ANSI 代码页读取无 BOM 的 .ps1，会破坏中文诊断甚至导致解析失败。"""
    assert BOOTSTRAP_PS1.read_bytes().startswith(b"\xef\xbb\xbf")
    assert not BOOTSTRAP_SH.read_bytes().startswith(b"\xef\xbb\xbf")
    windows = BOOTSTRAP_PS1.read_text(encoding="utf-8")
    assert "[Console]::OutputEncoding" in windows
    assert "PYTHONUTF8" in windows


@pytest.mark.skipif(POWERSHELL is None, reason="PowerShell is unavailable on this platform")
def test_windows_bootstrap_parses_under_powershell(tmp_path: Path) -> None:
    probe = tmp_path / "parse-probe.ps1"
    probe.write_text(
        "param([string]$Path)\n"
        "$errors = $null\n"
        "[void][System.Management.Automation.Language.Parser]::ParseFile("
        "$Path, [ref]$null, [ref]$errors)\n"
        "if ($errors) { $errors | ForEach-Object { $_.Message }; exit 1 }\n",
        encoding="ascii",
    )

    result = subprocess.run(
        [
            str(POWERSHELL),
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(probe),
            "-Path",
            str(BOOTSTRAP_PS1),
        ],
        capture_output=True,
        text=True, encoding="utf-8",
        check=False,
    )

    assert result.returncode == 0, result.stdout + result.stderr


@pytest.mark.skipif(POWERSHELL is None, reason="PowerShell is unavailable on this platform")
def test_powershell_uv_version_matcher_accepts_platform_suffix(tmp_path: Path) -> None:
    source = BOOTSTRAP_PS1.read_text(encoding="utf-8")
    function = re.search(
        r"(?ms)^function Test-UvVersion\(.*?^}\s*$",
        source,
    )
    assert function is not None
    probe = tmp_path / "uv-version-probe.ps1"
    probe.write_text(
        '$UvVersion = "0.11.7"\n'
        + function.group(0)
        + '\nWrite-Output (Test-UvVersion "uv 0.11.7 (Windows x86_64)")\n',
        encoding="utf-8",
    )

    result = subprocess.run(
        [str(POWERSHELL), "-NoProfile", "-ExecutionPolicy", "Bypass", "-File", str(probe)],
        capture_output=True,
        text=True, encoding="utf-8",
        check=False,
    )

    assert result.returncode == 0, result.stdout + result.stderr
    assert result.stdout.strip().lower() == "true"


@requires_posix_shell
def test_unix_bootstrap_creates_plugin_venv_with_existing_uv(tmp_path: Path) -> None:
    plugin_root = tmp_path / "plugin"
    script = plugin_root / "skills/setup/scripts/bootstrap.sh"
    script.parent.mkdir(parents=True)
    shutil.copy2(BOOTSTRAP_SH, script)
    (plugin_root / "pyproject.toml").write_text(
        '[project]\nname = "bootstrap-test"\nversion = "0"\nrequires-python = ">=3.12,<3.13"\n',
        encoding="utf-8",
    )
    (plugin_root / "uv.lock").write_text("version = 1\n", encoding="utf-8")
    fake_bin = tmp_path / "bin"
    fake_bin.mkdir()
    fake_uv = fake_bin / "uv"
    fake_uv.write_text(
        """#!/bin/sh
set -eu
printf '%s\\n' "$*" >> "$FAKE_UV_LOG"
if [ "${1:-}" = "--version" ]; then
  printf '%s\\n' 'uv 0.11.7 (Homebrew 2026-04-15 aarch64-apple-darwin)'
  exit 0
fi
if [ "${1:-}" = "python" ] && [ "${2:-}" = "find" ]; then
  exit 1
fi
if [ "${1:-}" = "sync" ]; then
  shift
  project=
  while [ "$#" -gt 0 ]; do
    if [ "$1" = "--project" ]; then
      shift
      project=$1
    fi
    shift
  done
  mkdir -p "$project/.venv/bin"
  printf '%s\\n' '#!/bin/sh' 'if [ "${1:-}" = "--version" ]; then echo "Python 3.12.13"; fi' 'exit 0' > "$project/.venv/bin/python"
  chmod +x "$project/.venv/bin/python"
fi
""",
        encoding="utf-8",
    )
    fake_uv.chmod(0o755)
    stale_local_uv = plugin_root / ".ai-sow-tools/bin/uv"
    stale_local_uv.parent.mkdir(parents=True)
    stale_local_uv.write_text(
        "#!/bin/sh\nprintf '%s\\n' 'uv 0.10.0'\n",
        encoding="utf-8",
    )
    stale_local_uv.chmod(0o755)
    fake_curl = fake_bin / "curl"
    fake_curl.write_text("#!/bin/sh\nexit 99\n", encoding="utf-8")
    fake_curl.chmod(0o755)
    log = tmp_path / "uv.log"
    env = {
        **os.environ,
        "FAKE_UV_LOG": str(log),
        "PATH": f"{fake_bin}:/usr/bin:/bin",
    }

    result = subprocess.run(
        ["/bin/sh", str(script)],
        capture_output=True,
        text=True, encoding="utf-8",
        check=False,
        env=env,
    )

    assert result.returncode == 0, result.stdout + result.stderr
    payload = json.loads(result.stdout.splitlines()[-1])
    assert payload["outcome"] == "OK"
    assert payload["uvSource"] == "PATH"
    assert payload["pythonVersion"].startswith("Python 3.12.")
    calls = log.read_text(encoding="utf-8")
    assert "python install 3.12" in calls
    assert f"sync --project {plugin_root} --locked --python 3.12" in calls
    assert (plugin_root / ".venv/bin/python").is_file()


@requires_posix_shell
def test_unix_bootstrap_installs_private_uv_when_path_has_none(tmp_path: Path) -> None:
    plugin_root = tmp_path / "plugin"
    script = plugin_root / "skills/setup/scripts/bootstrap.sh"
    script.parent.mkdir(parents=True)
    shutil.copy2(BOOTSTRAP_SH, script)
    (plugin_root / "pyproject.toml").write_text(
        '[project]\nname = "bootstrap-test"\nversion = "0"\nrequires-python = ">=3.12,<3.13"\n',
        encoding="utf-8",
    )
    (plugin_root / "uv.lock").write_text("version = 1\n", encoding="utf-8")

    fake_uv_source = tmp_path / "fake-uv"
    fake_uv_source.write_text(
        """#!/bin/sh
set -eu
if [ "${1:-}" = "--version" ]; then
  printf '%s\n' 'uv 0.11.7'
  exit 0
fi
if [ "${1:-}" = "python" ] && [ "${2:-}" = "find" ]; then
  exit 0
fi
if [ "${1:-}" = "sync" ]; then
  shift
  project=
  while [ "$#" -gt 0 ]; do
    if [ "$1" = "--project" ]; then
      shift
      project=$1
    fi
    shift
  done
  mkdir -p "$project/.venv/bin"
  printf '%s\n' '#!/bin/sh' 'if [ "${1:-}" = "--version" ]; then echo "Python 3.12.13"; fi' 'exit 0' > "$project/.venv/bin/python"
  chmod +x "$project/.venv/bin/python"
fi
""",
        encoding="utf-8",
    )
    fake_uv_source.chmod(0o755)

    installer = tmp_path / "installer.sh"
    installer.write_text(
        """#!/bin/sh
set -eu
mkdir -p "$UV_UNMANAGED_INSTALL"
cp "$FAKE_UV_SOURCE" "$UV_UNMANAGED_INSTALL/uv"
chmod +x "$UV_UNMANAGED_INSTALL/uv"
""",
        encoding="utf-8",
    )
    installer.chmod(0o755)

    fake_bin = tmp_path / "bin"
    fake_bin.mkdir()
    fake_curl = fake_bin / "curl"
    fake_curl.write_text(
        """#!/bin/sh
set -eu
output=
while [ "$#" -gt 0 ]; do
  if [ "$1" = "-o" ]; then
    shift
    output=$1
  fi
  shift
done
cp "$FAKE_INSTALLER" "$output"
""",
        encoding="utf-8",
    )
    fake_curl.chmod(0o755)
    env = {
        **os.environ,
        "FAKE_INSTALLER": str(installer),
        "FAKE_UV_SOURCE": str(fake_uv_source),
        "PATH": f"{fake_bin}:/usr/bin:/bin",
    }
    assert shutil.which("uv", path=env["PATH"]) is None

    result = subprocess.run(
        ["/bin/sh", str(script)],
        capture_output=True,
        text=True, encoding="utf-8",
        check=False,
        env=env,
    )

    assert result.returncode == 0, result.stdout + result.stderr
    payload = json.loads(result.stdout.splitlines()[-1])
    assert payload["outcome"] == "OK"
    assert payload["uvSource"] == "PLUGIN_LOCAL"
    assert (plugin_root / ".ai-sow-tools/bin/uv").is_file()
    assert (plugin_root / ".venv/bin/python").is_file()


@requires_posix_shell
def test_unix_bootstrap_failure_does_not_create_project_shell(tmp_path: Path) -> None:
    plugin_root = tmp_path / "plugin"
    script = plugin_root / "skills/setup/scripts/bootstrap.sh"
    script.parent.mkdir(parents=True)
    shutil.copy2(BOOTSTRAP_SH, script)
    fake_bin = tmp_path / "bin"
    fake_bin.mkdir()
    fake_uv = fake_bin / "uv"
    fake_uv.write_text(
        """#!/bin/sh
if [ "${1:-}" = "--version" ]; then
  printf '%s\\n' 'uv 0.11.7'
  exit 0
fi
if [ "${1:-}" = "python" ] && [ "${2:-}" = "find" ]; then
  exit 0
fi
if [ "${1:-}" = "sync" ]; then
  exit 1
fi
exit 0
""",
        encoding="utf-8",
    )
    fake_uv.chmod(0o755)
    project_root = tmp_path / "customer-project"
    project_root.mkdir()

    result = subprocess.run(
        [
            "/bin/sh",
            str(script),
            "--project-root",
            str(project_root),
            "--project-id",
            "empty-machine",
            "--name",
            "零预装环境",
        ],
        capture_output=True,
        text=True, encoding="utf-8",
        check=False,
        env={**os.environ, "PATH": f"{fake_bin}:/usr/bin:/bin"},
    )

    assert result.returncode == 2
    assert json.loads(result.stdout)["diagnostics"][0]["code"] == "DEPENDENCY_SYNC_FAILED"
    assert not (project_root / ".ai-sow").exists()


def run_setup(
    project_root: Path,
    *extra: str,
    no_site: bool = False,
) -> subprocess.CompletedProcess[str]:
    command = [sys.executable]
    if no_site:
        command.append("-S")
    command.extend(
        [
            str(SCRIPT),
            "--project-root",
            str(project_root),
            "--project-id",
            "bookstore-modernization",
            "--name",
            "在线书店 2.0",
            *extra,
        ]
    )
    return subprocess.run(
        command,
        capture_output=True,
        text=True, encoding="utf-8",
        check=False,
        cwd=project_root,
        env=os.environ,
    )


def test_bundled_template_round_trips_and_contains_authoritative_catalog() -> None:
    payload = TEMPLATE.read_bytes()
    workbook = openpyxl.load_workbook(BytesIO(payload), data_only=False)
    try:
        tables = {
            name: worksheet.tables[name]
            for worksheet in workbook.worksheets
            for name in worksheet.tables
        }
        catalog = tables["BaseUnitCatalogTable"]
        min_col, min_row, _, max_row = openpyxl.utils.range_boundaries(catalog.ref)
        assert max_row - min_row == 37
        family_column = [column.name for column in catalog.tableColumns].index("任务族名称")
        worksheet = next(
            sheet for sheet in workbook.worksheets if "BaseUnitCatalogTable" in sheet.tables
        )
        families = {
            worksheet.cell(row, min_col + family_column).value
            for row in range(min_row + 1, max_row + 1)
        }
        assert len(families) == 13
        saved = BytesIO()
        workbook.save(saved)
    finally:
        workbook.close()
    reopened = openpyxl.load_workbook(BytesIO(saved.getvalue()), data_only=False)
    reopened.close()


def test_setup_creates_exact_minimal_project_shell(tmp_path: Path) -> None:
    result = run_setup(tmp_path)
    assert result.returncode == 0, result.stdout + result.stderr
    payload = json.loads(result.stdout)
    assert payload["outcome"] == "OK"
    assert json.loads((tmp_path / ".ai-sow/project.json").read_text(encoding="utf-8")) == {
        "projectId": "bookstore-modernization",
        "name": "在线书店 2.0",
        "pluginVersion": "0.1.0-beta.1",
        "sowStandardVersion": "1.3",
    }
    assert (tmp_path / ".ai-sow/templates/sow-template.xlsx").read_bytes() == TEMPLATE.read_bytes()
    for relative in ("inputs", "work", "reviews", "data", "validation", "outputs"):
        assert (tmp_path / ".ai-sow" / relative).is_dir()


def test_complete_existing_project_is_read_only_idempotent(tmp_path: Path) -> None:
    assert run_setup(tmp_path).returncode == 0
    project = tmp_path / ".ai-sow/project.json"
    template = tmp_path / ".ai-sow/templates/sow-template.xlsx"
    before = (project.read_bytes(), template.read_bytes(), project.stat().st_mtime_ns)
    result = run_setup(tmp_path)
    assert result.returncode == 0, result.stdout
    assert json.loads(result.stdout)["outcome"] == "OK"
    assert (project.read_bytes(), template.read_bytes(), project.stat().st_mtime_ns) == before


@pytest.mark.parametrize("missing", ["templates/sow-template.xlsx", "reviews"])
def test_incomplete_existing_project_blocks_without_repair(tmp_path: Path, missing: str) -> None:
    assert run_setup(tmp_path).returncode == 0
    target = tmp_path / ".ai-sow" / missing
    if target.is_dir():
        target.rmdir()
    else:
        target.unlink()
    project_before = (tmp_path / ".ai-sow/project.json").read_bytes()
    result = run_setup(tmp_path)
    assert result.returncode == 2
    assert json.loads(result.stdout)["outcome"] == "BLOCKED"
    assert not target.exists()
    assert (tmp_path / ".ai-sow/project.json").read_bytes() == project_before


def test_existing_identity_conflict_blocks(tmp_path: Path) -> None:
    assert run_setup(tmp_path).returncode == 0
    project = tmp_path / ".ai-sow/project.json"
    value = json.loads(project.read_text(encoding="utf-8"))
    value["name"] = "其他项目"
    project.write_text(json.dumps(value, ensure_ascii=False), encoding="utf-8")
    result = run_setup(tmp_path)
    assert result.returncode == 2
    assert json.loads(result.stdout)["outcome"] == "BLOCKED"


def test_existing_valid_customized_project_template_is_read_only(tmp_path: Path) -> None:
    assert run_setup(tmp_path).returncode == 0
    template = tmp_path / ".ai-sow/templates/sow-template.xlsx"
    workbook = openpyxl.load_workbook(template, data_only=False)
    try:
        workbook.properties.title = "项目级定制模板"
        workbook.save(template)
    finally:
        workbook.close()
    before = template.read_bytes()

    result = run_setup(tmp_path)

    assert result.returncode == 0, result.stdout
    assert template.read_bytes() == before


def test_existing_corrupt_project_template_blocks_without_overwrite(tmp_path: Path) -> None:
    assert run_setup(tmp_path).returncode == 0
    template = tmp_path / ".ai-sow/templates/sow-template.xlsx"
    template.write_bytes(b"conflict")
    result = run_setup(tmp_path)
    assert result.returncode == 2
    assert template.read_bytes() == b"conflict"


@requires_symlinks
def test_fresh_setup_rejects_symlink_in_managed_path(tmp_path: Path) -> None:
    outside = tmp_path.parent / f"{tmp_path.name}-outside"
    outside.mkdir()
    (tmp_path / ".ai-sow").symlink_to(outside, target_is_directory=True)
    result = run_setup(tmp_path)
    assert result.returncode == 2
    assert json.loads(result.stdout)["outcome"] == "BLOCKED"
    assert list(outside.iterdir()) == []


def test_setup_reports_missing_python_dependencies(tmp_path: Path) -> None:
    result = run_setup(tmp_path, no_site=True)
    assert result.returncode == 2
    payload = json.loads(result.stdout)
    assert payload["outcome"] == "NEEDS_INPUT"
    assert "重新调用 setup" in payload["nextStep"]
    assert "uv sync --project" not in payload["nextStep"]
    assert not (tmp_path / ".ai-sow").exists()


def test_setup_rejects_removed_repair_option(tmp_path: Path) -> None:
    result = run_setup(tmp_path, "--repair")
    assert result.returncode == 2
    assert "unrecognized arguments" in result.stderr


def test_setup_rejects_invalid_project_id_without_partial_manifest(tmp_path: Path) -> None:
    result = run_setup(tmp_path, "--project-id", "Invalid")
    assert result.returncode == 2
    assert json.loads(result.stdout)["outcome"] == "BLOCKED"
    assert not (tmp_path / ".ai-sow/project.json").exists()
