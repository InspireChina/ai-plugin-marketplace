from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils.cell import range_boundaries


# Windows 控制台默认使用本地代码页（如 cp936），会把中文结构化输出写成非 UTF-8 字节。
# 调用方按 UTF-8 读取 stdout/stderr，这里显式固定编码，与 POSIX 行为保持一致。
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


CATALOG_HEADERS = {
    "任务族ID",
    "任务族名称",
    "基础单元ID",
    "基础单元名称",
    "计数口径",
    "包含内容",
    "不包含内容",
    "新建M档人天",
    "调整M档人天",
    "接入复用M档人天",
    "S标准",
    "M标准",
    "L标准",
    "X/拆分条件",
}
PARAMETER_HEADERS = {"参数代码", "名称", "值", "单位", "适用范围", "验证状态/说明"}
MODE_EFFORT_HEADERS = {
    "新建": "新建M档人天",
    "调整": "调整M档人天",
    "接入复用": "接入复用M档人天",
}
COMPLEXITIES = ("S", "M", "L")
CALIBRATED_PARAMETER_STATUSES = {"固定规则", "已校准", "已批准"}


def table_rows(workbook: Any, table_name: str) -> list[dict[str, Any]]:
    for worksheet in workbook.worksheets:
        if table_name not in worksheet.tables:
            continue
        table = worksheet.tables[table_name]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [worksheet.cell(min_row, column).value for column in range(min_col, max_col + 1)]
        return [
            {
                str(header): worksheet.cell(row, column).value
                for header, column in zip(headers, range(min_col, max_col + 1), strict=True)
            }
            for row in range(min_row + 1, max_row + 1)
        ]
    raise ValueError(f"template table is missing: {table_name}")


def require_headers(rows: list[dict[str, Any]], expected: set[str], table_name: str) -> None:
    if not rows or set(rows[0]) != expected:
        actual = set(rows[0]) if rows else set()
        raise ValueError(
            f"template table headers are invalid for {table_name}: "
            f"expected {sorted(expected)}, got {sorted(actual)}"
        )


def require_text(row: dict[str, Any], header: str, subject: str) -> str:
    value = row.get(header)
    if not isinstance(value, str) or not value.strip():
        raise ValueError(f"{subject} must define non-empty {header}")
    return value


def usable_mode(row: dict[str, Any], header: str, subject: str) -> bool:
    value = row.get(header)
    if value == "❌":
        return False
    if isinstance(value, bool) or not isinstance(value, (int, float)) or value <= 0:
        raise ValueError(f"{subject} {header} must be a positive number or ❌")
    return True


def validate_complexity_parameters(rows: list[dict[str, Any]]) -> None:
    seen: set[str] = set()
    for row in rows:
        code = row.get("参数代码")
        if not isinstance(code, str) or not code.strip():
            raise ValueError("project parameter must define a non-empty 参数代码")
        if not code.startswith("K_COMPLEXITY_"):
            continue
        level = code.removeprefix("K_COMPLEXITY_")
        if level not in COMPLEXITIES or level in seen:
            raise ValueError(f"complexity parameter is invalid or duplicated: {code}")
        value = row.get("值")
        if isinstance(value, bool) or not isinstance(value, (int, float)) or value <= 0:
            raise ValueError(f"complexity factor must be positive: {level}")
        if row.get("验证状态/说明") not in CALIBRATED_PARAMETER_STATUSES:
            raise ValueError(f"complexity factor is not calibrated: {level}")
        seen.add(level)
    if seen != set(COMPLEXITIES):
        raise ValueError("ProjectParameterTable must define calibrated S/M/L complexity factors")


def read_contract(template_path: Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(template_path, data_only=False, read_only=False)
    try:
        catalog_rows = table_rows(workbook, "BaseUnitCatalogTable")
        parameter_rows = table_rows(workbook, "ProjectParameterTable")
        require_headers(catalog_rows, CATALOG_HEADERS, "BaseUnitCatalogTable")
        require_headers(parameter_rows, PARAMETER_HEADERS, "ProjectParameterTable")

        base_units: dict[str, dict[str, Any]] = {}
        base_unit_names: set[str] = set()
        family_names: dict[str, str] = {}
        family_ids: dict[str, str] = {}
        task_options: list[list[str]] = []
        for index, row in enumerate(catalog_rows, start=1):
            subject = f"base-unit catalog row {index}"
            family_id = require_text(row, "任务族ID", subject)
            family_name = require_text(row, "任务族名称", subject)
            unit_id = require_text(row, "基础单元ID", subject)
            unit_name = require_text(row, "基础单元名称", subject)
            if unit_id in base_units:
                raise ValueError(f"base-unit catalog ID is duplicated: {unit_id}")
            if unit_name in base_unit_names:
                raise ValueError(f"base-unit catalog name is duplicated: {unit_name}")
            base_unit_names.add(unit_name)
            if family_id in family_names and family_names[family_id] != family_name:
                raise ValueError(f"task-family ID maps to multiple names: {family_id}")
            if family_name in family_ids and family_ids[family_name] != family_id:
                raise ValueError(f"task-family name maps to multiple IDs: {family_name}")
            family_names[family_id] = family_name
            family_ids[family_name] = family_id
            modes = [
                mode
                for mode, header in MODE_EFFORT_HEADERS.items()
                if usable_mode(row, header, subject)
            ]
            if not modes:
                raise ValueError(f"base-unit catalog must configure a work mode: {unit_id}")
            task_options.extend([unit_id, mode] for mode in modes)
            base_units[unit_id] = {
                "taskFamilyId": family_id,
                "taskFamily": family_name,
                "name": unit_name,
                "countRule": require_text(row, "计数口径", subject),
                "includes": require_text(row, "包含内容", subject),
                "excludes": require_text(row, "不包含内容", subject),
                "allowedWorkModes": modes,
                "complexityStandards": {
                    level: require_text(row, f"{level}标准", subject)
                    for level in COMPLEXITIES
                },
                "splitRule": require_text(row, "X/拆分条件", subject),
            }
        if len(base_units) != 37 or len(family_names) != 13:
            raise ValueError("template must define exactly 37 base units and 13 task families")
        validate_complexity_parameters(parameter_rows)
        return {
            "baseUnits": base_units,
            "taskOptions": task_options,
            "complexities": list(COMPLEXITIES),
        }
    finally:
        workbook.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="Read task rules from the project SOW template")
    parser.add_argument("--project-root", required=True, type=Path)
    root = parser.parse_args().project_root.resolve()
    try:
        contract = read_contract(root / ".ai-sow/templates/sow-template.xlsx")
    except Exception as error:
        print(json.dumps({"outcome": "BLOCKED", "summary": str(error)}, ensure_ascii=False))
        return 2
    print(json.dumps(contract, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    sys.exit(main())
