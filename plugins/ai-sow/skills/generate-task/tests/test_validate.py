from __future__ import annotations

import importlib.util
import json
import shutil
import stat
import subprocess
import sys
from collections import Counter
from pathlib import Path
from types import SimpleNamespace
from typing import Any

import openpyxl
import pytest


SKILL_ROOT = Path(__file__).parents[1]
VALIDATE = SKILL_ROOT / "scripts/validate.py"
READ_TEMPLATE = SKILL_ROOT / "scripts/read_template.py"
FIXTURE = SKILL_ROOT / "fixtures/estimate.valid.json"
TEMPLATE = SKILL_ROOT / "fixtures/sow-template.xlsx"
ESTIMATE_SCHEMA = SKILL_ROOT / "contracts/estimate.schema.json"


def write_json(root: Path, relative: str, payload: object) -> None:
    path = root / relative
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload), encoding="utf-8")


def prepare(root: Path) -> None:
    write_json(
        root,
        ".ai-sow/data/generate-story/delivery.json",
        {
            "stories": [
                {"storyId": "story-customer-profile"},
                {"storyId": "story-profile-api"},
            ],
            "integrations": [
                {
                    "integrationId": "integration-profile-api",
                    "storyId": "story-profile-api",
                    "owner": "INTERNAL",
                }
            ],
        },
    )
    write_json(root, ".ai-sow/data/analyze-as-is/asis.json", {
        "items": [],
        "effectiveStartItems": [
            {
                "effectiveStartItemId": "effective-start-customer-api",
                "topic": "APPLICATION",
                "itemType": "COMPONENT",
                "name": "现有客户档案 API 与页面框架",
                "summary": "当前客户档案接口和页面框架可作为本次交付的有效起点。",
                "sourceItemIds": [],
                "commitmentIds": [],
            },
        ],
    })
    write_json(root, ".ai-sow/data/generate-task/estimate.json", json.loads(FIXTURE.read_text()))
    template = root / ".ai-sow/templates/sow-template.xlsx"
    template.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(TEMPLATE, template)


def run(script: Path, root: Path) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [sys.executable, str(script), "--project-root", str(root)],
        capture_output=True,
        text=True,
        check=False,
    )


def diagnostic_codes(result: subprocess.CompletedProcess[str]) -> set[str]:
    return {
        item["code"]
        for item in json.loads(result.stdout)["diagnostics"]
    }


def test_reads_authoritative_task_options_from_project_template(tmp_path: Path) -> None:
    prepare(tmp_path)

    result = run(READ_TEMPLATE, tmp_path)

    assert result.returncode == 0, result.stderr
    payload = json.loads(result.stdout)
    assert len(payload["baseUnits"]) == 37
    assert len({entry["taskFamilyId"] for entry in payload["baseUnits"].values()}) == 13
    assert len(payload["taskOptions"]) == 86
    assert ["BU-UI-INTERACTION", "新建"] in payload["taskOptions"]
    assert ["BU-UI-INTERACTION", "接入复用"] not in payload["taskOptions"]
    assert payload["baseUnits"]["BU-UI-INTERACTION"] == {
        "taskFamilyId": "TF-FRONTEND",
        "taskFamily": "前端",
        "name": "界面与交互",
        "countRule": "一个可独立验收的界面，或一条紧密关联的交互流程",
        "includes": "界面结构与交互、状态管理、权限呈现、加载、空白和异常状态及开发自测",
        "excludes": "后端服务和另行开展的测试",
        "allowedWorkModes": ["新建", "调整"],
        "complexityStandards": {
            "S": "1～2 种主要状态；标准组件；简单校验；单一角色",
            "M": "3～5 种主要状态；条件联动；复杂表格、上传或多角色",
            "L": "复杂状态机；实时或离线处理；跨界面状态恢复；严格无障碍或多终端适配",
        },
        "splitRule": "包含多条可独立验收的交互流程，或跨多个应用",
    }
    assert payload["baseUnits"]["BU-USER-TRAINING"] == {
        "taskFamilyId": "TF-DELIVERY-HANDOVER",
        "taskFamily": "交付与移交",
        "name": "用户培训与使用材料",
        "countRule": "一个明确用户群体针对一项连贯能力的培训和材料交付",
        "includes": "培训需求确认、使用材料、演示或练习、培训实施、问答、出席记录及材料移交",
        "excludes": "运维交接、技术架构培训、业务内容翻译和长期培训运营",
        "allowedWorkModes": ["新建", "调整"],
        "complexityStandards": {
            "S": "单一角色；标准流程；一次短时讲解；无专门练习环境",
            "M": "一个用户群体包含多个相关角色；需要演示、练习和问答",
            "L": "多语言、Train-the-Trainer、专门练习环境或正式能力考核",
        },
        "splitRule": "多个独立用户群体或多个无关产品，必须拆分",
    }
    assert payload["complexities"] == ["S", "M", "L"]
    assert payload["complexityFactors"] == {"S": 0.6, "M": 1.0, "L": 1.5}


def test_rejects_invalid_inline_base_effort_value(tmp_path: Path) -> None:
    prepare(tmp_path)
    template = tmp_path / ".ai-sow/templates/sow-template.xlsx"
    workbook = openpyxl.load_workbook(template)
    try:
        worksheet = workbook["92-基础人天"]
        worksheet["H5"] = "待校准"
        workbook.save(template)
    finally:
        workbook.close()

    result = run(READ_TEMPLATE, tmp_path)

    assert result.returncode == 2
    assert "新建M档人天 must be a positive number or ❌" in json.loads(result.stdout)["summary"]


def test_rejects_catalog_without_complete_complexity_standard(tmp_path: Path) -> None:
    prepare(tmp_path)
    template = tmp_path / ".ai-sow/templates/sow-template.xlsx"
    workbook = openpyxl.load_workbook(template)
    try:
        workbook["92-基础人天"]["L5"] = None
        workbook.save(template)
    finally:
        workbook.close()

    result = run(READ_TEMPLATE, tmp_path)

    assert result.returncode == 2
    assert "M标准" in json.loads(result.stdout)["summary"]


def test_rejects_missing_complexity_factor_project_parameter(tmp_path: Path) -> None:
    prepare(tmp_path)
    template = tmp_path / ".ai-sow/templates/sow-template.xlsx"
    workbook = openpyxl.load_workbook(template)
    try:
        workbook["91-项目参数"]["C5"] = None
        workbook.save(template)
    finally:
        workbook.close()

    result = run(READ_TEMPLATE, tmp_path)

    assert result.returncode == 2
    assert "complexity factor must be positive: S" in json.loads(result.stdout)["summary"]


def test_rejects_uncalibrated_complexity_factor_project_parameter(tmp_path: Path) -> None:
    prepare(tmp_path)
    template = tmp_path / ".ai-sow/templates/sow-template.xlsx"
    workbook = openpyxl.load_workbook(template)
    try:
        workbook["91-项目参数"]["F5"] = "待样本校准"
        workbook.save(template)
    finally:
        workbook.close()

    result = run(READ_TEMPLATE, tmp_path)

    assert result.returncode == 2
    assert "complexity factor is not calibrated: S" in json.loads(result.stdout)["summary"]


def test_estimate_schema_uses_0_4_contract_urn() -> None:
    assert json.loads(ESTIMATE_SCHEMA.read_text())["$id"] == (
        "urn:ai-sow:generate-task:estimate:0.1"
    )


@pytest.mark.parametrize(
    "mutation",
    [
        lambda task: task.pop("workModeEvidence"),
        lambda task: task["workModeEvidence"].update(
            {"projectSideWorkTypes": []}
        ),
        lambda task: task["workModeEvidence"].pop("projectSideWorkCommitment"),
    ],
)
def test_schema_requires_structured_reuse_evidence(
    tmp_path: Path,
    mutation: Any,
) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    payload = json.loads(path.read_text())
    mutation(payload["tasks"][1])
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    assert "SCHEMA_INVALID" in diagnostic_codes(result)


def test_accepts_tasks_covering_all_stories(tmp_path: Path) -> None:
    prepare(tmp_path)

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 0, result.stderr
    assert json.loads(result.stdout)["outcome"] == "OK"


@pytest.mark.parametrize(
    ("field", "value"),
    [
        ("professionalDomain", "前端"),
        ("activity", "实现"),
        ("quantity", 1),
    ],
)
def test_schema_forbids_removed_task_fields(
    tmp_path: Path,
    field: str,
    value: object,
) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    payload = json.loads(path.read_text())
    payload["tasks"][0][field] = value
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    assert any(
        item["code"] == "SCHEMA_INVALID" and field in item["message"]
        for item in json.loads(result.stdout)["diagnostics"]
    )


def test_schema_forbids_removed_sit_estimates(tmp_path: Path) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    payload = json.loads(path.read_text())
    payload["sitEstimates"] = []
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    assert any(
        item["code"] == "SCHEMA_INVALID" and "sitEstimates" in item["message"]
        for item in json.loads(result.stdout)["diagnostics"]
    )


@pytest.mark.parametrize("work_mode", ["采用", "替换", "退役"])
def test_schema_rejects_removed_task_work_modes(
    tmp_path: Path,
    work_mode: str,
) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    payload = json.loads(path.read_text())
    payload["tasks"][0]["workMode"] = work_mode
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    assert any(
        item["code"] == "SCHEMA_INVALID" and work_mode in item["message"]
        for item in json.loads(result.stdout)["diagnostics"]
    )


def test_schema_requires_complexity_rationale_for_s_or_l(tmp_path: Path) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    payload = json.loads(path.read_text())
    del payload["tasks"][0]["complexityRationale"]
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    assert any(
        item["code"] == "SCHEMA_INVALID"
        and "complexityRationale" in item["message"]
        for item in json.loads(result.stdout)["diagnostics"]
    )


def test_schema_forbids_complexity_rationale_for_m(tmp_path: Path) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    payload = json.loads(path.read_text())
    payload["tasks"][1]["complexityRationale"] = "M 档无需说明。"
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    assert any(
        item["code"] == "SCHEMA_INVALID"
        and "complexityRationale" in item["message"]
        for item in json.loads(result.stdout)["diagnostics"]
    )


def test_allows_one_or_three_concrete_tasks_per_story(tmp_path: Path) -> None:
    prepare(tmp_path)
    write_json(
        tmp_path,
        ".ai-sow/data/generate-story/delivery.json",
        {
            "stories": [
                {"storyId": "story-checkout"},
                {"storyId": "story-reporting"},
            ],
            "integrations": [],
        },
    )
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    estimate = json.loads(path.read_text())
    checkout = dict(estimate["tasks"][0])
    checkout["storyId"] = "story-checkout"
    checkout["taskId"] = "task-checkout-page"
    checkout["name"] = "实现结账页面"
    reporting = dict(checkout)
    reporting["storyId"] = "story-reporting"
    reporting["taskId"] = "task-reporting-export"
    reporting["name"] = "实现报表导出"
    second_checkout = dict(checkout)
    second_checkout["taskId"] = "task-checkout-validation"
    second_checkout["name"] = "实现结账校验"
    third_checkout = dict(checkout)
    third_checkout["taskId"] = "task-checkout-confirmation"
    third_checkout["name"] = "实现结账确认"
    estimate["tasks"] = [checkout, second_checkout, third_checkout, reporting]
    path.write_text(json.dumps(estimate))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 0, result.stdout
    assert "quantity" not in estimate["tasks"][0]
    assert Counter(task["storyId"] for task in estimate["tasks"]) == {
        "story-checkout": 3,
        "story-reporting": 1,
    }


def test_rejects_story_without_a_task(tmp_path: Path) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    estimate = json.loads(path.read_text())
    estimate["tasks"] = estimate["tasks"][:1]
    path.write_text(json.dumps(estimate))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    assert any(
        item["code"] == "TASK_COVERAGE_MISSING"
        for item in json.loads(result.stdout)["diagnostics"]
    )


def test_rejects_normalized_duplicate_task_description_within_story(tmp_path: Path) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    estimate = json.loads(path.read_text())
    duplicate = dict(estimate["tasks"][0])
    duplicate["taskId"] = "task-customer-profile-duplicate"
    duplicate["name"] = "  实现客户档案页面和状态  "
    estimate["tasks"].append(duplicate)
    path.write_text(json.dumps(estimate))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    assert any(
        item["code"] == "TASK_DESCRIPTION_DUPLICATE"
        for item in json.loads(result.stdout)["diagnostics"]
    )


def test_rejects_two_release_cutover_tasks_for_one_story(tmp_path: Path) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    payload = json.loads(path.read_text())
    for suffix in ("primary", "secondary"):
        payload["tasks"].append(
            {
                "taskId": f"task-release-cutover-{suffix}",
                "storyId": "story-customer-profile",
                "name": f"制定并实施客户档案 {suffix} 发布与切换",
                "baseUnit": "BU-RELEASE-CUTOVER",
                "workMode": "新建",
                "workModeRationale": "本次交付需要形成统一窗口、回滚方案并完成实际切换。",
                "complexity": "M",
                "matchedEffectiveStartItemIds": [],
                "rationale": "该 Task 覆盖同一 Story 下的一次发布计划与实际切换。",
            }
        )
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    assert "RELEASE_CUTOVER_DUPLICATE" in diagnostic_codes(result)


def test_rejects_problem_diagnosis_and_remediation_for_same_story(
    tmp_path: Path,
) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    payload = json.loads(path.read_text())
    payload["tasks"].extend(
        [
            {
                "taskId": "task-profile-problem-diagnosis",
                "storyId": "story-customer-profile",
                "name": "诊断并恢复客户档案读取故障",
                "baseUnit": "BU-TECH-SUPPORT",
                "workMode": "新建",
                "workModeRationale": "已知故障仍需收集证据、诊断并形成恢复结论。",
                "complexity": "M",
                "matchedEffectiveStartItemIds": [],
                "rationale": "一个具有统一问题描述和处理结论的明确问题。",
            },
            {
                "taskId": "task-profile-root-cause-remediation",
                "storyId": "story-customer-profile",
                "name": "整改现有客户档案 API 与页面框架的已确认根因",
                "baseUnit": "BU-ROOT-CAUSE-REMEDIATION",
                "workMode": "新建",
                "workModeRationale": "根因已经确认，需要修改受影响的现有代码并回归。",
                "complexity": "M",
                "matchedEffectiveStartItemIds": ["effective-start-customer-api"],
                "rationale": "对同一已确认根因实施代码整改和针对性回归。",
            },
        ]
    )
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    assert "PROBLEM_TASK_OVERLAP" in diagnostic_codes(result)


def test_schema_enforces_task_id_prefix(tmp_path: Path) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    estimate = json.loads(path.read_text())
    estimate["tasks"][0]["taskId"] = "wrong-task"
    path.write_text(json.dumps(estimate))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    assert any(
        item["code"] == "SCHEMA_INVALID" and "wrong-task" in item["message"]
        for item in json.loads(result.stdout)["diagnostics"]
    )


def test_rejects_task_combination_not_configured_in_template(tmp_path: Path) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    payload = json.loads(path.read_text())
    payload["tasks"][0]["baseUnit"] = "不存在的单位"
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    assert any(
        item["code"] == "TASK_OPTION_NOT_CONFIGURED"
        for item in json.loads(result.stdout)["diagnostics"]
    )


def test_rejects_unknown_effective_start_reference(tmp_path: Path) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    payload = json.loads(path.read_text())
    payload["tasks"][0]["matchedEffectiveStartItemIds"] = ["effective-start-unknown"]
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    assert any(
        item["code"] == "EFFECTIVE_START_REF_UNKNOWN"
        for item in json.loads(result.stdout)["diagnostics"]
    )


def test_requires_effective_start_for_non_new_work_mode(tmp_path: Path) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    payload = json.loads(path.read_text())
    payload["tasks"][0]["workMode"] = "调整"
    payload["tasks"][0]["workModeEvidence"] = {
        "effectiveStartItemId": "effective-start-customer-api",
        "effectiveStartItemName": "现有客户档案 API 与页面框架",
    }
    payload["tasks"][0]["matchedEffectiveStartItemIds"] = []
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    assert any(
        item["code"] == "EFFECTIVE_START_REQUIRED"
        for item in json.loads(result.stdout)["diagnostics"]
    )


def test_rejects_irrelevant_effective_start_reference(tmp_path: Path) -> None:
    prepare(tmp_path)
    asis_path = tmp_path / ".ai-sow/data/analyze-as-is/asis.json"
    asis = json.loads(asis_path.read_text())
    asis["effectiveStartItems"][0]["name"] = "客户财务结算批处理"
    asis["effectiveStartItems"][0]["summary"] = "现有客户财务月结作业。"
    asis_path.write_text(json.dumps(asis))
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    payload = json.loads(path.read_text())
    payload["tasks"][0]["workMode"] = "调整"
    payload["tasks"][0]["workModeRationale"] = (
        "保留现有客户档案页面入口，只调整页面状态和校验规则。"
    )
    payload["tasks"][0]["workModeEvidence"] = {
        "effectiveStartItemId": "effective-start-customer-api",
        "effectiveStartItemName": "客户财务结算批处理",
    }
    payload["tasks"][0]["matchedEffectiveStartItemIds"] = [
        "effective-start-customer-api"
    ]
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    assert "EFFECTIVE_START_IRRELEVANT" in diagnostic_codes(result)


def test_rejects_adjusted_test_without_existing_test_asset(tmp_path: Path) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    payload = json.loads(path.read_text())
    payload["tasks"][0].update(
        {
            "name": "调整现有客户档案测试方案",
            "baseUnit": "BU-MANUAL-TESTING",
            "workMode": "调整",
            "workModeRationale": (
                "修改客户档案测试方案，但引用的 Effective Start 只有页面框架。"
            ),
            "workModeEvidence": {
                "effectiveStartItemId": "effective-start-customer-api",
                "effectiveStartItemName": "现有客户档案 API 与页面框架",
            },
            "matchedEffectiveStartItemIds": ["effective-start-customer-api"],
        }
    )
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    assert "WORK_MODE_ADJUSTMENT_ASSET_UNSPECIFIED" in diagnostic_codes(result)


def test_allows_adjustment_of_identified_existing_test_asset(tmp_path: Path) -> None:
    prepare(tmp_path)
    asis_path = tmp_path / ".ai-sow/data/analyze-as-is/asis.json"
    asis = json.loads(asis_path.read_text())
    asis["effectiveStartItems"][0]["name"] = "既有客户档案回归测试方案"
    asis["effectiveStartItems"][0]["summary"] = (
        "包含客户档案页面的回归测试范围和测试用例。"
    )
    asis_path.write_text(json.dumps(asis))
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    payload = json.loads(path.read_text())
    payload["tasks"][0].update(
        {
            "name": "调整既有客户档案回归测试方案",
            "baseUnit": "BU-MANUAL-TESTING",
            "workMode": "调整",
            "workModeRationale": (
                "保留既有客户档案回归测试方案，修改其中的测试范围和测试用例。"
            ),
            "workModeEvidence": {
                "effectiveStartItemId": "effective-start-customer-api",
                "effectiveStartItemName": "既有客户档案回归测试方案",
            },
            "matchedEffectiveStartItemIds": ["effective-start-customer-api"],
        }
    )
    payload["tasks"][1]["workModeRationale"] = (
        "既有客户档案回归测试方案保持不变；本项目负责并交付：认证、映射、适配。"
    )
    payload["tasks"][1]["workModeEvidence"]["effectiveStartItemName"] = (
        "既有客户档案回归测试方案"
    )
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 0, result.stdout


@pytest.mark.parametrize(
    ("rationale", "activity_types", "commitment"),
    [
        (
            "现有客户档案 API 与页面框架保持不变，本项目侧认证、映射和适配均不需要。",
            ["AUTHENTICATE", "MAP", "ADAPT"],
            "本项目负责并交付：认证、映射、适配",
        ),
        (
            "现有客户档案 API 与页面框架保持不变，配置不是本项目工作。",
            ["CONFIGURE"],
            "本项目负责并交付：配置",
        ),
        (
            "现有客户档案 API 与页面框架保持不变，本项目不会开展专项验证。",
            ["SPECIALIZED_VERIFY"],
            "本项目负责并交付：专项验证",
        ),
        (
            "现有客户档案 API 与页面框架保持不变，本项目不负责配置，只直接调用。",
            ["CONFIGURE"],
            "本项目负责并交付：配置",
        ),
        (
            "现有客户档案 API 与页面框架保持不变，认证、映射和适配全部由客户团队完成，本项目只直接调用。",
            ["AUTHENTICATE", "MAP", "ADAPT"],
            "本项目负责并交付：认证、映射、适配",
        ),
        (
            "现有客户档案 API 与页面框架保持不变，配置没有必要，本项目只直接调用。",
            ["CONFIGURE"],
            "本项目负责并交付：配置",
        ),
        (
            "现有客户档案 API 与页面框架保持不变，本项目既不配置，也不适配，只直接调用。",
            ["CONFIGURE", "ADAPT"],
            "本项目负责并交付：配置、适配",
        ),
    ],
)
def test_rejects_reuse_without_separately_estimable_integration_work(
    tmp_path: Path,
    rationale: str,
    activity_types: list[str],
    commitment: str,
) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    payload = json.loads(path.read_text())
    payload["tasks"][1]["workModeRationale"] = rationale
    payload["tasks"][1]["workModeEvidence"]["projectSideWorkTypes"] = activity_types
    payload["tasks"][1]["workModeEvidence"]["projectSideWorkCommitment"] = commitment
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    assert "WORK_MODE_REUSE_NOT_ESTIMABLE" in diagnostic_codes(result)


def test_allows_new_work_without_effective_start(tmp_path: Path) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    payload = json.loads(path.read_text())
    payload["tasks"][0]["workMode"] = "新建"
    payload["tasks"][0]["matchedEffectiveStartItemIds"] = []
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 0, result.stdout


def test_allows_new_work_with_effective_start(tmp_path: Path) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    payload = json.loads(path.read_text())
    payload["tasks"][0]["workMode"] = "新建"
    payload["tasks"][0]["matchedEffectiveStartItemIds"] = [
        "effective-start-customer-api",
    ]
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 0, result.stdout


def test_rejects_generic_work_mode_rationale(tmp_path: Path) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    payload = json.loads(path.read_text())
    payload["tasks"][0]["workModeRationale"] = "新建任务"
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    assert "WORK_MODE_RATIONALE_GENERIC" in diagnostic_codes(result)


def test_rejects_generic_or_copied_complexity_rationale(tmp_path: Path) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    payload = json.loads(path.read_text())
    payload["tasks"][0]["complexityRationale"] = (
        "1～2 种主要状态；标准组件；简单校验；单一角色"
    )
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    assert "COMPLEXITY_RATIONALE_GENERIC" in diagnostic_codes(result)


def test_allows_l_complexity_with_concrete_deviation_facts(tmp_path: Path) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    payload = json.loads(path.read_text())
    payload["tasks"][0]["complexity"] = "L"
    payload["tasks"][0]["complexityRationale"] = (
        "包含 8 种主要状态、跨界面恢复和无障碍键盘操作，均已在原型评审中确认。"
    )
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 0, result.stdout


def test_requires_effective_start_for_existing_object_new_work(tmp_path: Path) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    payload = json.loads(path.read_text())
    payload["tasks"][0].update(
        {
            "baseUnit": "BU-DATA-MIGRATION",
            "workMode": "新建",
            "workModeRationale": "为现有客户档案数据新做一次源到目标迁移。",
            "complexity": "M",
            "matchedEffectiveStartItemIds": [],
        }
    )
    payload["tasks"][0].pop("complexityRationale")
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    assert "EFFECTIVE_START_REQUIRED" in diagnostic_codes(result)


def test_requires_integration_id_for_integration_task(tmp_path: Path) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    payload = json.loads(path.read_text())
    del payload["tasks"][1]["integrationId"]
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    assert "INTEGRATION_ID_REQUIRED" in diagnostic_codes(result)


def test_forbids_integration_id_for_non_integration_task(tmp_path: Path) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    payload = json.loads(path.read_text())
    payload["tasks"][0]["integrationId"] = "integration-profile-api"
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    assert "INTEGRATION_ID_FORBIDDEN" in diagnostic_codes(result)


def test_rejects_unknown_integration_reference(tmp_path: Path) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    payload = json.loads(path.read_text())
    payload["tasks"][1]["integrationId"] = "integration-unknown"
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    assert "INTEGRATION_REF_UNKNOWN" in diagnostic_codes(result)


def test_rejects_integration_owner_and_base_unit_mismatch(tmp_path: Path) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    payload = json.loads(path.read_text())
    payload["tasks"][1]["baseUnit"] = "BU-EXTERNAL-INTEGRATION"
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    assert "INTEGRATION_OWNER_MISMATCH" in diagnostic_codes(result)


def test_rejects_integration_story_mismatch(tmp_path: Path) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-story/delivery.json"
    payload = json.loads(path.read_text())
    payload["integrations"][0]["storyId"] = "story-customer-profile"
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    assert "INTEGRATION_STORY_MISMATCH" in diagnostic_codes(result)


def test_rejects_integration_without_task_coverage(tmp_path: Path) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-story/delivery.json"
    payload = json.loads(path.read_text())
    payload["integrations"].append(
        {
            "integrationId": "integration-audit-event",
            "storyId": "story-profile-api",
            "owner": "INTERNAL",
        }
    )
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    assert "INTEGRATION_COVERAGE_MISSING" in diagnostic_codes(result)


def test_rejects_duplicate_integration_task_coverage(tmp_path: Path) -> None:
    prepare(tmp_path)
    path = tmp_path / ".ai-sow/data/generate-task/estimate.json"
    payload = json.loads(path.read_text())
    duplicate = dict(payload["tasks"][1])
    duplicate["taskId"] = "task-profile-api-integration-duplicate"
    duplicate["name"] = "接入内部客户档案 API 的第二条重复任务"
    payload["tasks"].append(duplicate)
    path.write_text(json.dumps(payload))

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    assert "INTEGRATION_COVERAGE_DUPLICATE" in diagnostic_codes(result)


@pytest.mark.parametrize("symlink_kind", ["directory", "report"])
def test_blocks_validation_output_symlink_escape(
    tmp_path: Path,
    symlink_kind: str,
) -> None:
    prepare(tmp_path)
    validation_path = tmp_path / ".ai-sow/validation/generate-task.json"
    outside = tmp_path.parent / f"{tmp_path.name}-outside-validation"
    outside.mkdir()
    if symlink_kind == "directory":
        validation_path.parent.symlink_to(outside, target_is_directory=True)
    else:
        validation_path.parent.mkdir(parents=True)
        validation_path.symlink_to(outside / "escaped.json")

    result = run(VALIDATE, tmp_path)

    assert result.returncode == 2
    payload = json.loads(result.stdout)
    assert payload["outcome"] == "BLOCKED"
    assert any(item["code"] == "OUTPUT_PATH_UNSAFE" for item in payload["diagnostics"])
    assert list(outside.iterdir()) == []


def test_portable_directory_snapshot_rejects_windows_reparse_point(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.syspath_prepend(str(VALIDATE.parent))
    spec = importlib.util.spec_from_file_location("generate_task_reparse", VALIDATE)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    snapshot = SimpleNamespace(
        st_mode=stat.S_IFDIR | 0o755,
        st_file_attributes=stat.FILE_ATTRIBUTE_REPARSE_POINT,
    )
    path = SimpleNamespace(stat=lambda *, follow_symlinks: snapshot)

    with pytest.raises(OSError, match="reparse point"):
        module._safe_directory_snapshot(path)


def test_portable_report_write_rejects_windows_reparse_point(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    validation_path = tmp_path / ".ai-sow/validation/generate-task.json"
    validation_path.parent.mkdir(parents=True)
    validation_path.write_text("original\n", encoding="utf-8")
    monkeypatch.syspath_prepend(str(VALIDATE.parent))
    spec = importlib.util.spec_from_file_location("generate_task_report_reparse", VALIDATE)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    original_stat = Path.stat

    def stat_with_reparse(path: Path, *, follow_symlinks: bool = True) -> object:
        snapshot = original_stat(path, follow_symlinks=follow_symlinks)
        if path == validation_path and not follow_symlinks:
            return SimpleNamespace(
                st_mode=snapshot.st_mode,
                st_dev=snapshot.st_dev,
                st_ino=snapshot.st_ino,
                st_file_attributes=stat.FILE_ATTRIBUTE_REPARSE_POINT,
            )
        return snapshot

    monkeypatch.setattr(Path, "stat", stat_with_reparse)

    with pytest.raises(OSError, match="reparse point"):
        module._write_validation_report_portable(
            tmp_path,
            validation_path,
            "replacement\n",
        )
    assert validation_path.read_text(encoding="utf-8") == "original\n"


@pytest.mark.parametrize("race_kind", ["directory", "report"])
@pytest.mark.parametrize("writer_backend", ["native", "portable"])
def test_blocks_validation_symlink_swap_after_safety_check(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
    race_kind: str,
    writer_backend: str,
) -> None:
    prepare(tmp_path)
    validation_path = tmp_path / ".ai-sow/validation/generate-task.json"
    validation_path.parent.mkdir(parents=True)
    outside = tmp_path.parent / f"{tmp_path.name}-outside-race"
    outside.mkdir()
    original_validation_dir = validation_path.parent.with_name("validation-before-race")
    monkeypatch.syspath_prepend(str(VALIDATE.parent))
    spec = importlib.util.spec_from_file_location("generate_task_race", VALIDATE)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    if writer_backend == "portable":
        monkeypatch.setattr(
            module,
            "write_validation_report",
            module._write_validation_report_portable,
        )
    original_check = module.validation_output_diagnostic

    def check_then_swap(project_root: Path, report_path: Path) -> dict[str, str] | None:
        result = original_check(project_root, report_path)
        assert result is None
        if race_kind == "directory":
            validation_path.parent.rename(original_validation_dir)
            validation_path.parent.symlink_to(outside, target_is_directory=True)
        else:
            validation_path.symlink_to(outside / "escaped.json")
        return result

    monkeypatch.setattr(module, "validation_output_diagnostic", check_then_swap)
    monkeypatch.setattr(
        sys,
        "argv",
        [str(VALIDATE), "--project-root", str(tmp_path)],
    )

    returncode = module.main()
    payload = json.loads(capsys.readouterr().out)

    assert returncode == 2
    assert payload["outcome"] == "BLOCKED"
    assert any(item["code"] == "OUTPUT_UNWRITABLE" for item in payload["diagnostics"])
    assert list(outside.iterdir()) == []
