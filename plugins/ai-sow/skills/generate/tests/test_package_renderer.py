from __future__ import annotations

import copy
import json
import re
import sys
from pathlib import Path

import openpyxl
import pytest


SKILL_ROOT = Path(__file__).parents[1]
SCRIPTS = SKILL_ROOT / "scripts"
FIXTURES = SKILL_ROOT / "fixtures"
ASSETS = SKILL_ROOT / "assets"
PUBLIC_WORKBOOK = SKILL_ROOT.parents[1] / "docs/reference/SOW估算与生成示例_v1.3.xlsx"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

from contracts import canonical_json_bytes, sha256_bytes  # noqa: E402
from office_engine import OfficeEngineError, require_office_engine  # noqa: E402
import package_renderer as package_renderer_module  # noqa: E402
from package_renderer import PackageRenderError, render_package  # noqa: E402
from workbook import audit_calculated_workbook, build_rows, write_workbook  # noqa: E402


def fixture(mode: str, name: str) -> dict[str, object]:
    return json.loads((FIXTURES / mode / name).read_text(encoding="utf-8"))


def render_fixture(root: Path, mode: str = "brownfield"):
    manifest = fixture(mode, "input-manifest.json")
    scope = fixture(mode, "scope.json")
    delivery = fixture(mode, "delivery.json")
    delivery["scopeSha256"] = sha256_bytes(canonical_json_bytes(scope))
    review = fixture(mode, "final-review.json")
    review["scopeSha256"] = sha256_bytes(canonical_json_bytes(scope))
    review["deliverySha256"] = sha256_bytes(canonical_json_bytes(delivery))
    result = render_package(
        generation_id="000001",
        template_path=ASSETS / "sow-template.xlsx",
        output_root=root,
        input_manifest=manifest,
        scope=scope,
        delivery=delivery,
        review=review,
    )
    return result, manifest, scope, delivery, review


def dual_feature_fixture() -> tuple[dict[str, object], dict[str, object]]:
    scope = fixture("greenfield", "scope.json")
    delivery = fixture("greenfield", "delivery.json")
    second_feature = copy.deepcopy(scope["features"][0])
    second_feature.update(
        {"featureId": "feature-refund-query", "name": "退款进度查询"}
    )
    scope["features"].append(second_feature)
    second_story = copy.deepcopy(delivery["stories"][0])
    second_story.update(
        {
            "storyId": "story-refund-query",
            "featureId": "feature-refund-query",
            "name": "[退款查询] 客户查看退款进度",
        }
    )
    delivery["stories"].append(second_story)
    second_task = copy.deepcopy(delivery["tasks"][0])
    second_task.update(
        {
            "taskId": "task-refund-query",
            "storyId": "story-refund-query",
            "name": "开发退款进度查询服务",
        }
    )
    delivery["tasks"].append(second_task)
    return scope, delivery


def workbook_string_contents(workbook: openpyxl.Workbook) -> list[str]:
    return [
        cell.value
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str)
    ]


def test_four_sheet_projection_uses_story_single_feature_as_subrequirement() -> None:
    scope = fixture("greenfield", "scope.json")
    delivery = fixture("greenfield", "delivery.json")

    rows = build_rows(
        scope,
        delivery,
        {
            "BU-ARCHITECTURE-DESIGN": "架构设计",
            "BU-BUSINESS-SERVICE-API": "业务服务/API",
            "BU-EXTERNAL-INTEGRATION": "外部系统对接",
        },
    )

    assert rows["SOWStoryTable"] == [
        {
            "需求": "退款服务",
            "子需求": "退款申请处理",
            "故事": "[退款申请] 客户提交退款",
            "UAT适用": "是",
            "验收条件": (
                "• 有效退款申请被受理并返回可追踪编号\n"
                "• 无效退款申请被拒绝并返回明确原因"
            ),
            "备注": (
                "假设（已明确）：测试环境按时可用；"
                "处置：环境延迟通过项目变更控制调整联调计划；"
                "估算边界：估算包含一次标准环境联调，不包含客户环境建设；"
                "变化触发：环境准备晚于约定日期或需供应商代建"
            ),
            "任务列表": "• [业务服务/API/新建/M] 新增退款申请提交接口",
        }
    ]


def test_story_projection_is_readable_and_minimal() -> None:
    scope = fixture("brownfield", "scope.json")
    delivery = fixture("brownfield", "delivery.json")

    rows = build_rows(
        scope,
        delivery,
        {
            "BU-ARCHITECTURE-DESIGN": "架构方案设计",
            "BU-BUSINESS-SERVICE-API": "业务服务与接口",
            "BU-EXTERNAL-INTEGRATION": "外部系统对接",
        },
    )

    workbook = openpyxl.load_workbook(ASSETS / "sow-template.xlsx", data_only=False)
    try:
        table = workbook["01-需求故事"].tables["SOWStoryTable"]
        formal_headers = [column.name for column in table.tableColumns]
    finally:
        workbook.close()
    assert formal_headers == [
        "需求",
        "子需求",
        "故事",
        "UAT适用",
        "验收条件",
        "备注",
        "任务列表",
        "故事人天",
        "校验结果",
    ]
    assert formal_headers[:6] == [
        "需求",
        "子需求",
        "故事",
        "UAT适用",
        "验收条件",
        "备注",
    ]
    assert formal_headers[6:] == ["任务列表", "故事人天", "校验结果"]

    story = rows["SOWStoryTable"][0]
    assert set(story) == {
        "需求",
        "子需求",
        "故事",
        "UAT适用",
        "验收条件",
        "备注",
        "任务列表",
    }
    assert all(line.startswith("• ") for line in story["验收条件"].splitlines())
    assert all(
        re.match(r"^• \[[^/]+/(?:新建|调整|接入复用)/(?:S|M|L)\] ", line)
        for line in story["任务列表"].splitlines()
    )


def test_task_projection_uses_catalog_name_and_three_labeled_reasons() -> None:
    scope = fixture("brownfield", "scope.json")
    delivery = fixture("brownfield", "delivery.json")
    delivery["tasks"][0]["complexity"] = "L"
    delivery["tasks"][0]["complexityRationale"] = "跨三个系统并含补偿。"

    rows = build_rows(
        scope,
        delivery,
        {
            "BU-ARCHITECTURE-DESIGN": "架构设计",
            "BU-BUSINESS-SERVICE-API": "业务服务/API",
            "BU-EXTERNAL-INTEGRATION": "外部系统对接",
        },
    )

    assert rows["TaskTable"][0] == {
        "所属故事": "[退款处理] 系统编排分项退款",
        "任务名称": "设计退款流程恢复目标与方案",
        "任务类型": "架构设计",
        "工作方式": "新建",
        "复杂度": "L",
        "备注": (
            "任务理由：以一项待明确恢复目标的方案设计实例计价。\n"
            "工作方式理由：往期 SOW 未包含退款流程恢复目标设计。\n"
            "复杂度理由：跨三个系统并含补偿。"
        ),
    }


def test_story_projection_leaves_notes_blank_without_linked_assumptions() -> None:
    scope = fixture("greenfield", "scope.json")
    delivery = fixture("greenfield", "delivery.json")
    scope["assumptions"][0]["featureIds"] = ["feature-unrelated"]

    rows = build_rows(
        scope,
        delivery,
        {
            "BU-ARCHITECTURE-DESIGN": "架构设计",
            "BU-BUSINESS-SERVICE-API": "业务服务/API",
            "BU-EXTERNAL-INTEGRATION": "外部系统对接",
        },
    )

    assert rows["SOWStoryTable"][0]["备注"] == ""


def test_story_projection_does_not_repeat_project_level_assumptions() -> None:
    scope, delivery = dual_feature_fixture()
    scope["assumptions"][0]["featureIds"] = [
        "feature-refund-processing",
        "feature-refund-query",
    ]

    rows = build_rows(
        scope,
        delivery,
        {
            "BU-ARCHITECTURE-DESIGN": "架构设计",
            "BU-BUSINESS-SERVICE-API": "业务服务/API",
            "BU-EXTERNAL-INTEGRATION": "外部系统对接",
        },
    )

    assert [row["备注"] for row in rows["SOWStoryTable"]] == ["", ""]


def test_story_projection_shows_feature_specific_note_only_once() -> None:
    scope = fixture("greenfield", "scope.json")
    delivery = fixture("greenfield", "delivery.json")
    second_story = copy.deepcopy(delivery["stories"][0])
    second_story.update(
        {
            "storyId": "story-refund-query",
            "name": "[退款查询] 客户查看退款进度",
        }
    )
    delivery["stories"].append(second_story)
    second_task = copy.deepcopy(delivery["tasks"][0])
    second_task.update(
        {
            "taskId": "task-refund-query",
            "storyId": "story-refund-query",
            "name": "开发退款进度查询服务",
        }
    )
    delivery["tasks"].append(second_task)

    rows = build_rows(
        scope,
        delivery,
        {
            "BU-ARCHITECTURE-DESIGN": "架构设计",
            "BU-BUSINESS-SERVICE-API": "业务服务/API",
            "BU-EXTERNAL-INTEGRATION": "外部系统对接",
        },
    )

    notes = [row["备注"] for row in rows["SOWStoryTable"]]
    assert notes[0].startswith("假设（已明确）：测试环境按时可用")
    assert notes[1] == ""


def test_story_projection_suppresses_same_note_copied_to_multiple_features() -> None:
    scope, delivery = dual_feature_fixture()
    second_assumption = copy.deepcopy(scope["assumptions"][0])
    second_assumption.update(
        {
            "assumptionId": "assumption-test-environment-copy",
            "featureIds": ["feature-refund-query"],
        }
    )
    scope["assumptions"].append(second_assumption)

    rows = build_rows(
        scope,
        delivery,
        {
            "BU-ARCHITECTURE-DESIGN": "架构设计",
            "BU-BUSINESS-SERVICE-API": "业务服务/API",
            "BU-EXTERNAL-INTEGRATION": "外部系统对接",
        },
    )

    assert [row["备注"] for row in rows["SOWStoryTable"]] == ["", ""]


def test_story_projection_does_not_repeat_homogeneous_uat_note() -> None:
    scope, delivery = dual_feature_fixture()
    scope["assumptions"][0].update(
        {
            "name": "统一 UAT 环境按时可用",
            "featureIds": ["feature-refund-processing", "feature-refund-query"],
        }
    )

    rows = build_rows(
        scope,
        delivery,
        {
            "BU-ARCHITECTURE-DESIGN": "架构设计",
            "BU-BUSINESS-SERVICE-API": "业务服务/API",
            "BU-EXTERNAL-INTEGRATION": "外部系统对接",
        },
    )

    assert [row["备注"] for row in rows["SOWStoryTable"]] == ["", ""]


def test_verified_workbook_formats_acceptance_criteria_and_task_list(
    tmp_path: Path,
) -> None:
    rendered, _manifest, _scope, _delivery, _review = render_fixture(tmp_path)
    workbook = openpyxl.load_workbook(rendered.workbook_path, data_only=True)
    try:
        story_sheet = workbook["01-需求故事"]
        assert story_sheet["E5"].value == (
            "• 通过审核的退款获得支付结果并可追踪\n"
            "• 失败支付分项可重试且最终状态保持一致"
        )
        assert story_sheet["G5"].value == (
            "• [架构方案设计/新建/M] 设计退款流程恢复目标与方案\n"
            "• [业务服务与接口/调整/M] 调整退款编排流程\n"
            "• [外部系统对接/新建/M] 接入支付平台退款接口"
        )
        assert story_sheet.row_dimensions[5].height >= 49
    finally:
        workbook.close()


@pytest.mark.parametrize("empty_collection", ["stories", "tasks"])
def test_workbook_rejects_logically_empty_estimate(
    tmp_path: Path, empty_collection: str
) -> None:
    scope = fixture("greenfield", "scope.json")
    delivery = fixture("greenfield", "delivery.json")
    delivery[empty_collection] = []
    output = tmp_path / "sow.xlsx"

    with pytest.raises(ValueError, match="at least one"):
        write_workbook(ASSETS / "sow-template.xlsx", scope, delivery, output)

    assert not output.exists()


def test_written_workbook_is_explicitly_a_candidate_before_recalculation(
    tmp_path: Path,
) -> None:
    scope = fixture("greenfield", "scope.json")
    delivery = fixture("greenfield", "delivery.json")

    audit = write_workbook(
        ASSETS / "sow-template.xlsx", scope, delivery, tmp_path / "sow.xlsx"
    )

    assert audit.trust_state == "CANDIDATE"
    assert (audit.story_count, audit.task_count) == (1, 1)
    assert audit.direct_days is None
    assert audit.engine_name is None


def test_written_workbook_allows_readable_vertical_print_pagination(
    tmp_path: Path,
) -> None:
    output = tmp_path / "sow.xlsx"
    write_workbook(
        ASSETS / "sow-template.xlsx",
        fixture("brownfield", "scope.json"),
        fixture("brownfield", "delivery.json"),
        output,
    )

    workbook = openpyxl.load_workbook(output, data_only=False, read_only=False)
    try:
        for worksheet in workbook.worksheets:
            assert worksheet.page_setup.fitToWidth == 1
            assert worksheet.page_setup.fitToHeight == 0
    finally:
        workbook.close()


def test_renderer_uses_scope_and_delivery_only(tmp_path: Path) -> None:
    rendered, _manifest, _scope, _delivery, _review = render_fixture(tmp_path)
    assert rendered.workbook_sha256
    assert rendered.notes_sha256
    assert rendered.files == ("output/sow-notes.md", "output/sow.xlsx")
    assert not any("validation/" in path or "reviews/" in path for path in rendered.files)


def test_renderer_returns_only_verified_calculated_workbook(tmp_path: Path) -> None:
    rendered, _manifest, _scope, _delivery, _review = render_fixture(tmp_path)

    assert rendered.workbook_audit.trust_state == "VERIFIED"
    assert rendered.workbook_audit.total_days == 9.5
    assert not (tmp_path / "output/sow.candidate.xlsx").exists()
    notes = Path(rendered.notes_path).read_text(encoding="utf-8")
    assert "模板参数校准状态" in notes
    assert "K_UAT：待样本校准" in notes
    assert "SIT_INT_SUPPORT：待样本校准" in notes
    assert "SIT_EXT_SUPPORT：待样本校准" in notes


def test_renderer_without_office_engine_cannot_create_formal_output(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.delenv("AI_SOW_OFFICE_BIN", raising=False)
    monkeypatch.setenv("PATH", "")
    manifest = fixture("brownfield", "input-manifest.json")
    scope = fixture("brownfield", "scope.json")
    delivery = fixture("brownfield", "delivery.json")
    review = fixture("brownfield", "final-review.json")

    with pytest.raises(PackageRenderError) as caught:
        render_package(
            generation_id="000002",
            template_path=ASSETS / "sow-template.xlsx",
            output_root=tmp_path / "generation",
            input_manifest=manifest,
            scope=scope,
            delivery=delivery,
            review=review,
        )

    assert caught.value.code == "OFFICE_ENGINE_UNAVAILABLE"
    assert not (tmp_path / "generation").exists()


def test_failed_office_roundtrip_preserves_prior_generation(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    generations = tmp_path / "generations"
    rendered, manifest, scope, delivery, review = render_fixture(generations / "000001")
    last_known_good_workbook = Path(rendered.workbook_path).read_bytes()
    last_known_good_notes = Path(rendered.notes_path).read_bytes()
    calls: list[tuple[bool, str]] = []

    def fail_roundtrip(candidate: Path, output: Path, _engine: object) -> None:
        calls.append((candidate.is_file(), output.name))
        raise OfficeEngineError("OFFICE_ENGINE_RECALCULATION_FAILED", "forced failure")

    monkeypatch.setattr(
        package_renderer_module, "recalculate_workbook", fail_roundtrip
    )

    with pytest.raises(PackageRenderError) as caught:
        render_package(
            generation_id="000002",
            template_path=ASSETS / "sow-template.xlsx",
            output_root=generations / "000002",
            input_manifest=manifest,
            scope=scope,
            delivery=delivery,
            review=review,
        )

    assert caught.value.code == "OFFICE_ENGINE_RECALCULATION_FAILED"
    assert calls == [(True, "sow.xlsx")]
    assert Path(rendered.workbook_path).read_bytes() == last_known_good_workbook
    assert Path(rendered.notes_path).read_bytes() == last_known_good_notes
    assert not (generations / "000002").exists()


def test_all_pass_with_notes_items_are_in_sow_notes(tmp_path: Path) -> None:
    rendered, _manifest, _scope, _delivery, review = render_fixture(tmp_path)
    notes = Path(rendered.notes_path).read_text(encoding="utf-8")
    for note in review["notes"]:
        assert notes.count(note["sowNotesText"]) == 1


def test_formula_like_user_text_remains_text(tmp_path: Path) -> None:
    manifest = fixture("greenfield", "input-manifest.json")
    scope = fixture("greenfield", "scope.json")
    scope["features"][0]["name"] = '=HYPERLINK("x")'
    delivery = fixture("greenfield", "delivery.json")
    delivery["scopeSha256"] = sha256_bytes(canonical_json_bytes(scope))
    review = fixture("greenfield", "final-review.json")
    review["scopeSha256"] = sha256_bytes(canonical_json_bytes(scope))
    review["deliverySha256"] = sha256_bytes(canonical_json_bytes(delivery))
    rendered = render_package(
        generation_id="000001",
        template_path=ASSETS / "sow-template.xlsx",
        output_root=tmp_path,
        input_manifest=manifest,
        scope=scope,
        delivery=delivery,
        review=review,
    )
    workbook = openpyxl.load_workbook(rendered.workbook_path, data_only=False)
    try:
        matches = [
            cell
            for row in workbook["01-需求故事"].iter_rows()
            for cell in row
            if cell.value == "'=HYPERLINK(\"x\")"
        ]
        assert len(matches) == 1
        assert matches[0].data_type == "s"
    finally:
        workbook.close()


def test_repeated_render_is_byte_identical(tmp_path: Path) -> None:
    first, manifest, scope, delivery, review = render_fixture(tmp_path / "first")
    second = render_package(
        generation_id="000001",
        template_path=ASSETS / "sow-template.xlsx",
        output_root=tmp_path / "second",
        input_manifest=copy.deepcopy(manifest),
        scope=copy.deepcopy(scope),
        delivery=copy.deepcopy(delivery),
        review=copy.deepcopy(review),
    )
    assert Path(first.workbook_path).read_bytes() == Path(second.workbook_path).read_bytes()
    assert Path(first.notes_path).read_bytes() == Path(second.notes_path).read_bytes()


def test_public_reference_workbook_matches_brownfield_fixture_semantics(
    tmp_path: Path,
) -> None:
    rendered, _manifest, scope, delivery, _review = render_fixture(tmp_path)
    expected = rendered.workbook_audit
    engine = require_office_engine()
    assert engine.name == expected.engine_name
    assert engine.version == expected.engine_version
    actual = audit_calculated_workbook(
        PUBLIC_WORKBOOK,
        ASSETS / "sow-template.xlsx",
        scope,
        delivery,
        engine,
    )

    assert actual == expected


def test_business_sheets_do_not_expose_stable_ids(tmp_path: Path) -> None:
    rendered, _manifest, scope, delivery, _review = render_fixture(tmp_path)
    forbidden = {
        str(value)
        for collection in (
            "epics",
            "features",
            "commitments",
            "effectiveStartItems",
            "designItems",
            "designDecisions",
            "integrations",
            "nfrs",
            "assumptions",
        )
        for item in scope[collection]
        for key, value in item.items()
        if key.endswith("Id") and isinstance(value, str)
    } | {
        str(value)
        for collection in ("stories", "acceptanceCriteria", "tasks", "dependencies")
        for item in delivery[collection]
        for key, value in item.items()
        if key.endswith("Id") and isinstance(value, str)
    }
    workbook = openpyxl.load_workbook(rendered.workbook_path, data_only=False)
    try:
        visible = workbook_string_contents(workbook)
        assert all(
            all(identifier not in content for content in visible)
            for identifier in forbidden
        )
    finally:
        workbook.close()


def test_business_sheets_do_not_expose_sources_story_paths_or_hashes(
    tmp_path: Path,
) -> None:
    manifest = fixture("brownfield", "input-manifest.json")
    scope = fixture("brownfield", "scope.json")
    delivery = fixture("brownfield", "delivery.json")
    delivery["stories"][0]["storyPath"] = "internal/epic/feature/story"
    delivery["scopeSha256"] = "f" * 64
    delivery["acceptanceCriteria"][0]["sourceRefs"][0]["locator"] = (
        "internal-source-locator"
    )
    review = fixture("brownfield", "final-review.json")
    review["scopeSha256"] = sha256_bytes(canonical_json_bytes(scope))
    review["deliverySha256"] = sha256_bytes(canonical_json_bytes(delivery))
    rendered = render_package(
        generation_id="000001",
        template_path=ASSETS / "sow-template.xlsx",
        output_root=tmp_path,
        input_manifest=manifest,
        scope=scope,
        delivery=delivery,
        review=review,
    )

    workbook = openpyxl.load_workbook(rendered.workbook_path, data_only=False)
    try:
        visible = workbook_string_contents(workbook)
        forbidden = {
            "internal/epic/feature/story",
            "internal-source-locator",
            "f" * 64,
            *(source_ref["sourceId"] for criterion in delivery["acceptanceCriteria"] for source_ref in criterion["sourceRefs"]),
            *(source_ref["anchorId"] for criterion in delivery["acceptanceCriteria"] for source_ref in criterion["sourceRefs"]),
            *(source_ref["sha256"] for criterion in delivery["acceptanceCriteria"] for source_ref in criterion["sourceRefs"]),
        }
        assert all(
            all(secret not in content for content in visible) for secret in forbidden
        )
        prefixed_leak = f"prefix:{next(iter(forbidden))}:suffix"
        workbook["90-估算标准"]["A1"] = prefixed_leak
        assert any(
            secret in content
            for secret in forbidden
            for content in workbook_string_contents(workbook)
        )
    finally:
        workbook.close()
