from __future__ import annotations

import copy
import json
import sys
from dataclasses import replace
from datetime import UTC, datetime
from pathlib import Path

import openpyxl
import pytest


SKILL_ROOT = Path(__file__).parents[1]
PLUGIN_ROOT = SKILL_ROOT.parents[1]
SCRIPTS = SKILL_ROOT / "scripts"
FIXTURES = SKILL_ROOT / "fixtures"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))
if str(PLUGIN_ROOT) not in sys.path:
    sys.path.insert(0, str(PLUGIN_ROOT))

from contracts import (  # noqa: E402
    canonical_json_bytes,
    load_schema_registry,
    sha256_bytes,
    validate_contract,
)
import orchestrator as orchestrator_module  # noqa: E402
from orchestrator import _publication_counts, run_mode  # noqa: E402
from models import Diagnostic  # noqa: E402
from runtime.project_io import ProjectFiles, ProjectIOError  # noqa: E402


NOW = lambda: datetime(2026, 9, 2, 8, 0, tzinfo=UTC)


def fixture(mode: str, name: str) -> dict[str, object]:
    return json.loads((FIXTURES / mode / name).read_text(encoding="utf-8"))


def write_json(path: Path, value: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(canonical_json_bytes(value))


def write_request(project: Path) -> str:
    request = fixture("greenfield", "request.json")
    inputs = project / "inputs"
    inputs.mkdir()
    for source in request["sources"]:
        source_path = inputs / f"{source['role'].lower()}.md"
        source_path.write_text(
            f"# {source['role']}\n\n退款申请、审核、结果通知与异常处理。\n",
            encoding="utf-8",
        )
        source["path"] = str(source_path)
    write_json(project / "request.json", request)
    return "request.json"


def decisions(candidate: dict[str, object], collections: dict[str, tuple[str, str]]):
    values = []
    for collection, (object_type, id_field) in collections.items():
        for item in candidate[collection]:
            values.append(
                {
                    "objectType": object_type,
                    "objectId": item[id_field],
                    "disposition": "NEW",
                    "meaningPreserved": False,
                    "rationale": "新对象分配稳定 ID。",
                }
            )
    return {"contract": "ai-sow-id-decisions-v1", "decisions": values}


SCOPE_COLLECTIONS = {
    "epics": ("EPIC", "epicId"),
    "features": ("FEATURE", "featureId"),
    "commitments": ("COMMITMENT", "commitmentId"),
    "effectiveStartItems": ("EFFECTIVE_START_ITEM", "effectiveStartItemId"),
    "designItems": ("DESIGN_ITEM", "designItemId"),
    "designDecisions": ("DESIGN_DECISION", "designDecisionId"),
    "integrations": ("INTEGRATION", "integrationId"),
    "nfrs": ("NFR", "nfrId"),
    "assumptions": ("ASSUMPTION", "assumptionId"),
}
DELIVERY_COLLECTIONS = {
    "stories": ("STORY", "storyId"),
    "acceptanceCriteria": ("ACCEPTANCE_CRITERION", "acceptanceCriterionId"),
    "tasks": ("TASK", "taskId"),
    "dependencies": ("DEPENDENCY", "dependencyId"),
}


def prepare_scope_files(project: Path, result: dict[str, object]) -> None:
    plan = result["runPlan"]
    bundle = fixture("greenfield", "scope.json")
    bundle["inputRevisionId"] = plan["targetRevisionId"]
    anchors = json.loads(
        (project / ".ai-sow/inputs/pending/anchors.json").read_text(encoding="utf-8")
    )
    first_anchor = {}
    for anchor in anchors:
        first_anchor.setdefault(anchor["sourceId"], anchor)
    for collection in SCOPE_COLLECTIONS:
        for item in bundle[collection]:
            item["sourceRefs"] = [
                {
                    "sourceId": first_anchor[ref["sourceId"]]["sourceId"],
                    "anchorId": first_anchor[ref["sourceId"]]["anchorId"],
                    "locator": first_anchor[ref["sourceId"]]["locator"],
                    "sha256": first_anchor[ref["sourceId"]]["sha256"],
                }
                for ref in item.get("sourceRefs", [])
            ]
    candidate = {
        "contract": "ai-sow-scope-slice-v1",
        "inputRevisionId": plan["targetRevisionId"],
        "impactPlanSha256": sha256_bytes(canonical_json_bytes(plan["impact"])),
        "replacesFeatureIds": (
            []
            if plan["impact"]["baselineGenerationId"] is None
            else list(plan["impact"]["affectedFeatureIds"])
        ),
        "newAnchorMappings": [],
        **{name: copy.deepcopy(bundle[name]) for name in SCOPE_COLLECTIONS},
        "responsibilityBoundaries": copy.deepcopy(bundle["responsibilityBoundaries"]),
    }
    write_json(project / "scope.json", candidate)
    write_json(project / "scope-ids.json", decisions(candidate, SCOPE_COLLECTIONS))


def prepare_delivery_files(project: Path, prepared: dict[str, object]) -> None:
    plan = json.loads(
        (project / ".ai-sow/work/run-plan.json").read_text(encoding="utf-8")
    )
    scope = json.loads(
        (project / ".ai-sow/work/scope.candidate.json").read_text(encoding="utf-8")
    )
    bundle = fixture("greenfield", "delivery.json")
    bundle["inputRevisionId"] = plan["targetRevisionId"]
    candidate = {
        "contract": "ai-sow-delivery-slice-v3",
        "inputRevisionId": plan["targetRevisionId"],
        "scopeSha256": sha256_bytes(canonical_json_bytes(scope)),
        "impactPlanSha256": sha256_bytes(canonical_json_bytes(plan["impact"])),
        "replacesFeatureIds": (
            []
            if plan["impact"]["baselineGenerationId"] is None
            else list(plan["impact"]["affectedFeatureIds"])
        ),
        **{name: copy.deepcopy(bundle[name]) for name in DELIVERY_COLLECTIONS},
    }
    write_json(project / "delivery.json", candidate)
    write_json(
        project / "delivery-ids.json", decisions(candidate, DELIVERY_COLLECTIONS)
    )


def publish_verified_current(project: Path, prepared: dict[str, object]) -> None:
    prepare_scope_files(project, prepared)
    scope_result = run_mode(
        project,
        "accept-scope",
        candidate="scope.json",
        ids="scope-ids.json",
        now=NOW,
    )
    assert scope_result["outcome"] == "READY_FOR_DELIVERY", scope_result
    prepare_delivery_files(project, prepared)
    delivery_result = run_mode(
        project,
        "accept-delivery",
        candidate="delivery.json",
        ids="delivery-ids.json",
        now=NOW,
    )
    assert delivery_result["outcome"] == "REVIEW_REQUIRED", delivery_result
    packet = run_mode(project, "prepare-review", now=NOW)
    plan = prepared["runPlan"]
    review = {
        "contract": "ai-sow-final-review-v1",
        "runId": plan["runId"],
        "inputRevisionId": plan["targetRevisionId"],
        "scopeSha256": sha256_bytes(
            (project / ".ai-sow/work/scope.candidate.json").read_bytes()
        ),
        "deliverySha256": sha256_bytes(
            (project / ".ai-sow/work/delivery.candidate.json").read_bytes()
        ),
        "packetSha256": sha256_bytes(
            (project / ".ai-sow/work/review-packet.json").read_bytes()
        ),
        "decision": "PASS",
        "notes": [],
        "questions": [],
    }
    write_json(project / "review.json", review)
    accepted = run_mode(
        project, "accept-review", review="review.json", now=NOW
    )
    assert accepted["outcome"] == "READY_TO_RENDER", accepted
    published = run_mode(project, "publish", now=NOW)
    assert published["outcome"] == "PUBLISHED", published


def test_prepare_initial_project_requests_full_scope(tmp_path: Path) -> None:
    request_path = write_request(tmp_path)
    result = run_mode(tmp_path, "prepare", request=request_path, now=NOW)
    assert result["outcome"] == "READY_FOR_SCOPE"
    assert result["runPlan"]["action"] == "FULL_COMPILE"
    assert (tmp_path / ".ai-sow/work/run-plan.json").is_file()


def test_prepare_freezes_current_template_for_entire_run(tmp_path: Path) -> None:
    request_path = write_request(tmp_path)
    prepared = run_mode(tmp_path, "prepare", request=request_path, now=NOW)

    plan = prepared["runPlan"]
    snapshot = tmp_path / plan["templateSnapshotPath"]
    assert snapshot.read_bytes() == (
        tmp_path / ".ai-sow/templates/sow-template.xlsx"
    ).read_bytes()
    assert plan["templateSha256"] == sha256_bytes(snapshot.read_bytes())


def test_run_template_resolver_rejects_an_arbitrary_matching_path(
    tmp_path: Path,
) -> None:
    request_path = write_request(tmp_path)
    prepared = run_mode(tmp_path, "prepare", request=request_path, now=NOW)
    plan = orchestrator_module._plan_from_value(prepared["runPlan"])
    alternate_path = ".ai-sow/work/alternate-template.xlsx"
    files = ProjectFiles.open(tmp_path)
    files.write_atomic(alternate_path, files.read_bytes(plan.template_snapshot_path))

    with pytest.raises(ProjectIOError) as caught:
        orchestrator_module._resolve_run_template(
            files, replace(plan, template_snapshot_path=alternate_path)
        )

    assert caught.value.code == "RUN_TEMPLATE_CHANGED"


def test_live_template_change_does_not_change_active_run(tmp_path: Path) -> None:
    request_path = write_request(tmp_path)
    prepared = run_mode(tmp_path, "prepare", request=request_path, now=NOW)
    snapshot = tmp_path / prepared["runPlan"]["templateSnapshotPath"]
    snapshot_before = snapshot.read_bytes()
    prepare_scope_files(tmp_path, prepared)
    scope_result = run_mode(
        tmp_path,
        "accept-scope",
        candidate="scope.json",
        ids="scope-ids.json",
        now=NOW,
    )
    assert scope_result["outcome"] == "READY_FOR_DELIVERY", scope_result
    prepare_delivery_files(tmp_path, prepared)

    template = tmp_path / ".ai-sow/templates/sow-template.xlsx"
    workbook = openpyxl.load_workbook(template)
    try:
        table = workbook["90-估算标准"].tables["BaseUnitCatalogTable"]
        min_col = openpyxl.utils.range_boundaries(table.ref)[0]
        min_row = openpyxl.utils.range_boundaries(table.ref)[1]
        workbook["90-估算标准"].cell(min_row, min_col).value = "无效目录列"
        workbook.save(template)
    finally:
        workbook.close()

    result = run_mode(
        tmp_path,
        "accept-delivery",
        candidate="delivery.json",
        ids="delivery-ids.json",
        now=NOW,
    )

    assert result["outcome"] == "REVIEW_REQUIRED", result
    assert snapshot.read_bytes() == snapshot_before


def test_changed_run_snapshot_blocks_before_compile(tmp_path: Path) -> None:
    request_path = write_request(tmp_path)
    prepared = run_mode(tmp_path, "prepare", request=request_path, now=NOW)
    snapshot = tmp_path / prepared["runPlan"]["templateSnapshotPath"]
    snapshot.write_bytes(snapshot.read_bytes() + b"changed")
    prepare_scope_files(tmp_path, prepared)
    assert run_mode(
        tmp_path,
        "accept-scope",
        candidate="scope.json",
        ids="scope-ids.json",
        now=NOW,
    )["outcome"] == "READY_FOR_DELIVERY"
    prepare_delivery_files(tmp_path, prepared)

    result = run_mode(
        tmp_path,
        "accept-delivery",
        candidate="delivery.json",
        ids="delivery-ids.json",
        now=NOW,
    )

    assert {item["code"] for item in result["diagnostics"]} == {
        "RUN_TEMPLATE_CHANGED"
    }


def test_publication_counts_separate_recompute_reuse_delete_and_final() -> None:
    previous_scope = {
        "features": [
            {"featureId": "feature-refund", "name": "退款"},
            {"featureId": "feature-ledger", "name": "台账"},
        ]
    }
    current_scope = {
        "features": [
            {"featureId": "feature-refund", "name": "退款升级"},
            {"featureId": "feature-ledger", "name": "台账"},
        ]
    }
    previous_delivery = {
        "stories": [
            {"storyId": "story-refund", "featureId": "feature-refund"},
            {"storyId": "story-ledger", "featureId": "feature-ledger"},
        ],
        "acceptanceCriteria": [
            {"acceptanceCriterionId": "ac-refund", "storyId": "story-refund"},
            {"acceptanceCriterionId": "ac-ledger", "storyId": "story-ledger"},
        ],
        "tasks": [
            {"taskId": "task-refund-update", "storyId": "story-refund", "v": 1},
            {"taskId": "task-refund-delete", "storyId": "story-refund", "v": 1},
            {"taskId": "task-ledger", "storyId": "story-ledger", "v": 1},
        ],
    }
    current_delivery = {
        "stories": [
            {"storyId": "story-refund", "featureId": "feature-refund", "v": 2},
            {"storyId": "story-ledger", "featureId": "feature-ledger"},
        ],
        "acceptanceCriteria": [
            {"acceptanceCriterionId": "ac-refund", "storyId": "story-refund", "v": 2},
            {"acceptanceCriterionId": "ac-ledger", "storyId": "story-ledger"},
        ],
        "tasks": [
            {"taskId": "task-refund-update", "storyId": "story-refund", "v": 2},
            {"taskId": "task-refund-new", "storyId": "story-refund", "v": 1},
            {"taskId": "task-ledger", "storyId": "story-ledger", "v": 1},
        ],
    }

    assert _publication_counts(
        previous_scope,
        previous_delivery,
        current_scope,
        current_delivery,
        ("feature-refund",),
    ) == {
        "features": {
            "affected": 1,
            "recomputed": 1,
            "reused": 1,
            "deleted": 0,
            "final": 2,
        },
        "stories": {
            "affected": 1,
            "recomputed": 1,
            "reused": 1,
            "deleted": 0,
            "final": 2,
        },
        "acceptanceCriteria": {
            "affected": 1,
            "recomputed": 1,
            "reused": 1,
            "deleted": 0,
            "final": 2,
        },
        "tasks": {
            "affected": 2,
            "recomputed": 2,
            "reused": 1,
            "deleted": 1,
            "final": 3,
        },
    }


def test_modes_are_ordered_and_fail_closed(tmp_path: Path) -> None:
    request_path = write_request(tmp_path)
    run_mode(tmp_path, "prepare", request=request_path, now=NOW)
    write_json(tmp_path / "delivery.json", {})
    write_json(tmp_path / "ids.json", {})
    result = run_mode(
        tmp_path,
        "accept-delivery",
        candidate="delivery.json",
        ids="ids.json",
        now=NOW,
    )
    assert result["outcome"] == "BLOCKED"
    assert result["diagnostics"][0]["code"] == "SCOPE_NOT_ACCEPTED"


def test_accept_scope_writes_complete_bundle_and_advances(tmp_path: Path) -> None:
    request_path = write_request(tmp_path)
    prepared = run_mode(tmp_path, "prepare", request=request_path, now=NOW)
    prepare_scope_files(tmp_path, prepared)
    result = run_mode(
        tmp_path,
        "accept-scope",
        candidate="scope.json",
        ids="scope-ids.json",
        now=NOW,
    )
    assert result["outcome"] == "READY_FOR_DELIVERY", result
    assert result["scopeSha256"] == sha256_bytes(
        (tmp_path / ".ai-sow/work/scope.candidate.json").read_bytes()
    )


def test_accept_delivery_binds_exact_scope_and_advances_to_review(tmp_path: Path) -> None:
    request_path = write_request(tmp_path)
    prepared = run_mode(tmp_path, "prepare", request=request_path, now=NOW)
    prepare_scope_files(tmp_path, prepared)
    scope_result = run_mode(
        tmp_path,
        "accept-scope",
        candidate="scope.json",
        ids="scope-ids.json",
        now=NOW,
    )
    assert scope_result["outcome"] == "READY_FOR_DELIVERY", scope_result
    prepare_delivery_files(tmp_path, prepared)
    result = run_mode(
        tmp_path,
        "accept-delivery",
        candidate="delivery.json",
        ids="delivery-ids.json",
        now=NOW,
    )
    assert result["outcome"] == "REVIEW_REQUIRED", result
    assert result["deliverySha256"] == sha256_bytes(
        (tmp_path / ".ai-sow/work/delivery.candidate.json").read_bytes()
    )


def test_accept_delivery_returns_one_self_contained_question_per_affected_task(
    tmp_path: Path,
) -> None:
    request_path = write_request(tmp_path)
    prepared = run_mode(tmp_path, "prepare", request=request_path, now=NOW)
    prepare_scope_files(tmp_path, prepared)
    assert run_mode(
        tmp_path,
        "accept-scope",
        candidate="scope.json",
        ids="scope-ids.json",
        now=NOW,
    )["outcome"] == "READY_FOR_DELIVERY"
    scope_path = tmp_path / ".ai-sow/work/scope.candidate.json"
    scope = json.loads(scope_path.read_text(encoding="utf-8"))
    scope["effectiveStartItems"].append(
        {
            "effectiveStartItemId": "effective-start-return-api",
            "matchLevel": "CAPABILITY",
            "name": "既有退货申请接口",
            "summary": "现有材料只能确认退货模块能力，不能确认该接口实例可直接使用。",
            "commitmentIds": [],
            "featureIds": ["feature-refund-processing"],
            "sourceRefs": [scope["designItems"][0]["sourceRefs"][0]],
        }
    )
    write_json(scope_path, scope)
    prepare_delivery_files(tmp_path, prepared)
    delivery_path = tmp_path / "delivery.json"
    delivery = json.loads(delivery_path.read_text(encoding="utf-8"))
    task = delivery["tasks"][0]
    task["workMode"] = "调整"
    task["workModeEvidence"] = {
        "effectiveStartItemId": "effective-start-return-api"
    }
    delivery["scopeSha256"] = sha256_bytes(canonical_json_bytes(scope))
    write_json(delivery_path, delivery)

    result = run_mode(
        tmp_path,
        "accept-delivery",
        candidate="delivery.json",
        ids="delivery-ids.json",
        now=NOW,
    )

    assert result["outcome"] == "BLOCKED"
    assert len(result["questions"]) == 1
    question = result["questions"][0]
    assert question["subjectIds"] == ["task-refund-service"]
    assert "既有退货申请接口" in question["question"]
    assert "是否已经存在" in question["question"]
    assert "直接修改或使用" in question["question"]
    assert "从头实现" in question["question"]
    assert validate_contract(
        question, "question.schema.json", load_schema_registry(SKILL_ROOT)
    ) == ()
    user_text = "\n".join(
        str(question[field])
        for field in ("question", "reason", "decisionImpact", "unansweredEffect")
    )
    for internal in (
        "CAPABILITY",
        "TASK_INSTANCE",
        "baseUnit",
        "Effective Start",
        "effective-start-return-api",
    ):
        assert internal not in user_text


def test_work_mode_question_deduplicates_related_diagnostics_for_one_task() -> None:
    scope = {
        "effectiveStartItems": [
            {
                "effectiveStartItemId": "effective-start-return-api",
                "name": "既有退货申请接口",
            }
        ]
    }
    delivery = {
        "tasks": [
            {
                "taskId": "task-return-api",
                "name": "调整退货申请接口",
                "workModeEvidence": {
                    "effectiveStartItemId": "effective-start-return-api"
                },
            }
        ]
    }
    diagnostics = (
        Diagnostic(
            code="TASK_WORK_MODE_REQUIRES_INSTANCE_START",
            message="测试诊断。",
            path="/tasks/task-return-api/workModeEvidence/effectiveStartItemId",
            details={},
        ),
        Diagnostic(
            code="TASK_WORK_MODE_BASE_UNIT_MISMATCH",
            message="测试诊断。",
            path="/tasks/task-return-api/workModeEvidence/effectiveStartItemId",
            details={},
        ),
    )

    questions = orchestrator_module._work_mode_evidence_questions(
        scope, delivery, diagnostics
    )

    assert [question["questionId"] for question in questions] == [
        "question-work-mode-task-return-api"
    ]


def test_prepare_identical_request_reuses_verified_current_outputs(
    tmp_path: Path,
) -> None:
    request_path = write_request(tmp_path)
    prepared = run_mode(tmp_path, "prepare", request=request_path, now=NOW)
    publish_verified_current(tmp_path, prepared)
    result = run_mode(tmp_path, "prepare", request=request_path, now=NOW)
    assert result["outcome"] == "REUSED", result
    assert result["workbookPath"].endswith("/output/sow.xlsx")
    assert not (tmp_path / ".ai-sow/inputs/pending").exists()


def test_prepare_rerenders_verified_output_from_previous_renderer_contract(
    tmp_path: Path,
) -> None:
    request_path = write_request(tmp_path)
    prepared = run_mode(tmp_path, "prepare", request=request_path, now=NOW)
    publish_verified_current(tmp_path, prepared)
    manifest_path = tmp_path / ".ai-sow/generations/000001/manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["rendererContract"] = "generation-renderer-v2"
    write_json(manifest_path, manifest)
    current_path = tmp_path / ".ai-sow/current.json"
    current = json.loads(current_path.read_text(encoding="utf-8"))
    current["generationManifestSha256"] = sha256_bytes(manifest_path.read_bytes())
    write_json(current_path, current)

    result = run_mode(tmp_path, "prepare", request=request_path, now=NOW)

    assert result["outcome"] == "READY_TO_RENDER", result
    assert result["runPlan"]["action"] == "RENDER_ONLY"
    assert result["runPlan"]["rendererContract"] == "generation-renderer-v7"


def test_prepare_full_compiles_from_legacy_v1_generation_evidence(
    tmp_path: Path,
) -> None:
    request_path = write_request(tmp_path)
    prepared = run_mode(tmp_path, "prepare", request=request_path, now=NOW)
    publish_verified_current(tmp_path, prepared)
    manifest_path = tmp_path / ".ai-sow/generations/000001/manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    del manifest["workbookVerification"]
    manifest["scopeCompilerContract"] = "scope-compiler-v1"
    manifest["deliveryCompilerContract"] = "delivery-compiler-v1"
    manifest["rendererContract"] = "generation-renderer-v1"
    manifest["changeCounts"] = {
        "features": {"added": 1, "updated": 0, "removed": 0},
        "recomputedStories": 1,
        "recomputedTasks": 1,
    }
    write_json(manifest_path, manifest)
    current_path = tmp_path / ".ai-sow/current.json"
    current = json.loads(current_path.read_text(encoding="utf-8"))
    current["generationManifestSha256"] = sha256_bytes(manifest_path.read_bytes())
    write_json(current_path, current)

    result = run_mode(tmp_path, "prepare", request=request_path, now=NOW)

    assert result["outcome"] == "READY_FOR_SCOPE", result
    assert result["runPlan"]["action"] == "FULL_COMPILE"
    assert "COMPILER_CONTRACT_CHANGED" in result["runPlan"]["impact"][
        "reasonCodes"
    ]


def test_prepare_adopts_new_bundled_template_when_project_copy_is_unchanged(
    tmp_path: Path, monkeypatch
) -> None:
    request_path = write_request(tmp_path)
    prepared = run_mode(tmp_path, "prepare", request=request_path, now=NOW)
    publish_verified_current(tmp_path, prepared)
    upgraded_template = tmp_path / "upgraded-template.xlsx"
    workbook = openpyxl.load_workbook(orchestrator_module.TEMPLATE_ASSET)
    try:
        workbook.properties.title = "bundled-template-upgrade"
        workbook.save(upgraded_template)
    finally:
        workbook.close()
    monkeypatch.setattr(orchestrator_module, "TEMPLATE_ASSET", upgraded_template)

    result = run_mode(tmp_path, "prepare", request=request_path, now=NOW)

    assert result["outcome"] == "READY_FOR_SCOPE", result
    assert result["runPlan"]["action"] == "FULL_COMPILE"
    assert (tmp_path / ".ai-sow/templates/sow-template.xlsx").read_bytes() == (
        upgraded_template.read_bytes()
    )


def test_previous_v6_bundled_template_remains_auto_upgradeable() -> None:
    assert (
        "587797c1cd369166d2273da45f07b045a7badd4caf9d5ae2b49ef6cee169b77d"
        in orchestrator_module.BUNDLED_TEMPLATE_SHA256_HISTORY
    )


def test_previous_v7_bundled_template_remains_auto_upgradeable() -> None:
    assert (
        "edec72f7a1d91a8fa5fbdd7ebc67a0b5ad313e1d2434a444741c88c86acf7030"
        in orchestrator_module.BUNDLED_TEMPLATE_SHA256_HISTORY
    )


def test_prepare_preserves_project_template_changed_after_last_publication(
    tmp_path: Path, monkeypatch
) -> None:
    request_path = write_request(tmp_path)
    prepared = run_mode(tmp_path, "prepare", request=request_path, now=NOW)
    publish_verified_current(tmp_path, prepared)
    project_template = tmp_path / ".ai-sow/templates/sow-template.xlsx"
    workbook = openpyxl.load_workbook(project_template)
    try:
        workbook.properties.title = "project-specific-template"
        workbook.save(project_template)
    finally:
        workbook.close()
    project_bytes = project_template.read_bytes()
    upgraded_template = tmp_path / "upgraded-template.xlsx"
    workbook = openpyxl.load_workbook(orchestrator_module.TEMPLATE_ASSET)
    try:
        workbook.properties.title = "bundled-template-upgrade"
        workbook.save(upgraded_template)
    finally:
        workbook.close()
    monkeypatch.setattr(orchestrator_module, "TEMPLATE_ASSET", upgraded_template)

    result = run_mode(tmp_path, "prepare", request=request_path, now=NOW)

    assert result["outcome"] == "READY_FOR_SCOPE", result
    assert result["runPlan"]["action"] == "FULL_COMPILE"
    assert project_template.read_bytes() == project_bytes


def test_prepare_preserves_project_template_customized_before_first_publication(
    tmp_path: Path, monkeypatch
) -> None:
    request_path = write_request(tmp_path)
    prepared = run_mode(tmp_path, "prepare", request=request_path, now=NOW)
    project_template = tmp_path / ".ai-sow/templates/sow-template.xlsx"
    workbook = openpyxl.load_workbook(project_template)
    try:
        workbook.properties.title = "project-specific-template"
        workbook.save(project_template)
    finally:
        workbook.close()
    publish_verified_current(tmp_path, prepared)
    project_bytes = project_template.read_bytes()
    upgraded_template = tmp_path / "upgraded-template.xlsx"
    workbook = openpyxl.load_workbook(orchestrator_module.TEMPLATE_ASSET)
    try:
        workbook.properties.title = "bundled-template-upgrade"
        workbook.save(upgraded_template)
    finally:
        workbook.close()
    monkeypatch.setattr(orchestrator_module, "TEMPLATE_ASSET", upgraded_template)

    result = run_mode(tmp_path, "prepare", request=request_path, now=NOW)

    assert result["outcome"] == "READY_FOR_SCOPE", result
    assert result["runPlan"]["action"] == "FULL_COMPILE"
    assert project_template.read_bytes() == project_bytes


def test_publish_without_office_engine_preserves_last_known_good(
    tmp_path: Path, monkeypatch
) -> None:
    request_path = write_request(tmp_path)
    prepared = run_mode(tmp_path, "prepare", request=request_path, now=NOW)
    publish_verified_current(tmp_path, prepared)
    previous_current = (tmp_path / ".ai-sow/current.json").read_bytes()

    template_path = tmp_path / ".ai-sow/templates/sow-template.xlsx"
    workbook = openpyxl.load_workbook(template_path)
    try:
        workbook.properties.title = "强制模板哈希变化"
        workbook.save(template_path)
    finally:
        workbook.close()
    rerender = run_mode(tmp_path, "prepare", request=request_path, now=NOW)
    assert rerender["outcome"] == "READY_FOR_SCOPE", rerender
    assert rerender["runPlan"]["action"] == "FULL_COMPILE"
    assert (tmp_path / ".ai-sow/current.json").read_bytes() == previous_current
    assert not (tmp_path / ".ai-sow/generations/000002").exists()


def test_review_modes_bind_packet_and_enable_render_only_after_pass(
    tmp_path: Path,
) -> None:
    request_path = write_request(tmp_path)
    prepared = run_mode(tmp_path, "prepare", request=request_path, now=NOW)
    prepare_scope_files(tmp_path, prepared)
    assert run_mode(
        tmp_path,
        "accept-scope",
        candidate="scope.json",
        ids="scope-ids.json",
        now=NOW,
    )["outcome"] == "READY_FOR_DELIVERY"
    prepare_delivery_files(tmp_path, prepared)
    assert run_mode(
        tmp_path,
        "accept-delivery",
        candidate="delivery.json",
        ids="delivery-ids.json",
        now=NOW,
    )["outcome"] == "REVIEW_REQUIRED"

    packet_result = run_mode(tmp_path, "prepare-review", now=NOW)
    assert packet_result["outcome"] == "REVIEW_REQUIRED", packet_result
    assert packet_result["reviewMaterialPath"] == ".ai-sow/work/review-material.md"
    plan = json.loads(
        (tmp_path / ".ai-sow/work/run-plan.json").read_text(encoding="utf-8")
    )
    review = {
        "contract": "ai-sow-final-review-v1",
        "runId": plan["runId"],
        "inputRevisionId": plan["targetRevisionId"],
        "scopeSha256": sha256_bytes(
            (tmp_path / ".ai-sow/work/scope.candidate.json").read_bytes()
        ),
        "deliverySha256": sha256_bytes(
            (tmp_path / ".ai-sow/work/delivery.candidate.json").read_bytes()
        ),
        "packetSha256": sha256_bytes(
            (tmp_path / ".ai-sow/work/review-packet.json").read_bytes()
        ),
        "decision": "PASS",
        "notes": [],
        "questions": [],
    }
    write_json(tmp_path / "review.json", review)
    result = run_mode(
        tmp_path, "accept-review", review="review.json", now=NOW
    )
    assert result["outcome"] == "READY_TO_RENDER", result
    assert run_mode(tmp_path, "status", now=NOW)["outcome"] == "READY_TO_RENDER"


def test_stale_scope_candidate_is_rejected(tmp_path: Path) -> None:
    request_path = write_request(tmp_path)
    prepared = run_mode(tmp_path, "prepare", request=request_path, now=NOW)
    prepare_scope_files(tmp_path, prepared)
    candidate = json.loads((tmp_path / "scope.json").read_text(encoding="utf-8"))
    candidate["impactPlanSha256"] = "0" * 64
    write_json(tmp_path / "scope.json", candidate)
    result = run_mode(
        tmp_path,
        "accept-scope",
        candidate="scope.json",
        ids="scope-ids.json",
        now=NOW,
    )
    assert result["outcome"] == "BLOCKED"
    assert "SCOPE_IMPACT_HASH_MISMATCH" in {
        item["code"] for item in result["diagnostics"]
    }
    assert not (tmp_path / ".ai-sow/work/scope.candidate.json").exists()


def test_status_is_derived_and_read_only(tmp_path: Path) -> None:
    request_path = write_request(tmp_path)
    run_mode(tmp_path, "prepare", request=request_path, now=NOW)
    before = {
        path.relative_to(tmp_path).as_posix(): path.read_bytes()
        for path in tmp_path.rglob("*")
        if path.is_file()
    }
    first = run_mode(tmp_path, "status", now=NOW)
    second = run_mode(tmp_path, "status", now=NOW)
    after = {
        path.relative_to(tmp_path).as_posix(): path.read_bytes()
        for path in tmp_path.rglob("*")
        if path.is_file()
    }
    assert first == second
    assert first["outcome"] == "READY_FOR_SCOPE"
    assert first["nextMode"] == "accept-scope"
    assert before == after


def test_cli_rejects_relative_project_root(capsys) -> None:
    from orchestrator import main

    exit_code = main(["--project-root", "relative", "--mode", "status"])
    payload = json.loads(capsys.readouterr().out)
    assert exit_code == 2
    assert payload["outcome"] == "BLOCKED"
    assert payload["diagnostics"][0]["code"] == "PROJECT_ROOT_NOT_ABSOLUTE"


def test_cli_argument_error_emits_one_json_object_without_stderr(capsys) -> None:
    from orchestrator import main

    exit_code = main(["--mode", "status"])
    captured = capsys.readouterr()
    payload = json.loads(captured.out)
    assert exit_code == 2
    assert payload["diagnostics"][0]["code"] == "CLI_ARGUMENTS_INVALID"
    assert captured.err == ""
