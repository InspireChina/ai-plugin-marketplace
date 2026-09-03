from __future__ import annotations

import copy
import hashlib
import json
import os
import re
import shutil
import sys
import zipfile
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


SKILL_ROOT = Path(__file__).parents[1]
SCRIPTS = SKILL_ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

from office_engine import (  # noqa: E402
    OfficeEngineError,
    deterministic_external_attr,
    discover_office_engine,
    require_office_engine,
)
import office_engine  # noqa: E402
import workbook as workbook_module  # noqa: E402
from workbook import write_workbook  # noqa: E402


def installed_soffice() -> Path:
    explicit = os.environ.get("AI_SOW_OFFICE_BIN")
    executable = explicit or shutil.which("soffice") or shutil.which("libreoffice")
    if executable is None:
        pytest.skip("当前测试环境未安装 LibreOffice")
    return Path(executable).resolve()


def fixture(mode: str, name: str) -> dict[str, object]:
    return json.loads(
        (SKILL_ROOT / "fixtures" / mode / name).read_text(encoding="utf-8")
    )


def validation_signatures(workbook) -> dict[str, list[tuple[object, ...]]]:
    return {
        worksheet.title: [
            (
                str(item.sqref),
                item.type,
                item.formula1,
                item.formula2,
                item.allow_blank,
                item.error,
                item.prompt,
            )
            for item in worksheet.data_validations.dataValidation
        ]
        for worksheet in workbook.worksheets
    }


def replace_zip_member(path: Path, member: str, old: bytes, new: bytes) -> None:
    temporary = path.with_suffix(".tampered.xlsx")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        temporary, "w", compression=zipfile.ZIP_DEFLATED
    ) as target:
        replaced = False
        for entry in source.infolist():
            payload = source.read(entry.filename)
            if entry.filename == member:
                assert payload.count(old) == 1
                payload = payload.replace(old, new)
                replaced = True
            target.writestr(entry, payload)
    assert replaced
    os.replace(temporary, path)


def replace_cached_value(path: Path, member: str, coordinate: str, new: bytes) -> None:
    temporary = path.with_suffix(".tampered.xlsx")
    pattern = re.compile(
        rb'(<c\b[^>]*\br="'
        + re.escape(coordinate.encode())
        + rb'"[^>]*>.*?<v>)([^<]*)(</v>.*?</c>)',
        re.DOTALL,
    )
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        temporary, "w", compression=zipfile.ZIP_DEFLATED
    ) as target:
        replaced = False
        for entry in source.infolist():
            payload = source.read(entry.filename)
            if entry.filename == member:
                payload, count = pattern.subn(rb"\g<1>" + new + rb"\g<3>", payload)
                assert count == 1
                replaced = True
            target.writestr(entry, payload)
    assert replaced
    os.replace(temporary, path)


def test_office_engine_prefers_explicit_supported_binary(monkeypatch) -> None:
    executable = installed_soffice()
    monkeypatch.setenv("AI_SOW_OFFICE_BIN", str(executable))

    engine = discover_office_engine()

    assert engine is not None
    assert engine.executable == str(executable)
    assert engine.name == "LibreOffice"
    assert "LibreOffice" in engine.version


def test_office_engine_falls_back_to_path_after_invalid_explicit_binary(
    monkeypatch,
) -> None:
    executable = installed_soffice()
    monkeypatch.setenv("AI_SOW_OFFICE_BIN", "/missing/ai-sow-office")
    monkeypatch.setattr(
        office_engine.shutil,
        "which",
        lambda name: str(executable) if name == "soffice" else None,
    )

    engine = discover_office_engine()

    assert engine is not None
    assert engine.executable == str(executable)


def test_missing_office_engine_is_not_a_verified_result(monkeypatch) -> None:
    monkeypatch.delenv("AI_SOW_OFFICE_BIN", raising=False)
    monkeypatch.setenv("PATH", "")

    with pytest.raises(OfficeEngineError) as caught:
        require_office_engine()

    assert caught.value.code == "OFFICE_ENGINE_UNAVAILABLE"


def test_real_office_roundtrip_is_isolated_and_byte_deterministic(
    tmp_path: Path, monkeypatch
) -> None:
    executable = installed_soffice()
    monkeypatch.setenv("AI_SOW_OFFICE_BIN", str(executable))
    candidate = tmp_path / "candidate.xlsx"
    write_workbook(
        SKILL_ROOT / "assets/sow-template.xlsx",
        fixture("brownfield", "scope.json"),
        fixture("brownfield", "delivery.json"),
        candidate,
    )
    candidate_sha = hashlib.sha256(candidate.read_bytes()).hexdigest()

    first = office_engine.recalculate_workbook(
        candidate, tmp_path / "first.xlsx", require_office_engine()
    )
    second = office_engine.recalculate_workbook(
        candidate, tmp_path / "second.xlsx", require_office_engine()
    )

    assert hashlib.sha256(candidate.read_bytes()).hexdigest() == candidate_sha
    assert hashlib.sha256(Path(first.output_path).read_bytes()).hexdigest() == (
        hashlib.sha256(Path(second.output_path).read_bytes()).hexdigest()
    )
    with zipfile.ZipFile(first.output_path, "r") as archive:
        app_metadata = archive.read("docProps/app.xml")
    assert b"AI SOW Office Engine" in app_metadata
    assert b"MacOSX_AARCH64" not in app_metadata
    assert b"LibreOfficeDev" not in app_metadata
    workbook = load_workbook(first.output_path, data_only=True, read_only=False)
    candidate_workbook = load_workbook(candidate, data_only=False, read_only=False)
    try:
        assert workbook["03-工作量汇总"]["B8"].value == 9.5
        assert validation_signatures(workbook) == validation_signatures(
            candidate_workbook
        )
        for worksheet in workbook.worksheets:
            assert worksheet.page_setup.fitToHeight == 0
        for worksheet in candidate_workbook.worksheets:
            assert worksheet.page_setup.fitToHeight == 0
    finally:
        workbook.close()
        candidate_workbook.close()


def test_calculated_workbook_audit_reads_full_authority_and_results(
    tmp_path: Path, monkeypatch
) -> None:
    executable = installed_soffice()
    monkeypatch.setenv("AI_SOW_OFFICE_BIN", str(executable))
    scope = fixture("brownfield", "scope.json")
    delivery = fixture("brownfield", "delivery.json")
    candidate = tmp_path / "candidate.xlsx"
    write_workbook(
        SKILL_ROOT / "assets/sow-template.xlsx",
        scope,
        delivery,
        candidate,
    )
    engine = require_office_engine()
    calculated = office_engine.recalculate_workbook(
        candidate, tmp_path / "calculated.xlsx", engine
    )

    audit = workbook_module.audit_calculated_workbook(
        Path(calculated.output_path),
        SKILL_ROOT / "assets/sow-template.xlsx",
        scope,
        delivery,
        engine,
    )

    assert audit.trust_state == "VERIFIED"
    assert (audit.story_count, audit.task_count) == (2, 4)
    assert audit.direct_days == 8
    assert audit.sit_days == 1
    assert audit.uat_days == 0.5
    assert audit.total_days == 9.5
    assert ("K_UAT", "待样本校准") in audit.parameter_statuses
    assert ("SIT_INT_SUPPORT", "待样本校准") in audit.parameter_statuses
    assert ("SIT_EXT_SUPPORT", "待样本校准") in audit.parameter_statuses
    assert audit.formula_errors == ()
    assert audit.engine_name == "LibreOffice"
    assert "LibreOffice" in str(audit.engine_version)


def test_project_effort_is_invariant_to_story_packaging(
    tmp_path: Path, monkeypatch
) -> None:
    executable = installed_soffice()
    monkeypatch.setenv("AI_SOW_OFFICE_BIN", str(executable))
    scope = fixture("greenfield", "scope.json")
    combined = fixture("greenfield", "delivery.json")
    task_prototype = combined["tasks"][0]
    combined["tasks"] = []
    for sequence in range(1, 5):
        task = copy.deepcopy(task_prototype)
        task["taskId"] = f"task-refund-service-{sequence}"
        task["name"] = f"实现退款服务实例 {sequence}"
        task["complexity"] = "L"
        task["complexityRationale"] = "该实例含复杂状态、补偿与严格证据要求。"
        combined["tasks"].append(task)

    split = copy.deepcopy(combined)
    split["stories"] = []
    split["acceptanceCriteria"] = []
    for sequence, task in enumerate(split["tasks"], 1):
        story_id = f"story-refund-processing-{sequence}"
        accepted_criterion_id = f"ac-refund-accepted-{sequence}"
        rejected_criterion_id = f"ac-refund-rejected-{sequence}"
        split["stories"].append(
            {
                "storyId": story_id,
                "featureId": "feature-refund-processing",
                "name": f"[退款申请] 客户提交退款 {sequence}",
                "uatRelevant": True,
            }
        )
        split["acceptanceCriteria"].extend(
            [
                {
                    "acceptanceCriterionId": accepted_criterion_id,
                    "storyId": story_id,
                    "name": f"第 {sequence} 个有效退款申请被受理并返回结果",
                },
                {
                    "acceptanceCriterionId": rejected_criterion_id,
                    "storyId": story_id,
                    "name": f"第 {sequence} 个无效退款申请被拒绝并返回原因",
                },
            ]
        )
        task["storyId"] = story_id
        task["acceptanceCriterionIds"] = [
            accepted_criterion_id,
            rejected_criterion_id,
        ]

    engine = require_office_engine()

    def calculated_audit(name: str, delivery: dict[str, object]):
        candidate = tmp_path / f"{name}.candidate.xlsx"
        write_workbook(SKILL_ROOT / "assets/sow-template.xlsx", scope, delivery, candidate)
        output = office_engine.recalculate_workbook(
            candidate, tmp_path / f"{name}.xlsx", engine
        )
        return workbook_module.audit_calculated_workbook(
            Path(output.output_path),
            SKILL_ROOT / "assets/sow-template.xlsx",
            scope,
            delivery,
            engine,
        )

    combined_audit = calculated_audit("combined", combined)
    split_audit = calculated_audit("split", split)

    assert combined_audit.direct_days == split_audit.direct_days
    assert combined_audit.uat_days == split_audit.uat_days
    assert combined_audit.total_days == split_audit.total_days


def test_calculated_workbook_audit_rejects_formula_changed_after_roundtrip(
    tmp_path: Path, monkeypatch
) -> None:
    executable = installed_soffice()
    monkeypatch.setenv("AI_SOW_OFFICE_BIN", str(executable))
    scope = fixture("brownfield", "scope.json")
    delivery = fixture("brownfield", "delivery.json")
    candidate = tmp_path / "candidate.xlsx"
    write_workbook(
        SKILL_ROOT / "assets/sow-template.xlsx",
        scope,
        delivery,
        candidate,
    )
    engine = require_office_engine()
    calculated = Path(
        office_engine.recalculate_workbook(
            candidate, tmp_path / "calculated.xlsx", engine
        ).output_path
    )
    replace_zip_member(
        calculated,
        "xl/worksheets/sheet2.xml",
        b"ROUND($G5*$H5,1)",
        b"ROUND($G5*$H5*2,1)",
    )

    with pytest.raises(ValueError, match="formula"):
        workbook_module.audit_calculated_workbook(
            calculated,
            SKILL_ROOT / "assets/sow-template.xlsx",
            scope,
            delivery,
            engine,
        )


def test_calculated_workbook_audit_rejects_catalog_row_without_work_mode(
    tmp_path: Path, monkeypatch
) -> None:
    executable = installed_soffice()
    monkeypatch.setenv("AI_SOW_OFFICE_BIN", str(executable))
    template = tmp_path / "template.xlsx"
    shutil.copyfile(SKILL_ROOT / "assets/sow-template.xlsx", template)
    workbook = load_workbook(template)
    try:
        worksheet = workbook["90-估算标准"]
        table = worksheet.tables["BaseUnitCatalogTable"]
        min_col, min_row, max_col, _max_row = range_boundaries(table.ref)
        headers = {
            str(worksheet.cell(min_row, column).value): column
            for column in range(min_col, max_col + 1)
        }
        for header in ("新建M档人天", "调整M档人天", "接入复用M档人天"):
            worksheet.cell(min_row + 1, headers[header]).value = "❌"
        workbook.save(template)
    finally:
        workbook.close()
    scope = fixture("brownfield", "scope.json")
    delivery = fixture("brownfield", "delivery.json")
    candidate = tmp_path / "candidate.xlsx"
    write_workbook(template, scope, delivery, candidate)
    engine = require_office_engine()
    calculated = Path(
        office_engine.recalculate_workbook(
            candidate, tmp_path / "calculated.xlsx", engine
        ).output_path
    )

    with pytest.raises(ValueError, match="no available work mode"):
        workbook_module.audit_calculated_workbook(
            calculated, template, scope, delivery, engine
        )


def test_calculated_workbook_audit_rejects_summary_formula_with_stale_cache(
    tmp_path: Path, monkeypatch
) -> None:
    executable = installed_soffice()
    monkeypatch.setenv("AI_SOW_OFFICE_BIN", str(executable))
    scope = fixture("brownfield", "scope.json")
    delivery = fixture("brownfield", "delivery.json")
    candidate = tmp_path / "candidate.xlsx"
    write_workbook(
        SKILL_ROOT / "assets/sow-template.xlsx",
        scope,
        delivery,
        candidate,
    )
    engine = require_office_engine()
    calculated = Path(
        office_engine.recalculate_workbook(
            candidate, tmp_path / "calculated.xlsx", engine
        ).output_path
    )
    replace_zip_member(
        calculated,
        "xl/worksheets/sheet3.xml",
        b"B5+B6+B7",
        b"B5+B6",
    )

    with pytest.raises(ValueError, match="formula"):
        workbook_module.audit_calculated_workbook(
            calculated,
            SKILL_ROOT / "assets/sow-template.xlsx",
            scope,
            delivery,
            engine,
        )


def test_calculated_workbook_audit_rejects_unchanged_formula_with_forged_cache(
    tmp_path: Path, monkeypatch
) -> None:
    executable = installed_soffice()
    monkeypatch.setenv("AI_SOW_OFFICE_BIN", str(executable))
    scope = fixture("brownfield", "scope.json")
    delivery = fixture("brownfield", "delivery.json")
    candidate = tmp_path / "candidate.xlsx"
    write_workbook(
        SKILL_ROOT / "assets/sow-template.xlsx",
        scope,
        delivery,
        candidate,
    )
    engine = require_office_engine()
    calculated = Path(
        office_engine.recalculate_workbook(
            candidate, tmp_path / "calculated.xlsx", engine
        ).output_path
    )
    replace_cached_value(
        calculated,
        "xl/worksheets/sheet2.xml",
        "G5",
        b"999",
    )

    with pytest.raises(ValueError, match="cached formula result"):
        workbook_module.audit_calculated_workbook(
            calculated,
            SKILL_ROOT / "assets/sow-template.xlsx",
            scope,
            delivery,
            engine,
        )


def test_calculated_workbook_audit_rejects_changed_data_validation(
    tmp_path: Path, monkeypatch
) -> None:
    executable = installed_soffice()
    monkeypatch.setenv("AI_SOW_OFFICE_BIN", str(executable))
    scope = fixture("brownfield", "scope.json")
    delivery = fixture("brownfield", "delivery.json")
    candidate = tmp_path / "candidate.xlsx"
    write_workbook(
        SKILL_ROOT / "assets/sow-template.xlsx",
        scope,
        delivery,
        candidate,
    )
    engine = require_office_engine()
    calculated = Path(
        office_engine.recalculate_workbook(
            candidate, tmp_path / "calculated.xlsx", engine
        ).output_path
    )
    replace_zip_member(
        calculated,
        "xl/worksheets/sheet2.xml",
        b'"S,M,L"',
        b'"S,M"',
    )

    with pytest.raises(ValueError, match="data validation"):
        workbook_module.audit_calculated_workbook(
            calculated,
            SKILL_ROOT / "assets/sow-template.xlsx",
            scope,
            delivery,
            engine,
        )


def test_calculated_workbook_audit_rejects_changed_visible_style(
    tmp_path: Path, monkeypatch
) -> None:
    executable = installed_soffice()
    monkeypatch.setenv("AI_SOW_OFFICE_BIN", str(executable))
    scope = fixture("brownfield", "scope.json")
    delivery = fixture("brownfield", "delivery.json")
    candidate = tmp_path / "candidate.xlsx"
    write_workbook(
        SKILL_ROOT / "assets/sow-template.xlsx", scope, delivery, candidate
    )
    engine = require_office_engine()
    calculated = Path(
        office_engine.recalculate_workbook(
            candidate, tmp_path / "calculated.xlsx", engine
        ).output_path
    )
    replace_zip_member(
        calculated,
        "xl/worksheets/sheet2.xml",
        b'<c r="A5" s="6"',
        b'<c r="A5" s="4"',
    )

    with pytest.raises(ValueError, match="visible style"):
        workbook_module.audit_calculated_workbook(
            calculated,
            SKILL_ROOT / "assets/sow-template.xlsx",
            scope,
            delivery,
            engine,
        )


def test_calculated_workbook_audit_rejects_changed_row_height(
    tmp_path: Path, monkeypatch
) -> None:
    executable = installed_soffice()
    monkeypatch.setenv("AI_SOW_OFFICE_BIN", str(executable))
    scope = fixture("brownfield", "scope.json")
    delivery = fixture("brownfield", "delivery.json")
    candidate = tmp_path / "candidate.xlsx"
    write_workbook(
        SKILL_ROOT / "assets/sow-template.xlsx", scope, delivery, candidate
    )
    engine = require_office_engine()
    calculated = Path(
        office_engine.recalculate_workbook(
            candidate, tmp_path / "calculated.xlsx", engine
        ).output_path
    )
    replace_zip_member(
        calculated,
        "xl/worksheets/sheet2.xml",
        b'<row r="5" customFormat="false" ht="64"',
        b'<row r="5" customFormat="false" ht="80"',
    )

    with pytest.raises(ValueError, match="row height"):
        workbook_module.audit_calculated_workbook(
            calculated,
            SKILL_ROOT / "assets/sow-template.xlsx",
            scope,
            delivery,
            engine,
        )


def test_calculated_workbook_audit_rejects_unreadable_print_pagination(
    tmp_path: Path, monkeypatch
) -> None:
    executable = installed_soffice()
    monkeypatch.setenv("AI_SOW_OFFICE_BIN", str(executable))
    scope = fixture("brownfield", "scope.json")
    delivery = fixture("brownfield", "delivery.json")
    candidate = tmp_path / "candidate.xlsx"
    write_workbook(
        SKILL_ROOT / "assets/sow-template.xlsx", scope, delivery, candidate
    )
    engine = require_office_engine()
    calculated = Path(
        office_engine.recalculate_workbook(
            candidate, tmp_path / "calculated.xlsx", engine
        ).output_path
    )
    replace_zip_member(
        calculated,
        "xl/worksheets/sheet2.xml",
        b'fitToHeight="0"',
        b'fitToHeight="1"',
    )

    with pytest.raises(ValueError, match="print layout"):
        workbook_module.audit_calculated_workbook(
            calculated,
            SKILL_ROOT / "assets/sow-template.xlsx",
            scope,
            delivery,
            engine,
        )


def test_zip_external_attributes_are_host_independent() -> None:
    windows_archive_bit = 0x20
    unix_regular_file = 0o100644 << 16

    assert deterministic_external_attr(windows_archive_bit) == (
        deterministic_external_attr(unix_regular_file)
    )
