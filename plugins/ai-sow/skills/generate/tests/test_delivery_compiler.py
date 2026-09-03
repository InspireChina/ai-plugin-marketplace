from __future__ import annotations

import copy
import json
import sys
from pathlib import Path

import pytest
import openpyxl


SKILL_ROOT = Path(__file__).parents[1]
FIXTURES = SKILL_ROOT / "fixtures"
ASSETS = SKILL_ROOT / "assets"
SCRIPTS = SKILL_ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

from contracts import canonical_json_bytes, sha256_bytes  # noqa: E402
from delivery_compiler import (  # noqa: E402
    compile_delivery as _compile_delivery,
    read_template_catalog,
)
from models import ImpactPlan  # noqa: E402
from scope_compiler import impact_plan_sha256  # noqa: E402


COLLECTION_TYPES = {
    "stories": ("STORY", "storyId"),
    "acceptanceCriteria": ("ACCEPTANCE_CRITERION", "acceptanceCriterionId"),
    "tasks": ("TASK", "taskId"),
    "dependencies": ("DEPENDENCY", "dependencyId"),
}


def fixture(mode: str, name: str) -> dict[str, object]:
    return json.loads((FIXTURES / mode / name).read_text(encoding="utf-8"))


def criterion_source_ref_inventory(value: dict[str, object]) -> set[tuple[str, str, str]]:
    return {
        (str(ref["sourceId"]), str(ref["anchorId"]), str(ref["sha256"]))
        for criterion in value["acceptanceCriteria"]
        for ref in criterion.get("sourceRefs", [])
    }


def compile_delivery(
    scope: dict[str, object],
    previous: dict[str, object] | None,
    candidate: dict[str, object],
    ids: dict[str, object],
    plan: ImpactPlan,
    source_ref_inventory: set[tuple[str, str, str]] | object | None = None,
    catalog=None,
):
    if catalog is None:
        catalog = source_ref_inventory
        source_ref_inventory = None
    inventory = (
        source_ref_inventory
        if isinstance(source_ref_inventory, set)
        else criterion_source_ref_inventory(candidate)
    )
    return _compile_delivery(scope, previous, candidate, ids, plan, inventory, catalog)


def plan_for(scope: dict[str, object], action: str = "FULL_COMPILE") -> ImpactPlan:
    return ImpactPlan(
        action=action,
        baseline_generation_id=None if action == "FULL_COMPILE" else "000001",
        baseline_revision_id=None if action == "FULL_COMPILE" else "000001",
        changed_source_ids=(),
        changed_anchor_ids=(),
        affected_feature_ids=tuple(item["featureId"] for item in scope["features"]),
        escalation="FULL" if action == "FULL_COMPILE" else "FEATURE",
        reason_codes=("NO_CURRENT_GENERATION",) if action == "FULL_COMPILE" else ("INPUT_CHANGED",),
    )


def delivery_slice(
    bundle: dict[str, object], scope: dict[str, object], plan: ImpactPlan
) -> dict[str, object]:
    scope_sha = sha256_bytes(canonical_json_bytes(scope))
    return {
        "contract": "ai-sow-delivery-slice-v3",
        "inputRevisionId": scope["inputRevisionId"],
        "scopeSha256": scope_sha,
        "impactPlanSha256": impact_plan_sha256(plan),
        "replacesFeatureIds": (
            [] if plan.baseline_generation_id is None else list(plan.affected_feature_ids)
        ),
        **{
            collection: copy.deepcopy(bundle[collection])
            for collection in COLLECTION_TYPES
        },
    }


def id_decisions(candidate: dict[str, object], previous=None) -> dict[str, object]:
    previous_ids = set()
    if previous is not None:
        for collection, (_kind, id_field) in COLLECTION_TYPES.items():
            previous_ids.update(item[id_field] for item in previous[collection])
    decisions = []
    for collection, (kind, id_field) in COLLECTION_TYPES.items():
        for item in candidate[collection]:
            object_id = item[id_field]
            preserved = object_id in previous_ids
            decision = {
                "objectType": kind,
                "objectId": object_id,
                "disposition": "UNCHANGED" if preserved else "NEW",
                "meaningPreserved": preserved,
                "rationale": "语义不变，保留原 ID。" if preserved else "新交付对象分配稳定 ID。",
            }
            if preserved:
                decision["previousId"] = object_id
            decisions.append(decision)
    return {"contract": "ai-sow-id-decisions-v1", "decisions": decisions}


def compile_fixture(mode: str = "brownfield"):
    scope = fixture(mode, "scope.json")
    delivery = fixture(mode, "delivery.json")
    delivery["scopeSha256"] = sha256_bytes(canonical_json_bytes(scope))
    plan = plan_for(scope)
    candidate = delivery_slice(delivery, scope, plan)
    return compile_delivery(
        scope,
        None,
        candidate,
        id_decisions(candidate),
        plan,
        read_template_catalog(ASSETS / "sow-template.xlsx"),
    )


def compile_with_effective_start(
    scope: dict[str, object], delivery: dict[str, object]
):
    delivery["scopeSha256"] = sha256_bytes(canonical_json_bytes(scope))
    plan = plan_for(scope)
    candidate = delivery_slice(delivery, scope, plan)
    return compile_delivery(
        scope,
        None,
        candidate,
        id_decisions(candidate),
        plan,
        read_template_catalog(ASSETS / "sow-template.xlsx"),
    )


def diagnostic_codes(result) -> set[str]:
    return {diagnostic.code for diagnostic in result.diagnostics}


def compile_semantic_mutation(mutation: str):
    scope = fixture("greenfield", "scope.json")
    delivery = fixture("greenfield", "delivery.json")
    if mutation == "story-starts-with-complete":
        delivery["stories"][0]["name"] = "完成退款申请处理"
    elif mutation == "story-with-one-ac":
        delivery["acceptanceCriteria"] = [delivery["acceptanceCriteria"][0]]
        delivery["tasks"][0]["acceptanceCriterionIds"] = ["ac-refund-submit"]
    elif mutation == "task-with-two-interface-objects":
        delivery["tasks"][0]["name"] = "新增退款申请创建与提交接口"
    elif mutation == "task-with-generic-service-title":
        delivery["tasks"][0]["name"] = "开发退款申请受理服务"
    elif mutation == "task-crosses-story-ac":
        other_story = {
            "storyId": "story-refund-result",
            "featureId": "feature-refund-processing",
            "name": "[退款申请] 系统查询处理结果",
            "uatRelevant": True,
        }
        other_criteria = [
            {
                "acceptanceCriterionId": "ac-refund-result-query",
                "storyId": "story-refund-result",
                "name": "退款处理结果可被查询",
                "sourceRefs": copy.deepcopy(delivery["acceptanceCriteria"][0]["sourceRefs"]),
            },
            {
                "acceptanceCriterionId": "ac-refund-result-reason",
                "storyId": "story-refund-result",
                "name": "退款处理失败时可查询明确原因",
                "sourceRefs": copy.deepcopy(delivery["acceptanceCriteria"][0]["sourceRefs"]),
            },
        ]
        delivery["stories"].append(other_story)
        delivery["acceptanceCriteria"].extend(other_criteria)
        delivery["tasks"][0]["acceptanceCriterionIds"].append(
            "ac-refund-result-query"
        )
    else:
        raise ValueError(f"unknown mutation: {mutation}")
    return compile_with_effective_start(scope, delivery)


@pytest.mark.parametrize(
    ("mutation", "expected"),
    [
        ("story-with-one-ac", "DELIVERY_STORY_AC_INSUFFICIENT"),
        ("task-crosses-story-ac", "DELIVERY_TASK_AC_CROSS_STORY"),
    ],
)
def test_delivery_semantic_failures(mutation: str, expected: str) -> None:
    assert expected in diagnostic_codes(compile_semantic_mutation(mutation))


@pytest.mark.parametrize(
    ("mutation", "language_specific_diagnostic"),
    [
        ("story-starts-with-complete", "DELIVERY_STORY_TITLE_FORMULAIC"),
        ("task-with-two-interface-objects", "DELIVERY_TASK_INSTANCE_AMBIGUOUS"),
        ("task-with-generic-service-title", "DELIVERY_TASK_NAME_GENERIC"),
    ],
)
def test_mechanical_validation_does_not_infer_semantics_from_chinese_titles(
    mutation: str, language_specific_diagnostic: str
) -> None:
    assert language_specific_diagnostic not in diagnostic_codes(
        compile_semantic_mutation(mutation)
    )


def test_catalog_accepts_any_non_empty_unique_family_and_base_unit_count(
    tmp_path: Path,
) -> None:
    template = tmp_path / "one-base-unit.xlsx"
    workbook = openpyxl.load_workbook(ASSETS / "sow-template.xlsx")
    try:
        worksheet = workbook["90-估算标准"]
        table = worksheet.tables["BaseUnitCatalogTable"]
        min_col, min_row, max_col, _max_row = openpyxl.utils.range_boundaries(table.ref)
        table.ref = (
            f"{openpyxl.utils.get_column_letter(min_col)}{min_row}:"
            f"{openpyxl.utils.get_column_letter(max_col)}{min_row + 1}"
        )
        workbook.save(template)
    finally:
        workbook.close()

    catalog = read_template_catalog(template)

    assert len(catalog.base_units) == 1


def compile_with_mutated_ac_source(mutation: str):
    scope = fixture("greenfield", "scope.json")
    delivery = fixture("greenfield", "delivery.json")
    source_ref = {
        "sourceId": "prd-refund",
        "anchorId": "anchor-return-submit-state",
        "locator": "heading:退款申请",
        "sha256": "a" * 64,
    }
    delivery["acceptanceCriteria"][0]["sourceRefs"] = [source_ref]
    if mutation == "unknown-anchor":
        source_ref["anchorId"] = "anchor-not-in-revision"
    elif mutation == "wrong-hash":
        source_ref["sha256"] = "b" * 64
    else:
        raise ValueError(f"unknown mutation: {mutation}")
    plan = plan_for(scope)
    candidate = delivery_slice(delivery, scope, plan)
    return compile_delivery(
        scope,
        None,
        candidate,
        id_decisions(candidate),
        plan,
        source_ref_inventory={
            ("prd-refund", "anchor-return-submit-state", "a" * 64),
            (
                "prd-refund",
                "anchor-a1076b3d69271346",
                "9d4bacb4376916256e3e9f7f63f189bf9dd20499747f653562b6f5585841d1e8",
            ),
        },
        catalog=read_template_catalog(ASSETS / "sow-template.xlsx"),
    )


@pytest.mark.parametrize("mutation", ["unknown-anchor", "wrong-hash"])
def test_delivery_rejects_unresolvable_ac_source_ref(mutation: str) -> None:
    result = compile_with_mutated_ac_source(mutation)

    assert "DELIVERY_AC_SOURCE_REF_INVALID" in diagnostic_codes(result)


def test_design_required_item_becomes_design_task_under_implementation_story() -> None:
    result = compile_fixture()
    assert not result.diagnostics
    task = next(
        item
        for item in result.bundle["tasks"]
        if item["taskId"] == "task-refund-recovery-design"
    )
    assert task["taskKind"] == "DESIGN"
    assert task["storyId"] == "story-refund-processing"
    assert task["nfrIds"] == ["nfr-recovery-objective"]
    assert all("storyType" not in story for story in result.bundle["stories"])


def test_initial_full_delivery_rejects_nonempty_replacement_set() -> None:
    scope = fixture("greenfield", "scope.json")
    delivery = fixture("greenfield", "delivery.json")
    delivery["scopeSha256"] = sha256_bytes(canonical_json_bytes(scope))
    plan = plan_for(scope)
    candidate = delivery_slice(delivery, scope, plan)
    candidate["replacesFeatureIds"] = [scope["features"][0]["featureId"]]

    result = compile_delivery(
        scope,
        None,
        candidate,
        id_decisions(candidate),
        plan,
        read_template_catalog(ASSETS / "sow-template.xlsx"),
    )

    assert "DELIVERY_REPLACEMENT_CLOSURE_MISMATCH" in diagnostic_codes(result)


def test_template_catalog_validates_but_does_not_copy_effort() -> None:
    catalog = read_template_catalog(ASSETS / "sow-template.xlsx")
    result = compile_fixture("greenfield")
    assert len(catalog.base_units) == 37
    assert all(
        not hasattr(rule, "person_days") and not hasattr(rule, "base_effort")
        for rule in catalog.base_units.values()
    )
    assert all(
        "personDays" not in task and "baseEffort" not in task
        for task in result.bundle["tasks"]
    )


def test_template_catalog_counts_one_independently_estimable_interface() -> None:
    catalog = read_template_catalog(ASSETS / "sow-template.xlsx")
    rule = catalog.base_units["BU-BUSINESS-SERVICE-API"]

    assert rule.count_rule == "一个可独立开发、测试和估算的业务操作或接口"
    assert rule.complexity_standards == {
        "S": "单一查询或写入接口；简单规则；单一事务；标准权限",
        "M": "单一接口涉及多个业务对象；多条业务规则；异步处理或多角色权限",
        "L": "单一接口需要跨服务编排、复杂状态机、幂等补偿、高并发或严格一致性",
    }
    assert rule.split_rule == "包含多个可独立验收的业务操作或接口，或需要跨多个服务分别交付"


def test_replaced_slice_drops_old_story_ac_task_and_preserves_unaffected() -> None:
    scope = fixture("brownfield", "scope.json")
    delivery = fixture("brownfield", "delivery.json")
    delivery["scopeSha256"] = sha256_bytes(canonical_json_bytes(scope))
    previous = copy.deepcopy(delivery)
    previous["stories"].append(
        {
            "storyId": "story-old-notification",
            "featureId": "feature-refund-notification",
            "name": "[退款通知] 系统处理旧通知",
            "uatRelevant": True,
        }
    )
    previous["acceptanceCriteria"].append(
        {
            "acceptanceCriterionId": "ac-old-notification",
            "storyId": "story-old-notification",
            "name": "旧通知可用",
        }
    )
    previous["acceptanceCriteria"].append(
        {
            "acceptanceCriterionId": "ac-old-notification-retry",
            "storyId": "story-old-notification",
            "name": "旧通知失败后可重试",
        }
    )
    previous["tasks"].append(
        {
            "taskId": "task-old-notification",
            "storyId": "story-old-notification",
            "taskKind": "IMPLEMENTATION",
            "name": "实现旧通知",
            "baseUnit": "BU-BUSINESS-SERVICE-API",
            "workMode": "新建",
            "workModeRationale": "旧任务。",
            "complexity": "M",
            "acceptanceCriterionIds": [
                "ac-old-notification",
                "ac-old-notification-retry",
            ],
            "designItemIds": ["design-item-refund-orchestration"],
            "integrationIds": [],
            "nfrIds": [],
            "rationale": "旧任务应删除。",
        }
    )
    plan = plan_for(scope, "SLICE_COMPILE")
    candidate = delivery_slice(delivery, scope, plan)
    result = compile_delivery(
        scope,
        previous,
        candidate,
        id_decisions(candidate, previous),
        plan,
        read_template_catalog(ASSETS / "sow-template.xlsx"),
    )
    all_ids = {
        item[id_field]
        for collection, (_kind, id_field) in COLLECTION_TYPES.items()
        for item in result.bundle[collection]
    }
    assert not result.diagnostics
    assert {
        "story-old-notification",
        "ac-old-notification",
        "ac-old-notification-retry",
        "task-old-notification",
    }.isdisjoint(all_ids)


def test_unknown_template_base_unit_blocks_delivery() -> None:
    scope = fixture("greenfield", "scope.json")
    delivery = fixture("greenfield", "delivery.json")
    delivery["tasks"][0]["baseUnit"] = "BU-UNKNOWN"
    plan = plan_for(scope)
    candidate = delivery_slice(delivery, scope, plan)
    result = compile_delivery(
        scope,
        None,
        candidate,
        id_decisions(candidate),
        plan,
        read_template_catalog(ASSETS / "sow-template.xlsx"),
    )
    assert "DELIVERY_BASE_UNIT_UNKNOWN" in diagnostic_codes(result)


def test_capability_start_cannot_support_adjustment() -> None:
    scope = fixture("brownfield", "scope.json")
    delivery = fixture("brownfield", "delivery.json")
    item = scope["effectiveStartItems"][0]
    item["matchLevel"] = "CAPABILITY"
    item.pop("baseUnit", None)

    result = compile_with_effective_start(scope, delivery)

    assert "TASK_WORK_MODE_REQUIRES_INSTANCE_START" in diagnostic_codes(result)


def test_instance_start_base_unit_must_match_task() -> None:
    scope = fixture("brownfield", "scope.json")
    delivery = fixture("brownfield", "delivery.json")
    item = scope["effectiveStartItems"][0]
    item["matchLevel"] = "TASK_INSTANCE"
    item["baseUnit"] = "BU-DATA-MODEL"

    result = compile_with_effective_start(scope, delivery)

    assert "TASK_WORK_MODE_BASE_UNIT_MISMATCH" in diagnostic_codes(result)


def reuse_start(scope: dict[str, object]) -> dict[str, object]:
    start = {
        "effectiveStartItemId": "effective-start-payment-refund",
        "matchLevel": "TASK_INSTANCE",
        "baseUnit": "BU-EXTERNAL-INTEGRATION",
        "name": "既有支付退款对接",
        "summary": "项目启动时已有可用的支付退款对接。",
        "commitmentIds": [],
        "featureIds": ["feature-refund-processing"],
        "sourceRefs": [
            {
                "sourceId": "prior-sow-refund",
                "anchorId": "anchor-prior-payment-refund-integration",
                "locator": "table:系统能力/支付退款对接",
                "sha256": "c622222222222222222222222222222222222222222222222222222222222222",
            }
        ],
    }
    scope["effectiveStartItems"].append(start)
    return start


def reuse_existing_task(delivery: dict[str, object]) -> None:
    task = next(
        item
        for item in delivery["tasks"]
        if item["taskId"] == "task-payment-refund-integration"
    )
    task["workMode"] = "接入复用"
    task["workModeRationale"] = "复用既有支付退款对接能力并完成本项目配置。"
    task["workModeEvidence"] = {
        "effectiveStartItemId": "effective-start-payment-refund",
        "projectSideWorkTypes": ["CONFIGURE"],
        "projectSideWorkCommitment": "配置本项目的支付退款接入并完成专项验证。",
    }


def test_capability_start_cannot_support_reuse() -> None:
    scope = fixture("brownfield", "scope.json")
    delivery = fixture("brownfield", "delivery.json")
    item = reuse_start(scope)
    item["matchLevel"] = "CAPABILITY"
    item.pop("baseUnit", None)
    reuse_existing_task(delivery)

    result = compile_with_effective_start(scope, delivery)

    assert "TASK_WORK_MODE_REQUIRES_INSTANCE_START" in diagnostic_codes(result)


def test_reuse_start_base_unit_must_match_task() -> None:
    scope = fixture("brownfield", "scope.json")
    delivery = fixture("brownfield", "delivery.json")
    item = reuse_start(scope)
    item["baseUnit"] = "BU-DATA-MODEL"
    reuse_existing_task(delivery)

    result = compile_with_effective_start(scope, delivery)

    assert "TASK_WORK_MODE_BASE_UNIT_MISMATCH" in diagnostic_codes(result)


def test_story_with_more_than_four_tasks_blocks_delivery() -> None:
    scope = fixture("greenfield", "scope.json")
    delivery = fixture("greenfield", "delivery.json")
    prototype = delivery["tasks"][0]
    delivery["tasks"] = []
    for sequence in range(1, 6):
        task = copy.deepcopy(prototype)
        task["taskId"] = f"task-refund-service-{sequence}"
        task["name"] = f"实现退款申请服务实例 {sequence}"
        delivery["tasks"].append(task)
    plan = plan_for(scope)
    candidate = delivery_slice(delivery, scope, plan)

    result = compile_delivery(
        scope,
        None,
        candidate,
        id_decisions(candidate),
        plan,
        read_template_catalog(ASSETS / "sow-template.xlsx"),
    )

    assert "DELIVERY_STORY_TASK_LIMIT_EXCEEDED" in diagnostic_codes(result)


def test_story_requires_at_least_two_acceptance_criteria() -> None:
    scope = fixture("greenfield", "scope.json")
    delivery = fixture("greenfield", "delivery.json")
    first_criterion = delivery["acceptanceCriteria"][0]
    delivery["acceptanceCriteria"] = [first_criterion]
    delivery["tasks"][0]["acceptanceCriterionIds"] = [
        first_criterion["acceptanceCriterionId"]
    ]
    plan = plan_for(scope)
    candidate = delivery_slice(delivery, scope, plan)

    result = compile_delivery(
        scope,
        None,
        candidate,
        id_decisions(candidate),
        plan,
        read_template_catalog(ASSETS / "sow-template.xlsx"),
    )

    assert "DELIVERY_STORY_AC_INSUFFICIENT" in diagnostic_codes(result)


def test_story_names_must_be_unique_before_workbook_projection() -> None:
    scope = fixture("brownfield", "scope.json")
    delivery = fixture("brownfield", "delivery.json")
    delivery["stories"][1]["name"] = delivery["stories"][0]["name"]
    plan = plan_for(scope)
    candidate = delivery_slice(delivery, scope, plan)

    result = compile_delivery(
        scope,
        None,
        candidate,
        id_decisions(candidate),
        plan,
        read_template_catalog(ASSETS / "sow-template.xlsx"),
    )

    assert "DELIVERY_STORY_NAME_DUPLICATE" in diagnostic_codes(result)


def test_dependency_cycle_blocks_delivery() -> None:
    scope = fixture("brownfield", "scope.json")
    delivery = fixture("brownfield", "delivery.json")
    delivery["dependencies"].append(
        {
            "dependencyId": "dependency-build-before-design",
            "predecessorTaskId": "task-refund-orchestration",
            "successorTaskId": "task-refund-recovery-design",
            "rationale": "合成循环。",
        }
    )
    plan = plan_for(scope)
    candidate = delivery_slice(delivery, scope, plan)
    result = compile_delivery(
        scope,
        None,
        candidate,
        id_decisions(candidate),
        plan,
        read_template_catalog(ASSETS / "sow-template.xlsx"),
    )
    assert "DELIVERY_DEPENDENCY_CYCLE" in diagnostic_codes(result)
