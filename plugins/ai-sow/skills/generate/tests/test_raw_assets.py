from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula


SKILL_ROOT = Path(__file__).parents[1]
ASSETS = SKILL_ROOT / "assets"
PRD_TEMPLATE = ASSETS / "prd-template.md"
HLD_TEMPLATE = ASSETS / "hld-template.md"
QUESTIONNAIRE = ASSETS / "greenfield-questionnaire.md"
SOW_TEMPLATE = ASSETS / "sow-template.xlsx"
RENDERER_BASELINE = SKILL_ROOT / "contracts/renderer-fingerprint-baseline.json"
REQUIRED_PRD_SECTIONS = {
    "项目背景与问题",
    "目标与成功指标",
    "In Scope",
    "Out of Scope",
    "用户与角色",
    "核心业务场景",
    "Feature、业务结果、业务规则与验收意图",
    "优先级与阶段",
    "业务约束、依赖与假设",
    "业务数据、合规与外部参与方",
}
REQUIRED_HLD_SECTIONS = {
    "系统上下文与责任",
    "目标架构",
    "关键业务流",
    "跨系统 Integration",
    "数据、迁移、保留与安全分类",
    "NFR",
    "环境、部署与切换",
    "关键技术决策",
    "待设计事项",
}


def markdown_headings(path: Path) -> set[str]:
    return {
        match.group(1).strip()
        for line in path.read_text(encoding="utf-8").splitlines()
        if (match := re.match(r"^#{2,6}\s+(.+)$", line))
    }


def test_prd_and_hld_assets_express_every_required_semantic_section() -> None:
    assert REQUIRED_PRD_SECTIONS <= markdown_headings(PRD_TEMPLATE)
    assert REQUIRED_HLD_SECTIONS <= markdown_headings(HLD_TEMPLATE)


def test_templates_do_not_request_internal_compilation_artifacts() -> None:
    prd = PRD_TEMPLATE.read_text(encoding="utf-8")
    hld = HLD_TEMPLATE.read_text(encoding="utf-8")
    for forbidden in ("Coverage Matrix", "Story 分解", "Task 分解"):
        assert forbidden not in prd
    for forbidden in ("Coverage Matrix", "字段级接口", "类设计"):
        assert forbidden not in hld


def test_greenfield_questionnaire_is_minimal() -> None:
    text = QUESTIONNAIRE.read_text(encoding="utf-8")
    assert {"责任边界", "环境准备", "第三方依赖", "数据迁移责任"} <= set(
        re.findall(r"^##\s+(.+)$", text, re.MULTILINE)
    )
    assert text.count("？") == 4


def table_names(workbook) -> set[str]:
    return {
        table_name
        for worksheet in workbook.worksheets
        for table_name in worksheet.tables
    }


def table_headers(workbook, table_name: str) -> list[str]:
    for worksheet in workbook.worksheets:
        if table_name in worksheet.tables:
            return [column.name for column in worksheet.tables[table_name].tableColumns]
    raise AssertionError(f"missing table: {table_name}")


def formula_headers(workbook, table_name: str) -> set[str]:
    for worksheet in workbook.worksheets:
        if table_name not in worksheet.tables:
            continue
        table = worksheet.tables[table_name]
        header_row = worksheet[table.ref.split(":", 1)[0]].row
        result: set[str] = set()
        for column in table.tableColumns:
            cell = worksheet.cell(header_row + 1, column.id)
            if cell.data_type == "f" and isinstance(cell.value, (str, ArrayFormula)):
                result.add(column.name)
        return result
    raise AssertionError(f"missing table: {table_name}")


def test_bundled_sow_template_is_pinned_four_sheet_calculation_authority() -> None:
    assert hashlib.sha256(SOW_TEMPLATE.read_bytes()).hexdigest() == (
        "51f88c98a6f68fb2b95b58c28b95a7d68897df38d685532ef89a5de19727bac9"
    )
    workbook = load_workbook(SOW_TEMPLATE, data_only=False, read_only=False)
    try:
        assert workbook.sheetnames == [
            "01-需求故事",
            "02-任务清单",
            "03-工作量汇总",
            "90-估算标准",
        ]
        assert workbook["01-需求故事"]["A2"].value == (
            "每个 Story 一行；标题使用自然的角色/对象动作。"
            "每条验收条件以“• ”开头并独占一行；"
            "任务列表显示“[任务类型/工作方式/复杂度]”。"
            "备注只显示对象特有风险、假设、边界或不确定性，"
            "通用事项留在说明中。灰色列由公式维护。"
        )
        assert table_names(workbook) == {
            "SOWStoryTable",
            "TaskTable",
            "ProjectSummaryTable",
            "ProjectParameterTable",
            "BaseUnitCatalogTable",
        }
        assert table_headers(workbook, "SOWStoryTable") == [
            "需求",
            "子需求",
            "故事",
            "UAT适用",
            "验收条件",
            "备注",
            "任务列表",
            "故事人天",
            "校验结果",
        ]
        assert table_headers(workbook, "TaskTable") == [
            "所属故事",
            "任务名称",
            "任务类型",
            "工作方式",
            "复杂度",
            "备注",
            "M档标准人天",
            "复杂度系数",
            "任务人天",
            "SIT支持人天",
            "校验结果",
        ]
        assert formula_headers(workbook, "SOWStoryTable") == {
            "任务列表",
            "故事人天",
            "校验结果",
        }
        assert formula_headers(workbook, "TaskTable") == {
            "M档标准人天",
            "复杂度系数",
            "任务人天",
            "SIT支持人天",
            "校验结果",
        }
        task_sheet = workbook["02-任务清单"]
        assert "'01-需求故事'!$C:$C" in task_sheet["K5"].value
        assert "'01-需求故事'!$J:$J" not in task_sheet["K5"].value
        summary = workbook["03-工作量汇总"]
        assert summary["B5"].value == "=SUM('02-任务清单'!$I:$I)"
        assert (
            'SUMPRODUCT((SOWStoryTable[UAT适用]="是")*'
            "SUMIF(TaskTable[所属故事],SOWStoryTable[故事],TaskTable[任务人天]))"
            in summary["B7"].value
        )
        assert "故事路径" not in summary["B7"].value
    finally:
        workbook.close()


def test_renderer_fingerprint_binds_projection_and_office_implementation() -> None:
    baseline = json.loads(RENDERER_BASELINE.read_text(encoding="utf-8"))
    assert baseline["rendererContract"] == "generation-renderer-v7"
    assert baseline["files"] == {
        name: hashlib.sha256((SKILL_ROOT / name).read_bytes()).hexdigest()
        for name in (
            "scripts/package_renderer.py",
            "scripts/workbook.py",
            "scripts/office_engine.py",
            "scripts/story_notes.py",
        )
    }
