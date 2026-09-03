from __future__ import annotations

import copy
import datetime as dt
import math
import re
import tempfile
import unicodedata
import xml.etree.ElementTree as ET
from contextlib import ExitStack
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.workbook.properties import CalcProperties
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.table import TableFormula

from models import WorkbookAudit
from office_engine import normalize_xlsx, recalculate_workbook
from story_notes import story_note_projection


FORMAL_SHEETS = (
    "01-需求故事",
    "02-任务清单",
    "03-工作量汇总",
    "90-估算标准",
)
TABLES = ("SOWStoryTable", "TaskTable")
FORMAL_TABLES = {
    "SOWStoryTable",
    "TaskTable",
    "ProjectSummaryTable",
    "BaseUnitCatalogTable",
    "ProjectParameterTable",
}
FORMULA_HEADERS = {
    "SOWStoryTable": {"任务列表", "故事人天", "校验结果"},
    "TaskTable": {"M档标准人天", "复杂度系数", "任务人天", "SIT支持人天", "校验结果"},
}
TABLE_HEADERS = {
    "SOWStoryTable": [
        "需求",
        "子需求",
        "故事",
        "UAT适用",
        "验收条件",
        "备注",
        "任务列表",
        "故事人天",
        "校验结果",
    ],
    "TaskTable": [
        "所属故事",
        "任务名称",
        "任务类型",
        "工作方式",
        "复杂度",
        "备注",
        "M档标准人天",
        "复杂度系数",
        "任务人天",
        "SIT支持人天",
        "校验结果",
    ],
}
SUMMARY_HEADERS = ["工作量项", "人天"]
CATALOG_HEADERS = [
    "任务族ID",
    "任务族名称",
    "基础单元ID",
    "基础单元名称",
    "计数口径",
    "包含内容",
    "不包含内容",
    "新建M档人天",
    "调整M档人天",
    "接入复用M档人天",
    "S标准",
    "M标准",
    "L标准",
    "X/拆分条件",
]
PARAMETER_HEADERS = ["参数代码", "名称", "值", "单位", "适用范围", "验证状态/说明"]
PROTECTED_SHEETS = {"01-需求故事", "02-任务清单"}
RISKY_TEXT = re.compile(r"^[=+\-@]")
BARE_TEXTJOIN = re.compile(r"(?<![\w.])TEXTJOIN\(")
DETERMINISTIC_TIME = dt.datetime(2000, 1, 1, 0, 0, 0)
WRAPPED_LINE_HEIGHT = 15
WRAPPED_ROW_PADDING = 4
MAX_EXCEL_ROW_HEIGHT = 409.5
FORMULA_ERROR_PREFIXES = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "Err:")
SUMMARY_LABELS = (
    "直接开发人天",
    "SIT支持人天",
    "UAT支持人天",
    "总开发人天",
)


def safe_text(value: object) -> object:
    if isinstance(value, str) and RISKY_TEXT.match(value):
        return "'" + value
    return value


def formula_text(value: object) -> str:
    """Return a formula's text without discarding legacy array metadata."""
    if isinstance(value, ArrayFormula):
        value = value.text
    if not isinstance(value, str):
        raise TypeError("formula value is not text")
    return value


def normalize_table_formula(formula: object) -> str:
    """Serialize table references and future functions in OOXML form."""
    formula = formula_text(formula)
    parts = formula.split('"')
    for index in range(0, len(parts), 2):
        parts[index] = parts[index].replace("@", "[#This Row],")
        parts[index] = BARE_TEXTJOIN.sub("_xlfn.TEXTJOIN(", parts[index])
    return '"'.join(parts)


def comparable_formula(formula: object) -> str:
    """Normalize equivalent formula spelling used by Excel and LibreOffice."""
    value = formula_text(formula).replace("_xlfn.", "")
    return re.sub(r"\b(TRUE|FALSE)\(\)", r"\1", value)


def require_unique_names(entries: list[dict[str, Any]], label: str) -> None:
    projected_names: dict[str, str] = {}
    for entry in entries:
        name = entry.get("name")
        if not isinstance(name, str) or not name.strip():
            raise ValueError(f"{label} display name is blank")
        projected = str(safe_text(name))
        key = unicodedata.normalize("NFC", projected).casefold()
        if key in projected_names:
            raise ValueError(
                f"{label} display name is duplicated after Excel projection: "
                f"{projected_names[key]} / {name}"
            )
        projected_names[key] = name


def build_rows(
    scope: dict[str, Any],
    delivery: dict[str, Any],
    base_unit_names: dict[str, str],
) -> dict[str, list[dict[str, object]]]:
    if not delivery["stories"]:
        raise ValueError("formal workbook requires at least one Story")
    if not delivery["tasks"]:
        raise ValueError("formal workbook requires at least one Task")

    epics = {entry["epicId"]: entry for entry in scope["epics"]}
    features = {entry["featureId"]: entry for entry in scope["features"]}
    stories = {entry["storyId"]: entry for entry in delivery["stories"]}

    for label, entries in (
        ("Epic", scope["epics"]),
        ("Feature", scope["features"]),
        ("Story", delivery["stories"]),
        ("Task", delivery["tasks"]),
    ):
        require_unique_names(entries, label)

    acceptance_names_by_story: dict[str, list[str]] = {}
    for criterion in delivery["acceptanceCriteria"]:
        acceptance_names_by_story.setdefault(criterion["storyId"], []).append(
            criterion["name"]
        )

    task_display_names_by_story: dict[str, list[str]] = {}
    for task in delivery["tasks"]:
        base_unit = task["baseUnit"]
        if base_unit not in base_unit_names:
            raise ValueError(f"template base-unit name is missing: {base_unit}")
        task_display_names_by_story.setdefault(task["storyId"], []).append(
            f"• [{base_unit_names[base_unit]}/{task['workMode']}/{task['complexity']}] "
            f"{task['name']}"
        )

    story_notes, _story_note_inventory = story_note_projection(scope, delivery)

    story_rows: list[dict[str, object]] = []
    for story in delivery["stories"]:
        feature = features[story["featureId"]]
        epic = epics[feature["epicId"]]
        story_name = str(safe_text(story["name"]))
        story_rows.append(
            {
                "需求": safe_text(epic["name"]),
                "子需求": safe_text(feature["name"]),
                "故事": story_name,
                "UAT适用": "是" if story["uatRelevant"] else "否",
                "验收条件": "\n".join(
                    f"• {name}"
                    for name in acceptance_names_by_story.get(story["storyId"], [])
                ),
                "备注": story_notes.get(story["storyId"], ""),
                "任务列表": "\n".join(
                    task_display_names_by_story.get(story["storyId"], [])
                ),
            }
        )

    task_rows: list[dict[str, object]] = []
    for task in delivery["tasks"]:
        story = stories[task["storyId"]]
        story_name = str(safe_text(story["name"]))
        base_unit = task["baseUnit"]
        if base_unit not in base_unit_names:
            raise ValueError(f"template base-unit name is missing: {base_unit}")
        notes = [
            f"任务理由：{task['rationale']}",
            f"工作方式理由：{task['workModeRationale']}",
        ]
        if task.get("complexityRationale"):
            notes.append(f"复杂度理由：{task['complexityRationale']}")
        task_rows.append(
            {
                "所属故事": story_name,
                "任务名称": task["name"],
                "任务类型": base_unit_names[base_unit],
                "工作方式": task["workMode"],
                "复杂度": task["complexity"],
                "备注": "\n".join(notes),
            }
        )

    return {"SOWStoryTable": story_rows, "TaskTable": task_rows}


def table_index(workbook: Any) -> dict[str, tuple[Any, Any]]:
    found: dict[str, tuple[Any, Any]] = {}
    for worksheet in workbook.worksheets:
        for table_name in worksheet.tables:
            if table_name in found:
                raise ValueError(f"duplicate template table: {table_name}")
            found[table_name] = (worksheet, worksheet.tables[table_name])
    missing = sorted(set(TABLES) - set(found))
    if missing:
        raise ValueError(f"template tables are missing: {missing}")
    return found


def base_unit_name_map(workbook: Any) -> dict[str, str]:
    index = table_index(workbook)
    if "BaseUnitCatalogTable" not in index:
        raise ValueError("template base-unit catalog is missing")
    worksheet, table = index["BaseUnitCatalogTable"]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = [
        worksheet.cell(min_row, column).value
        for column in range(min_col, max_col + 1)
    ]
    if "基础单元ID" not in headers or "基础单元名称" not in headers:
        raise ValueError("template base-unit name projection columns are missing")
    id_column = min_col + headers.index("基础单元ID")
    name_column = min_col + headers.index("基础单元名称")
    result: dict[str, str] = {}
    used_names: set[str] = set()
    for row in range(min_row + 1, max_row + 1):
        unit_id = worksheet.cell(row, id_column).value
        name = worksheet.cell(row, name_column).value
        if not isinstance(unit_id, str) or not unit_id.strip():
            raise ValueError("template base-unit ID is blank")
        if not isinstance(name, str) or not name.strip():
            raise ValueError(f"template base-unit name is blank: {unit_id}")
        if unit_id in result:
            raise ValueError(f"template base-unit ID is duplicated: {unit_id}")
        if name in used_names:
            raise ValueError(f"template base-unit name is duplicated: {name}")
        result[unit_id] = name
        used_names.add(name)
    return result


def table_records(workbook: Any, table_name: str) -> list[dict[str, object]]:
    worksheet, table = table_index(workbook)[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = [
        worksheet.cell(min_row, column).value
        for column in range(min_col, max_col + 1)
    ]
    if not all(isinstance(header, str) and header for header in headers):
        raise ValueError(f"invalid table header: {table_name}")
    return [
        {
            str(header): worksheet.cell(row, min_col + offset).value
            for offset, header in enumerate(headers)
        }
        for row in range(min_row + 1, max_row + 1)
    ]


def require_number(value: object, label: str) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"calculated workbook value is not numeric: {label}")
    if not math.isfinite(float(value)):
        raise ValueError(f"calculated workbook value is not finite: {label}")
    return float(value)


def formula_errors(workbook: Any) -> tuple[str, ...]:
    errors: list[str] = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if cell.data_type == "e" or (
                    isinstance(value, str)
                    and value.startswith(FORMULA_ERROR_PREFIXES)
                ):
                    errors.append(f"{worksheet.title}!{cell.coordinate}:{value}")
    return tuple(errors)


def verify_formula_cache_results(
    formula_workbook: Any,
    cached_workbook: Any,
    reference_formula_workbook: Any,
    reference_cached_workbook: Any,
) -> None:
    """Compare every formula cache with a fresh Office calculation.

    Python deliberately does not reimplement the workbook's estimation rules.
    Instead, the authoritative template is projected and independently
    recalculated by the same supported Office engine; every resulting formula
    cache must then match that reference calculation.
    """
    for sheet_name in FORMAL_SHEETS:
        formula_sheet = formula_workbook[sheet_name]
        cached_sheet = cached_workbook[sheet_name]
        reference_formula_sheet = reference_formula_workbook[sheet_name]
        reference_cached_sheet = reference_cached_workbook[sheet_name]
        coordinates = {
            cell.coordinate
            for row in formula_sheet.iter_rows()
            for cell in row
            if cell.data_type == "f" and isinstance(cell.value, (str, ArrayFormula))
        }
        reference_coordinates = {
            cell.coordinate
            for row in reference_formula_sheet.iter_rows()
            for cell in row
            if cell.data_type == "f" and isinstance(cell.value, (str, ArrayFormula))
        }
        if coordinates != reference_coordinates:
            raise ValueError(f"formula inventory mismatch: {sheet_name}")
        for coordinate in sorted(coordinates):
            actual_formula = formula_sheet[coordinate].value
            reference_formula = reference_formula_sheet[coordinate].value
            if comparable_formula(actual_formula) != comparable_formula(
                reference_formula
            ):
                raise ValueError(
                    f"formula mismatch against reference: {sheet_name}!{coordinate}"
                )
            actual = cached_sheet[coordinate].value
            expected = reference_cached_sheet[coordinate].value
            if (
                isinstance(actual, (int, float))
                and not isinstance(actual, bool)
                and isinstance(expected, (int, float))
                and not isinstance(expected, bool)
            ):
                matches = math.isclose(float(actual), float(expected), abs_tol=1e-9)
            else:
                matches = actual == expected
            if not matches:
                raise ValueError(
                    f"cached formula result mismatch: {sheet_name}!{coordinate}"
                )


def clear_orphan_table_formulas(workbook: Any) -> None:
    for worksheet in workbook.worksheets:
        table_ranges = [
            range_boundaries(worksheet.tables[name].ref)
            for name in worksheet.tables
        ]
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type != "f" or not isinstance(cell.value, str):
                    continue
                if "@" not in cell.value and "[#This Row]," not in cell.value:
                    continue
                inside_table = any(
                    min_col <= cell.column <= max_col
                    and min_row < cell.row <= max_row
                    for min_col, min_row, max_col, max_row in table_ranges
                )
                if not inside_table:
                    cell.value = None


def copy_style(source: Any, target: Any) -> None:
    target.font = copy.copy(source.font)
    target.fill = copy.copy(source.fill)
    target.border = copy.copy(source.border)
    target.alignment = copy.copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy.copy(source.protection)


def wrapped_line_count(value: str, column_width: float) -> int:
    lines = value.splitlines() or [""]
    return sum(
        max(
            1,
            math.ceil(
                sum(
                    2 if unicodedata.east_asian_width(character) in {"W", "F"} else 1
                    for character in line
                )
                / column_width
            ),
        )
        for line in lines
    )


def effective_cell_width(worksheet: Any, cell: Any) -> float:
    for merged in worksheet.merged_cells.ranges:
        if cell.coordinate in merged:
            return sum(
                worksheet.column_dimensions[get_column_letter(column)].width or 8.43
                for column in range(merged.min_col, merged.max_col + 1)
            )
    return worksheet.column_dimensions[get_column_letter(cell.column)].width or 8.43


def fill_table(workbook: Any, table_name: str, rows: list[dict[str, object]]) -> None:
    worksheet, table = table_index(workbook)[table_name]
    min_col, min_row, max_col, old_max_row = range_boundaries(table.ref)
    prototype_row = min_row + 1
    headers = [
        worksheet.cell(min_row, column).value
        for column in range(min_col, max_col + 1)
    ]
    if not all(isinstance(header, str) for header in headers):
        raise ValueError(f"invalid table header: {table_name}")
    expected_headers = TABLE_HEADERS[table_name]
    if headers != expected_headers:
        raise ValueError(
            f"template header mismatch in {table_name}: "
            f"expected {expected_headers}, got {headers}"
        )
    metadata_headers = [column.name for column in table.tableColumns]
    if metadata_headers != expected_headers:
        raise ValueError(
            f"template table metadata mismatch in {table_name}: "
            f"expected {expected_headers}, got {metadata_headers}"
        )

    prototypes = [
        worksheet.cell(prototype_row, column)
        for column in range(min_col, max_col + 1)
    ]
    formulas = {
        offset: (normalize_table_formula(cell.value), isinstance(cell.value, ArrayFormula))
        for offset, cell in enumerate(prototypes)
        if cell.data_type == "f" and isinstance(cell.value, (str, ArrayFormula))
    }
    expected_formula_headers = FORMULA_HEADERS.get(table_name, set())
    actual_formula_headers = {headers[offset] for offset in formulas}
    if actual_formula_headers != expected_formula_headers:
        raise ValueError(f"formula prototype mismatch in {table_name}")
    for column_offset, column in enumerate(table.tableColumns):
        specification = formulas.get(column_offset)
        if specification is None:
            column.calculatedColumnFormula = None
            continue
        formula, is_array = specification
        column.calculatedColumnFormula = TableFormula(
            array=True if is_array else None,
            attr_text=formula.removeprefix("="),
        )

    physical_rows = rows if rows else [{}]
    clear_through = max(old_max_row, min_row + len(physical_rows), prototype_row)
    for row in range(prototype_row, clear_through + 1):
        for column in range(min_col, max_col + 1):
            worksheet.cell(row, column).value = None

    prototype_height = worksheet.row_dimensions[prototype_row].height
    for offset, payload in enumerate(physical_rows, start=1):
        row = min_row + offset
        if prototype_height is not None:
            worksheet.row_dimensions[row].height = prototype_height
        for column_offset, header in enumerate(headers):
            cell = worksheet.cell(row, min_col + column_offset)
            copy_style(prototypes[column_offset], cell)
            if column_offset in formulas:
                formula, is_array = formulas[column_offset]
                translated = Translator(
                    formula,
                    origin=prototypes[column_offset].coordinate,
                ).translate_formula(cell.coordinate)
                cell.value = (
                    ArrayFormula(ref=cell.coordinate, text=translated)
                    if is_array
                    else translated
                )
            else:
                value = None if not rows else safe_text(payload.get(header, ""))
                cell.value = value
                if isinstance(value, str):
                    cell.data_type = "s"
        wrapped_lines = max(
            (
                wrapped_line_count(visible_value, effective_cell_width(worksheet, cell))
                for column_offset, cell in enumerate(
                    worksheet[row][min_col - 1 : max_col]
                )
                for visible_value in [
                    payload.get(headers[column_offset])
                    if column_offset in formulas
                    else cell.value
                ]
                if cell.alignment.wrap_text and isinstance(visible_value, str)
            ),
            default=1,
        )
        worksheet.row_dimensions[row].height = min(
            MAX_EXCEL_ROW_HEIGHT,
            max(
                prototype_height or 15,
                wrapped_lines * WRAPPED_LINE_HEIGHT + WRAPPED_ROW_PADDING,
            ),
        )

    new_max_row = min_row + len(physical_rows)
    for row in range(new_max_row + 1, worksheet.max_row + 1):
        trailing_cells = [
            worksheet.cell(row, column)
            for column in range(min_col, max_col + 1)
        ]
        if any(cell.value not in (None, "") for cell in trailing_cells):
            break
        for cell in trailing_cells:
            cell._style = None
        if worksheet.row_dimensions[row].height == prototype_height:
            worksheet.row_dimensions[row].height = None
    table.ref = (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{new_max_row}"
    )
    if table.autoFilter is not None:
        table.autoFilter.ref = table.ref


def projection_contract(workbook: Any) -> dict[str, dict[str, object]]:
    contract: dict[str, dict[str, object]] = {}
    for table_name, (worksheet, table) in table_index(workbook).items():
        if table_name not in TABLES:
            continue
        min_col, min_row, max_col, _ = range_boundaries(table.ref)
        headers = [
            worksheet.cell(min_row, column).value
            for column in range(min_col, max_col + 1)
        ]
        if not all(isinstance(header, str) for header in headers):
            raise ValueError(f"invalid table header: {table_name}")
        formulas: dict[str, tuple[str, str, bool]] = {}
        styles: dict[str, tuple[object, ...]] = {}
        for offset, header in enumerate(headers):
            cell = worksheet.cell(min_row + 1, min_col + offset)
            styles[str(header)] = style_signature(cell)
            if cell.data_type == "f" and isinstance(cell.value, (str, ArrayFormula)):
                formula = normalize_table_formula(cell.value)
                if "_xlfn._xlws." in formula:
                    raise ValueError(
                        f"unsupported dynamic worksheet formula in {table_name}.{header}"
                    )
                formulas[str(header)] = (
                    cell.coordinate,
                    formula,
                    isinstance(cell.value, ArrayFormula),
                )
        contract[table_name] = {
            "headers": headers,
            "formulas": formulas,
            "styles": styles,
        }
    return contract


def style_signature(cell: Any) -> tuple[object, ...]:
    return (
        copy.copy(cell.font),
        copy.copy(cell.fill),
        copy.copy(cell.border),
        copy.copy(cell.alignment),
        cell.number_format,
        copy.copy(cell.protection),
    )


def _visible_color(color: Any) -> tuple[object, ...] | None:
    if color is None:
        return None
    value = getattr(color, color.type, None)
    if color.type == "rgb" and isinstance(value, str):
        value = value[-6:].upper()
    return (
        color.type,
        value,
        round(float(color.tint or 0), 8),
    )


def _visible_side(side: Any) -> tuple[object, ...] | None:
    if side is None or (side.style is None and side.color is None):
        return None
    return (side.style, _visible_color(side.color))


def visible_style_signature(cell: Any) -> tuple[object, ...]:
    """Compare rendered appearance while tolerating Office font substitution."""
    font = cell.font
    fill = cell.fill
    border = cell.border
    alignment = cell.alignment
    return (
        (
            bool(font.bold),
            bool(font.italic),
            float(font.sz) if font.sz is not None else None,
            font.underline,
            bool(font.strike),
            _visible_color(font.color),
        ),
        (
            fill.patternType,
            _visible_color(fill.fgColor),
            (
                _visible_color(fill.bgColor)
                if fill.patternType not in {None, "solid"}
                else None
            ),
        ),
        tuple(
            _visible_side(getattr(border, name))
            for name in ("left", "right", "top", "bottom", "diagonal")
        ),
        (
            alignment.horizontal or "general",
            alignment.vertical or "bottom",
            int(alignment.textRotation or 0),
            bool(alignment.wrapText),
            bool(alignment.shrinkToFit),
            float(alignment.indent or 0),
        ),
        cell.number_format,
        (bool(cell.protection.locked), bool(cell.protection.hidden)),
    )


def verify_visible_layout(workbook: Any, expected_workbook: Any) -> None:
    for sheet_name in FORMAL_SHEETS:
        worksheet = workbook[sheet_name]
        expected_sheet = expected_workbook[sheet_name]
        max_row = max(worksheet.max_row, expected_sheet.max_row)
        max_column = max(worksheet.max_column, expected_sheet.max_column)
        for row in range(1, max_row + 1):
            actual_height = (
                worksheet.row_dimensions[row].height
                or worksheet.sheet_format.defaultRowHeight
            )
            expected_height = (
                expected_sheet.row_dimensions[row].height
                or expected_sheet.sheet_format.defaultRowHeight
            )
            if actual_height != expected_height:
                raise ValueError(f"row height mismatch: {sheet_name}!{row}")
            for column in range(1, max_column + 1):
                actual = worksheet.cell(row, column)
                expected = expected_sheet.cell(row, column)
                if visible_style_signature(actual) != visible_style_signature(expected):
                    raise ValueError(
                        f"visible style mismatch: {sheet_name}!{actual.coordinate}"
                    )


def verify_print_layout(workbook: Any) -> None:
    """Reject pagination that makes long sheets unreadable when printed."""
    for sheet_name in FORMAL_SHEETS:
        page_setup = workbook[sheet_name].page_setup
        if page_setup.fitToWidth != 1 or page_setup.fitToHeight != 0:
            raise ValueError(f"print layout mismatch: {sheet_name}")


def verify_workbook(
    path: Path,
    expected: dict[str, list[dict[str, object]]],
    contract: dict[str, dict[str, object]],
    *,
    require_recalculation: bool = True,
    verify_styles: bool = True,
) -> None:
    workbook = openpyxl.load_workbook(path, data_only=False, read_only=False)
    try:
        if tuple(workbook.sheetnames) != FORMAL_SHEETS:
            raise ValueError("formal workbook sheet contract changed")
        index = table_index(workbook)
        if require_recalculation and workbook.calculation.calcMode != "auto":
            raise ValueError("workbook recalculation is not enabled")
        for table_name, rows in expected.items():
            worksheet, table = index[table_name]
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            if max_row - min_row != max(1, len(rows)):
                raise ValueError(f"table row count mismatch: {table_name}")
            headers = [
                worksheet.cell(min_row, column).value
                for column in range(min_col, max_col + 1)
            ]
            if headers != TABLE_HEADERS[table_name]:
                raise ValueError(f"table header mismatch: {table_name}")
            specification = contract[table_name]
            if headers != specification["headers"]:
                raise ValueError(f"table headers changed: {table_name}")
            formulas = specification["formulas"]
            styles = specification["styles"]
            assert isinstance(formulas, dict) and isinstance(styles, dict)
            physical_rows = rows if rows else [{}]
            for row_offset, payload in enumerate(physical_rows, start=1):
                for column_offset, header in enumerate(headers):
                    cell = worksheet.cell(min_row + row_offset, min_col + column_offset)
                    if verify_styles and style_signature(cell) != styles[header]:
                        raise ValueError(
                            f"prototype style changed in {table_name}.{header}"
                        )
                    if header in formulas:
                        origin, prototype, is_array = formulas[header]
                        expected_formula = Translator(
                            prototype,
                            origin=origin,
                        ).translate_formula(cell.coordinate)
                        actual_formula = (
                            formula_text(cell.value)
                            if isinstance(cell.value, (str, ArrayFormula))
                            else None
                        )
                        formula_kind_matches = (
                            isinstance(cell.value, ArrayFormula)
                            and cell.value.ref
                            in {cell.coordinate, f"{cell.coordinate}:{cell.coordinate}"}
                            if is_array
                            else isinstance(cell.value, str)
                        )
                        if (
                            comparable_formula(actual_formula or "")
                            != comparable_formula(expected_formula)
                            or cell.data_type != "f"
                            or not formula_kind_matches
                        ):
                            raise ValueError(
                                f"formula mismatch in {table_name}.{header}"
                            )
                    else:
                        expected_value = (
                            None if not rows else safe_text(payload.get(header, ""))
                        )
                        if expected_value == "":
                            expected_value = None
                        if cell.value != expected_value:
                            raise ValueError(
                                f"projected value mismatch in {table_name}.{header}"
                            )
                        if isinstance(expected_value, str) and cell.data_type != "s":
                            raise ValueError(
                                f"projected text type mismatch in {table_name}.{header}"
                            )
            calculated_headers = {
                column.name
                for column in table.tableColumns
                if column.calculatedColumnFormula is not None
                and column.calculatedColumnFormula.text
            }
            if calculated_headers != FORMULA_HEADERS[table_name]:
                raise ValueError(f"calculated column mismatch in {table_name}")
            formula_columns = {
                column.name: column.calculatedColumnFormula
                for column in table.tableColumns
                if column.calculatedColumnFormula is not None
            }
            for header, (_, _, is_array) in formulas.items():
                metadata_formula = formula_columns[header]
                if (
                    comparable_formula("=" + str(metadata_formula.text))
                    != comparable_formula(formulas[header][1])
                ):
                    raise ValueError(
                        f"calculated column formula mismatch in {table_name}.{header}"
                    )
                if (metadata_formula.array is True) != is_array:
                    raise ValueError(
                        f"calculated column array mismatch in {table_name}.{header}"
                    )
            if table.autoFilter is not None and table.autoFilter.ref != table.ref:
                raise ValueError(f"autoFilter range mismatch: {table_name}")
        for sheet_name in PROTECTED_SHEETS:
            if not workbook[sheet_name].protection.sheet:
                raise ValueError(f"worksheet protection is missing: {sheet_name}")
    finally:
        workbook.close()


def verify_static_authority(workbook: Any, template_workbook: Any) -> None:
    """Verify immutable catalog, parameter and summary inputs/formulas."""
    workbook_index = table_index(workbook)
    template_index = table_index(template_workbook)
    for table_name in sorted(FORMAL_TABLES - set(TABLES)):
        worksheet, table = workbook_index[table_name]
        template_sheet, template_table = template_index[table_name]
        if table.ref != template_table.ref:
            raise ValueError(f"static table range mismatch: {table_name}")
        bounds = range_boundaries(table.ref)
        template_bounds = range_boundaries(template_table.ref)
        if bounds != template_bounds:
            raise ValueError(f"static table bounds mismatch: {table_name}")
        if [column.name for column in table.tableColumns] != [
            column.name for column in template_table.tableColumns
        ]:
            raise ValueError(f"static table metadata mismatch: {table_name}")
        min_col, min_row, max_col, max_row = bounds
        for row in range(min_row, max_row + 1):
            for column in range(min_col, max_col + 1):
                actual = worksheet.cell(row, column)
                expected = template_sheet.cell(row, column)
                if expected.data_type == "f" and isinstance(
                    expected.value, (str, ArrayFormula)
                ):
                    if actual.data_type != "f" or not isinstance(
                        actual.value, (str, ArrayFormula)
                    ):
                        raise ValueError(f"static formula is missing: {table_name}")
                    if comparable_formula(actual.value) != comparable_formula(
                        expected.value
                    ):
                        raise ValueError(f"static formula mismatch: {table_name}")
                elif actual.value != expected.value:
                    raise ValueError(f"static value mismatch: {table_name}")


def verify_worksheet_authority(workbook: Any, template_workbook: Any) -> None:
    """Verify worksheet controls that remain authoritative after Office roundtrip."""
    for sheet_name in FORMAL_SHEETS:
        worksheet = workbook[sheet_name]
        template_sheet = template_workbook[sheet_name]
        if ET.tostring(worksheet.data_validations.to_tree()) != ET.tostring(
            template_sheet.data_validations.to_tree()
        ):
            raise ValueError(f"data validation metadata mismatch: {sheet_name}")
        if ET.tostring(worksheet.protection.to_tree()) != ET.tostring(
            template_sheet.protection.to_tree()
        ):
            raise ValueError(f"worksheet protection metadata mismatch: {sheet_name}")


def audit_calculated_workbook(
    path: Path,
    template_path: Path,
    scope: dict[str, Any],
    delivery: dict[str, Any],
    engine: Any,
) -> WorkbookAudit:
    """Verify projected inputs and reread every calculation authority/result.

    The office engine, rather than Python, remains responsible for evaluating
    formulas. This function only proves that the verified output still contains
    the approved projection and that all authoritative cached results are usable.
    """
    stack = ExitStack()
    try:
        temporary_root = Path(
            stack.enter_context(tempfile.TemporaryDirectory(prefix="ai-sow-audit-"))
        )
        expected_layout_path = temporary_root / "expected-layout.xlsx"
        reference_path = temporary_root / "reference.xlsx"
        write_workbook(template_path, scope, delivery, expected_layout_path)
        recalculate_workbook(expected_layout_path, reference_path, engine)
        expected_layout_workbook = openpyxl.load_workbook(
            expected_layout_path, data_only=False, read_only=False
        )
        reference_formula_workbook = openpyxl.load_workbook(
            reference_path, data_only=False, read_only=False
        )
        reference_cached_workbook = openpyxl.load_workbook(
            reference_path, data_only=True, read_only=False
        )
        template_workbook = openpyxl.load_workbook(
            template_path, data_only=False, read_only=False
        )
        formula_workbook = openpyxl.load_workbook(
            path, data_only=False, read_only=False
        )
        cached_workbook = openpyxl.load_workbook(
            path, data_only=True, read_only=False
        )
        for workbook in (
            expected_layout_workbook,
            reference_formula_workbook,
            reference_cached_workbook,
            template_workbook,
            formula_workbook,
            cached_workbook,
        ):
            stack.callback(workbook.close)
        for workbook in (formula_workbook, cached_workbook):
            if tuple(workbook.sheetnames) != FORMAL_SHEETS:
                raise ValueError("formal workbook sheet contract changed")
            if set(table_index(workbook)) != FORMAL_TABLES:
                raise ValueError("formal workbook table contract changed")

        expected = build_rows(
            scope,
            delivery,
            base_unit_name_map(formula_workbook),
        )
        verify_workbook(
            path,
            expected,
            projection_contract(template_workbook),
            require_recalculation=False,
            verify_styles=False,
        )
        verify_static_authority(formula_workbook, template_workbook)
        verify_worksheet_authority(formula_workbook, template_workbook)
        verify_visible_layout(formula_workbook, expected_layout_workbook)
        verify_print_layout(formula_workbook)
        formula_index = table_index(formula_workbook)
        cached_index = table_index(cached_workbook)
        cached_errors = formula_errors(cached_workbook)
        if cached_errors:
            raise ValueError(
                "calculated workbook contains formula errors: "
                + ", ".join(cached_errors)
            )
        reference_errors = formula_errors(reference_cached_workbook)
        if reference_errors:
            raise ValueError(
                "reference calculation contains formula errors: "
                + ", ".join(reference_errors)
            )
        verify_formula_cache_results(
            formula_workbook,
            cached_workbook,
            reference_formula_workbook,
            reference_cached_workbook,
        )

        for table_name in TABLES:
            formula_sheet, formula_table = formula_index[table_name]
            cached_sheet, cached_table = cached_index[table_name]
            formula_bounds = range_boundaries(formula_table.ref)
            cached_bounds = range_boundaries(cached_table.ref)
            if formula_bounds != cached_bounds:
                raise ValueError(f"calculated table range changed: {table_name}")
            min_col, min_row, max_col, max_row = formula_bounds
            headers = [
                formula_sheet.cell(min_row, column).value
                for column in range(min_col, max_col + 1)
            ]
            if headers != TABLE_HEADERS[table_name]:
                raise ValueError(f"calculated table header mismatch: {table_name}")
            if max_row - min_row != len(expected[table_name]):
                raise ValueError(f"calculated table row count mismatch: {table_name}")
            for row_offset, payload in enumerate(expected[table_name], start=1):
                for column_offset, header in enumerate(headers):
                    formula_cell = formula_sheet.cell(
                        min_row + row_offset, min_col + column_offset
                    )
                    cached_cell = cached_sheet.cell(
                        min_row + row_offset, min_col + column_offset
                    )
                    if header in FORMULA_HEADERS[table_name]:
                        if formula_cell.data_type != "f" or not isinstance(
                            formula_cell.value, (str, ArrayFormula)
                        ):
                            raise ValueError(
                                f"calculated formula is missing: {table_name}.{header}"
                            )
                    else:
                        expected_value = safe_text(payload.get(str(header), ""))
                        if expected_value == "":
                            expected_value = None
                        if formula_cell.value != expected_value:
                            raise ValueError(
                                f"calculated projection changed: {table_name}.{header}"
                            )
                        if cached_cell.value != expected_value:
                            raise ValueError(
                                f"cached projection changed: {table_name}.{header}"
                            )

        story_records = table_records(cached_workbook, "SOWStoryTable")
        task_records = table_records(cached_workbook, "TaskTable")
        task_names_by_story: dict[str, list[str]] = {}
        task_days: list[float] = []
        for record in task_records:
            story_path = record["所属故事"]
            task_name = record["任务名称"]
            if not isinstance(story_path, str) or not isinstance(task_name, str):
                raise ValueError("calculated task projection is incomplete")
            task_names_by_story.setdefault(story_path, []).append(task_name)
            for header in ("M档标准人天", "复杂度系数", "任务人天", "SIT支持人天"):
                value = require_number(record[header], f"TaskTable.{header}")
                if value < 0:
                    raise ValueError(f"calculated task value is negative: {header}")
                if header == "任务人天":
                    task_days.append(value)
            if record["校验结果"] != "通过":
                raise ValueError("calculated task validation did not pass")

        for record in story_records:
            story_name = record["故事"]
            if not isinstance(story_name, str) or not story_name:
                raise ValueError("calculated story name is missing")
            matching_tasks = [
                task
                for task in task_records
                if task.get("所属故事") == story_name
            ]
            expected_task_list = "\n".join(
                f"• [{task['任务类型']}/{task['工作方式']}/{task['复杂度']}] "
                f"{task['任务名称']}"
                for task in matching_tasks
            )
            if not expected_task_list or record["任务列表"] != expected_task_list:
                raise ValueError("calculated story task list changed")
            require_number(record["故事人天"], "SOWStoryTable.故事人天")
            if record["校验结果"] != "通过":
                raise ValueError("calculated story validation did not pass")

        summary_formula_records = table_records(
            formula_workbook, "ProjectSummaryTable"
        )
        summary_records = table_records(cached_workbook, "ProjectSummaryTable")
        if [*summary_records[0].keys()] != SUMMARY_HEADERS:
            raise ValueError("summary table header contract changed")
        if [record["工作量项"] for record in summary_records] != list(SUMMARY_LABELS):
            raise ValueError("summary row contract changed")
        for record in summary_formula_records:
            if not isinstance(record["人天"], (str, ArrayFormula)):
                raise ValueError("summary formula is missing")
        summary_values = [
            require_number(record["人天"], f"ProjectSummaryTable.{record['工作量项']}")
            for record in summary_records
        ]
        direct_days, sit_days, uat_days, total_days = summary_values
        if not math.isclose(direct_days, sum(task_days), abs_tol=1e-9):
            raise ValueError("summary direct days do not match task results")
        if not math.isclose(total_days, direct_days + sit_days + uat_days, abs_tol=1e-9):
            raise ValueError("summary total days do not match component results")

        catalog_records = table_records(cached_workbook, "BaseUnitCatalogTable")
        if not catalog_records or [*catalog_records[0].keys()] != CATALOG_HEADERS:
            raise ValueError("base-unit catalog contract changed")
        catalog_ids: set[str] = set()
        for record in catalog_records:
            for header in CATALOG_HEADERS:
                if record[header] in (None, ""):
                    raise ValueError(f"base-unit catalog value is blank: {header}")
            unit_id = record["基础单元ID"]
            if not isinstance(unit_id, str) or unit_id in catalog_ids:
                raise ValueError("base-unit catalog ID is invalid or duplicated")
            catalog_ids.add(unit_id)
            available_modes = 0
            for header in ("新建M档人天", "调整M档人天", "接入复用M档人天"):
                value = record[header]
                if value == "❌":
                    continue
                if require_number(value, f"BaseUnitCatalogTable.{unit_id}.{header}") <= 0:
                    raise ValueError("base-unit catalog person-days must be positive")
                available_modes += 1
            if available_modes == 0:
                raise ValueError(
                    f"base-unit catalog row has no available work mode: {unit_id}"
                )

        parameter_records = table_records(cached_workbook, "ProjectParameterTable")
        if [*parameter_records[0].keys()] != PARAMETER_HEADERS:
            raise ValueError("project parameter header contract changed")
        parameter_codes: set[str] = set()
        parameter_statuses: list[tuple[str, str]] = []
        for record in parameter_records:
            code = record["参数代码"]
            status = record["验证状态/说明"]
            if not isinstance(code, str) or not code or code in parameter_codes:
                raise ValueError("project parameter code is invalid or duplicated")
            if not isinstance(status, str) or not status.strip():
                raise ValueError(f"project parameter status is missing: {code}")
            for header in PARAMETER_HEADERS[1:-1]:
                if record[header] in (None, ""):
                    raise ValueError(f"project parameter value is blank: {code}.{header}")
            require_number(record["值"], f"ProjectParameterTable.{code}.值")
            parameter_codes.add(code)
            parameter_statuses.append((code, status))

        return WorkbookAudit(
            trust_state="VERIFIED",
            story_count=len(story_records),
            task_count=len(task_records),
            direct_days=direct_days,
            sit_days=sit_days,
            uat_days=uat_days,
            total_days=total_days,
            parameter_statuses=tuple(parameter_statuses),
            formula_errors=(),
            engine_name=str(engine.name),
            engine_version=str(engine.version),
        )
    finally:
        stack.close()


def write_workbook(
    template_path: Path,
    scope: dict[str, Any],
    delivery: dict[str, Any],
    output_path: Path,
) -> WorkbookAudit:
    workbook = openpyxl.load_workbook(
        template_path,
        data_only=False,
        read_only=False,
    )
    if workbook.calculation is None:
        workbook.calculation = CalcProperties()
    try:
        table_index(workbook)
        rows = build_rows(scope, delivery, base_unit_name_map(workbook))
        contract = projection_contract(workbook)
        clear_orphan_table_formulas(workbook)
        for table_name in TABLES:
            fill_table(workbook, table_name, rows[table_name])
        # A blank fitToHeight is interpreted as one page by LibreOffice when
        # fit-to-page is enabled, which compresses long Task sheets until the
        # text is unreadable. Zero means unlimited vertical pages while the
        # template's one-page-wide layout remains authoritative.
        for worksheet in workbook.worksheets:
            worksheet.page_setup.fitToHeight = 0
        workbook.calculation.calcMode = "auto"
        workbook.calculation.calcOnSave = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.fullCalcOnLoad = True
        workbook.properties.created = DETERMINISTIC_TIME
        workbook.properties.modified = DETERMINISTIC_TIME
        workbook.save(output_path)
    finally:
        workbook.close()
    normalize_xlsx(output_path)
    verify_workbook(output_path, rows, contract)
    return WorkbookAudit(
        trust_state="CANDIDATE",
        story_count=len(rows["SOWStoryTable"]),
        task_count=len(rows["TaskTable"]),
        direct_days=None,
        sit_days=None,
        uat_days=None,
        total_days=None,
        parameter_statuses=(),
        formula_errors=(),
        engine_name=None,
        engine_version=None,
    )
