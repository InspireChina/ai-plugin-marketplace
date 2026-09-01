from __future__ import annotations

import hashlib
from collections import defaultdict
from collections.abc import Mapping, Sequence
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils.cell import range_boundaries

from contracts import (
    canonical_json_bytes,
    load_schema_registry,
    sha256_bytes,
    validate_contract,
    validate_id_decisions,
)
from models import (
    BaseUnitRule,
    DeliveryCompilation,
    Diagnostic,
    ImpactPlan,
    TemplateCatalog,
)
from scope_compiler import impact_plan_sha256


SKILL_ROOT = Path(__file__).resolve().parents[1]
SCHEMA_REGISTRY = load_schema_registry(SKILL_ROOT)
COLLECTION_TYPES = {
    "stories": ("STORY", "storyId"),
    "acceptanceCriteria": ("ACCEPTANCE_CRITERION", "acceptanceCriterionId"),
    "tasks": ("TASK", "taskId"),
    "dependencies": ("DEPENDENCY", "dependencyId"),
}
CATALOG_HEADERS = {
    "任务族ID",
    "任务族名称",
    "基础单元ID",
    "基础单元名称",
    "计数口径",
    "包含内容",
    "不包含内容",
    "新建M档人天",
    "调整M档人天",
    "接入复用M档人天",
    "S标准",
    "M标准",
    "L标准",
    "X/拆分条件",
}
PARAMETER_HEADERS = {"参数代码", "名称", "值", "单位", "适用范围", "验证状态/说明"}
MODE_EFFORT_HEADERS = {
    "新建": "新建M档人天",
    "调整": "调整M档人天",
    "接入复用": "接入复用M档人天",
}
COMPLEXITIES = ("S", "M", "L")
CALIBRATED_PARAMETER_STATUSES = {"固定规则", "已校准", "已批准"}
CLARIFICATION_FIELDS = frozenset({"name", "description", "rationale"})


def _diagnostic(code: str, message: str, path: str = "") -> Diagnostic:
    return Diagnostic(code=code, message=message, path=path, details={})


def _sort_diagnostics(values: Sequence[Diagnostic]) -> tuple[Diagnostic, ...]:
    return tuple(sorted(values, key=lambda item: (item.path, item.code, item.message)))


def _mappings(value: object) -> list[Mapping[str, object]]:
    return [item for item in value if isinstance(item, Mapping)] if isinstance(value, list) else []


def _ids(value: object) -> tuple[str, ...]:
    return tuple(item for item in value if isinstance(item, str)) if isinstance(value, list) else ()


def _table_rows(workbook: Any, table_name: str) -> list[dict[str, Any]]:
    for worksheet in workbook.worksheets:
        if table_name not in worksheet.tables:
            continue
        table = worksheet.tables[table_name]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [
            worksheet.cell(min_row, column).value
            for column in range(min_col, max_col + 1)
        ]
        return [
            {
                str(header): worksheet.cell(row, column).value
                for header, column in zip(
                    headers, range(min_col, max_col + 1), strict=True
                )
            }
            for row in range(min_row + 1, max_row + 1)
        ]
    raise ValueError(f"template table is missing: {table_name}")


def _require_headers(rows: list[dict[str, Any]], expected: set[str], table_name: str) -> None:
    actual = set(rows[0]) if rows else set()
    if actual != expected:
        raise ValueError(f"template table headers are invalid: {table_name}")


def _require_text(row: Mapping[str, object], header: str, subject: str) -> str:
    value = row.get(header)
    if not isinstance(value, str) or not value.strip():
        raise ValueError(f"{subject} must define non-empty {header}")
    return value


def _usable_mode(row: Mapping[str, object], header: str, subject: str) -> bool:
    value = row.get(header)
    if value == "❌":
        return False
    if isinstance(value, bool) or not isinstance(value, (int, float)) or value <= 0:
        raise ValueError(f"{subject} {header} must be a positive number or ❌")
    return True


def _validate_complexity_parameters(rows: list[dict[str, Any]]) -> None:
    seen: set[str] = set()
    for row in rows:
        code = row.get("参数代码")
        if not isinstance(code, str) or not code.startswith("K_COMPLEXITY_"):
            continue
        level = code.removeprefix("K_COMPLEXITY_")
        value = row.get("值")
        if (
            level not in COMPLEXITIES
            or level in seen
            or isinstance(value, bool)
            or not isinstance(value, (int, float))
            or value <= 0
            or row.get("验证状态/说明") not in CALIBRATED_PARAMETER_STATUSES
        ):
            raise ValueError(f"complexity parameter is invalid: {code}")
        seen.add(level)
    if seen != set(COMPLEXITIES):
        raise ValueError("template must define calibrated S/M/L complexity factors")


def read_template_catalog(template_path: Path) -> TemplateCatalog:
    payload = Path(template_path).read_bytes()
    workbook = openpyxl.load_workbook(template_path, data_only=False, read_only=False)
    try:
        catalog_rows = _table_rows(workbook, "BaseUnitCatalogTable")
        parameter_rows = _table_rows(workbook, "ProjectParameterTable")
        _require_headers(catalog_rows, CATALOG_HEADERS, "BaseUnitCatalogTable")
        _require_headers(parameter_rows, PARAMETER_HEADERS, "ProjectParameterTable")
        base_units: dict[str, BaseUnitRule] = {}
        family_ids: set[str] = set()
        for index, row in enumerate(catalog_rows, 1):
            subject = f"base-unit catalog row {index}"
            base_unit = _require_text(row, "基础单元ID", subject)
            if base_unit in base_units:
                raise ValueError(f"base-unit catalog ID is duplicated: {base_unit}")
            modes = tuple(
                mode
                for mode, header in MODE_EFFORT_HEADERS.items()
                if _usable_mode(row, header, subject)
            )
            if not modes:
                raise ValueError(f"base-unit has no work mode: {base_unit}")
            family_id = _require_text(row, "任务族ID", subject)
            family_ids.add(family_id)
            base_units[base_unit] = BaseUnitRule(
                base_unit=base_unit,
                name=_require_text(row, "基础单元名称", subject),
                task_family_id=family_id,
                task_family=_require_text(row, "任务族名称", subject),
                count_rule=_require_text(row, "计数口径", subject),
                includes=_require_text(row, "包含内容", subject),
                excludes=_require_text(row, "不包含内容", subject),
                allowed_work_modes=modes,
                allowed_complexities=("S", "M", "L"),
                complexity_standards={
                    level: _require_text(row, f"{level}标准", subject)
                    for level in COMPLEXITIES
                },
                split_rule=_require_text(row, "X/拆分条件", subject),
            )
        if len(base_units) != 37 or len(family_ids) != 13:
            raise ValueError("template must define exactly 37 base units and 13 task families")
        _validate_complexity_parameters(parameter_rows)
        return TemplateCatalog(
            template_sha256=hashlib.sha256(payload).hexdigest(),
            base_units=base_units,
        )
    finally:
        workbook.close()


def _all_objects(bundle: Mapping[str, object] | None) -> dict[tuple[str, str], Mapping[str, object]]:
    result: dict[tuple[str, str], Mapping[str, object]] = {}
    if bundle is None:
        return result
    for collection, (object_type, id_field) in COLLECTION_TYPES.items():
        for item in _mappings(bundle.get(collection)):
            object_id = item.get(id_field)
            if isinstance(object_id, str):
                result[(object_type, object_id)] = item
    return result


def _without_clarifications(value: Mapping[str, object]) -> dict[str, object]:
    return {key: item for key, item in value.items() if key not in CLARIFICATION_FIELDS}


def _validate_id_ledger(
    candidate: Mapping[str, object],
    previous: Mapping[str, object] | None,
    id_decisions: object,
) -> list[Diagnostic]:
    diagnostics = list(validate_id_decisions(id_decisions, SCHEMA_REGISTRY))
    if not isinstance(id_decisions, Mapping):
        return diagnostics
    candidate_objects = _all_objects(candidate)
    previous_objects = _all_objects(previous)
    decisions = {
        (str(item.get("objectType")), str(item.get("objectId"))): item
        for item in _mappings(id_decisions.get("decisions"))
    }
    for identity, item in candidate_objects.items():
        decision = decisions.get(identity)
        if decision is None:
            diagnostics.append(
                _diagnostic(
                    "DELIVERY_ID_DECISION_MISSING",
                    "Delivery 切片中的每个对象必须有 ID 决定。",
                    f"/{identity[0]}/{identity[1]}",
                )
            )
            continue
        disposition = decision.get("disposition")
        previous_id = decision.get("previousId")
        previous_item = (
            previous_objects.get((identity[0], str(previous_id)))
            if isinstance(previous_id, str)
            else None
        )
        if disposition == "NEW" and identity in previous_objects:
            diagnostics.append(
                _diagnostic("DELIVERY_ID_NEW_ALREADY_EXISTS", "NEW 不得复用既有 ID。")
            )
        elif disposition == "UNCHANGED" and (
            previous_item is None
            or canonical_json_bytes(previous_item) != canonical_json_bytes(item)
        ):
            diagnostics.append(
                _diagnostic(
                    "DELIVERY_ID_UNCHANGED_SEMANTICS_DIFFER",
                    "UNCHANGED 对象的规范语义必须完全相同。",
                )
            )
        elif disposition == "CLARIFIED" and (
            previous_item is None
            or canonical_json_bytes(_without_clarifications(previous_item))
            != canonical_json_bytes(_without_clarifications(item))
        ):
            diagnostics.append(
                _diagnostic(
                    "DELIVERY_ID_CLARIFICATION_CHANGED_MEANING",
                    "CLARIFIED 不得改变交付含义。",
                )
            )
        elif disposition == "CHANGED" and previous_item is None:
            diagnostics.append(
                _diagnostic(
                    "DELIVERY_ID_CHANGED_PREVIOUS_MISSING",
                    "CHANGED 必须引用存在的旧对象。",
                )
            )
    for identity in decisions:
        if identity not in candidate_objects:
            diagnostics.append(
                _diagnostic(
                    "DELIVERY_ID_DECISION_ORPHAN",
                    "ID 决定引用了切片中不存在的对象。",
                )
            )
    return diagnostics


def _scope_indices(scope: Mapping[str, object]) -> dict[str, object]:
    features = {
        str(item.get("featureId")): item for item in _mappings(scope.get("features"))
    }
    return {
        "features": features,
        "designItems": {
            str(item.get("designItemId")): item
            for item in _mappings(scope.get("designItems"))
        },
        "integrations": {
            str(item.get("integrationId")): item
            for item in _mappings(scope.get("integrations"))
        },
        "nfrs": {
            str(item.get("nfrId")): item for item in _mappings(scope.get("nfrs"))
        },
        "effectiveStartItems": {
            str(item.get("effectiveStartItemId")): item
            for item in _mappings(scope.get("effectiveStartItems"))
        },
    }


def _has_cycle(task_ids: set[str], edges: set[tuple[str, str]]) -> bool:
    successors: defaultdict[str, set[str]] = defaultdict(set)
    indegree = {task_id: 0 for task_id in task_ids}
    for predecessor, successor in edges:
        if successor not in successors[predecessor]:
            successors[predecessor].add(successor)
            indegree[successor] = indegree.get(successor, 0) + 1
    ready = sorted(task_id for task_id, degree in indegree.items() if degree == 0)
    visited = 0
    while ready:
        task_id = ready.pop(0)
        visited += 1
        for successor in sorted(successors.get(task_id, set())):
            indegree[successor] -= 1
            if indegree[successor] == 0:
                ready.append(successor)
                ready.sort()
    return visited != len(task_ids)


def _validate_delivery(
    bundle: Mapping[str, object],
    scope: Mapping[str, object],
    catalog: TemplateCatalog,
) -> list[Diagnostic]:
    diagnostics = list(
        validate_contract(bundle, "delivery-bundle.schema.json", SCHEMA_REGISTRY)
    )
    indices = _scope_indices(scope)
    features = indices["features"]
    stories = {
        str(item.get("storyId")): item for item in _mappings(bundle.get("stories"))
    }
    criteria = {
        str(item.get("acceptanceCriterionId")): item
        for item in _mappings(bundle.get("acceptanceCriteria"))
    }
    tasks = {str(item.get("taskId")): item for item in _mappings(bundle.get("tasks"))}

    for feature_id, feature in features.items():
        decision = feature.get("scopeDecision")
        requires_story = isinstance(decision, Mapping) and decision.get("decision") == "IN_SCOPE"
        if requires_story and not any(
            feature_id in _ids(story.get("featureIds")) for story in stories.values()
        ):
            diagnostics.append(
                _diagnostic(
                    "DELIVERY_FEATURE_STORY_MISSING",
                    "每个 IN_SCOPE Feature 必须由至少一个 Story 交付。",
                    f"/features/{feature_id}",
                )
            )
    for story_id, story in stories.items():
        if any(feature_id not in features for feature_id in _ids(story.get("featureIds"))):
            diagnostics.append(
                _diagnostic("DELIVERY_STORY_FEATURE_UNKNOWN", "Story 引用了未知 Feature。")
            )
        story_criteria = [
            item for item in criteria.values() if item.get("storyId") == story_id
        ]
        if not story_criteria:
            diagnostics.append(
                _diagnostic("DELIVERY_STORY_AC_MISSING", "每个 Story 必须至少有一条 AC。")
            )
        sequences = sorted(
            item.get("sequence") for item in story_criteria if isinstance(item.get("sequence"), int)
        )
        if sequences and sequences != list(range(1, len(sequences) + 1)):
            diagnostics.append(
                _diagnostic("DELIVERY_AC_SEQUENCE_INVALID", "Story 内 AC 顺序必须从 1 连续递增。")
            )

    covered_criteria: set[str] = set()
    integration_tasks: defaultdict[str, list[str]] = defaultdict(list)
    design_required = {
        identifier
        for key in ("integrations", "nfrs")
        for identifier, item in indices[key].items()
        if isinstance(item, Mapping) and item.get("status") == "DESIGN_REQUIRED"
    }
    designed: set[str] = set()
    edges: set[tuple[str, str]] = set()
    for task_id, task in tasks.items():
        story_id = task.get("storyId")
        story = stories.get(str(story_id))
        if story is None:
            diagnostics.append(
                _diagnostic("DELIVERY_TASK_STORY_UNKNOWN", "Task 引用了未知 Story。")
            )
            continue
        for criterion_id in _ids(task.get("acceptanceCriterionIds")):
            criterion = criteria.get(criterion_id)
            if criterion is None or criterion.get("storyId") != story_id:
                diagnostics.append(
                    _diagnostic(
                        "DELIVERY_TASK_AC_WRONG_STORY",
                        "Task 只能覆盖同一 Story 的 AC。",
                    )
                )
            else:
                covered_criteria.add(criterion_id)
        for field, known in (
            ("designItemIds", indices["designItems"]),
            ("integrationIds", indices["integrations"]),
            ("nfrIds", indices["nfrs"]),
        ):
            if any(identifier not in known for identifier in _ids(task.get(field))):
                diagnostics.append(
                    _diagnostic(
                        "DELIVERY_SCOPE_REF_UNKNOWN",
                        "Task 引用了 Scope 中不存在的设计对象。",
                    )
                )
        for integration_id in _ids(task.get("integrationIds")):
            integration_tasks[integration_id].append(task_id)
        if task.get("taskKind") == "DESIGN":
            task_design_refs = set(_ids(task.get("integrationIds"))) | set(
                _ids(task.get("nfrIds"))
            )
            if not task_design_refs & design_required:
                diagnostics.append(
                    _diagnostic(
                        "DELIVERY_DESIGN_TASK_UNBOUND",
                        "DESIGN Task 必须落实一个 DESIGN_REQUIRED 对象。",
                    )
                )
            designed.update(task_design_refs & design_required)
        base_unit = task.get("baseUnit")
        rule = catalog.base_units.get(str(base_unit))
        if rule is None:
            diagnostics.append(
                _diagnostic(
                    "DELIVERY_BASE_UNIT_UNKNOWN",
                    "Task 基础单元不在当前模板目录中。",
                    f"/tasks/{task_id}/baseUnit",
                )
            )
        else:
            if task.get("workMode") not in rule.allowed_work_modes:
                diagnostics.append(
                    _diagnostic(
                        "DELIVERY_WORK_MODE_UNAVAILABLE",
                        "模板未为该基础单元配置所选工作模式。",
                    )
                )
            if task.get("complexity") not in rule.allowed_complexities:
                diagnostics.append(
                    _diagnostic("DELIVERY_COMPLEXITY_INVALID", "Task 复杂度不受支持。")
                )
        if any(token in str(task.get("name", "")) for token in ("不限", "其他支持", "持续支持")):
            diagnostics.append(
                _diagnostic(
                    "DELIVERY_OPEN_ENDED_SUPPORT_FORBIDDEN",
                    "Task 不得表达无边界的持续或其他支持。",
                )
            )
        evidence = task.get("workModeEvidence")
        matched = task.get("matchedEffectiveStartItemId")
        if isinstance(evidence, Mapping):
            effective = indices["effectiveStartItems"].get(str(matched))
            if (
                effective is None
                or evidence.get("effectiveStartItemId") != matched
                or evidence.get("effectiveStartItemName") != effective.get("name")
            ):
                diagnostics.append(
                    _diagnostic(
                        "DELIVERY_WORK_MODE_EVIDENCE_INVALID",
                        "工作模式证据必须精确匹配 Effective Start。",
                    )
                )
        for predecessor in _ids(task.get("dependsOnTaskIds")):
            edges.add((predecessor, task_id))

    missing_coverage = set(criteria) - covered_criteria
    if missing_coverage:
        diagnostics.append(
            _diagnostic(
                "DELIVERY_AC_TASK_COVERAGE_MISSING",
                "每条 AC 必须由同 Story 的至少一个 Task 覆盖。",
            )
        )
    for integration_id in indices["integrations"]:
        if len(integration_tasks.get(integration_id, [])) != 1:
            diagnostics.append(
                _diagnostic(
                    "DELIVERY_INTEGRATION_TASK_COUNT_INVALID",
                    "每个 Integration 必须恰好由一个 Task 负责。",
                    f"/integrations/{integration_id}",
                )
            )
    if design_required - designed:
        diagnostics.append(
            _diagnostic(
                "DELIVERY_DESIGN_TASK_MISSING",
                "每个 DESIGN_REQUIRED Integration/NFR 必须由 Design Task 落实。",
            )
        )

    declared_edges: set[tuple[str, str]] = set()
    for dependency in _mappings(bundle.get("dependencies")):
        predecessor = dependency.get("predecessorTaskId")
        successor = dependency.get("successorTaskId")
        if not isinstance(predecessor, str) or not isinstance(successor, str):
            continue
        if predecessor not in tasks or successor not in tasks or predecessor == successor:
            diagnostics.append(
                _diagnostic(
                    "DELIVERY_DEPENDENCY_TASK_UNKNOWN",
                    "依赖必须连接两个不同的已知 Task。",
                )
            )
        declared_edges.add((predecessor, successor))
    if edges != declared_edges:
        diagnostics.append(
            _diagnostic(
                "DELIVERY_DEPENDENCY_MISMATCH",
                "Task 前置列表必须与 Dependency 集合完全一致。",
            )
        )
    if _has_cycle(set(tasks), edges | declared_edges):
        diagnostics.append(
            _diagnostic("DELIVERY_DEPENDENCY_CYCLE", "Task 依赖不得形成循环。")
        )
    return diagnostics


def _merge_delivery(
    previous: Mapping[str, object] | None,
    candidate: Mapping[str, object],
    impact: ImpactPlan,
    diagnostics: list[Diagnostic],
) -> dict[str, object]:
    if previous is None or impact.action == "FULL_COMPILE":
        collections = {
            collection: list(_mappings(candidate.get(collection)))
            for collection in COLLECTION_TYPES
        }
    else:
        replaced = set(_ids(candidate.get("replacesFeatureIds")))
        if replaced != set(impact.affected_feature_ids):
            diagnostics.append(
                _diagnostic(
                    "DELIVERY_REPLACEMENT_CLOSURE_MISMATCH",
                    "Delivery 切片替换 Feature 必须与最终 ImpactPlan 完全一致。",
                )
            )
        old_stories = list(_mappings(previous.get("stories")))
        removed_story_ids: set[str] = set()
        remove_story_indexes: list[int] = []
        candidate_story_ids = {
            str(item.get("storyId")) for item in _mappings(candidate.get("stories"))
        }
        for index, story in enumerate(old_stories):
            linked = set(_ids(story.get("featureIds")))
            if linked & replaced and not linked <= replaced:
                diagnostics.append(
                    _diagnostic(
                        "DELIVERY_SHARED_STORY_REQUIRES_WIDENING",
                        "共享 Story 跨越受影响闭包，必须扩大 ImpactPlan。",
                    )
                )
            if linked & replaced or str(story.get("storyId")) in candidate_story_ids:
                remove_story_indexes.append(index)
                removed_story_ids.add(str(story.get("storyId")))
        insertion = min(remove_story_indexes, default=len(old_stories))
        retained_stories = [
            item for index, item in enumerate(old_stories) if index not in remove_story_indexes
        ]
        stories = (
            retained_stories[:insertion]
            + list(_mappings(candidate.get("stories")))
            + retained_stories[insertion:]
        )
        old_criteria = list(_mappings(previous.get("acceptanceCriteria")))
        removed_ac_ids = {
            str(item.get("acceptanceCriterionId"))
            for item in old_criteria
            if item.get("storyId") in removed_story_ids
        }
        old_tasks = list(_mappings(previous.get("tasks")))
        removed_task_ids = {
            str(item.get("taskId"))
            for item in old_tasks
            if item.get("storyId") in removed_story_ids
        }

        def replace_rows(
            collection: str,
            old: list[Mapping[str, object]],
            remove_ids: set[str],
        ) -> list[object]:
            _kind, id_field = COLLECTION_TYPES[collection]
            candidate_rows = list(_mappings(candidate.get(collection)))
            candidate_ids = {str(item.get(id_field)) for item in candidate_rows}
            indexes = [
                index
                for index, item in enumerate(old)
                if str(item.get(id_field)) in remove_ids | candidate_ids
            ]
            insert_at = min(indexes, default=len(old))
            retained = [item for index, item in enumerate(old) if index not in indexes]
            return retained[:insert_at] + candidate_rows + retained[insert_at:]

        criteria_rows = replace_rows("acceptanceCriteria", old_criteria, removed_ac_ids)
        task_rows = replace_rows("tasks", old_tasks, removed_task_ids)
        old_dependencies = list(_mappings(previous.get("dependencies")))
        removed_dependency_ids = {
            str(item.get("dependencyId"))
            for item in old_dependencies
            if item.get("predecessorTaskId") in removed_task_ids
            or item.get("successorTaskId") in removed_task_ids
        }
        dependency_rows = replace_rows(
            "dependencies", old_dependencies, removed_dependency_ids
        )
        collections = {
            "stories": stories,
            "acceptanceCriteria": criteria_rows,
            "tasks": task_rows,
            "dependencies": dependency_rows,
        }
    return {
        "contract": "ai-sow-delivery-bundle-v1",
        "inputRevisionId": candidate.get("inputRevisionId"),
        "scopeSha256": candidate.get("scopeSha256"),
        **collections,
    }


def _metrics(previous: Mapping[str, object] | None, current: Mapping[str, object]) -> dict[str, object]:
    before = _all_objects(previous)
    after = _all_objects(current)
    before_ids = set(before)
    after_ids = set(after)
    preserved = sorted(
        identity
        for identity in before_ids & after_ids
        if canonical_json_bytes(before[identity]) == canonical_json_bytes(after[identity])
    )
    changed = sorted((before_ids & after_ids) - set(preserved))
    created = sorted(after_ids - before_ids)
    removed = sorted(before_ids - after_ids)

    def labels(values: Sequence[tuple[str, str]]) -> tuple[str, ...]:
        return tuple(f"{kind}:{identifier}" for kind, identifier in values)

    return {
        "createdIds": labels(created),
        "preservedIds": labels(preserved),
        "removedIds": labels(removed),
        "changedIds": labels(changed),
        "createdCount": len(created),
        "preservedCount": len(preserved),
        "removedCount": len(removed),
        "changedCount": len(changed),
    }


def compile_delivery(
    scope: Mapping[str, object],
    previous_delivery: Mapping[str, object] | None,
    delivery_slice: Mapping[str, object],
    id_decisions: object,
    impact: ImpactPlan,
    template_catalog: TemplateCatalog,
) -> DeliveryCompilation:
    diagnostics = list(
        validate_contract(delivery_slice, "delivery-slice.schema.json", SCHEMA_REGISTRY)
    )
    diagnostics.extend(validate_contract(scope, "scope-bundle.schema.json", SCHEMA_REGISTRY))
    scope_sha = sha256_bytes(canonical_json_bytes(scope))
    if delivery_slice.get("inputRevisionId") != scope.get("inputRevisionId"):
        diagnostics.append(
            _diagnostic(
                "DELIVERY_INPUT_REVISION_MISMATCH",
                "Delivery 切片未绑定 Scope 使用的 input revision。",
            )
        )
    if delivery_slice.get("scopeSha256") != scope_sha:
        diagnostics.append(
            _diagnostic("DELIVERY_SCOPE_HASH_MISMATCH", "Delivery 切片未绑定当前 Scope。")
        )
    if delivery_slice.get("impactPlanSha256") != impact_plan_sha256(impact):
        diagnostics.append(
            _diagnostic(
                "DELIVERY_IMPACT_HASH_MISMATCH",
                "Delivery 切片未绑定最终 ImpactPlan。",
            )
        )
    diagnostics.extend(
        _validate_id_ledger(delivery_slice, previous_delivery, id_decisions)
    )
    bundle = _merge_delivery(previous_delivery, delivery_slice, impact, diagnostics)
    diagnostics.extend(_validate_delivery(bundle, scope, template_catalog))
    return DeliveryCompilation(
        bundle=bundle,
        bundle_sha256=sha256_bytes(canonical_json_bytes(bundle)),
        metrics=_metrics(previous_delivery, bundle),
        diagnostics=_sort_diagnostics(diagnostics),
    )
